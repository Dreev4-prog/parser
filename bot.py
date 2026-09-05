from __future__ import annotations

import asyncio
import csv
import copy
import json
from collections import Counter
from contextvars import ContextVar
import html
import logging
import os
import re
import shutil
import statistics
import tempfile
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote
from zoneinfo import ZoneInfo

# v4.8.3 Reliable Core: the main Telegram parser must not keep the historical
# process-wide emergency pause after a single 403. Dedicated workers already use
# this profile. Apply it before importing traffic.py/parser.py so local fallback
# scans behave the same way: 403 = no hard sleep, 429 = short bounded pause,
# one penalty level with fast recovery. Aggregate Redis concurrency limits remain.
os.environ["DIST_TRAFFIC_SHARED_COOLDOWN"] = "0"
os.environ["TRAFFIC_MAX_PENALTY_LEVEL"] = "1"
os.environ["TRAFFIC_403_COOLDOWN_SECONDS"] = "0"
os.environ["TRAFFIC_429_COOLDOWN_SECONDS"] = "3"
os.environ["TRAFFIC_MAX_COOLDOWN_SECONDS"] = "3"
os.environ["TRAFFIC_RECOVERY_SUCCESS_COUNT"] = "10"
os.environ["TRAFFIC_RECOVERY_QUIET_SECONDS"] = "10"

# v4.9.1 FOUR-LANE QUEUE GUARANTEE.
# The Telegram parser service is the owner of the four user-facing scan lanes.
# Railway variables must not be able to silently switch this service back to a
# one/two-lane distributed parser profile. Dedicated Date/Page/View entrypoints
# set STABLE_SINGLE_SERVICE_MODE=0 inside their own processes before imports, so
# they keep using Redis exactly as before. Trial and paid scans share these same
# four FIFO lanes.
GUARANTEED_LOCAL_PARSER_LANES = 4
os.environ["STABLE_SINGLE_SERVICE_MODE"] = "1"
os.environ["MULTIUSER_STABLE_MODE"] = "1"
os.environ["MULTIUSER_LOCAL_WORKERS"] = str(GUARANTEED_LOCAL_PARSER_LANES)

from aiogram import BaseMiddleware, Bot, Dispatcher, F
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError, TelegramRetryAfter
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BotCommand, BotCommandScopeChat, BotCommandScopeDefault, CallbackQuery, FSInputFile,
    InlineKeyboardButton, InlineKeyboardMarkup, MenuButtonCommands, Message,
)
from sqlalchemy import delete, func, select, text, update
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app_version import APP_VERSION
from categories import CATEGORIES, GROUPS, categories_for_group, group_root_key
from db import DATABASE_BACKEND, SessionLocal, init_db
from filters import (
    apply_listing_settings,
    base_filter,
    below_market_rows,
    dedupe_rows,
    disappearing_rows,
    frequent_rows,
    price_drop_rows,
    sort_rows,
    unique_rows,
)
from models import (
    AppSetting, BotUser, CategoryScanState, FreeRadarEvent, Listing, ListingIntegrity, ParserRun, PriceHistory, RadarProduct, RadarObservation, RadarSnapshot, ScanListing,
    ScanObservation, ScanViewHistory, SelectedCategory, StableCategoryJob, StablePageCheckpoint,
    SubscriptionPayment, SubscriptionPlan, UserScan, UserSettings, ViewHistory,
    VintedScan, VintedScanCategory, VintedScanItem,
)
from product_identity import TYPE_DISPLAY, ProductIdentity, recognize_product
from organic_velocity import (
    ORGANIC_HIGH_BASELINE_VIEWS, ORGANIC_HIGH_BATCH_SIZE,
    ORGANIC_HIGH_CHECKPOINT_MINUTES, ORGANIC_HIGH_POLL_SECONDS,
    ORGANIC_HIGH_REQUIRED_CHECKPOINTS, apply_organic_measurement,
)
from traffic import TRAFFIC
from i18n import LANG_EN, LANG_RU, get_user_language, language_name, set_user_language, translate_text
from distributed import (
    COORDINATOR, DISTRIBUTED_WORKERS, DISTRIBUTED_MODE_SOURCE,
    DISTRIBUTED_CONFIG_ERROR, IS_RAILWAY, REDIS_URL, STABLE_SINGLE_SERVICE_MODE,
)
from scan_control import ScanStopRequested, wait_for_task_or_stop
from scan_selection import MAX_SELECTED_CATEGORIES, bulk_group_selection, toggle_selection, validate_scan_category_keys
from commerce import (
    PAYMENT_POLL_SECONDS, PaymentProviderError, admin_stats, cached_access_until,
    create_subscription_payment, current_access_mode, find_users, get_payment,
    get_plan, get_plans, get_user as get_commerce_user, grant_access_days,
    has_access, initialize_commerce, is_banned_cached, pending_payments,
    provider_enabled, providers_status, recent_payments, recent_users,
    referral_admin_stats, referral_promo_enabled, referral_user_stats,
    refresh_payment, revoke_access, set_access_mode, set_banned, toggle_plan,
    set_referral_promo_enabled,
    touch_user, update_plan_price, set_onboarding_completed, user_payments,
    subscription_notice_candidates, mark_subscription_notice,
)
from parser import (
    MAX_PAGES_PER_CATEGORY,
    PAGE_DELAY_SECONDS,
    STOP_AFTER_EMPTY_TODAY_PAGES,
    SCAN_TRANSPORT,
    KleinanzeigenParser,
    ParsedListing,
    ViewCountResult,
    TemporaryAccessError,
    is_today_text,
    posted_date_moscow,
    profile_page_dates,
    page_url,
    private_provider_url,
)
from view_manager import REMOTE_VIEW_MANAGER, REMOTE_VIEW_WORKER_ENABLED
from radar import (
    RADAR_PAGE_SIZE, RADAR_SCAN_TOP_LIMIT, bump_resurrection_integrity_sweep_once,
    prepare_bump_resurrection_sweep_once, prepare_verified_organic_velocity_once, prepare_unified_48h_ranking_once, get_fast_sold_info, get_fast_sold_infos,
    get_radar_product, is_radar_favorite, list_radar_products, radar_categories, radar_stats,
    purge_nonorganic_analytics, record_autoscan_hot_detailed, record_user_scan_radar3_baselines, radar_v3_category_allowed,
    record_verified_velocity_signals, refresh_radar_scores, verify_listing_organic_now,
    prepare_radar_v3_once, repair_radar_v3_historical_scores_once, repair_radar_v3_live_retention_once, radar_v3_due_external_ids, radar_v3_claim_due_external_ids, radar_v3_release_claims, radar_v3_record_refreshed, radar_v3_expire_observations, radar_v3_expire_stale_products, radar_v3_rollover_successful_category,
    search_radar_products, toggle_radar_favorite,
)
from page_manager import (
    PAGE_PREFETCH_EXTRA_PAGES, PAGE_PREFETCH_LOW_WATER_PAGES, PAGE_PREFETCH_WINDOW_PAGES,
    REMOTE_PAGE_MANAGER, REMOTE_PAGE_WORKER_ENABLED, rolling_prefetch_range,
)
from date_manager import (
    DATE_MAX_AGE_DAYS, REMOTE_DATE_MANAGER, REMOTE_DATE_WORKER_ENABLED,
)
from stable_engine import (
    load_date_index, load_page_checkpoint, mark_category_job, record_page_failure,
    save_date_index, save_page_checkpoint,
)
from vinted_lab import (
    VINTED_QUEUE, cancel_scan as cancel_vinted_scan, create_scan as create_vinted_scan,
    enqueue_scan as enqueue_vinted_scan, fetch_catalog_tree as fetch_vinted_catalog_tree,
    flatten_catalog_tree as flatten_vinted_catalog_tree, get_scan as get_vinted_scan,
    list_scans as list_vinted_scans, list_scan_items as list_vinted_scan_items,
    scan_progress_snapshot as vinted_scan_progress, catalog_like_delta as vinted_catalog_like_delta,
)
from vinted_session_store import (
    clear_vinted_session, create_session_ticket, get_session_service,
    load_vinted_session_json, load_vinted_session_meta,
)
from vinted_radar import (
    VINTED_RADAR_HISTORY_DAYS, VINTED_RADAR_INTERVAL_MINUTES, VINTED_RADAR_LIVE_HOURS,
    build_radar_snapshot as build_vinted_radar_snapshot,
    disable_radar as disable_vinted_radar,
    enable_radar as enable_vinted_radar,
    get_radar_entry as get_vinted_radar_entry,
    maybe_start_due_round as maybe_start_vinted_radar_round,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("kleinanzeigen-bot")

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
ADMIN_IDS = {int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()}

# v4.9.0 Free Trial Launch. The promotion is intentionally a product/access
# layer only: parser workers and scan algorithms remain untouched.
FREE_TRIAL_SETTING_KEY = "free_trial_enabled"
FREE_TRIAL_ENABLED_DEFAULT = os.getenv("FREE_TRIAL_ENABLED", "1").strip().lower() in {"1", "true", "yes", "on"}
FREE_TRIAL_SCAN_LIMIT = max(0, min(10, int(os.getenv("FREE_TRIAL_SCAN_LIMIT", "2"))))
FREE_TRIAL_MAX_CATEGORIES = 1
FREE_TRIAL_MAX_PAGES = 25
# v4.11.6 Free Radar Preview: non-subscribers can inspect a small live sample
# without gaining Search/Categories/My Radar or pagination access.
FREE_RADAR_PREVIEW_LIMIT = 5
_trial_credit_guard = asyncio.Lock()

# v4.11.5 DT Radar AutoScan Stability. AutoScan is a low-priority product-feed producer:
# it scans only product-oriented leaf categories, reuses one parser/browser session across
# the whole round, yields to foreground users between categories, and applies bounded
# backoff after partial/system failures. Non-product/service categories remain available
# to the normal parser; they are excluded only from the automatic Radar seeding round.
RADAR_AUTOSCAN_SETTING_KEY = "dt_radar_autoscan_v1"
RADAR_AUTOSCAN_POLICY_VERSION = 7
RADAR_AUTOSCAN_DEPTH = 20
# Radar 3.0 observes live demand only on today's market. Yesterday context was retired in v4.21.5.
RADAR_CONTEXT_DEPTH = 0
RADAR_CONTEXT_ENABLED = False


def _radar_layer_depth(state: dict | None = None) -> int:
    """Return the configured page depth for the active Radar layer."""
    layer = str((state or {}).get("layer") or "fresh")
    return RADAR_AUTOSCAN_DEPTH


RADAR_AUTOSCAN_USER_ID = -411000001
RADAR_AUTOSCAN_DEFAULT_TIME = "05:00"
RADAR_AUTOSCAN_POLL_SECONDS = 10
RADAR_AUTOSCAN_HISTORY_LIMIT = 20
RADAR_AUTOSCAN_TIME_CHOICES = ("03:00", "05:00", "08:00", "12:00", "18:00", "23:00")
# v4.21.14: these watchdog settings were referenced by runtime code but had no
# module-level definitions in 4.21.12/4.21.14, causing a startup NameError.
# Keep Railway overrides optional and clamp obviously unsafe values.
def _radar_env_seconds(name: str, default: int, minimum: int) -> int:
    try:
        return max(int(minimum), int(str(os.getenv(name, str(default))).strip()))
    except (TypeError, ValueError):
        log.warning("Invalid %s; using default=%ss", name, default)
        return int(default)

RADAR_AUTOSCAN_CATEGORY_TIMEOUT_SECONDS = _radar_env_seconds(
    "RADAR_AUTOSCAN_CATEGORY_TIMEOUT_SECONDS", 480, 120
)
RADAR_AUTOSCAN_VIEW_RECOVERY_TIMEOUT_SECONDS = _radar_env_seconds(
    "RADAR_AUTOSCAN_VIEW_RECOVERY_TIMEOUT_SECONDS", 240, 60
)
RADAR_AUTOSCAN_PARTIAL_BACKOFF_BASE_SECONDS = 3.0
RADAR_AUTOSCAN_SYSTEM_BACKOFF_BASE_SECONDS = 8.0
RADAR_AUTOSCAN_MAX_BACKOFF_SECONDS = 30.0
RADAR_AUTOSCAN_SUCCESS_GAP_SECONDS = 0.25
RADAR_AUTOSCAN_SAFE_VIEW_CONCURRENCY = 4
# v4.22.6 Fast Today + Exact Tail. A tiny unresolved exact-view tail no longer
# forces a full 20-page category rescan. Unknown counters stay NULL and therefore
# cannot seed Radar; only the verified majority creates baselines. A materially
# incomplete batch remains fail-closed and still goes to the normal retry list.
RADAR_AUTOSCAN_MIN_VIEW_COVERAGE_PCT = 99.0
RADAR_AUTOSCAN_VIEW_SOFT_TAIL_MIN = 3
RADAR_AUTOSCAN_VIEW_SOFT_TAIL_MAX = 8
RADAR_AUTOSCAN_VIEW_TAIL_RETRY_MAX = 12
RADAR_AUTOSCAN_LAUNCH_WATCHDOG_SECONDS = 20


def _radar_autoscan_view_tail_budget(requested: int) -> int:
    requested = max(0, int(requested or 0))
    if requested <= 0:
        return 0
    one_percent = max(1, (requested + 99) // 100)
    return min(
        RADAR_AUTOSCAN_VIEW_SOFT_TAIL_MAX,
        max(RADAR_AUTOSCAN_VIEW_SOFT_TAIL_MIN, one_percent),
    )


def _radar_autoscan_views_usable(requested: int, verified: int) -> bool:
    requested = max(0, int(requested or 0))
    verified = max(0, min(requested, int(verified or 0))) if requested else 0
    if requested <= 0 or verified >= requested:
        return True
    missing = requested - verified
    coverage_pct = (verified / requested) * 100.0
    return (
        coverage_pct >= RADAR_AUTOSCAN_MIN_VIEW_COVERAGE_PCT
        and missing <= _radar_autoscan_view_tail_budget(requested)
    )
# v4.12.0 Daily Radar Growth Loop. One factual daily digest turns the live Radar
# database into a recurring acquisition/retention surface. The digest is enabled by
# default, sends at 20:00 Moscow time, and stores its last send date in AppSetting so
# Railway restarts cannot duplicate the same day's campaign.
RADAR_DAILY_DIGEST_SETTING_KEY = "dt_radar_daily_digest_v1"
RADAR_DAILY_DIGEST_DEFAULT_TIME = "20:00"
RADAR_DAILY_DIGEST_TIME_CHOICES = ("12:00", "18:00", "20:00", "22:00")
RADAR_DAILY_DIGEST_TIME_RE = re.compile(r"^(?:[01]?\d|2[0-3]):[0-5]\d$")
RADAR_DAILY_DIGEST_POLL_SECONDS = 30
# v4.12.2: admin controls must never look dead while Radar tables are busy. UI reads
# use short bounded waits and can fall back to the last successfully cached metrics.
RADAR_DAILY_DIGEST_STATE_TIMEOUT_SECONDS = 4.0
RADAR_DAILY_DIGEST_UI_METRICS_TIMEOUT_SECONDS = 8.0
RADAR_DAILY_DIGEST_SEND_METRICS_TIMEOUT_SECONDS = 20.0
RADAR_DAILY_DIGEST_RECIPIENTS_TIMEOUT_SECONDS = 8.0
_RADAR_DAILY_DIGEST_SEND_LOCK = asyncio.Lock()

_radar_autoscan_guard = asyncio.Lock()
# Serializes the actual round runner. The scheduler and a manual admin kick may race,
# but only one parser round is ever allowed to execute at a time.
_radar_autoscan_run_guard = asyncio.Lock()
_radar_autoscan_wakeup = asyncio.Event()
_radar_autoscan_kick_task: asyncio.Task | None = None
# Process-local interrupt used for a real hard stop. Persistent state remains the
# source of truth across Railway restarts, while this event aborts the in-flight
# category/cooldown immediately in the current process.
_radar_autoscan_stop_event = asyncio.Event()


class RadarAutoScanStopped(Exception):
    pass


class RadarAutoScanCategoryTimeout(Exception):
    pass


# v4.6.5: per-user RU/EN presentation layer.  The parser/business logic remains
# language-neutral; all outgoing Telegram text/inline-button labels are localized
# at the Bot API boundary.  Admin chats intentionally stay Russian.
_UI_LANGUAGE: ContextVar[str | None] = ContextVar("dt_ui_language", default=None)


def _copy_model_with_updates(obj, **updates):
    if not updates:
        return obj
    copier = getattr(obj, "model_copy", None)
    if callable(copier):
        return copier(update=updates)
    copier = getattr(obj, "copy", None)
    if callable(copier):
        try:
            return copier(update=updates)
        except TypeError:
            pass
    for key, value in updates.items():
        try:
            setattr(obj, key, value)
        except Exception:
            pass
    return obj


def _localize_inline_markup(markup, language: str | None):
    if language != LANG_EN or not isinstance(markup, InlineKeyboardMarkup):
        return markup
    changed = False
    rows = []
    for row in markup.inline_keyboard:
        new_row = []
        for button in row:
            text = getattr(button, "text", None)
            translated = translate_text(text, language) if isinstance(text, str) else text
            if translated != text:
                changed = True
                button = _copy_model_with_updates(button, text=translated)
            new_row.append(button)
        rows.append(new_row)
    return InlineKeyboardMarkup(inline_keyboard=rows) if changed else markup


def _localize_telegram_method(method, language: str | None):
    if language != LANG_EN:
        return method
    updates = {}
    text = getattr(method, "text", None)
    if isinstance(text, str):
        translated = translate_text(text, language)
        if translated != text:
            updates["text"] = translated
    caption = getattr(method, "caption", None)
    if isinstance(caption, str):
        translated = translate_text(caption, language)
        if translated != caption:
            updates["caption"] = translated
    markup = getattr(method, "reply_markup", None)
    translated_markup = _localize_inline_markup(markup, language)
    if translated_markup is not markup:
        updates["reply_markup"] = translated_markup
    return _copy_model_with_updates(method, **updates)


class LocalizedBot(Bot):
    """Bot wrapper that localizes both foreground and background sends."""

    async def __call__(self, method, request_timeout=None):
        language = _UI_LANGUAGE.get()
        chat_id = getattr(method, "chat_id", None)
        numeric_chat_id = None
        try:
            numeric_chat_id = int(chat_id) if chat_id is not None else None
        except (TypeError, ValueError):
            numeric_chat_id = None

        # v4.6.6: admin accounts are normal users outside the admin surface.
        # Do not force their entire private chat to Russian here. The language
        # middleware below forces Russian only while an admin is actually inside
        # /admin, admin callbacks or AdminInput FSM screens.
        if language is None and numeric_chat_id is not None and numeric_chat_id > 0:
            try:
                language = await get_user_language(numeric_chat_id)
            except Exception:
                log.debug("Could not resolve UI language for chat=%s", numeric_chat_id, exc_info=True)

        method = _localize_telegram_method(method, language)
        return await super().__call__(method, request_timeout=request_timeout)
_PROJECT_DIR = Path(__file__).resolve().parent
MENU_IMAGE_PATH = _PROJECT_DIR / "dt_parser_menu.png"
if not MENU_IMAGE_PATH.exists():
    MENU_IMAGE_PATH = _PROJECT_DIR / "assets" / "dt_parser_menu.png"
# v4.6.7: after the first upload Telegram gives us a reusable file_id. Reusing
# it avoids re-uploading the menu image every time the user returns home.
_MENU_IMAGE_FILE_ID: str | None = None

BERLIN = ZoneInfo("Europe/Berlin")
MOSCOW = ZoneInfo("Europe/Moscow")
AVAILABILITY_CHECK_LIMIT = max(1, int(os.getenv("AVAILABILITY_CHECK_LIMIT", "150")))
AVAILABILITY_CONCURRENCY = max(1, min(8, int(os.getenv("AVAILABILITY_CONCURRENCY", "4"))))

# v3.8 Stable Scan Engine. Category/date work is shared across users and
# verified page/date boundaries are checkpointed in PostgreSQL, so a worker
# restart or recovery pass never has to rediscover healthy pages from scratch.
STABLE_SCAN_ENGINE = os.getenv("STABLE_SCAN_ENGINE", "1").strip().lower() not in {"0", "false", "no", "off"}
STABLE_PAGE_RETRIES = max(1, min(5, int(os.getenv("STABLE_PAGE_RETRIES", "3"))))
STABLE_PAGE_RETRY_SECONDS = max(0.2, min(10.0, float(os.getenv("STABLE_PAGE_RETRY_SECONDS", "1.2"))))
# v4.10.1: one persistently repeated nationwide page must not abort an otherwise
# healthy category. After the normal retries + one clean BrowserContext recycle,
# skip only repeated-content defects and replace the missing verified depth with
# later nationwide pages (or, at the public cap, one verified regional page).
# Keep the allowance bounded so a genuinely stuck/looping feed still fails partial.
DIRECT_REPEATED_RECOVERY_LIMIT = max(1, min(5, int(os.getenv("DIRECT_REPEATED_RECOVERY_LIMIT", "3"))))
# v4.3.x Multi-User Stable keeps the proven single Railway service / Stable Engine
# path, but allows a small local worker pool. All jobs share ONE Chromium process
# while every KleinanzeigenParser owns an isolated BrowserContext.
# v4.9.1: the main bot is intentionally pinned to the production-safe four-lane
# profile above. Keep these as explicit constants too, so a future refactor cannot
# accidentally re-read a stale Railway value after startup.
MULTIUSER_STABLE_MODE = True
MULTIUSER_LOCAL_WORKERS = GUARANTEED_LOCAL_PARSER_LANES
# v4.3.3: process-wide foreground public-view lane. v4.3.0 still inherited the
# old global traffic limit of three view requests, so three simultaneous scans
# were forced to share only three slots. Six keeps two fast official-counter
# requests available per default scan worker while the traffic manager still
# reserves separate capacity for category-page work and serializes Chromium fallback.
MULTIUSER_VIEW_POOL_SIZE = max(2, min(12, int(os.getenv("MULTIUSER_VIEW_POOL_SIZE", "9"))))
# v4.3.3: the lightweight official counter can start faster than the old
# 0.20s process-wide cadence. Chromium has its own slower browser limiter.
MULTIUSER_VIEW_MIN_INTERVAL_SECONDS = max(0.05, min(0.50, float(
    os.getenv("MULTIUSER_VIEW_MIN_INTERVAL_SECONDS", "0.05")
)))
SCAN_CATEGORY_HARD_TIMEOUT_SECONDS = max(300.0, min(3600.0, float(
    os.getenv("SCAN_CATEGORY_HARD_TIMEOUT_SECONDS", "1200")
)))
if STABLE_SINGLE_SERVICE_MODE:
    import parser as _stable_parser_module
    _stable_parser_module.SCAN_TRANSPORT = "browser"
    _stable_parser_module.SHARED_BROWSER_RUNTIME = bool(MULTIUSER_STABLE_MODE)
    SCAN_TRANSPORT = "browser"
# v4.0.1: tolerate a bounded number of pages with zero usable chronology evidence.
# They are recorded for diagnostics/recovery but no longer force a whole recent
# nationwide scan into the 16-region hidden fallback.
STABLE_WEAK_PAGE_GAP_LIMIT = max(1, min(8, int(os.getenv("STABLE_WEAK_PAGE_GAP_LIMIT", "3"))))
# v4.2.5: every completed primary scan gets a fresh baseline view measurement.
# The work is deferred until AFTER category/date page collection, so it does not
# slow the date locator page-by-page. Ignore the legacy Railway flag entirely:
# older deployments often still have PRIMARY_SCAN_INLINE_VIEWS=0 persisted, which
# would otherwise silently re-enable the exact v4.2.4 regression we are fixing.
PRIMARY_SCAN_INLINE_VIEWS = True
os.environ["PRIMARY_SCAN_INLINE_VIEWS"] = "1"

# v2.6 Multi-User Core. User launches go into a queue. Only a limited number
# of jobs are processed at once, while category scans are shared globally.
MAX_CONCURRENT_JOBS = max(1, min(12, int(os.getenv("MAX_CONCURRENT_JOBS", "5"))))
if STABLE_SINGLE_SERVICE_MODE:
    # v4.9.1 hard guarantee: exactly four local user-facing parser consumers.
    # The fifth and later launch stays in scan_queue FIFO until one consumer is free.
    # Trial and paid scans use this same queue; there is no separate trial cap.
    MAX_CONCURRENT_JOBS = GUARANTEED_LOCAL_PARSER_LANES
MAX_QUEUE_SIZE = max(10, int(os.getenv("MAX_QUEUE_SIZE", "200")))
QUEUE_START_NOTIFY_AFTER_SECONDS = max(0, min(300, int(os.getenv("QUEUE_START_NOTIFY_AFTER_SECONDS", "8"))))
PARSER_WORKER_CONCURRENCY = max(1, min(8, int(os.getenv("PARSER_WORKER_CONCURRENCY", "1"))))
if STABLE_SINGLE_SERVICE_MODE and MULTIUSER_STABLE_MODE:
    # TRAFFIC is process-wide and acts as the global Views Pool. Keep request
    # pressure bounded, but let three foreground scans share six official-counter
    # HTTP slots instead of the old process-wide limit of three. The global cap
    # deliberately includes TRAFFIC.reserved_scan_slots, so category/date traffic
    # keeps its own capacity even while all three users are measuring views.
    TRAFFIC.base_scan_limit = max(TRAFFIC.base_scan_limit, MAX_CONCURRENT_JOBS)
    TRAFFIC.base_view_limit = MULTIUSER_VIEW_POOL_SIZE
    TRAFFIC.view_min_interval = MULTIUSER_VIEW_MIN_INTERVAL_SECONDS
    TRAFFIC.base_global_limit = max(
        TRAFFIC.base_global_limit,
        MULTIUSER_VIEW_POOL_SIZE + TRAFFIC.reserved_scan_slots,
    )
    TRAFFIC.background_during_scans = 0
# v4.1.7: always bootstrap one in-process browser reserve when Redis is configured.
# A previous deployment can leave a heartbeat alive for ~20 seconds; treating that
# stale heartbeat as an external fleet caused the new deployment to skip its own
# worker and then drop to zero workers after the stale key expired. Dedicated fleet
# replicas can still be added later; Redis job locks keep execution safe.
EMBEDDED_FLEET_FALLBACK = os.getenv("EMBEDDED_FLEET_FALLBACK", "1").strip().lower() not in {
    "0", "false", "no", "off",
}
EMBEDDED_FLEET_READY_WAIT_SECONDS = max(0, min(15, int(os.getenv("EMBEDDED_FLEET_READY_WAIT_SECONDS", "3"))))
DISTRIBUTED_STALE_QUEUE_SECONDS = max(60, min(3600, int(os.getenv("DISTRIBUTED_STALE_QUEUE_SECONDS", "600"))))
DISTRIBUTED_QUEUE_UI_SECONDS = max(2.0, min(15.0, float(os.getenv("DISTRIBUTED_QUEUE_UI_SECONDS", "3"))))
CATEGORY_CACHE_TTL_SECONDS = max(0, int(os.getenv("CATEGORY_CACHE_TTL_SECONDS", "300")))
if STABLE_SINGLE_SERVICE_MODE:
    # Keep the baseline deterministic: each queued user gets the parser path.
    # PostgreSQL verified-page checkpoints still protect restarts/retries.
    CATEGORY_CACHE_TTL_SECONDS = 0
SHARE_ACTIVE_CATEGORY_SCANS = os.getenv(
    "SHARE_ACTIVE_CATEGORY_SCANS",
    "1" if STABLE_SCAN_ENGINE else ("0" if SCAN_TRANSPORT in {"browser", "hybrid"} else "1")
).strip().lower() not in {"0", "false", "no", "off"}
if STABLE_SINGLE_SERVICE_MODE and MULTIUSER_STABLE_MODE:
    # Each user's job must keep its own BrowserContext. A slow date lookup for one
    # user must never make another user await the same in-flight category task.
    SHARE_ACTIVE_CATEGORY_SCANS = False
JOB_PARSER: ContextVar[KleinanzeigenParser | None] = ContextVar("dtparser_job_parser", default=None)
STATUS_UPDATE_INTERVAL_SECONDS = max(0.5, float(os.getenv("STATUS_UPDATE_INTERVAL_SECONDS", "1.5")))
# v3.3.0 job-level resilience on top of parser HTTP retries. A category that
# throws an unexpected transient exception gets one controlled retry before the
# user receives a partial result.
SCAN_CATEGORY_ATTEMPTS = max(1, min(3, int(os.getenv("SCAN_CATEGORY_ATTEMPTS", "2"))))
if STABLE_SINGLE_SERVICE_MODE:
    SCAN_CATEGORY_ATTEMPTS = 1
SCAN_CATEGORY_RETRY_SECONDS = max(1.0, min(30.0, float(os.getenv("SCAN_CATEGORY_RETRY_SECONDS", "4"))))
# v3.7.2: a parser result that is structurally partial is automatically retried
# inside the same per-user parser session. Verified page checkpoints are reused,
# so recovery refetches weak/missing areas instead of making the user press a button.
SCAN_AUTO_RECOVERY_PASSES = max(0, min(3, int(os.getenv("SCAN_AUTO_RECOVERY_PASSES", "3"))))
if STABLE_SINGLE_SERVICE_MODE:
    # Page-level retry + one clean browser-context recycle replaces repeated whole
    # category passes. This makes failures deterministic and keeps scans short.
    SCAN_AUTO_RECOVERY_PASSES = 0
SCAN_AUTO_RECOVERY_DELAY_SECONDS = max(0.5, min(15.0, float(os.getenv("SCAN_AUTO_RECOVERY_DELAY_SECONDS", "2"))))
SUBSCRIPTION_NOTICE_POLL_SECONDS = max(60, int(os.getenv("SUBSCRIPTION_NOTICE_POLL_SECONDS", "300")))

# Public view counts are collected inline while category pages are scanned.
# Recent values are cached so shared/multi-user scans do not reopen the same ad.
VIEW_COUNT_CACHE_TTL_SECONDS = max(60, int(os.getenv("VIEW_COUNT_CACHE_TTL_SECONDS", "1800")))
VIEW_COUNT_CONCURRENCY = max(1, min(10, int(os.getenv("VIEW_COUNT_CONCURRENCY", "3"))))
# v4.2.2 Accurate Views Core. Ranking/filter measurements never reuse the old
# direct-only cache: each measurement must produce a freshly verified public
# counter or remain unknown.
ACCURATE_VIEWS_MODE = os.getenv("ACCURATE_VIEWS_MODE", "1").strip().lower() not in {"0", "false", "no", "off"}
VIEW_ANALYTICS_CORE_VERSION = "4.2.2"
# A control measurement must be fresh. This tiny window is only for coalescing
# truly simultaneous checks of the same IDs across users/scans; it is not a
# normal cache and cannot replace a 3/6/12h measurement.
VIEW_MEASUREMENT_REUSE_SECONDS = max(0, min(60, int(os.getenv("VIEW_MEASUREMENT_REUSE_SECONDS", "20"))))
VIEW_COUNT_EXPORT_MODES = {"newest", "all", "unique", "below_market"}

# v3.1 keeps the v3.0.7 Popularity Tracker. Every completed scan gets automatic public-view
# checkpoints. They are persisted, so a Railway restart does not lose the plan.
OBSERVATION_HOURS = (3, 6, 12)
OBSERVATION_SCHEDULE_HOURS = (0, 3, 6, 12) if not PRIMARY_SCAN_INLINE_VIEWS else OBSERVATION_HOURS
OBSERVATION_POLL_SECONDS = max(15, int(os.getenv("OBSERVATION_POLL_SECONDS", "30")))
OBSERVATION_CONCURRENCY = max(1, min(4, int(os.getenv("OBSERVATION_CONCURRENCY", "1"))))
OBSERVATION_LATE_GRACE_MINUTES = max(5, int(os.getenv("OBSERVATION_LATE_GRACE_MINUTES", "45")))

# v3.2.8 My Scans inbox. Completed scans stay in the main list for 24 hours,
# then only their UI card is archived. Underlying scan/listing/history data remains
# intact for Popular Now, exports and future analytics.
SCAN_ARCHIVE_AFTER_HOURS = 24
SCAN_ARCHIVE_PAGE_SIZE = 8
SCAN_ARCHIVE_SWEEP_SECONDS = 15 * 60
ARCHIVABLE_SCAN_STATUSES = ("done", "partial", "cancelled", "failed")

GROWTH_TOP_LIMIT = 50
GROWTH_TELEGRAM_LIMIT = 10

# v2.5 incremental scan tuning (kept in v2.6). A full scan is forced once per category per Berlin day.
# Later runs stop after the parser crosses the previous head checkpoint and then
# sees a small safety overlap of already-known pages.
INCREMENTAL_STOP_AFTER_KNOWN_PAGES = max(1, int(os.getenv("INCREMENTAL_STOP_AFTER_KNOWN_PAGES", "2")))
INCREMENTAL_MIN_KNOWN_RATIO = min(1.0, max(0.5, float(os.getenv("INCREMENTAL_MIN_KNOWN_RATIO", "0.80"))))
INCREMENTAL_MIN_PAGES = max(1, int(os.getenv("INCREMENTAL_MIN_PAGES", "2")))
INCREMENTAL_HEAD_SIZE = max(3, min(20, int(os.getenv("INCREMENTAL_HEAD_SIZE", "8"))))
INCREMENTAL_OVERLAP_PAGES = max(0, int(os.getenv("INCREMENTAL_OVERLAP_PAGES", "1")))

MODE_LABELS = {
    "newest": "🆕 Самые новые",
    "all": "📚 Все",
    "unique": "💎 Уникальные",
    "frequent": "🔥 Часто публикуемые",
    "below_market": "💰 Ниже рынка",
    "fast_disappearing": "⚡ Быстро исчезающие",
    "price_drop": "📉 Снижение цены",
}
PRICE_LABELS = {
    "any": "Любая цена",
    "50_plus": "50+ €",
    "100_plus": "100+ €",
    "200_plus": "200+ €",
    "500_plus": "500+ €",
    # Legacy values are kept so old saved settings/scans remain readable.
    "0_50": "0–50 €",
    "50_100": "50–100 €",
    "100_200": "100–200 €",
    "200_500": "200–500 €",
}


def price_filter_label(value: str | None) -> str:
    value = (value or "any").strip()
    if value in PRICE_LABELS:
        return PRICE_LABELS[value]
    if value.startswith("custom:"):
        parts = value.split(":", 2)
        if len(parts) == 3:
            lo = parts[1].strip()
            hi = parts[2].strip()
            if lo and hi:
                return f"{lo}–{hi} €"
            if lo:
                return f"{lo}+ €"
            if hi:
                return f"до {hi} €"
    return value or "Любая цена"


def parse_scan_price_input(text: str | None) -> str | None:
    raw = (text or "").strip().lower().replace("€", "").replace("eur", "")
    raw = re.sub(r"\s+", "", raw)
    if not raw:
        return None
    m = re.fullmatch(r"(\d{1,7})[-–—](\d{1,7})", raw)
    if m:
        lo, hi = int(m.group(1)), int(m.group(2))
        if lo <= hi:
            return f"custom:{lo}:{hi}"
        return None
    m = re.fullmatch(r"(\d{1,7})\+", raw)
    if m:
        return f"custom:{int(m.group(1))}:"
    m = re.fullmatch(r"[-–—](\d{1,7})", raw)
    if m:
        return f"custom::{int(m.group(1))}"
    m = re.fullmatch(r"до(\d{1,7})", raw)
    if m:
        return f"custom::{int(m.group(1))}"
    m = re.fullmatch(r"от(\d{1,7})", raw)
    if m:
        return f"custom:{int(m.group(1))}:"
    return None
SORT_LABELS = {"newest": "Сначала новые", "price_asc": "Цена ↑", "price_desc": "Цена ↓"}
PAGE_LIMIT_CHOICES = (15, 25, 50)
# Conservative baseline for user-facing ETA. A full 50-page category starts
# around 2 minutes, then the estimate is recalculated from the real page rate.
PAGE_LIMIT_BASE_ETA_SECONDS = {15: 45, 25: 60, 50: 120}

# Kleinanzeigen's nationwide public search feed exposes a bounded pagination
# window. v4.3.31 changes 15/25/50 to a MAXIMUM nationwide target-date depth:
# once the selected day ends, or the public nationwide window ends, the scan is
# complete. Regional hidden-fill is disabled by default because it turns one old
# date scan into many independent regional date searches and was the largest
# remaining source of latency. It can still be re-enabled explicitly for testing.
PUBLIC_SEARCH_PAGE_CAP = max(10, min(50, int(os.getenv("PUBLIC_SEARCH_PAGE_CAP", "50"))))
REGIONAL_HIDDEN_FILL_ENABLED = os.getenv("REGIONAL_HIDDEN_FILL_ENABLED", "0").strip().lower() not in {
    "0", "false", "no", "off",
}
# v4.3.32 smart hybrid: keep ordinary nationwide scans region-free, but if the
# requested date itself is deeper than Kleinanzeigen's public nationwide window,
# regional sharding is required to avoid a false zero. This fallback is enabled
# by default and only activates for a verified `too_deep` date result.
AUTO_REGIONAL_FALLBACK_TOO_DEEP = os.getenv("AUTO_REGIONAL_FALLBACK_TOO_DEEP", "1").strip().lower() not in {
    "0", "false", "no", "off",
}
DATE_JUMP_PROBE_DELAY_SECONDS = max(0.0, min(1.0, float(os.getenv("DATE_JUMP_PROBE_DELAY_SECONDS", "0.18"))))

# Hidden implementation detail: these location shards cover Germany without
# intentionally overlapping. They are not shown to end users; the UI remains
# category + date + 15/25/50 pages. Smaller feeds are tried first because older
# dates are more likely to remain inside the public 50-page window.
GERMAN_STATE_SEGMENTS = (
    ("Bremen", "bremen", 1),
    ("Saarland", "saarland", 285),
    ("Hamburg", "hamburg", 9409),
    ("Mecklenburg-Vorpommern", "mecklenburg-vorpommern", 61),
    ("Thüringen", "thueringen", 3548),
    ("Sachsen-Anhalt", "sachsen-anhalt", 2165),
    ("Brandenburg", "brandenburg", 7711),
    ("Schleswig-Holstein", "schleswig-holstein", 408),
    ("Berlin", "berlin", 3331),
    ("Rheinland-Pfalz", "rheinland-pfalz", 4938),
    ("Sachsen", "sachsen", 3799),
    ("Hessen", "hessen", 4279),
    ("Niedersachsen", "niedersachsen", 2428),
    ("Baden-Württemberg", "baden-wuerttemberg", 7970),
    ("Bayern", "bayern", 5510),
    ("Nordrhein-Westfalen", "nordrhein-westfalen", 928),
)

# v4.3.28: when a historical nationwide scan is likely to need regional depth,
# pre-warm several independent regional date locators while Page Worker is still
# collecting the nationwide pages. This removes the visible second multi-minute
# date-search staircase without letting remote hints become source of truth.
HIDDEN_DATE_PREWARM_ENABLED = os.getenv("HIDDEN_DATE_PREWARM_ENABLED", "0").strip().lower() not in {"0", "false", "no", "off"}
HIDDEN_DATE_PREWARM_WINDOW = max(1, min(4, int(os.getenv("HIDDEN_DATE_PREWARM_WINDOW", "2"))))
HIDDEN_DATE_PREWARM_CONCURRENCY = max(1, min(2, int(os.getenv("HIDDEN_DATE_PREWARM_CONCURRENCY", "1"))))

# v4.3.36 Four-User Worker Fleet profile. Old-date scans still use the trusted
# foreground verifier, while a four-replica Date Worker fleet can keep four
# independent regional locators in flight. Capacity is added through Railway
# replicas rather than higher per-worker request concurrency. Remote hints remain
# hints only; every accepted page still goes through exact verification/fallback.
REGIONAL_DATE_PIPELINE_ENABLED = os.getenv("REGIONAL_DATE_PIPELINE_ENABLED", "1").strip().lower() not in {
    "0", "false", "no", "off",
}
REGIONAL_DATE_PIPELINE_WINDOW = max(1, min(10, int(os.getenv("REGIONAL_DATE_PIPELINE_WINDOW", "8"))))
# v4.3.37 safety net: even a stale/wide remote hint must never make the
# foreground verifier walk tens of target pages backwards one-by-one.  The
# Date Manager now refines wide direct-target brackets first; if a pathological
# hint still survives, stop the linear walk-back after a few exact pages and
# fall through to the proven local exponential/binary locator.
REMOTE_DATE_MAX_LINEAR_WALKBACK = max(2, min(12, int(os.getenv("REMOTE_DATE_MAX_LINEAR_WALKBACK", "6"))))
REGIONAL_DATE_PIPELINE_CONCURRENCY = max(1, min(4, int(os.getenv("REGIONAL_DATE_PIPELINE_CONCURRENCY", "4"))))

def _regional_category_url(base_url: str, slug: str, location_id: int) -> str:
    m = re.match(r"^(https://www\.kleinanzeigen\.de/.+)/(c\d+)$", base_url.rstrip("/"))
    if not m:
        raise ValueError(f"Unsupported Kleinanzeigen category URL: {base_url}")
    return f"{m.group(1)}/{slug}/{m.group(2)}l{int(location_id)}"


class SettingsInput(StatesGroup):
    include_words = State()
    exclude_words = State()
    min_views = State()
    view_test_url = State()


class ScanInput(StatesGroup):
    target_date = State()
    custom_price = State()


class RadarInput(StatesGroup):
    search = State()
    price = State()


class AdminInput(StatesGroup):
    user_search = State()
    plan_price = State()
    custom_days = State()
    broadcast_content = State()
    daily_radar_time = State()


def allowed(user_id: int) -> bool:
    return has_access(int(user_id), ADMIN_IDS)


@dataclass(slots=True)
class TrialStatus:
    enabled: bool
    eligible: bool
    used: int
    remaining: int
    limit: int = FREE_TRIAL_SCAN_LIMIT


async def free_trial_enabled() -> bool:
    async with SessionLocal() as session:
        row = await session.get(AppSetting, FREE_TRIAL_SETTING_KEY)
        if row is None:
            return bool(FREE_TRIAL_ENABLED_DEFAULT)
        return (row.value or "").strip().lower() in {"1", "true", "yes", "on"}


async def set_free_trial_enabled(enabled: bool) -> bool:
    value = "1" if enabled else "0"
    async with SessionLocal() as session:
        row = await session.get(AppSetting, FREE_TRIAL_SETTING_KEY)
        if row is None:
            row = AppSetting(key=FREE_TRIAL_SETTING_KEY, value=value, updated_at=datetime.utcnow())
            session.add(row)
        else:
            row.value = value
            row.updated_at = datetime.utcnow()
        await session.commit()
    return bool(enabled)


async def get_trial_status(user_id: int) -> TrialStatus:
    uid = int(user_id)
    enabled = await free_trial_enabled()
    if uid in ADMIN_IDS or is_banned_cached(uid) or current_access_mode() != "subscription":
        return TrialStatus(enabled=enabled, eligible=False, used=0, remaining=0)
    async with SessionLocal() as session:
        user = await session.get(BotUser, uid)
        if user is None:
            used = 0
            payments_count = 0
            access_until = None
        else:
            used = max(0, int(getattr(user, "trial_scans_used", 0) or 0))
            payments_count = max(0, int(user.payments_count or 0))
            access_until = user.access_until
    has_active_subscription = bool(access_until and access_until > datetime.utcnow())
    remaining = max(0, FREE_TRIAL_SCAN_LIMIT - used)
    eligible = bool(
        enabled and FREE_TRIAL_SCAN_LIMIT > 0 and remaining > 0
        and payments_count == 0 and not has_active_subscription
    )
    return TrialStatus(enabled=enabled, eligible=eligible, used=used, remaining=remaining)


async def trial_or_paid_scan_access(user_id: int) -> bool:
    if allowed(user_id):
        return True
    return (await get_trial_status(user_id)).eligible


def free_radar_preview_allowed(user_id: int) -> bool:
    """Public read-only DT Radar preview for subscription mode.

    Active subscribers/admins already have full access through ``allowed``.
    Banned users and admin-only mode never receive the preview.
    """
    uid = int(user_id)
    if allowed(uid) or is_banned_cached(uid):
        return False
    return current_access_mode() == "subscription"


async def _consume_trial_credit(user_id: int) -> TrialStatus | None:
    """Atomically reserve one trial launch immediately before queueing it."""
    uid = int(user_id)
    async with _trial_credit_guard:
        status = await get_trial_status(uid)
        if not status.eligible:
            return None
        async with SessionLocal() as session:
            # PostgreSQL row lock keeps the two-credit limit correct even if the
            # Telegram service is later scaled to multiple replicas.
            user = (await session.execute(
                select(BotUser).where(BotUser.user_id == uid).with_for_update()
            )).scalar_one_or_none()
            if user is None:
                now = datetime.utcnow()
                user = BotUser(user_id=uid, joined_at=now, last_seen_at=now, trial_scans_used=0)
                session.add(user)
                await session.flush()
            used = max(0, int(getattr(user, "trial_scans_used", 0) or 0))
            if int(user.payments_count or 0) > 0 or used >= FREE_TRIAL_SCAN_LIMIT:
                return None
            if user.access_until and user.access_until > datetime.utcnow():
                return None
            user.trial_scans_used = used + 1
            await session.commit()
            remaining = max(0, FREE_TRIAL_SCAN_LIMIT - int(user.trial_scans_used or 0))
            return TrialStatus(enabled=True, eligible=remaining > 0, used=int(user.trial_scans_used or 0), remaining=remaining)


async def _refund_trial_credit_for_scan(scan_id: int) -> bool:
    """Return a reserved credit only when a trial job never actually started."""
    async with _trial_credit_guard:
        async with SessionLocal() as session:
            scan = (await session.execute(
                select(UserScan).where(UserScan.id == int(scan_id)).with_for_update()
            )).scalar_one_or_none()
            if scan is None or not bool(getattr(scan, "is_trial", False)) or bool(getattr(scan, "trial_credit_refunded", False)):
                return False
            user = (await session.execute(
                select(BotUser).where(BotUser.user_id == int(scan.user_id)).with_for_update()
            )).scalar_one_or_none()
            if user is not None:
                user.trial_scans_used = max(0, int(getattr(user, "trial_scans_used", 0) or 0) - 1)
            scan.trial_credit_refunded = True
            await session.commit()
            return True


async def free_trial_stats() -> dict[str, int | bool]:
    enabled = await free_trial_enabled()
    async with SessionLocal() as session:
        used_one = int((await session.execute(
            select(func.count(BotUser.user_id)).where(BotUser.trial_scans_used >= 1)
        )).scalar_one() or 0)
        used_all = int((await session.execute(
            select(func.count(BotUser.user_id)).where(BotUser.trial_scans_used >= FREE_TRIAL_SCAN_LIMIT)
        )).scalar_one() or 0) if FREE_TRIAL_SCAN_LIMIT > 0 else 0
        converted = int((await session.execute(
            select(func.count(BotUser.user_id)).where(
                BotUser.trial_scans_used >= 1, BotUser.payments_count >= 1
            )
        )).scalar_one() or 0)
    return {"enabled": enabled, "used_one": used_one, "used_all": used_all, "converted": converted}


async def record_free_radar_event(
    user_id: int, event_type: str, *, mode: str = "", feature: str = "",
    product_id: int | None = None, item_count: int = 0,
) -> None:
    """Persist one anonymous-by-default free Radar funnel action.

    The event is intentionally best-effort: analytics must never block the user's
    Radar experience. Paid/admin users are not recorded here because this funnel
    measures only the public preview. Usernames/names are read later from BotUser,
    where normal bot activity already keeps them up to date.
    """
    uid = int(user_id)
    if not free_radar_preview_allowed(uid):
        return
    try:
        async with SessionLocal() as session:
            session.add(FreeRadarEvent(
                user_id=uid,
                event_type=str(event_type or "")[:32],
                mode=str(mode or "")[:24],
                feature=str(feature or "")[:40],
                product_id=(int(product_id) if product_id is not None else None),
                item_count=max(0, int(item_count or 0)),
                created_at=datetime.utcnow(),
            ))
            await session.commit()
    except Exception:
        log.exception("Could not record free Radar event user=%s event=%s", uid, event_type)


async def free_radar_funnel_stats(since: datetime | None = None) -> dict[str, int]:
    """Return distinct-user funnel counters for the free Radar preview."""
    async with SessionLocal() as session:
        def event_users(event_type: str):
            q = select(func.count(func.distinct(FreeRadarEvent.user_id))).where(
                FreeRadarEvent.event_type == event_type
            )
            if since is not None:
                q = q.where(FreeRadarEvent.created_at >= since)
            return q

        opened = int((await session.execute(event_users("radar_open"))).scalar_one() or 0)
        digest_open = int((await session.execute(event_users("daily_digest_open"))).scalar_one() or 0)
        best = int((await session.execute(event_users("best_open"))).scalar_one() or 0)
        mode_opened = int((await session.execute(event_users("mode_open"))).scalar_one() or 0)
        viewed_item = int((await session.execute(event_users("preview_item"))).scalar_one() or 0)
        upgrade = int((await session.execute(event_users("upgrade_click"))).scalar_one() or 0)

        completed_q = select(FreeRadarEvent.user_id).where(
            FreeRadarEvent.event_type == "preview_item",
            FreeRadarEvent.product_id.is_not(None),
        )
        if since is not None:
            completed_q = completed_q.where(FreeRadarEvent.created_at >= since)
        completed_q = completed_q.group_by(FreeRadarEvent.user_id, FreeRadarEvent.mode).having(
            func.count(func.distinct(FreeRadarEvent.product_id)) >= FREE_RADAR_PREVIEW_LIMIT
        ).subquery()
        completed = int((await session.execute(
            select(func.count(func.distinct(completed_q.c.user_id))).select_from(completed_q)
        )).scalar_one() or 0)

        visitors_q = select(
            FreeRadarEvent.user_id.label("user_id"),
            func.min(FreeRadarEvent.created_at).label("first_open_at"),
        ).where(FreeRadarEvent.event_type == "radar_open")
        if since is not None:
            visitors_q = visitors_q.where(FreeRadarEvent.created_at >= since)
        visitors_q = visitors_q.group_by(FreeRadarEvent.user_id).subquery()
        converted = int((await session.execute(
            select(func.count(func.distinct(visitors_q.c.user_id))).select_from(
                visitors_q.join(SubscriptionPayment, SubscriptionPayment.user_id == visitors_q.c.user_id)
            ).where(
                SubscriptionPayment.status == "paid",
                SubscriptionPayment.paid_at.is_not(None),
                SubscriptionPayment.paid_at >= visitors_q.c.first_open_at,
            )
        )).scalar_one() or 0)

    return {
        "opened": opened,
        "daily_digest_open": digest_open,
        "best": best,
        "mode_opened": mode_opened,
        "viewed_item": viewed_item,
        "completed_five": completed,
        "upgrade_click": upgrade,
        "converted": converted,
    }


async def free_radar_recent_visitors(page: int = 0, page_size: int = 6) -> tuple[list[dict], int]:
    """Return recent free-preview visitors with compact per-user behavior."""
    page = max(0, int(page or 0))
    page_size = max(1, min(20, int(page_size or 6)))
    async with SessionLocal() as session:
        activity = select(
            FreeRadarEvent.user_id.label("user_id"),
            func.min(FreeRadarEvent.created_at).label("first_event_at"),
            func.max(FreeRadarEvent.created_at).label("last_event_at"),
        ).group_by(FreeRadarEvent.user_id).subquery()
        total = int((await session.execute(
            select(func.count()).select_from(activity)
        )).scalar_one() or 0)
        rows = (await session.execute(
            select(
                activity.c.user_id, activity.c.first_event_at, activity.c.last_event_at,
                BotUser.username, BotUser.first_name, BotUser.trial_scans_used,
                BotUser.payments_count, BotUser.paid_total_usdt,
            ).select_from(
                activity.outerjoin(BotUser, BotUser.user_id == activity.c.user_id)
            ).order_by(activity.c.last_event_at.desc()).offset(page * page_size).limit(page_size)
        )).all()
        ids = [int(row.user_id) for row in rows]
        events = []
        paid_rows = []
        if ids:
            events = (await session.execute(
                select(FreeRadarEvent).where(FreeRadarEvent.user_id.in_(ids)).order_by(FreeRadarEvent.created_at.asc())
            )).scalars().all()
            paid_rows = (await session.execute(
                select(SubscriptionPayment.user_id, SubscriptionPayment.paid_at).where(
                    SubscriptionPayment.user_id.in_(ids),
                    SubscriptionPayment.status == "paid",
                    SubscriptionPayment.paid_at.is_not(None),
                )
            )).all()

    by_user: dict[int, dict] = {}
    for row in rows:
        uid = int(row.user_id)
        by_user[uid] = {
            "user_id": uid,
            "username": row.username,
            "first_name": row.first_name,
            "first_event_at": row.first_event_at,
            "last_event_at": row.last_event_at,
            "trial_scans_used": max(0, int(row.trial_scans_used or 0)),
            "payments_count": max(0, int(row.payments_count or 0)),
            "paid_total_usdt": float(row.paid_total_usdt or 0.0),
            "radar_opens": 0,
            "best_opens": 0,
            "mode_opens": {"hot": 0, "rising": 0, "ai": 0},
            "preview_products": {"hot": set(), "rising": set(), "ai": set()},
            "upgrade_clicks": 0,
            "locked_features": set(),
            "converted_after_radar": False,
        }
    for paid_row in paid_rows:
        data = by_user.get(int(paid_row.user_id))
        if data is None or paid_row.paid_at is None:
            continue
        first_event_at = data.get("first_event_at")
        if first_event_at is not None and paid_row.paid_at >= first_event_at:
            data["converted_after_radar"] = True
    for event in events:
        data = by_user.get(int(event.user_id))
        if data is None:
            continue
        event_type = str(event.event_type or "")
        mode = str(event.mode or "")
        if event_type == "radar_open":
            data["radar_opens"] += 1
        elif event_type == "best_open":
            data["best_opens"] += 1
        elif event_type == "mode_open" and mode in data["mode_opens"]:
            data["mode_opens"][mode] += 1
        elif event_type == "preview_item" and mode in data["preview_products"] and event.product_id is not None:
            data["preview_products"][mode].add(int(event.product_id))
        elif event_type == "upgrade_click":
            data["upgrade_clicks"] += 1
        elif event_type == "locked_feature" and event.feature:
            data["locked_features"].add(str(event.feature))
    return [by_user[int(row.user_id)] for row in rows], total


def read_only_history_allowed(user_id: int) -> bool:
    """Allow expired subscribers to keep access to their own saved scan history.

    This deliberately does not grant parser/network work. Banned users remain blocked,
    and admin-only mode keeps its original closed semantics.
    """
    uid = int(user_id)
    if uid in ADMIN_IDS or allowed(uid):
        return True
    if is_banned_cached(uid):
        return False
    return current_access_mode() == "subscription"


def main_keyboard(
    selected_count: int = 0, *, admin: bool = False, auto_observations: bool | None = None,
    access_active: bool = True, trial_remaining: int = 0, referral_enabled: bool = False,
) -> InlineKeyboardMarkup:
    """Product-style home screen with Scan and Radar as equal primary actions."""
    if auto_observations is True:
        auto_label = "⏱ Автозамеры · ✅ ВКЛ"
    elif auto_observations is False:
        auto_label = "⏱ Автозамеры · ⛔ ВЫКЛ"
    else:
        auto_label = "⏱ Автозамеры"
    if access_active:
        rows = [
            [InlineKeyboardButton(text="▶️ НОВЫЙ СКАН", callback_data="start_scan")],
            [InlineKeyboardButton(text="📡 DT RADAR 3.0", callback_data="radar_home")],
            [InlineKeyboardButton(text="🔥 Популярное", callback_data="popular_now"),
             InlineKeyboardButton(text="📊 Мои сканы", callback_data="my_scans")],
            [InlineKeyboardButton(text=f"🗂 Категории · {selected_count}/{MAX_SELECTED_CATEGORIES}", callback_data="groups"),
             InlineKeyboardButton(text="⚙️ Настройки", callback_data="settings")],
            [InlineKeyboardButton(text="📥 Очередь", callback_data="queue_status"),
             InlineKeyboardButton(text=auto_label, callback_data="auto_obs_menu")],
            [InlineKeyboardButton(text="💎 Подписка", callback_data="subscription")],
        ]
    elif trial_remaining > 0:
        rows = [
            [InlineKeyboardButton(text=f"🎁 Бесплатный скан · осталось {trial_remaining}", callback_data="start_scan")],
            [InlineKeyboardButton(text="📡 DT RADAR 3.0 · 🎁", callback_data="radar_home")],
            [InlineKeyboardButton(text=f"🗂 Категория · {min(selected_count, FREE_TRIAL_MAX_CATEGORIES)}/{FREE_TRIAL_MAX_CATEGORIES}", callback_data="groups"),
             InlineKeyboardButton(text="⚙️ Настройки", callback_data="settings")],
            [InlineKeyboardButton(text="📊 Мои сканы", callback_data="my_scans"),
             InlineKeyboardButton(text="📥 Очередь", callback_data="queue_status")],
            [InlineKeyboardButton(text="💎 Полный доступ", callback_data="subscription")],
        ]
    else:
        rows = [
            [InlineKeyboardButton(text="🔒 НОВЫЙ СКАН", callback_data="start_scan")],
            [InlineKeyboardButton(text="📡 DT RADAR 3.0 · 🎁", callback_data="radar_home")],
            [InlineKeyboardButton(text="📊 Мои сканы", callback_data="my_scans")],
            [InlineKeyboardButton(text="💎 Продлить подписку", callback_data="subscription")],
        ]
    if referral_enabled:
        rows.append([InlineKeyboardButton(text="🎁 Получить день бесплатно", callback_data="referral")])
    if admin:
        rows.append([InlineKeyboardButton(text="🛠 Админ-панель", callback_data="adminhome")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def post_scan_keyboard(scan_id: int | None = None, *, recheck: bool = False) -> InlineKeyboardMarkup:
    """Short result actions: TOP-12 remains instant, TOP-50 is available on demand."""
    rows = []
    if scan_id is not None:
        rows.append([
            InlineKeyboardButton(text="🔥 TOP-12", callback_data=f"scantop:{scan_id}"),
            InlineKeyboardButton(text="📋 TOP-50", callback_data=f"scantop50:{scan_id}:0"),
        ])
        rows.append([InlineKeyboardButton(text="📊 Открыть скан", callback_data=f"scan:{scan_id}")])
        if recheck and not STABLE_SCAN_ENGINE:
            rows.append([InlineKeyboardButton(text="🔄 Допроверить категории", callback_data=f"scanrecheck:{scan_id}")])
        rows.append([InlineKeyboardButton(text="🔄 Повторить скан", callback_data=f"scanrepeat:{scan_id}")])
    else:
        rows.append([InlineKeyboardButton(text="▶️ Новый скан", callback_data="start_scan")])
    rows.append([InlineKeyboardButton(text="🏠 Меню", callback_data="post_home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def partial_recheck_keyboard(scan_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Допроверить проблемные категории", callback_data=f"scanrecheck:{scan_id}")],
        [InlineKeyboardButton(text="📊 Открыть сохранённый скан", callback_data=f"scan:{scan_id}")],
    ])


def scan_detail_keyboard(scan_id: int, *, archived: bool = False, recheck: bool = False) -> InlineKeyboardMarkup:
    """Compact scan actions with both quick TOP-12 and full TOP-50."""
    back_text = "⬅️ Архив" if archived else "⬅️ Мои сканы"
    back_callback = "scan_archive:0" if archived else "my_scans"
    rows = [
        [InlineKeyboardButton(text="🔥 TOP-12", callback_data=f"scantop:{scan_id}"),
         InlineKeyboardButton(text="📋 TOP-50", callback_data=f"scantop50:{scan_id}:0")],
        [InlineKeyboardButton(text="🚀 Топ роста", callback_data=f"scangrowth:{scan_id}:3")],
        [InlineKeyboardButton(text="👁 Обновить", callback_data=f"scanviews:{scan_id}"),
         InlineKeyboardButton(text="📊 XLSX", callback_data=f"scanexport:{scan_id}")],
    ]
    if recheck and not STABLE_SCAN_ENGINE:
        rows.append([InlineKeyboardButton(text="🔄 Допроверить категории", callback_data=f"scanrecheck:{scan_id}")])
    rows += [
        [InlineKeyboardButton(text="🔄 Повторить", callback_data=f"scanrepeat:{scan_id}"),
         InlineKeyboardButton(text="🕘 История", callback_data=f"scanhistory:{scan_id}")],
        [InlineKeyboardButton(text=back_text, callback_data=back_callback),
         InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


def growth_period_keyboard(scan_id: int, active_hours: int = 3, category_key: str | None = None) -> InlineKeyboardMarkup:
    prefix = f"pcg:{scan_id}:{category_key}:" if category_key else f"scangrowth:{scan_id}:"
    export_prefix = f"pce:{scan_id}:{category_key}:" if category_key else f"scangrowthexport:{scan_id}:"
    def b(hours: int) -> InlineKeyboardButton:
        label = f"{hours}ч"
        if hours == active_hours:
            label = "✅ " + label
        return InlineKeyboardButton(text=label, callback_data=f"{prefix}{hours}")

    back_callback = f"popularcat:{category_key}" if category_key else f"scan:{scan_id}"
    return InlineKeyboardMarkup(inline_keyboard=[
        [b(3), b(6), b(12)],
        [InlineKeyboardButton(text="📊 Скачать TOP-50", callback_data=f"{export_prefix}{active_hours}")],
        [InlineKeyboardButton(
            text="👁 Обновить последний скан" if category_key else "👁 Обновить сейчас",
            callback_data=f"scanviews:{scan_id}",
        )],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data=back_callback)],
    ])


def popular_categories_keyboard(items: list[tuple[str, UserScan]]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if not items:
        rows.append([InlineKeyboardButton(text="▶️ Сделать первый скан", callback_data="start_scan")])
    for key, scan in items[:30]:
        cat = CATEGORIES.get(key)
        if cat is None:
            continue
        icon = GROUPS.get(cat.group).icon if cat.group in GROUPS else "📂"
        rows.append([InlineKeyboardButton(text=f"{icon} {cat.name[:38]}", callback_data=f"popularcat:{key}")])
    rows.append([InlineKeyboardButton(text="🏠 Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def popular_category_keyboard(scan_id: int, category_key: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👁 По просмотрам", callback_data=f"pcv:{scan_id}:{category_key}")],
        [InlineKeyboardButton(text="🚀 3ч", callback_data=f"pcg:{scan_id}:{category_key}:3"),
         InlineKeyboardButton(text="🚀 6ч", callback_data=f"pcg:{scan_id}:{category_key}:6"),
         InlineKeyboardButton(text="🚀 12ч", callback_data=f"pcg:{scan_id}:{category_key}:12")],
        [InlineKeyboardButton(text="⬅️ Категории", callback_data="popular_now"),
         InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ])


RADAR_STATUS_ICON = {
    "hot": "🔥",
    "rising": "📈",
    "stable": "🟡",
    "cooling": "💤",
    "historical": "🕒",
}
RADAR_STATUS_LABEL = {
    "hot": "Hot · спрос доказан",
    "rising": "Strong · спрос подтверждается",
    "stable": "Early · ранний сигнал",
    "cooling": "Остывает",
    "historical": "История · сигнал устарел",
}
RADAR_TYPE_LABEL = {
    "hot_product": "🔥 Hot Product",
    "hidden_gem": "💎 Hidden Gem",
    "emerging": "🚀 Emerging",
    "saturated": "⚫ Saturated",
    "spark": "⚡ Signal",
    "fast_sold": "⚡ Fast Sold",
    "observed_demand": "👁 DT Observed Demand",
}


def _radar_price_filter(data: dict | None) -> str:
    value = str((data or {}).get("radar_price_filter") or "any").strip()
    return value or "any"


def _radar_product_price_text(product) -> str:
    lo = getattr(product, "min_price_eur", None)
    hi = getattr(product, "max_price_eur", None)
    if lo is None and hi is None:
        return "—"
    if lo is None:
        return f"до {int(hi)} €"
    if hi is None or int(lo) == int(hi):
        return f"{int(lo)} €"
    return f"{int(lo)}–{int(hi)} €"


def _radar_context_back(data: dict | None) -> tuple[str, str]:
    data = data or {}
    kind = str(data.get("radar_context_kind") or "")
    page = max(0, int(data.get("radar_context_page") or 0))
    if kind == "category":
        category_key = str(data.get("radar_context_category") or "")
        mode = str(data.get("radar_context_mode") or "category_new")
        prefix = "radarcatbest" if mode == "category_best" else "radarcat"
        if category_key:
            return f"{prefix}:{category_key}:{page}", "⬅️ Назад к категории"
    if kind == "search":
        return f"radarsearchpage:{page}", "⬅️ К результатам"
    if kind == "list":
        mode = str(data.get("radar_context_mode") or "hot")
        return f"radarlist:{mode}:{page}", "⬅️ К списку"
    return "radar_home", "⬅️ DT Radar"


def radar_price_keyboard(current_filter: str = "any") -> InlineKeyboardMarkup:
    current = (current_filter or "any").strip()
    def b(label: str, value: str) -> InlineKeyboardButton:
        mark = "✅ " if current == value else ""
        return InlineKeyboardButton(text=f"{mark}{label}", callback_data=f"radarprice:set:{value}")
    return InlineKeyboardMarkup(inline_keyboard=[
        [b("Любая", "any"), b("до 50 €", "0_50")],
        [b("50–100 €", "50_100"), b("100–200 €", "100_200")],
        [b("200–500 €", "200_500"), b("500+ €", "500_plus")],
        [InlineKeyboardButton(text="✍️ Свой диапазон", callback_data="radarprice:custom")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="radarprice:back")],
    ])


def radar_home_keyboard(user_id: int | None = None) -> InlineKeyboardMarkup:
    # Paid users see all four destinations.  Free preview keeps the same shape so
    # the value of the full product is visible without granting the locked data.
    full = bool(user_id is not None and allowed(int(user_id)))
    search_cb = "radarsearch" if full else "radar_locked:search"
    cats_cb = "radarcats:0" if full else "radar_locked:categories"
    favorites_cb = "radarlist:favorites:0" if full else "radar_locked:favorites"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔥 Лучшие сейчас", callback_data="radarbest"),
         InlineKeyboardButton(text="🔎 Поиск" + ("" if full else " · 🔒"), callback_data=search_cb)],
        [InlineKeyboardButton(text="🗂 Категории" + ("" if full else " · 🔒"), callback_data=cats_cb),
         InlineKeyboardButton(text="⭐ Мой Radar" + ("" if full else " · 🔒"), callback_data=favorites_cb)],
        [InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ])


def radar_best_keyboard(user_id: int | None = None) -> InlineKeyboardMarkup:
    full = bool(user_id is not None and allowed(int(user_id)))
    fast_cb = "radarlist:fastsold:0" if full else "radar_locked:fastsold"
    rows = [
        [InlineKeyboardButton(text="🔥 Горячие", callback_data="radarlist:hot:0"),
         InlineKeyboardButton(text="🚀 Набирают", callback_data="radarlist:rising:0")],
        [InlineKeyboardButton(text="⚡ Быстро исчезли" + ("" if full else " · 🔒"), callback_data=fast_cb)],
    ]
    if full:
        rows.append([InlineKeyboardButton(text="🏆 Рекорды Radar", callback_data="radarlist:alltime:0")])
    else:
        rows.append([InlineKeyboardButton(text="🏆 Рекорды Radar · 🔒", callback_data="radar_locked:records")])
        rows.append([InlineKeyboardButton(text="💎 Открыть полный DT Radar", callback_data="radar_upgrade:best")])
    rows.append([InlineKeyboardButton(text="⬅️ DT Radar", callback_data="radar_home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def radar_search_keyboard(items, *, page: int, total: int, price_filter: str = "any") -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for product in items:
        icon = RADAR_STATUS_ICON.get(str(product.status or ""), "📡")
        title = " ".join(str(product.title or "Товар").split())
        if len(title) > 32:
            title = title[:31].rstrip() + "…"
        rows.append([InlineKeyboardButton(
            text=f"{icon} {int(product.current_score or 0)} · {title}",
            callback_data=f"radaritem:{int(product.id)}",
        )])
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"radarsearchpage:{page-1}"))
    if (page + 1) * RADAR_PAGE_SIZE < total:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"radarsearchpage:{page+1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton(
        text=f"💶 Цена: {price_filter_label(price_filter)}", callback_data="radarprice:open"
    )])
    rows.append([InlineKeyboardButton(text="🔎 Новый поиск", callback_data="radarsearch")])
    rows.append([InlineKeyboardButton(text="⬅️ DT Radar", callback_data="radar_home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def radar_locked_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💎 Открыть полный DT Radar", callback_data="radar_upgrade:locked")],
        [InlineKeyboardButton(text="🔥 Посмотреть 5 бесплатных находок", callback_data="radarbest")],
        [InlineKeyboardButton(text="⬅️ DT Radar", callback_data="radar_home")],
    ])


def radar_preview_list_keyboard(
    items, *, mode: str, total: int, trial_remaining: int = 0,
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for product in items:
        icon = RADAR_STATUS_ICON.get(str(product.status or ""), "📡")
        title = " ".join(str(product.title or "Товар").split())
        if len(title) > 32:
            title = title[:31].rstrip() + "…"
        rows.append([InlineKeyboardButton(
            text=f"{icon} {int(product.current_score or 0)} · {title}",
            callback_data=f"radarpreviewitem:{mode}:{int(product.id)}",
        )])
    if total > len(items):
        rows.append([InlineKeyboardButton(
            text=f"🔒 Ещё {max(0, int(total) - len(items))} · открыть полный Radar",
            callback_data="radar_upgrade:preview",
        )])
    else:
        rows.append([InlineKeyboardButton(text="💎 Открыть полный DT Radar", callback_data="radar_upgrade:preview")])
    if int(trial_remaining) > 0:
        rows.append([InlineKeyboardButton(
            text=f"🎁 Бесплатный скан · осталось {int(trial_remaining)}", callback_data="start_scan"
        )])
    rows.append([InlineKeyboardButton(text="⬅️ Лучшие сейчас", callback_data="radarbest")])
    return InlineKeyboardMarkup(inline_keyboard=rows)



def radar_list_keyboard(
    items, *, mode: str, page: int, total: int, category_key: str | None = None,
    price_filter: str = "any",
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for product in items:
        icon = "⚡" if mode == "fastsold" else RADAR_STATUS_ICON.get(str(product.status or ""), "📡")
        title = " ".join(str(product.title or "Товар").split())
        if len(title) > 32:
            title = title[:31].rstrip() + "…"
        shown_score = int(product.peak_score or 0) if mode == "alltime" else int(product.current_score or 0)
        rows.append([InlineKeyboardButton(
            text=f"{icon} {shown_score} · {title}",
            callback_data=f"radaritem:{int(product.id)}",
        )])
    nav: list[InlineKeyboardButton] = []
    category_prefix = "radarcatbest" if mode == "category_best" else "radarcat"
    if page > 0:
        if category_key:
            nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"{category_prefix}:{category_key}:{page-1}"))
        else:
            nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"radarlist:{mode}:{page-1}"))
    if (page + 1) * RADAR_PAGE_SIZE < total:
        if category_key:
            nav.append(InlineKeyboardButton(text="➡️", callback_data=f"{category_prefix}:{category_key}:{page+1}"))
        else:
            nav.append(InlineKeyboardButton(text="➡️", callback_data=f"radarlist:{mode}:{page+1}"))
    if nav:
        rows.append(nav)
    if category_key:
        if mode == "category_best":
            rows.append([InlineKeyboardButton(text="🆕 Сначала новые", callback_data=f"radarcat:{category_key}:0")])
        else:
            rows.append([InlineKeyboardButton(text="🔥 Сначала лучшие", callback_data=f"radarcatbest:{category_key}:0")])
        rows.append([InlineKeyboardButton(
            text=f"💶 Цена: {price_filter_label(price_filter)}", callback_data="radarprice:open"
        )])
        cat = CATEGORIES.get(category_key)
        group = GROUPS.get(cat.group) if cat is not None else None
        rows.append([InlineKeyboardButton(
            text=_button_text(f"⬅️ {group.name}" if group is not None else "⬅️ Категории"),
            callback_data=f"radargroup:{cat.group}" if cat is not None and cat.group in GROUPS else "radarcats:0",
        )])
    else:
        back_callback = "radarbest" if mode in {"hot", "rising", "ai", "fastsold", "alltime"} else "radar_home"
        back_text = "⬅️ Лучшие сейчас" if back_callback == "radarbest" else "⬅️ DT Radar"
        rows.append([InlineKeyboardButton(text=back_text, callback_data=back_callback)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _radar_category_stats(items: list[tuple[str, int, int, int]]) -> dict[str, tuple[int, int, int]]:
    """category -> (all accepted products, new today, max DT Score)."""
    return {
        str(key): (int(total or 0), int(new_today or 0), int(score or 0))
        for key, total, new_today, score in items
    }


def radar_groups_keyboard(items: list[tuple[str, int, int, int]]) -> InlineKeyboardMarkup:
    """First Radar category level: only the large Kleinanzeigen sections.

    Product counts are aggregated from leaf categories.  DT Score is intentionally
    kept for the product list instead of turning the navigator into another stats
    table.
    """
    stats = _radar_category_stats(items)
    group_counts: dict[str, int] = {key: 0 for key in GROUPS}
    for category_key, (count, _new_today, _score) in stats.items():
        cat = CATEGORIES.get(category_key)
        if cat is None or cat.is_group or cat.group not in GROUPS:
            continue
        group_counts[cat.group] = int(group_counts.get(cat.group, 0)) + int(count)

    rows: list[list[InlineKeyboardButton]] = []
    for group_key, group in GROUPS.items():
        count = int(group_counts.get(group_key, 0))
        rows.append([InlineKeyboardButton(
            text=_button_text(f"{group.icon} {group.name} · {count}"),
            callback_data=f"radargroup:{group_key}",
        )])
    rows.append([InlineKeyboardButton(text="⬅️ DT Radar", callback_data="radar_home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def radar_group_keyboard(group_key: str, items: list[tuple[str, int, int, int]]) -> InlineKeyboardMarkup:
    """Second Radar category level: leaf subcategories inside one large section."""
    stats = _radar_category_stats(items)
    rows: list[list[InlineKeyboardButton]] = []
    for cat in categories_for_group(group_key):
        if cat.is_group:
            continue
        count, new_today, _score = stats.get(cat.key, (0, 0, 0))
        suffix = f" · {count}" + (f" · 🆕 {new_today}" if new_today else "")
        rows.append([InlineKeyboardButton(
            text=_button_text(f"📂 {cat.name}{suffix}"),
            callback_data=f"radarcat:{cat.key}:0",
        )])
    rows.append([InlineKeyboardButton(text="⬅️ Все разделы", callback_data="radarcats:0")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def radar_product_keyboard(
    product_id: int, *, favorite: bool = False, listing_url: str | None = None,
    preview_mode: str | None = None, return_callback: str | None = None,
    return_text: str | None = None,
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if preview_mode is None:
        rows.append([InlineKeyboardButton(
            text="★ Убрать из Моего Radar" if favorite else "⭐ Добавить в Мой Radar",
            callback_data=f"radarfav:{int(product_id)}",
        )])
    if listing_url:
        rows.append([InlineKeyboardButton(text="🔗 Открыть актуальное объявление", url=listing_url)])
    if preview_mode is not None:
        rows.append([InlineKeyboardButton(text="💎 Открыть полный DT Radar", callback_data="radar_upgrade:item")])
        rows.append([InlineKeyboardButton(text="⬅️ К бесплатным находкам", callback_data=f"radarlist:{preview_mode}:0")])
    else:
        rows.append([InlineKeyboardButton(
            text=return_text or "⬅️ DT Radar", callback_data=return_callback or "radar_home"
        )])
    return InlineKeyboardMarkup(inline_keyboard=rows)


class BotChatAdapter:
    """Small duck-typed adapter so export code can send to a chat after a queue job."""

    def __init__(self, bot: Bot, chat_id: int, *, prefix: str = "", reply_markup: InlineKeyboardMarkup | None = None):
        self.bot = bot
        self.chat_id = chat_id
        self.prefix = prefix.strip()
        self.reply_markup = reply_markup

    async def answer(self, text: str, **kwargs):
        if self.prefix:
            text = f"{self.prefix}\n\n{text}"
        if self.reply_markup is not None:
            kwargs["reply_markup"] = self.reply_markup
        kwargs.setdefault("parse_mode", ParseMode.HTML)
        return await self.bot.send_message(self.chat_id, text, **kwargs)

    async def answer_document(self, document, *, caption: str | None = None, **kwargs):
        full_caption = caption or ""
        if self.prefix:
            full_caption = f"{self.prefix}\n\n{full_caption}" if full_caption else self.prefix
        if self.reply_markup is not None:
            kwargs["reply_markup"] = self.reply_markup
        kwargs.setdefault("parse_mode", ParseMode.HTML)
        return await self.bot.send_document(self.chat_id, document, caption=full_caption, **kwargs)


def _button_text(text: str, limit: int = 62) -> str:
    """Keep long one-row category labels readable in Telegram clients."""
    clean = " ".join((text or "").split())
    return clean if len(clean) <= limit else clean[: max(1, limit - 1)].rstrip() + "…"


def _selected_group_children(group_key: str, selected_keys: set[str]) -> list:
    return [
        cat for cat in categories_for_group(group_key)
        if not cat.is_group and cat.key in selected_keys
    ]


def groups_keyboard(selected_keys: set[str], *, max_selected: int = MAX_SELECTED_CATEGORIES) -> InlineKeyboardMarkup:
    # v4.3.4: one long vertical list. A selected section shows not only a count,
    # but also the chosen subcategory names so the user can see the selection
    # without opening every section again.
    rows: list[list[InlineKeyboardButton]] = []
    for group in GROUPS.values():
        chosen = _selected_group_children(group.key, selected_keys)
        if chosen:
            preview_names = [cat.name for cat in chosen[:2]]
            preview = ", ".join(preview_names)
            if len(chosen) > 2:
                preview += f" +{len(chosen) - 2}"
            label = _button_text(f"✅ {group.icon} {group.name} · {len(chosen)} · {preview}")
        else:
            label = _button_text(f"▫️ {group.icon} {group.name}")
        rows.append([InlineKeyboardButton(text=label, callback_data=f"grp:{group.key}")])

    counter_icon = "⚠️" if len(selected_keys) > max_selected else "✅"
    rows.append([InlineKeyboardButton(
        text=f"{counter_icon} Выбрано: {len(selected_keys)}/{max_selected}",
        callback_data="selected",
    )])
    rows.append([InlineKeyboardButton(text="🧹 Очистить выбор", callback_data="clear_all")])
    rows.append([InlineKeyboardButton(text="⬅️ Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def category_keyboard(
    group_key: str, selected_keys: set[str], *, max_selected: int = MAX_SELECTED_CATEGORIES
) -> InlineKeyboardMarkup:
    # "Весь раздел" and bulk-select are intentionally gone. Only explicit
    # subcategory choices are shown, so every scan has a clear scope.
    cats = [cat for cat in categories_for_group(group_key) if not cat.is_group]
    chosen_here = [cat for cat in cats if cat.key in selected_keys]
    header = f"✅ В разделе выбрано: {len(chosen_here)} · всего {len(selected_keys)}/{max_selected}"
    rows = [[InlineKeyboardButton(text=_button_text(header), callback_data="selected")]]
    for cat in cats:
        marker = "✅" if cat.key in selected_keys else "▫️"
        rows.append([InlineKeyboardButton(text=_button_text(f"{marker} {cat.name}"), callback_data=f"cat:{cat.key}")])
    rows.append([InlineKeyboardButton(text="▶️ Новый скан", callback_data="start_scan")])
    rows.append([InlineKeyboardButton(text="⬅️ К разделам", callback_data="groups")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _mode_button(s: UserSettings, mode: str) -> InlineKeyboardButton:
    label = MODE_LABELS[mode]
    if s.output_mode == mode:
        label = "✅ " + label
    return InlineKeyboardButton(text=label, callback_data=f"quickmode:{mode}")


def min_views_label(value: int | None) -> str:
    threshold = max(0, int(value or 0))
    return "Без порога" if threshold == 0 else f"{threshold}+"


def settings_keyboard(s: UserSettings) -> InlineKeyboardMarkup:
    mode_label = MODE_LABELS.get(s.output_mode, s.output_mode)
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"Режим: {mode_label}", callback_data="set_mode")],
        [InlineKeyboardButton(text=f"👁 {min_views_label(getattr(s, 'min_views', 0))}", callback_data="set_min_views")],
        [InlineKeyboardButton(text=f"🧠 Дубли · {'Вкл' if s.smart_dedupe else 'Выкл'}", callback_data="toggle_dedupe"),
         InlineKeyboardButton(text=f"🧹 Шум · {'Вкл' if s.clean_noise else 'Выкл'}", callback_data="toggle_noise")],
        [InlineKeyboardButton(text=f"↕️ {SORT_LABELS.get(s.sort_mode, s.sort_mode)}", callback_data="set_sort")],
        [InlineKeyboardButton(text="🔎 Ключевые слова", callback_data="set_include"),
         InlineKeyboardButton(text="🚫 Исключения", callback_data="set_exclude")],
        [InlineKeyboardButton(text="🌐 Язык", callback_data="language_settings")],
        [InlineKeyboardButton(text="▶️ Новый скан", callback_data="start_scan")],
        [InlineKeyboardButton(text="ℹ️ Что выбрать?", callback_data="mode_help"),
         InlineKeyboardButton(text="♻️ Сбросить", callback_data="reset_settings")],
        [InlineKeyboardButton(text="⬅️ Меню", callback_data="home")],
    ])


def scan_price_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💶 Любая цена", callback_data="scanprice:any")],
        [InlineKeyboardButton(text="50+ €", callback_data="scanprice:50_plus"),
         InlineKeyboardButton(text="100+ €", callback_data="scanprice:100_plus")],
        [InlineKeyboardButton(text="200+ €", callback_data="scanprice:200_plus"),
         InlineKeyboardButton(text="500+ €", callback_data="scanprice:500_plus")],
        [InlineKeyboardButton(text="✍️ Свой диапазон", callback_data="scanprice:custom")],
        [InlineKeyboardButton(text="⬅️ Другая дата", callback_data="start_scan"),
         InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ])


def page_limit_keyboard(*, trial: bool = False) -> InlineKeyboardMarkup:
    if trial:
        return InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="15 стр.", callback_data="scanpages:15"),
             InlineKeyboardButton(text="🎁 25 стр.", callback_data="scanpages:25")],
            [InlineKeyboardButton(text="🔒 50 стр. · подписка", callback_data="subscription")],
            [InlineKeyboardButton(text="💶 Изменить цену", callback_data="scanprice_menu")],
            [InlineKeyboardButton(text="⬅️ Другая дата", callback_data="start_scan"),
             InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
        ])
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="15 стр.", callback_data="scanpages:15"),
         InlineKeyboardButton(text="25 стр.", callback_data="scanpages:25"),
         InlineKeyboardButton(text="50 стр.", callback_data="scanpages:50")],
        [InlineKeyboardButton(text="💶 Изменить цену", callback_data="scanprice_menu")],
        [InlineKeyboardButton(text="⬅️ Другая дата", callback_data="start_scan"),
         InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ])


def scan_date_keyboard() -> InlineKeyboardMarkup:
    """Five-day date picker. Manual/older dates are intentionally unavailable."""
    today = datetime.now(MOSCOW).date()
    days = [today - timedelta(days=offset) for offset in range(DATE_MAX_AGE_DAYS + 1)]
    labels: list[InlineKeyboardButton] = []
    for offset, day in enumerate(days):
        if offset == 0:
            label = f"📅 Сегодня · {day:%d.%m}"
        elif offset == 1:
            label = f"↩️ Вчера · {day:%d.%m}"
        else:
            label = f"{day:%d.%m}"
        labels.append(InlineKeyboardButton(text=label, callback_data=f"scan_date:{day.isoformat()}"))
    rows: list[list[InlineKeyboardButton]] = []
    for index in range(0, len(labels), 2):
        rows.append(labels[index:index + 2])
    rows.append([InlineKeyboardButton(text="⬅️ Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def choice_keyboard(prefix: str, options: list[tuple[str, str]]) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=label, callback_data=f"{prefix}:{value}")] for value, label in options]
    rows.append([InlineKeyboardButton(text="⬅️ К настройкам", callback_data="settings")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def min_views_keyboard(current: int = 0) -> InlineKeyboardMarkup:
    presets = (0, 10, 25, 50, 100)
    rows: list[list[InlineKeyboardButton]] = []
    row: list[InlineKeyboardButton] = []
    for value in presets:
        label = "Без порога" if value == 0 else f"{value}+"
        if int(current or 0) == value:
            label = "✅ " + label
        row.append(InlineKeyboardButton(text=label, callback_data=f"minviews:{value}"))
        if len(row) == 3:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="✏️ Своё значение", callback_data="minviews:custom")])
    rows.append([InlineKeyboardButton(text="⬅️ К настройкам", callback_data="settings")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def get_settings(user_id: int) -> UserSettings:
    async with SessionLocal() as session:
        s = await session.get(UserSettings, user_id)
        if s is None:
            s = UserSettings(user_id=user_id)
            session.add(s)
            await session.commit()
            await session.refresh(s)
        return s


async def get_settings_for_scan(scan: UserScan) -> UserSettings:
    s = await get_settings(int(scan.user_id))
    # Price is a snapshot property of the scan since v4.3.7. Other filters keep
    # their existing behavior, but opening an old scan must never silently change
    # its price range after the user starts a different scan later.
    s.price_filter = (getattr(scan, "price_filter", "any") or "any").strip()
    return s


async def update_setting(user_id: int, field: str, value) -> UserSettings:
    async with SessionLocal() as session:
        s = await session.get(UserSettings, user_id)
        if s is None:
            s = UserSettings(user_id=user_id)
            session.add(s)
        setattr(s, field, value)
        await session.commit()
        await session.refresh(s)
        return s


async def set_auto_observations(user_id: int, enabled: bool) -> UserSettings:
    """Enable/disable future +3/+6/+12h checkpoints for one user.

    Turning the feature off also removes every unfinished automatic checkpoint
    already queued for that user's saved scans. Completed history is preserved.
    Manual view refreshes are independent and always remain available.
    """
    s = await update_setting(user_id, "auto_observations", bool(enabled))
    if not enabled:
        async with db_write_lock:
            async with SessionLocal() as session:
                user_scan_ids = select(UserScan.id).where(UserScan.user_id == user_id)
                await session.execute(
                    delete(ScanObservation).where(
                        ScanObservation.scan_id.in_(user_scan_ids),
                        ScanObservation.status != "done",
                    )
                )
                await session.commit()
    return s


async def auto_observations_enabled(user_id: int) -> bool:
    s = await get_settings(user_id)
    return bool(getattr(s, "auto_observations", False))


async def reset_user_settings(user_id: int) -> UserSettings:
    async with SessionLocal() as session:
        old = await session.get(UserSettings, user_id)
        if old:
            await session.delete(old)
            await session.commit()
    return await get_settings(user_id)


async def get_selected(user_id: int) -> set[str]:
    async with SessionLocal() as session:
        result = await session.execute(select(SelectedCategory.category_key).where(SelectedCategory.user_id == user_id))
        # v4.3.4: root/"whole section" categories remain in CATEGORIES only for
        # backwards compatibility with old saved scans. They are no longer a
        # selectable user option and therefore must not leak back into a new run.
        return {
            x for x in result.scalars().all()
            if x in CATEGORIES and not CATEGORIES[x].is_group
        }


def _scan_category_keys(scan: UserScan) -> list[str]:
    return [x for x in (scan.category_keys or "").split(",") if x]


def _scan_title(keys: list[str]) -> str:
    names = [CATEGORIES[k].name for k in keys if k in CATEGORIES]
    if not names:
        return "Скан Kleinanzeigen"
    if len(names) == 1:
        return names[0]
    if len(names) == 2:
        return f"{names[0]} + {names[1]}"
    return f"{names[0]} + ещё {len(names) - 1}"


@dataclass
class GrowthMetric:
    listing: Listing
    base_views: int
    current_views: int
    delta: int
    elapsed_hours: float
    per_hour: float
    observed_at: datetime


def _validate_scan_category_count(category_keys: list[str]) -> list[str]:
    return validate_scan_category_keys(category_keys, set(CATEGORIES))


async def create_user_scan(
    user_id: int, job_uid: str, category_keys: list[str], page_limit: int, target_date: str,
    price_filter: str = "any", *, is_trial: bool = False,
) -> UserScan:
    category_keys = _validate_scan_category_count(category_keys)
    scan = UserScan(
        job_uid=job_uid,
        user_id=user_id,
        title=_scan_title(category_keys),
        category_keys=",".join(category_keys),
        page_limit=page_limit,
        target_date=target_date,
        price_filter=(price_filter or "any"),
        status="queued",
        total_categories=len(category_keys),
        is_trial=bool(is_trial),
        trial_credit_refunded=False,
    )
    async with SessionLocal() as session:
        session.add(scan)
        await session.commit()
        await session.refresh(scan)
        return scan


async def invalidate_untrusted_view_analytics_once() -> int:
    """One-time reset of counters produced before Accurate Views Core.

    Existing listing/scan rows are preserved. Only view values and growth history
    are invalidated because mixing old unverified numbers with verified v4.2.2
    measurements would create false TOP/growth signals.
    """
    if not ACCURATE_VIEWS_MODE:
        return 0
    async with db_write_lock:
        async with SessionLocal() as session:
            setting = await session.get(AppSetting, "view_analytics_core_version")
            if setting is not None and (setting.value or "").strip() == VIEW_ANALYTICS_CORE_VERSION:
                return 0

            listing_result = await session.execute(
                update(Listing).values(view_count=None, views_checked_at=None)
            )
            await session.execute(update(ScanListing).values(initial_view_count=None))
            await session.execute(delete(ViewHistory))
            await session.execute(delete(ScanViewHistory))
            await session.execute(delete(ScanObservation))

            if setting is None:
                setting = AppSetting(
                    key="view_analytics_core_version",
                    value=VIEW_ANALYTICS_CORE_VERSION,
                    updated_at=datetime.utcnow(),
                )
                session.add(setting)
            else:
                setting.value = VIEW_ANALYTICS_CORE_VERSION
                setting.updated_at = datetime.utcnow()
            await session.commit()
            return int(listing_result.rowcount or 0)


async def get_persisted_active_scan(user_id: int) -> UserScan | None:
    """Return the user's unfinished scan from PostgreSQL, if any."""
    async with SessionLocal() as session:
        result = await session.execute(
            select(UserScan)
            .where(
                UserScan.user_id == int(user_id),
                UserScan.status.in_(["queued", "running", "cancelling"]),
                UserScan.finished_at.is_(None),
            )
            .order_by(UserScan.created_at.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()


async def attach_user_scan_message(scan_id: int, chat_id: int, status_message_id: int) -> None:
    async with SessionLocal() as session:
        scan = await session.get(UserScan, int(scan_id))
        if scan is None:
            return
        scan.chat_id = int(chat_id)
        scan.status_message_id = int(status_message_id)
        await session.commit()


async def record_user_scan_retry(scan_id: int | None, error_text: str) -> None:
    if scan_id is None:
        return
    async with SessionLocal() as session:
        scan = await session.get(UserScan, int(scan_id))
        if scan is None:
            return
        scan.retry_count = int(scan.retry_count or 0) + 1
        scan.last_error = (error_text or "")[:1000] or None
        await session.commit()


async def get_user_scan(user_id: int, scan_id: int) -> UserScan | None:
    async with SessionLocal() as session:
        result = await session.execute(select(UserScan).where(UserScan.id == scan_id, UserScan.user_id == user_id))
        return result.scalar_one_or_none()


ACTIVE_SCAN_STATUSES = ("queued", "running", "cancelling")


async def get_active_user_scan(user_id: int) -> UserScan | None:
    """PostgreSQL source-of-truth for one user's active scan.

    In distributed mode the Telegram process and parser workers do not share RAM,
    so active_jobs cannot be authoritative. Keeping this check in PostgreSQL also
    prevents duplicate launches when the bot process restarts.
    """
    async with SessionLocal() as session:
        result = await session.execute(
            select(UserScan)
            .where(
                UserScan.user_id == int(user_id),
                UserScan.status.in_(ACTIVE_SCAN_STATUSES),
                UserScan.finished_at.is_(None),
            )
            .order_by(UserScan.created_at.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()


async def user_has_active_scan(user_id: int) -> bool:
    if DISTRIBUTED_WORKERS:
        return (await get_active_user_scan(user_id)) is not None
    async with job_guard:
        existing = active_jobs.get(user_id)
        return bool(existing and existing.state in {"queued", "running"} and not existing.cancel_requested)


async def queued_scan_count() -> int:
    if DISTRIBUTED_WORKERS:
        async with SessionLocal() as session:
            value = (await session.execute(
                select(func.count(UserScan.id)).where(
                    UserScan.status == "queued", UserScan.finished_at.is_(None)
                )
            )).scalar_one()
            return int(value or 0)
    async with job_guard:
        return len(queued_job_ids)


async def queue_is_full() -> bool:
    return (await queued_scan_count()) >= MAX_QUEUE_SIZE


async def archive_expired_scans(user_id: int | None = None) -> int:
    """Archive completed scan cards 24h after completion without deleting data."""
    now = datetime.utcnow()
    cutoff = now - timedelta(hours=SCAN_ARCHIVE_AFTER_HOURS)
    conditions = [
        UserScan.archived_at.is_(None),
        UserScan.status.in_(ARCHIVABLE_SCAN_STATUSES),
        func.coalesce(UserScan.finished_at, UserScan.created_at) <= cutoff,
    ]
    if user_id is not None:
        conditions.append(UserScan.user_id == user_id)
    async with db_write_lock:
        async with SessionLocal() as session:
            result = await session.execute(
                update(UserScan).where(*conditions).values(archived_at=now)
            )
            await session.commit()
            return int(result.rowcount or 0)


async def archive_active_finished_scans(user_id: int) -> int:
    """Manual inbox cleanup: archive every finished visible scan immediately."""
    now = datetime.utcnow()
    async with db_write_lock:
        async with SessionLocal() as session:
            result = await session.execute(
                update(UserScan)
                .where(
                    UserScan.user_id == user_id,
                    UserScan.archived_at.is_(None),
                    UserScan.status.in_(ARCHIVABLE_SCAN_STATUSES),
                )
                .values(archived_at=now)
            )
            await session.commit()
            return int(result.rowcount or 0)


async def get_user_scans(user_id: int, limit: int = 10, *, archive: bool = True) -> list[UserScan]:
    if archive:
        await archive_expired_scans(user_id)
    async with SessionLocal() as session:
        result = await session.execute(
            select(UserScan)
            .where(UserScan.user_id == user_id, UserScan.archived_at.is_(None))
            .order_by(UserScan.created_at.desc())
            .limit(limit)
        )
        return list(result.scalars().all())


async def get_user_archive(user_id: int, page: int = 0, page_size: int = SCAN_ARCHIVE_PAGE_SIZE) -> tuple[list[UserScan], int]:
    await archive_expired_scans(user_id)
    page = max(0, int(page))
    async with SessionLocal() as session:
        total = int((await session.execute(
            select(func.count(UserScan.id)).where(
                UserScan.user_id == user_id, UserScan.archived_at.is_not(None)
            )
        )).scalar_one())
        result = await session.execute(
            select(UserScan)
            .where(UserScan.user_id == user_id, UserScan.archived_at.is_not(None))
            .order_by(UserScan.archived_at.desc(), UserScan.created_at.desc())
            .offset(page * page_size)
            .limit(page_size)
        )
        return list(result.scalars().all()), total


async def get_archive_count(user_id: int, *, archive: bool = True) -> int:
    if archive:
        await archive_expired_scans(user_id)
    async with SessionLocal() as session:
        return int((await session.execute(
            select(func.count(UserScan.id)).where(
                UserScan.user_id == user_id, UserScan.archived_at.is_not(None)
            )
        )).scalar_one())


async def get_user_scans_overview(user_id: int, limit: int = 10) -> tuple[list[UserScan], int]:
    """One archive sweep, then load inbox + archive count in parallel."""
    await archive_expired_scans(user_id)
    scans, archive_count = await asyncio.gather(
        get_user_scans(user_id, limit, archive=False),
        get_archive_count(user_id, archive=False),
    )
    return scans, archive_count


async def get_user_popular_categories(user_id: int, limit_scans: int = 100) -> list[tuple[str, UserScan]]:
    """Return one menu entry per category using its latest successful scan only."""
    async with SessionLocal() as session:
        result = await session.execute(
            select(UserScan)
            .where(
                UserScan.user_id == user_id,
                UserScan.status == "done",
            )
            .order_by(UserScan.finished_at.desc(), UserScan.created_at.desc())
            .limit(limit_scans)
        )
        scans = list(result.scalars().all())
    latest: dict[str, UserScan] = {}
    for scan in scans:
        for key in _scan_category_keys(scan):
            if key in CATEGORIES and key not in latest:
                latest[key] = scan
    return list(latest.items())


async def get_user_category_scans(user_id: int, category_key: str) -> list[UserScan]:
    """Return successful scans containing category_key, newest first.

    This remains available for history/admin features. «Популярное сейчас» uses
    only the first item (the latest successful scan).
    """
    if category_key not in CATEGORIES:
        return []
    async with SessionLocal() as session:
        result = await session.execute(
            select(UserScan)
            .where(
                UserScan.user_id == user_id,
                UserScan.status == "done",
            )
            .order_by(UserScan.finished_at.desc(), UserScan.created_at.desc())
        )
        scans = list(result.scalars().all())
    return [scan for scan in scans if category_key in _scan_category_keys(scan)]


async def get_latest_scan_for_category(user_id: int, category_key: str) -> UserScan | None:
    scans = await get_user_category_scans(user_id, category_key)
    return scans[0] if scans else None


async def get_category_scan_rows(user_id: int, category_key: str) -> tuple[list[Listing], list[UserScan]]:
    """Return listings from the latest successful scan of a category only."""
    scan = await get_latest_scan_for_category(user_id, category_key)
    if scan is None:
        return [], []
    async with SessionLocal() as session:
        result = await session.execute(
            select(Listing)
            .join(ScanListing, Listing.external_id == ScanListing.external_id)
            .where(
                ScanListing.scan_id == scan.id,
                Listing.category_key == category_key,
                Listing.is_promoted.is_(False),
                Listing.is_price_reduced.is_(False),
            )
        )
        rows = list(result.scalars().all())
    return rows, [scan]


def _category_scan_dates(scans: list[UserScan], max_items: int = 8) -> str:
    dates = sorted({scan.target_date for scan in scans if scan.target_date}, reverse=True)
    if not dates:
        return "—"
    shown = [_date_label(value) for value in dates[:max_items]]
    if len(dates) > max_items:
        shown.append(f"+ ещё {len(dates) - max_items}")
    return ", ".join(shown)


async def get_scan_rows(scan_id: int) -> list[tuple[Listing, ScanListing]]:
    async with SessionLocal() as session:
        result = await session.execute(
            select(Listing, ScanListing)
            .join(ScanListing, Listing.external_id == ScanListing.external_id)
            .where(
                ScanListing.scan_id == scan_id,
                Listing.is_promoted.is_(False),
                Listing.is_price_reduced.is_(False),
            )
        )
        return list(result.all())


async def ensure_scan_observation_plan(
    scan_id: int, finished_at: datetime | None = None, *, include_baseline: bool = True
) -> None:
    """Create the immediate baseline (new scans) and +3/+6/+12h plan once."""
    async with db_write_lock:
        async with SessionLocal() as session:
            scan = await session.get(UserScan, scan_id)
            if scan is None or scan.status not in {"done", "partial"}:
                return
            # Free-trial launches deliberately demonstrate the live scan itself,
            # but never schedule the paid +3/+6/+12h observation workload.
            if bool(getattr(scan, "is_trial", False)):
                return
            scan_settings = await session.get(UserSettings, int(scan.user_id))
            if scan_settings is None or not bool(getattr(scan_settings, "auto_observations", False)):
                return
            base = finished_at or scan.finished_at or scan.created_at
            # Drop obsolete unfinished +1h/+24h checkpoints left by older versions.
            # Completed history remains available, but it is no longer scheduled or shown.
            await session.execute(
                delete(ScanObservation).where(
                    ScanObservation.scan_id == scan_id,
                    ScanObservation.target_hours.notin_(OBSERVATION_SCHEDULE_HOURS),
                    ScanObservation.status.in_(["pending", "error", "missed"]),
                )
            )
            schedule_hours = OBSERVATION_SCHEDULE_HOURS if include_baseline else OBSERVATION_HOURS
            existing = await session.execute(
                select(ScanObservation.target_hours).where(
                    ScanObservation.scan_id == scan_id,
                    ScanObservation.target_hours.in_(schedule_hours),
                )
            )
            have = {int(x) for x in existing.scalars().all()}
            for hours in schedule_hours:
                if hours in have:
                    continue
                session.add(ScanObservation(
                    scan_id=scan_id,
                    target_hours=hours,
                    due_at=base + timedelta(hours=hours),
                    status="pending",
                ))
            await session.commit()


async def get_scan_observation_statuses(scan_id: int) -> dict[int, str]:
    async with SessionLocal() as session:
        result = await session.execute(
            select(ScanObservation.target_hours, ScanObservation.status)
            .where(ScanObservation.scan_id == scan_id)
            .order_by(ScanObservation.target_hours)
        )
        return {int(hours): status for hours, status in result.all()}


async def cleanup_obsolete_observation_plans() -> int:
    """Remove unfinished +1h/+24h jobs created by pre-v3.2.8 versions."""
    async with db_write_lock:
        async with SessionLocal() as session:
            result = await session.execute(
                delete(ScanObservation).where(
                    ScanObservation.target_hours.notin_(OBSERVATION_SCHEDULE_HOURS),
                    ScanObservation.status != "done",
                )
            )
            await session.commit()
            return int(result.rowcount or 0)


async def cleanup_disabled_observation_plans() -> int:
    """Remove unfinished automatic checkpoints for users who keep auto-measurements off."""
    async with db_write_lock:
        async with SessionLocal() as session:
            disabled_scan_ids = (
                select(UserScan.id)
                .join(UserSettings, UserSettings.user_id == UserScan.user_id)
                .where(UserSettings.auto_observations.is_(False))
            )
            result = await session.execute(
                delete(ScanObservation).where(
                    ScanObservation.scan_id.in_(disabled_scan_ids),
                    ScanObservation.status != "done",
                )
            )
            await session.commit()
            return int(result.rowcount or 0)


async def backfill_recent_observation_plans() -> int:
    """Attach the new plan to still-relevant scans from the last 24h."""
    cutoff = datetime.utcnow() - timedelta(hours=24)
    async with SessionLocal() as session:
        result = await session.execute(
            select(UserScan.id, UserScan.finished_at)
            .where(
                UserScan.status.in_(["done", "partial"]),
                UserScan.finished_at.is_not(None),
                UserScan.finished_at >= cutoff,
            )
        )
        rows = list(result.all())
    for scan_id, finished_at in rows:
        await ensure_scan_observation_plan(int(scan_id), finished_at, include_baseline=False)
    return len(rows)


async def _safe_record_scan_hot(scan_id: int) -> None:
    """Feed today's completed user scans into Radar 3.0 as baseline-only evidence."""
    try:
        seeded = await record_user_scan_radar3_baselines(int(scan_id))
        if seeded:
            log.info("DT Radar 3.0 user-scan baselines scan=%s seeded=%s", scan_id, seeded)
    except Exception:
        log.exception("DT Radar 3.0 user-scan baseline merge failed scan=%s", scan_id)


async def finalize_user_scan(job: "ScanJob", *, cancelled: bool = False) -> None:
    if job.scan_id is None:
        return
    now = datetime.utcnow()
    async with db_write_lock:
        async with SessionLocal() as session:
            scan = await session.get(UserScan, job.scan_id)
            if scan is None:
                return
            scan.status = "cancelled" if cancelled else ("partial" if job.incomplete_categories else "done")
            scan.finished_at = now
            scan.completed_categories = job.completed_categories
            scan.total_categories = len(job.category_keys)
            scan.new_count = job.total_new
            scan.target_complete = bool(not cancelled and job.incomplete_categories == 0)
            scan.scan_note = " | ".join((job.scan_notes or [])[:4])[:500]
            scan.incomplete_category_keys = ",".join(
                key for key in job.category_keys if key in (job.incomplete_category_keys or set())
            )
            quality_scores = [int(x) for x in (job.quality_scores or []) if x is not None]
            scan.quality_score = round(sum(quality_scores) / len(quality_scores)) if quality_scores else 0
            scan.quality_note = " | ".join((job.quality_notes or [])[:4])[:500]
            if not cancelled and job.incomplete_categories == 0:
                scan.last_error = None
            if cancelled:
                await session.commit()
                return

            matched_ids = sorted(job.matched_ids or set())
            if matched_ids:
                result = await session.execute(select(Listing).where(
                    Listing.external_id.in_(matched_ids),
                    Listing.category_key.in_(job.category_keys),
                    Listing.posted_date_msk == job.target_date,
                    Listing.is_promoted.is_(False),
                    Listing.is_price_reduced.is_(False),
                ))
                rows = list(result.scalars().all())
            else:
                rows = []

            # Freeze the user's active parser settings into this saved scan.
            # The raw crawl stays in Listing for analytics, while ScanListing
            # contains only what the user asked to see at scan completion.
            scan_settings = await session.get(UserSettings, job.user_id)
            if scan_settings is None:
                scan_settings = UserSettings(user_id=job.user_id)
            scan_settings.price_filter = job.price_filter or "any"
            effective_settings = scan_settings
            if job.incomplete_categories:
                # v4.8.6: a partial crawl deliberately skips the final views phase.
                # Do not turn hundreds of confirmed rows into a fake zero merely
                # because a min-views filter cannot yet be evaluated. Preserve a
                # provisional snapshot with every non-view filter still applied.
                effective_settings = copy.copy(scan_settings)
                effective_settings.min_views = 0
            rows = apply_listing_settings(
                rows, effective_settings, exact_date_scan=True, apply_output_mode=True
            )

            scan.result_count = len(rows)
            fresh_view_cutoff = now - timedelta(minutes=2)
            fresh_views = {
                row.external_id: int(row.view_count)
                for row in rows
                if row.view_count is not None
                and row.views_checked_at is not None
                and row.views_checked_at >= fresh_view_cutoff
            }
            scan.viewed_count = len(fresh_views)
            scan.last_view_refresh_at = now if scan.viewed_count else None

            await session.execute(delete(ScanListing).where(ScanListing.scan_id == scan.id))
            for row in rows:
                initial_view = fresh_views.get(row.external_id)
                session.add(ScanListing(
                    scan_id=scan.id,
                    external_id=row.external_id,
                    initial_view_count=initial_view,
                    captured_at=now,
                ))
                if initial_view is not None:
                    session.add(ScanViewHistory(
                        scan_id=scan.id,
                        external_id=row.external_id,
                        view_count=int(initial_view),
                        recorded_at=now,
                        target_hours=0,
                    ))
            await session.commit()

    if not cancelled and job.incomplete_categories == 0:
        # v4.10.0 DT Radar: every completed scan contributes its TOP real-view
        # products to the global persistent knowledge base. This is DB-only and
        # never performs additional Kleinanzeigen requests.
        # Never hold the user's completion card on Radar bookkeeping. Missed
        # background merges are recovered by the one-time/history backfill.
        asyncio.create_task(
            _safe_record_scan_hot(job.scan_id),
            name=f"dt-radar-scan-{job.scan_id}",
        )

        # v4.8.6: automatic view observations belong only to a fully confirmed
        # snapshot. A partial scan must not start background view rounds behind
        # the integrity gate.
        await ensure_scan_observation_plan(job.scan_id, now)


async def update_scan_view_refresh(
    scan_id: int, target_hours: int | None = None, *, fresh_after: datetime | None = None
) -> int:
    """Store one real view observation round for a saved scan.

    v3.1.6 never creates a new point from arbitrary stale values already sitting
    in ``listings``. ``fresh_after`` is the start of the current measurement
    (minus the tiny simultaneous-request reuse window).
    """
    now = datetime.utcnow()
    async with db_write_lock:
        async with SessionLocal() as session:
            scan = await session.get(UserScan, scan_id)
            if scan is None:
                return 0
            query = (
                select(Listing.external_id, Listing.view_count, ScanListing)
                .join(ScanListing, Listing.external_id == ScanListing.external_id)
                .where(
                    ScanListing.scan_id == scan_id,
                    Listing.is_promoted.is_(False),
                    Listing.is_price_reduced.is_(False),
                )
            )
            if fresh_after is not None:
                query = query.where(
                    Listing.views_checked_at.is_not(None),
                    Listing.views_checked_at >= fresh_after,
                )
            elif target_hours is not None:
                # Compatibility safety for older callers.
                query = query.where(
                    Listing.views_checked_at.is_not(None),
                    Listing.views_checked_at >= now - timedelta(minutes=2),
                )
            result = await session.execute(query)
            values = list(result.all())
            recorded = 0
            for external_id, view_count, membership in values:
                if view_count is None:
                    continue
                if target_hours == 0 and membership.initial_view_count is None:
                    membership.initial_view_count = int(view_count)
                    membership.captured_at = now
                session.add(ScanViewHistory(
                    scan_id=scan_id,
                    external_id=external_id,
                    view_count=int(view_count),
                    recorded_at=now,
                    target_hours=target_hours,
                ))
                recorded += 1
            if recorded > 0:
                scan.viewed_count = recorded
                scan.last_view_refresh_at = now
            await session.commit()
            return recorded


async def get_scan_history_rounds(scan_id: int, limit: int = 12) -> list[tuple[datetime, int, int]]:
    """Return observation rounds, including v2.8 scan snapshots as a baseline."""
    async with SessionLocal() as session:
        live_result = await session.execute(
            select(
                ScanViewHistory.recorded_at,
                func.count(ScanViewHistory.id),
                func.sum(ScanViewHistory.view_count),
            )
            .join(Listing, Listing.external_id == ScanViewHistory.external_id)
            .where(
                ScanViewHistory.scan_id == scan_id,
                Listing.is_promoted.is_(False),
                Listing.is_price_reduced.is_(False),
            )
            .group_by(ScanViewHistory.recorded_at)
        )
        baseline_result = await session.execute(
            select(
                ScanListing.captured_at,
                func.count(ScanListing.id),
                func.sum(ScanListing.initial_view_count),
            )
            .join(Listing, Listing.external_id == ScanListing.external_id)
            .where(
                ScanListing.scan_id == scan_id,
                ScanListing.initial_view_count.is_not(None),
                Listing.is_promoted.is_(False),
                Listing.is_price_reduced.is_(False),
            )
            .group_by(ScanListing.captured_at)
        )

    merged: dict[datetime, tuple[int, int]] = {}
    for dt, count, total in baseline_result.all():
        merged[dt] = (int(count or 0), int(total or 0))
    # Real v2.9 rounds win on identical timestamps.
    for dt, count, total in live_result.all():
        merged[dt] = (int(count or 0), int(total or 0))
    ordered = sorted(merged.items(), key=lambda item: item[0], reverse=True)[:limit]
    return [(dt, count, total) for dt, (count, total) in ordered]


async def get_latest_manual_growth_summary(scan_id: int) -> tuple[int, int, int]:
    """Return (grew_count, max_delta, total_delta) for the latest real round.

    The comparison is against the immediately previous observation point for
    each listing (or the initial scan snapshot when this is the first refresh).
    """
    pairs = await get_scan_rows(scan_id)
    if not pairs:
        return 0, 0, 0
    baseline = {
        listing.external_id: (snap.captured_at, int(snap.initial_view_count))
        for listing, snap in pairs if snap.initial_view_count is not None
    }
    ids = list(baseline)
    if not ids:
        return 0, 0, 0
    async with SessionLocal() as session:
        result = await session.execute(
            select(
                ScanViewHistory.external_id,
                ScanViewHistory.view_count,
                ScanViewHistory.recorded_at,
            )
            .where(
                ScanViewHistory.scan_id == scan_id,
                ScanViewHistory.external_id.in_(ids),
            )
            .order_by(ScanViewHistory.external_id, ScanViewHistory.recorded_at)
        )
        points = list(result.all())

    by_id: dict[str, list[tuple[datetime, int]]] = {}
    for external_id, (captured_at, initial_views) in baseline.items():
        by_id[external_id] = [(captured_at, initial_views)]
    for external_id, view_count, recorded_at in points:
        by_id.setdefault(external_id, []).append((recorded_at, int(view_count)))

    grew = 0
    max_delta = 0
    total_delta = 0
    for series in by_id.values():
        series.sort(key=lambda item: item[0])
        # Deduplicate an identical timestamp/value pair defensively.
        compact: list[tuple[datetime, int]] = []
        for point in series:
            if not compact or point != compact[-1]:
                compact.append(point)
        if len(compact) < 2:
            continue
        delta = compact[-1][1] - compact[-2][1]
        if delta > 0:
            grew += 1
            total_delta += delta
            max_delta = max(max_delta, delta)
    return grew, max_delta, total_delta


async def get_scan_growth_rows(
    scan_id: int, period_hours: int, category_key: str | None = None
) -> tuple[list[GrowthMetric], int]:
    """Return TOP growth, sorted by real absolute view increase.

    If an automatic checkpoint exists for the requested horizon, compare it to
    the initial scan snapshot. Otherwise fall back to the closest manual history
    so old scans remain useful.
    """
    period_hours = period_hours if period_hours in set(OBSERVATION_HOURS) else 3
    pairs = await get_scan_rows(scan_id)
    if category_key:
        pairs = [p for p in pairs if p[0].category_key == category_key]
    listings = {listing.external_id: listing for listing, _ in pairs}
    if not listings:
        return [], 0

    async with SessionLocal() as session:
        result = await session.execute(
            select(
                ScanViewHistory.external_id,
                ScanViewHistory.view_count,
                ScanViewHistory.recorded_at,
                ScanViewHistory.target_hours,
            )
            .where(
                ScanViewHistory.scan_id == scan_id,
                ScanViewHistory.external_id.in_(list(listings)),
            )
            .order_by(ScanViewHistory.external_id, ScanViewHistory.recorded_at)
        )
        points = list(result.all())

    baseline: dict[str, tuple[datetime, int]] = {}
    for listing, snap in pairs:
        if snap.initial_view_count is not None:
            baseline[listing.external_id] = (snap.captured_at, int(snap.initial_view_count))

    rounds = {snap.captured_at for _, snap in pairs if snap.initial_view_count is not None}
    for _, _, recorded_at, _ in points:
        rounds.add(recorded_at)

    # Prefer the exact scheduled checkpoint for +N hours.
    exact: dict[str, tuple[datetime, int]] = {}
    for external_id, view_count, recorded_at, target_hours in points:
        if target_hours == period_hours:
            exact[external_id] = (recorded_at, int(view_count))

    growth: list[GrowthMetric] = []
    if exact:
        for external_id, (current_at, current_views) in exact.items():
            base = baseline.get(external_id)
            listing = listings.get(external_id)
            if base is None or listing is None:
                continue
            base_at, base_views = base
            elapsed_hours = (current_at - base_at).total_seconds() / 3600
            if elapsed_hours <= 0:
                continue
            delta = current_views - base_views
            if delta <= 0:
                continue
            growth.append(GrowthMetric(
                listing=listing, base_views=base_views, current_views=current_views,
                delta=delta, elapsed_hours=elapsed_hours, per_hour=delta / elapsed_hours,
                observed_at=current_at,
            ))
    else:
        # Compatibility fallback for manual v2.9/v3.0 snapshots.
        by_id: dict[str, list[tuple[datetime, int]]] = {}
        for external_id, (base_at, base_views) in baseline.items():
            by_id.setdefault(external_id, []).append((base_at, base_views))
        for external_id, view_count, recorded_at, _ in points:
            point = (recorded_at, int(view_count))
            series = by_id.setdefault(external_id, [])
            if point not in series:
                series.append(point)
        for external_id, series in by_id.items():
            series.sort(key=lambda point: point[0])
            if len(series) < 2:
                continue
            current_at, current_views = series[-1]
            target_at = current_at - timedelta(hours=period_hours)
            before = [point for point in series[:-1] if point[0] <= target_at]
            base_at, base_views = before[-1] if before else series[0]
            elapsed_hours = (current_at - base_at).total_seconds() / 3600
            if elapsed_hours < (2 / 60):
                continue
            delta = current_views - base_views
            listing = listings.get(external_id)
            if delta <= 0 or listing is None:
                continue
            growth.append(GrowthMetric(
                listing=listing, base_views=base_views, current_views=current_views,
                delta=delta, elapsed_hours=elapsed_hours, per_hour=delta / elapsed_hours,
                observed_at=current_at,
            ))

    # User requested the ranking by actual added views, not by tiny-window velocity.
    growth.sort(key=lambda item: (item.delta, item.per_hour, item.current_views), reverse=True)
    return growth[:GROWTH_TOP_LIMIT], len(rounds)


async def get_category_growth_rows(
    user_id: int, category_key: str, period_hours: int
) -> tuple[list[GrowthMetric], int, int]:
    """Return growth metrics for the latest successful scan of a category only."""
    scan = await get_latest_scan_for_category(user_id, category_key)
    if scan is None:
        return [], 0, 0
    growth, rounds = await get_scan_growth_rows(
        scan.id, period_hours, category_key=category_key
    )
    return growth[:GROWTH_TOP_LIMIT], 1, rounds


def _scan_list_button(scan: UserScan) -> InlineKeyboardButton:
    icon = (
        "✅" if scan.status == "done"
        else "⚠️" if scan.status == "partial"
        else "⏳" if scan.status in {"queued", "running"}
        else "⏹" if scan.status in {"cancelling", "cancelled"}
        else "❌" if scan.status == "failed"
        else "⚪️"
    )
    target_label = _date_label(scan.target_date) if scan.target_date else _moscow_text(scan.finished_at or scan.created_at)[:10]
    label = f"{icon} {scan.title[:22]} · {target_label[:5]}"
    return InlineKeyboardButton(text=label, callback_data=f"scan:{scan.id}")


def my_scans_keyboard(scans: list[UserScan], archive_count: int = 0) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for scan in scans[:SCAN_ARCHIVE_PAGE_SIZE]:
        rows.append([_scan_list_button(scan)])
    if any(scan.status in ARCHIVABLE_SCAN_STATUSES for scan in scans):
        rows.append([InlineKeyboardButton(
            text="🧹 Очистить и переместить в архив", callback_data="archive_my_scans"
        )])
    rows.append([InlineKeyboardButton(text=f"📦 Архив · {archive_count}", callback_data="scan_archive:0")])
    rows.append([InlineKeyboardButton(text="▶️ Новый скан", callback_data="start_scan")])
    rows.append([InlineKeyboardButton(text="🏠 Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def scan_archive_keyboard(scans: list[UserScan], page: int, total: int) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = [[_scan_list_button(scan)] for scan in scans]
    max_page = max(0, (max(0, total - 1)) // SCAN_ARCHIVE_PAGE_SIZE)
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"scan_archive:{page - 1}"))
    nav.append(InlineKeyboardButton(text=f"{page + 1}/{max_page + 1}", callback_data="archive_noop"))
    if page < max_page:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"scan_archive:{page + 1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton(text="⬅️ Мои сканы", callback_data="my_scans")])
    rows.append([InlineKeyboardButton(text="🏠 Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _replace_selected_categories(user_id: int, keys: set[str]) -> set[str]:
    clean = [key for key in CATEGORIES if key in keys][:MAX_SELECTED_CATEGORIES]
    async with SessionLocal() as session:
        await session.execute(delete(SelectedCategory).where(SelectedCategory.user_id == user_id))
        session.add_all([SelectedCategory(user_id=user_id, category_key=key) for key in clean])
        await session.commit()
    return set(clean)


async def toggle_category(user_id: int, key: str) -> tuple[set[str], bool]:
    """Toggle one category without ever allowing a new selection above 2.

    Returns (selected, limit_reached). Removing a category is always allowed.
    Choosing a group-root replaces its child selections; choosing a child replaces
    the root, so those operations do not consume an extra slot unnecessarily.
    """
    cat = CATEGORIES[key]
    selected = await get_selected(user_id)
    updated, limit_reached = toggle_selection(
        selected,
        key,
        is_group=bool(cat.is_group),
        root_key=group_root_key(cat.group),
        child_keys={c.key for c in categories_for_group(cat.group) if not c.is_group},
    )
    if limit_reached:
        return selected, True
    return await _replace_selected_categories(user_id, updated), False


async def toggle_group_children(user_id: int, group_key: str) -> tuple[set[str], bool]:
    """Bulk-select only the remaining free slots, or clear this group's children."""
    child_keys = [c.key for c in categories_for_group(group_key) if not c.is_group]
    selected = await get_selected(user_id)
    updated, limit_reached = bulk_group_selection(
        selected, child_keys, root_key=group_root_key(group_key)
    )
    return await _replace_selected_categories(user_id, updated), limit_reached


async def clear_selected(user_id: int) -> None:
    async with SessionLocal() as session:
        await session.execute(delete(SelectedCategory).where(SelectedCategory.user_id == user_id))
        await session.commit()


def _identity_kwargs(title: str, category: str) -> tuple[ProductIdentity, dict]:
    identity = recognize_product(title, category)
    values = {
        "identity_key": identity.key or None,
        "identity_label": identity.label or None,
        "identity_brand": identity.brand or None,
        "identity_model": identity.model or None,
        "identity_variant": identity.variant or None,
        "identity_product_type": identity.product_type or None,
        "identity_storage_gb": identity.storage_gb,
        "identity_ram_gb": identity.ram_gb,
        "identity_specs": identity.specs or None,
        "identity_confidence": identity.confidence,
    }
    return identity, values


def _apply_identity(row: Listing, title: str, category: str) -> ProductIdentity:
    identity, values = _identity_kwargs(title, category)
    for field, value in values.items():
        setattr(row, field, value)
    return identity


def _identity_display(row: Listing) -> str:
    if (row.identity_confidence or 0) >= 70 and row.identity_label:
        return row.identity_label
    return row.title


def _parse_iso_day(value: str | None):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw[:10], "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _first_seen_moscow_day(value: datetime | None):
    if value is None:
        return None
    aware = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
    return aware.astimezone(MOSCOW).date()


def _resurfaced_listing_reason(row: Listing, incoming_day) -> str:
    """Return strong same-external-id resurrection evidence, never a view-count guess."""
    if incoming_day is None:
        return ""
    original_day = _parse_iso_day(getattr(row, "first_posted_date_msk", None))
    previous_day = _parse_iso_day(getattr(row, "posted_date_msk", None))
    stable_day = original_day or previous_day
    if stable_day is not None and incoming_day > stable_day:
        return "resurfaced_posted_date_shift"
    first_seen_day = _first_seen_moscow_day(getattr(row, "first_seen_at", None))
    # DT cannot have observed a normal listing before its claimed publication day.
    if first_seen_day is not None and incoming_day > first_seen_day:
        return "resurfaced_after_first_seen"
    return ""


def _apply_organic_view_baseline(row: Listing, views: int, measured_at: datetime) -> str:
    """Apply v4.15.7 view-provenance rules to one clean exact measurement.

    Initial counters >=400 are never trusted as demand. They become a baseline and
    need two later clean checkpoints; only the post-baseline delta may then vote.
    """
    transition = apply_organic_measurement(row, int(views), measured_at)
    if transition in {"high_baseline_started", "high_checkpoint_1", "high_verified"}:
        log.info(
            "Verified Organic Velocity external_id=%s transition=%s baseline=%s views=%s checkpoints=%s",
            getattr(row, "external_id", ""), transition,
            getattr(row, "organic_baseline_views", None), int(views),
            int(getattr(row, "organic_verified_checkpoints", 0) or 0),
        )
    return transition


async def backfill_product_identities() -> int:
    """Fill v3.0 identity fields for listings collected by older versions."""
    async with db_write_lock:
        async with SessionLocal() as session:
            result = await session.execute(select(Listing).where(Listing.identity_confidence.is_(None)))
            rows = list(result.scalars().all())
            for row in rows:
                _apply_identity(row, row.title, row.category)
            if rows:
                await session.commit()
            return len(rows)


async def upsert_page_items(
    category_key: str, category_name: str, items: list[ParsedListing]
) -> tuple[list[ParsedListing], int, int, set[str]]:
    """Persist one clean page and flag historical non-organic contamination.

    v4.15.2 keeps paid-promotion and price-reduction flags sticky. A later clean
    card must not make the accumulated views organic again. If a price drops
    between two observations, that listing is immediately excluded even when the
    search-card template temporarily omits the crossed old price.
    """
    if not items:
        return [], 0, 0, set()
    unique = {item.external_id: item for item in items}
    ids = list(unique)
    nonorganic_ids: set[str] = set()
    async with db_write_lock:
        async with SessionLocal() as session:
            await _lock_listing_integrity_ids(session, ids)
            result = await session.execute(select(Listing).where(Listing.external_id.in_(ids)))
            existing = {row.external_id: row for row in result.scalars().all()}
            integrity_rows = list((await session.execute(
                select(ListingIntegrity).where(ListingIntegrity.external_id.in_(ids))
            )).scalars().all())
            integrity = {row.external_id: row for row in integrity_rows}
            now = datetime.utcnow()
            new_items: list[ParsedListing] = []
            enriched_count = 0
            for external_id, item in unique.items():
                row = existing.get(external_id)
                integrity_row = integrity.get(external_id)
                registry_blocked = bool(
                    integrity_row is not None
                    and (bool(integrity_row.is_promoted) or bool(integrity_row.is_price_reduced))
                )
                item_reduced = bool(getattr(item, "is_price_reduced", False))
                if registry_blocked and not item_reduced:
                    nonorganic_ids.add(str(external_id))
                    if row is not None:
                        row.is_promoted = bool(row.is_promoted or integrity_row.is_promoted)
                        row.is_price_reduced = bool(row.is_price_reduced or integrity_row.is_price_reduced)
                    continue
                if row is None:
                    _identity, identity_values = _identity_kwargs(item.title, category_name)
                    initial_posted_day = posted_date_moscow(item.posted_text)
                    first_seen_day = _first_seen_moscow_day(now)
                    initial_history_status = (
                        "trusted_new" if initial_posted_day is not None and initial_posted_day == first_seen_day else "unknown"
                    )
                    session.add(Listing(
                        external_id=item.external_id, category_key=category_key, category=category_name,
                        title=item.title, price_text=item.price_text, price_eur=item.price_eur,
                        posted_text=item.posted_text, posted_date_msk=(posted_date_moscow(item.posted_text).isoformat() if posted_date_moscow(item.posted_text) else None),
                        first_posted_date_msk=(posted_date_moscow(item.posted_text).isoformat() if posted_date_moscow(item.posted_text) else None),
                        url=item.url, first_seen_at=now, last_seen_at=now,
                        is_active=True, is_promoted=False, is_price_reduced=item_reduced, disappeared_at=None,
                        organic_history_status=initial_history_status,
                        **identity_values,
                    ))
                    if item.price_text:
                        session.add(PriceHistory(
                            external_id=item.external_id, price_text=item.price_text,
                            price_eur=item.price_eur, recorded_at=now,
                        ))
                    if item_reduced:
                        if integrity_row is None:
                            integrity_row = ListingIntegrity(
                                external_id=str(external_id), is_price_reduced=True,
                                first_detected_at=now, last_detected_at=now,
                            )
                            session.add(integrity_row)
                            integrity[str(external_id)] = integrity_row
                        else:
                            integrity_row.is_price_reduced = True
                            integrity_row.last_detected_at = now
                        nonorganic_ids.add(str(external_id))
                    else:
                        new_items.append(item)
                    continue

                old_price_text = row.price_text
                old_price_eur = row.price_eur
                incoming_day = posted_date_moscow(item.posted_text)
                if not getattr(row, "first_posted_date_msk", None):
                    # Preserve the oldest defensible publication day for future id-resurrection checks.
                    previous_day = _parse_iso_day(getattr(row, "posted_date_msk", None))
                    first_seen_day = _first_seen_moscow_day(getattr(row, "first_seen_at", None))
                    seed_day = previous_day if previous_day is not None and (first_seen_day is None or previous_day <= first_seen_day) else None
                    row.first_posted_date_msk = seed_day.isoformat() if seed_day is not None else None
                resurrection_reason = _resurfaced_listing_reason(row, incoming_day)
                if (
                    not resurrection_reason
                    and str(getattr(row, "organic_history_status", "unknown") or "unknown") == "unknown"
                    and incoming_day is not None
                    and incoming_day == _first_seen_moscow_day(getattr(row, "first_seen_at", None))
                ):
                    row.organic_history_status = "trusted_new"
                if resurrection_reason:
                    row.is_promoted = True
                    if integrity_row is None:
                        integrity_row = ListingIntegrity(
                            external_id=str(external_id), is_promoted=True, promotion_reason=resurrection_reason,
                            first_detected_at=now, last_detected_at=now,
                        )
                        session.add(integrity_row)
                        integrity[str(external_id)] = integrity_row
                    else:
                        integrity_row.is_promoted = True
                        integrity_row.promotion_reason = resurrection_reason
                        integrity_row.last_detected_at = now
                    nonorganic_ids.add(str(external_id))
                    log.warning(
                        "Bump resurrection detected external_id=%s reason=%s previous=%s incoming=%s first_seen=%s",
                        external_id, resurrection_reason, getattr(row, "first_posted_date_msk", None),
                        incoming_day.isoformat() if incoming_day else "", getattr(row, "first_seen_at", None),
                    )
                already_nonorganic = bool(getattr(row, "is_promoted", False)) or bool(
                    getattr(row, "is_price_reduced", False)
                )
                if item_reduced:
                    row.is_price_reduced = True
                    already_nonorganic = True
                    nonorganic_ids.add(str(external_id))
                    if integrity_row is None:
                        integrity_row = ListingIntegrity(
                            external_id=str(external_id), is_price_reduced=True,
                            first_detected_at=now, last_detected_at=now,
                        )
                        session.add(integrity_row)
                        integrity[str(external_id)] = integrity_row
                    else:
                        integrity_row.is_price_reduced = True
                        integrity_row.last_detected_at = now
                price_dropped = bool(
                    old_price_eur is not None
                    and item.price_eur is not None
                    and int(item.price_eur) < int(old_price_eur)
                )
                if price_dropped:
                    row.is_price_reduced = True
                    integrity_row = integrity.get(str(external_id))
                    if integrity_row is None:
                        integrity_row = ListingIntegrity(
                            external_id=str(external_id), is_price_reduced=True,
                            first_detected_at=now, last_detected_at=now,
                        )
                        session.add(integrity_row)
                        integrity[str(external_id)] = integrity_row
                    else:
                        integrity_row.is_price_reduced = True
                        integrity_row.last_detected_at = now
                    nonorganic_ids.add(str(external_id))
                    already_nonorganic = True

                # Never erase a previously parsed price because of one weak HTML response.
                if item.price_text is not None:
                    if old_price_text is None and item.price_text and not already_nonorganic:
                        enriched_count += 1
                    if (old_price_text, old_price_eur) != (item.price_text, item.price_eur):
                        if old_price_text is not None or old_price_eur is not None:
                            session.add(PriceHistory(
                                external_id=external_id, price_text=old_price_text,
                                price_eur=old_price_eur, recorded_at=now - timedelta(microseconds=1),
                            ))
                        session.add(PriceHistory(
                            external_id=external_id, price_text=item.price_text,
                            price_eur=item.price_eur, recorded_at=now,
                        ))
                    row.price_text = item.price_text
                    row.price_eur = item.price_eur
                row.category_key = category_key
                row.category = category_name
                row.title = item.title
                _apply_identity(row, item.title, category_name)
                row.posted_text = item.posted_text
                parsed_day = posted_date_moscow(item.posted_text)
                if parsed_day and not getattr(row, "first_posted_date_msk", None):
                    row.first_posted_date_msk = parsed_day.isoformat()
                row.posted_date_msk = parsed_day.isoformat() if parsed_day else row.posted_date_msk
                row.url = item.url
                row.last_seen_at = now
                row.is_active = True
                # is_promoted / is_price_reduced are intentionally NOT reset here.
                row.disappeared_at = None
                if bool(getattr(row, "is_promoted", False)) or bool(getattr(row, "is_price_reduced", False)):
                    if str(external_id) not in integrity:
                        integrity_row = ListingIntegrity(
                            external_id=str(external_id),
                            is_promoted=bool(getattr(row, "is_promoted", False)),
                            is_price_reduced=bool(getattr(row, "is_price_reduced", False)),
                            first_detected_at=now, last_detected_at=now,
                        )
                        session.add(integrity_row)
                        integrity[str(external_id)] = integrity_row
                    nonorganic_ids.add(str(external_id))

            await session.commit()

    if nonorganic_ids:
        await purge_nonorganic_analytics(nonorganic_ids)
    clean_known = max(0, (len(unique) - len(new_items)) - len(nonorganic_ids))
    return new_items, clean_known, enriched_count, nonorganic_ids


async def _mark_nonorganic_flag(
    external_ids: list[str] | set[str], *, field: str, reason: str = ""
) -> int:
    ids = {str(x).strip() for x in external_ids if str(x).strip()}
    if not ids:
        return 0
    changed = 0
    now = datetime.utcnow()
    async with db_write_lock:
        async with SessionLocal() as session:
            await _lock_listing_integrity_ids(session, ids)
            result = await session.execute(select(Listing).where(Listing.external_id.in_(list(ids))))
            rows = {row.external_id: row for row in result.scalars().all()}
            integrity_rows = list((await session.execute(
                select(ListingIntegrity).where(ListingIntegrity.external_id.in_(list(ids)))
            )).scalars().all())
            registry = {row.external_id: row for row in integrity_rows}
            for external_id in ids:
                integrity_row = registry.get(external_id)
                if integrity_row is None:
                    integrity_row = ListingIntegrity(
                        external_id=external_id, first_detected_at=now, last_detected_at=now
                    )
                    session.add(integrity_row)
                    registry[external_id] = integrity_row
                integrity_row.last_detected_at = now
                if field == "is_promoted" and reason:
                    integrity_row.promotion_reason = str(reason)[:80]
                if not bool(getattr(integrity_row, field, False)):
                    setattr(integrity_row, field, True)
                    changed += 1
                row = rows.get(external_id)
                if row is not None and not bool(getattr(row, field, False)):
                    setattr(row, field, True)
            await session.commit()
    # Purge even when the flag was already set: a stale Radar/AI signal from an
    # older release may still exist and this operation is idempotent.
    await purge_nonorganic_analytics(ids)
    return changed


async def mark_promoted_listings(
    external_ids: list[str] | set[str], *, reason: str = "search_promotion_marker"
) -> int:
    """Permanently exclude a listing once paid visibility is observed."""
    return await _mark_nonorganic_flag(external_ids, field="is_promoted", reason=reason)


async def mark_price_reduced_listings(external_ids: list[str] | set[str]) -> int:
    """Permanently exclude a listing once crossed/reduced pricing is observed."""
    return await _mark_nonorganic_flag(external_ids, field="is_price_reduced")


def berlin_date_key() -> str:
    """Legacy worker-day key kept in Moscow time to match scan date selection.

    The helper existed conceptually in the category-state/statistics code, but was
    accidentally dropped during the v4 refactor.  Missing it made a completely
    successful category scan crash *after* all pages had already been parsed.
    """
    return datetime.now(MOSCOW).date().isoformat()


def berlin_today_utc_bounds() -> tuple[datetime, datetime]:
    now = datetime.now(MOSCOW)
    start_local = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_local = start_local + timedelta(days=1)
    return (
        start_local.astimezone(timezone.utc).replace(tzinfo=None),
        end_local.astimezone(timezone.utc).replace(tzinfo=None),
    )


async def today_rows(*, include_price_reduced: bool = False) -> list[Listing]:
    start_utc, end_utc = berlin_today_utc_bounds()
    conditions = [
        Listing.first_seen_at >= start_utc, Listing.first_seen_at < end_utc,
        Listing.is_promoted.is_(False),
    ]
    if not include_price_reduced:
        conditions.append(Listing.is_price_reduced.is_(False))
    async with SessionLocal() as session:
        result = await session.execute(select(Listing).where(*conditions))
        return list(result.scalars().all())


async def filtered_rows(user_id: int) -> tuple[UserSettings, list[Listing]]:
    s = await get_settings(user_id)
    rows = await today_rows()
    rows = base_filter(
        rows, period=None, price_filter=s.price_filter, clean_noise=s.clean_noise,
        include_words=s.include_words or "", exclude_words=s.exclude_words or "",
    )
    if s.smart_dedupe:
        rows = dedupe_rows(rows)
    if s.output_mode == "unique":
        rows = unique_rows(rows)
    return s, sort_rows(rows, s.sort_mode)


def _temp_csv(name: str) -> tuple[Path, csv.writer, object]:
    temp_dir = Path(tempfile.mkdtemp(prefix="kleinanzeigen_"))
    path = temp_dir / name
    f = path.open("w", encoding="utf-8-sig", newline="")
    return path, csv.writer(f, delimiter=";"), f


def _price_display(price_text: str | None, price_eur: int | None) -> str:
    if price_text:
        return price_text
    if price_eur is not None:
        return f"{price_eur} €"
    return "—"


def _moscow_text(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.replace(tzinfo=timezone.utc).astimezone(MOSCOW).strftime("%d.%m.%Y %H:%M")


def _berlin_text(dt: datetime | None) -> str:
    # Backward-compatible helper name; all user-facing timestamps are Moscow time.
    return _moscow_text(dt)


def _moscow_today_iso() -> str:
    return datetime.now(MOSCOW).date().isoformat()


def _date_label(value: str | None) -> str:
    if not value:
        return "—"
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return value


def _parse_scan_date_input(text: str | None) -> str | None:
    raw = (text or "").strip().replace("/", ".").replace("-", ".")
    today = datetime.now(MOSCOW).date()
    try:
        if re.fullmatch(r"\d{1,2}", raw):
            day = int(raw)
            value = today.replace(day=day)
        elif re.fullmatch(r"\d{1,2}\.\d{1,2}", raw):
            day, month = map(int, raw.split("."))
            value = datetime(today.year, month, day).date()
        elif re.fullmatch(r"\d{1,2}\.\d{1,2}\.\d{4}", raw):
            value = datetime.strptime(raw, "%d.%m.%Y").date()
        else:
            return None
    except ValueError:
        return None
    if value > today:
        return None
    if value < today - timedelta(days=DATE_MAX_AGE_DAYS):
        return None
    return value.isoformat()


def _export_group_meta(row) -> tuple[int, int, str]:
    """Return section/category order for a result row without changing its inner ranking."""
    group_order = {key: i for i, key in enumerate(GROUPS)}
    category_order = {key: i for i, key in enumerate(CATEGORIES)}

    category_key = getattr(row, "category_key", None)
    cat = CATEGORIES.get(category_key) if category_key else None
    if cat is None:
        category_name = str(getattr(row, "category", "") or "")
        cat = next((item for item in CATEGORIES.values() if item.name == category_name), None)

    if cat is None:
        return (len(group_order) + 1, len(category_order) + 1, "")
    return (
        group_order.get(cat.group, len(group_order)),
        category_order.get(cat.key, len(category_order)),
        cat.group,
    )


def _group_export_rows(rows: list) -> list:
    """Group multi-section exports as section -> subcategory, preserving ranking inside each subcategory."""
    decorated = list(enumerate(rows))
    decorated.sort(key=lambda pair: (*_export_group_meta(pair[1])[:2], pair[0]))
    return [row for _, row in decorated]


def _export_section_name(row) -> str:
    """Human-readable top-level Kleinanzeigen section for universal XLSX exports."""
    category_key = getattr(row, "category_key", None)
    cat = CATEGORIES.get(category_key) if category_key else None
    if cat is None:
        category_name = str(getattr(row, "category", "") or "")
        cat = next((item for item in CATEGORIES.values() if item.name == category_name and not item.is_group), None)
    if cat is None:
        return "Другое"
    group = GROUPS.get(cat.group)
    return group.name if group is not None else cat.group


def _write_universal_xlsx(
    filename: str,
    title: str,
    headers: list[str],
    data_rows: list[list],
    *,
    hyperlink_headers: set[str] | None = None,
    integer_headers: set[str] | None = None,
    decimal_headers: set[str] | None = None,
) -> Path:
    """Create a mobile-friendly XLSX that renders consistently across iOS/Android/desktop."""
    out_dir = Path(tempfile.mkdtemp(prefix="kleinanzeigen_xlsx_"))
    path = out_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    max_col = max(1, len(headers))
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34

    for values in data_rows:
        ws.append(values)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=max_col).column_letter}{max(2, ws.max_row)}"
    ws.sheet_view.showGridLines = False

    hyperlink_headers = hyperlink_headers or set()
    integer_headers = integer_headers or set()
    decimal_headers = decimal_headers or set()
    header_to_col = {str(cell.value): cell.column for cell in ws[2] if cell.value is not None}

    for header in hyperlink_headers:
        col = header_to_col.get(header)
        if not col:
            continue
        for row_num in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    for header in integer_headers:
        col = header_to_col.get(header)
        if col:
            for row_num in range(3, ws.max_row + 1):
                ws.cell(row=row_num, column=col).number_format = "0"
    for header in decimal_headers:
        col = header_to_col.get(header)
        if col:
            for row_num in range(3, ws.max_row + 1):
                ws.cell(row=row_num, column=col).number_format = "0.00"

    # Make section changes obvious without inserting blank rows that would break filters.
    section_col = header_to_col.get("Раздел")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    last_section = None
    for row_num in range(3, ws.max_row + 1):
        current_section = ws.cell(row=row_num, column=section_col).value if section_col else None
        if section_col and current_section != last_section:
            for col_num in range(1, max_col + 1):
                ws.cell(row=row_num, column=col_num).fill = section_fill
                ws.cell(row=row_num, column=col_num).font = Font(bold=True)
        last_section = current_section
        for col_num in range(1, max_col + 1):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(vertical="top", wrap_text=True)

    width_by_header = {
        "#": 6, "Раздел": 28, "Категория": 28, "Название": 48, "Товар": 48,
        "Группа товара": 34, "Пример названия": 46, "Цена": 15, "Цена, €": 12,
        "Цена примера": 15, "👁 Просмотры": 15, "Публикаций": 13, "Мин. цена, €": 14,
        "Медиана, €": 14, "Макс. цена, €": 14, "Точность группы, %": 18,
        "Последнее": 18, "Дата (МСК)": 15, "Как показано на Kleinanzeigen": 25,
        "Медиана группы, €": 18, "Ниже медианы, %": 18, "Образцов": 12,
        "Точность": 15, "Время жизни, мин": 18, "Окно проверки, мин": 19,
        "Впервые замечено": 19, "Обнаружено исчезновение": 22, "Старая цена, €": 15,
        "Новая цена, €": 15, "Снижение, €": 14, "Снижение, %": 14,
        "Зафиксировано": 18, "Ссылка": 48, "Ссылка-пример": 48,
    }
    for cell in ws[2]:
        width = width_by_header.get(str(cell.value), 16)
        ws.column_dimensions[cell.column_letter].width = width

    # Keep rows compact enough for phone viewers while still allowing wrapped titles.
    for row_num in range(3, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 32

    wb.save(path)
    return path


def write_listing_xlsx(rows: list[Listing], mode: str) -> Path:
    now = datetime.now(MOSCOW)
    headers = [
        "Раздел", "Категория", "Название", "Цена, €", "👁 Просмотры",
        "Дата (МСК)", "Как показано на Kleinanzeigen", "Ссылка",
    ]
    data = [[
        _export_section_name(row), row.category, row.title,
        row.price_eur if row.price_eur is not None else None,
        row.view_count if row.view_count is not None else None,
        _date_label(row.posted_date_msk), row.posted_text or "", row.url,
    ] for row in rows]
    return _write_universal_xlsx(
        f"kleinanzeigen_{mode}_{now:%Y-%m-%d_%H-%M}.xlsx",
        f"Kleinanzeigen · {MODE_LABELS.get(mode, mode)} · {len(rows)} объявлений",
        headers, data,
        hyperlink_headers={"Ссылка"}, integer_headers={"Цена, €", "👁 Просмотры"},
    )


def write_frequent_xlsx(rows) -> Path:
    now = datetime.now(MOSCOW)
    headers = [
        "Раздел", "Категория", "Группа товара", "Пример названия", "Цена примера",
        "Публикаций", "Мин. цена, €", "Медиана, €", "Макс. цена, €",
        "Точность группы, %", "Последнее", "Ссылка-пример",
    ]
    data = [[
        _export_section_name(row), row.category, row.product_key, row.example_title,
        row.example_price_text or "—", row.count, row.min_price, row.median_price,
        row.max_price, row.confidence, row.newest_posted, row.example_url,
    ] for row in rows]
    return _write_universal_xlsx(
        f"kleinanzeigen_chasto_publikuemye_{now:%Y-%m-%d_%H-%M}.xlsx",
        f"Kleinanzeigen · Часто публикуемые · {len(rows)} групп",
        headers, data, hyperlink_headers={"Ссылка-пример"},
        integer_headers={"Публикаций", "Мин. цена, €", "Медиана, €", "Макс. цена, €", "Точность группы, %"},
    )


def write_market_xlsx(rows) -> Path:
    now = datetime.now(MOSCOW)
    headers = [
        "Раздел", "Категория", "Название", "Цена, €", "👁 Просмотры",
        "Медиана группы, €", "Ниже медианы, %", "Образцов", "Точность группы, %", "Дата", "Ссылка",
    ]
    data = [[
        _export_section_name(row), row.category, row.title,
        row.price_eur, getattr(row, "view_count", None), row.median_price,
        row.discount_pct, row.samples, row.confidence, row.posted_text, row.url,
    ] for row in rows]
    return _write_universal_xlsx(
        f"kleinanzeigen_nizhe_rynka_{now:%Y-%m-%d_%H-%M}.xlsx",
        f"Kleinanzeigen · Ниже рынка · {len(rows)} объявлений",
        headers, data, hyperlink_headers={"Ссылка"},
        integer_headers={"Цена, €", "👁 Просмотры", "Медиана группы, €", "Образцов", "Точность группы, %"},
        decimal_headers={"Ниже медианы, %"},
    )


def write_disappearing_xlsx(rows) -> Path:
    now = datetime.now(MOSCOW)
    headers = [
        "Раздел", "Категория", "Название", "Цена", "Время жизни, мин",
        "Окно проверки, мин", "Точность", "Впервые замечено", "Обнаружено исчезновение", "Ссылка",
    ]
    data = [[
        _export_section_name(row), row.category, row.title, row.price_text or "—",
        row.lifespan_minutes, row.detection_gap_minutes, row.confidence,
        _berlin_text(row.first_seen_at), _berlin_text(row.disappeared_at), row.url,
    ] for row in rows]
    return _write_universal_xlsx(
        f"kleinanzeigen_bystro_ischezayushchie_{now:%Y-%m-%d_%H-%M}.xlsx",
        f"Kleinanzeigen · Быстро исчезающие · {len(rows)} объявлений",
        headers, data, hyperlink_headers={"Ссылка"},
        integer_headers={"Время жизни, мин", "Окно проверки, мин"},
    )


def write_price_drop_xlsx(rows) -> Path:
    now = datetime.now(MOSCOW)
    headers = [
        "Раздел", "Категория", "Название", "Старая цена, €", "Новая цена, €",
        "Снижение, €", "Снижение, %", "Зафиксировано", "Ссылка",
    ]
    data = [[
        _export_section_name(row), row.category, row.title, row.previous_price, row.current_price,
        row.drop_eur, row.drop_pct, _berlin_text(row.changed_at), row.url,
    ] for row in rows]
    return _write_universal_xlsx(
        f"kleinanzeigen_snizhenie_ceny_{now:%Y-%m-%d_%H-%M}.xlsx",
        f"Kleinanzeigen · Снижение цены · {len(rows)} объявлений",
        headers, data, hyperlink_headers={"Ссылка"},
        integer_headers={"Старая цена, €", "Новая цена, €", "Снижение, €"},
        decimal_headers={"Снижение, %"},
    )


def write_listing_csv(rows: list[Listing], mode: str) -> Path:
    now = datetime.now(MOSCOW)
    path, writer, f = _temp_csv(f"kleinanzeigen_{mode}_{now:%Y-%m-%d_%H-%M}.csv")
    try:
        writer.writerow([
            "Категория", "Название", "Цена, €", "👁 Просмотры",
            "Дата (МСК)", "Как показано на Kleinanzeigen", "Ссылка"
        ])
        for row in rows:
            writer.writerow([
                row.category, row.title,
                row.price_eur if row.price_eur is not None else "",
                row.view_count if row.view_count is not None else "",
                _date_label(row.posted_date_msk), row.posted_text or "", row.url,
            ])
    finally:
        f.close()
    return path


def write_frequent_csv(rows) -> Path:
    now = datetime.now(MOSCOW)
    path, writer, f = _temp_csv(f"kleinanzeigen_chasto_publikuemye_{now:%Y-%m-%d_%H-%M}.csv")
    try:
        writer.writerow([
            "Категория", "Группа товара", "Пример названия", "Цена примера",
            "Публикаций", "Мин. цена, €", "Медиана, €", "Макс. цена, €",
            "Точность группы, %", "Последнее", "Ссылка-пример",
        ])
        for row in rows:
            writer.writerow([
                row.category, row.product_key, row.example_title, row.example_price_text or "—", row.count,
                row.min_price if row.min_price is not None else "",
                row.median_price if row.median_price is not None else "",
                row.max_price if row.max_price is not None else "",
                row.confidence, row.newest_posted, row.example_url,
            ])
    finally:
        f.close()
    return path


def write_market_csv(rows) -> Path:
    now = datetime.now(MOSCOW)
    path, writer, f = _temp_csv(f"kleinanzeigen_nizhe_rynka_{now:%Y-%m-%d_%H-%M}.csv")
    try:
        writer.writerow([
            "Категория", "Название", "Цена, €", "👁 Просмотры", "Медиана группы, €",
            "Ниже медианы, %", "Образцов", "Точность группы, %", "Дата", "Ссылка",
        ])
        for row in rows:
            writer.writerow([
                row.category, row.title, row.price_eur,
                getattr(row, "view_count", None) if getattr(row, "view_count", None) is not None else "",
                row.median_price, row.discount_pct, row.samples, row.confidence, row.posted_text, row.url,
            ])
    finally:
        f.close()
    return path


def write_disappearing_csv(rows) -> Path:
    now = datetime.now(MOSCOW)
    path, writer, f = _temp_csv(f"kleinanzeigen_bystro_ischezayushchie_{now:%Y-%m-%d_%H-%M}.csv")
    try:
        writer.writerow([
            "Категория", "Название", "Цена", "Время жизни, мин",
            "Окно проверки, мин", "Точность", "Впервые замечено", "Обнаружено исчезновение", "Ссылка",
        ])
        for row in rows:
            writer.writerow([
                row.category, row.title, row.price_text or "—", row.lifespan_minutes,
                row.detection_gap_minutes, row.confidence,
                _berlin_text(row.first_seen_at), _berlin_text(row.disappeared_at), row.url,
            ])
    finally:
        f.close()
    return path


def write_price_drop_csv(rows) -> Path:
    now = datetime.now(MOSCOW)
    path, writer, f = _temp_csv(f"kleinanzeigen_snizhenie_ceny_{now:%Y-%m-%d_%H-%M}.csv")
    try:
        writer.writerow([
            "Категория", "Название", "Старая цена, €", "Новая цена, €",
            "Снижение, €", "Снижение, %", "Зафиксировано", "Ссылка",
        ])
        for row in rows:
            writer.writerow([
                row.category, row.title, row.previous_price, row.current_price,
                row.drop_eur, row.drop_pct, _berlin_text(row.changed_at), row.url,
            ])
    finally:
        f.close()
    return path


async def histories_for(rows: list[Listing]) -> list[PriceHistory]:
    ids = [row.external_id for row in rows]
    if not ids:
        return []
    async with SessionLocal() as session:
        result = await session.execute(
            select(PriceHistory).where(PriceHistory.external_id.in_(ids)).order_by(PriceHistory.recorded_at.asc())
        )
        return list(result.scalars().all())


async def refresh_availability(rows: list[Listing]) -> tuple[int, int, int]:
    """Check a bounded batch of tracked public ad links for availability.

    Returns checked, newly_disappeared, unknown. This is intentionally bounded
    and low-concurrency so the analytics mode does not hammer the site.
    """
    candidates = [r for r in rows if r.is_active and r.url][:AVAILABILITY_CHECK_LIMIT]
    if not candidates:
        return 0, 0, 0

    parser = KleinanzeigenParser()
    sem = asyncio.Semaphore(AVAILABILITY_CONCURRENCY)

    async def check(row: Listing):
        async with sem:
            result = await parser.check_listing_active(row.url)
            return row.external_id, result

    try:
        results = await asyncio.gather(*(check(row) for row in candidates))
    finally:
        await parser.close()

    disappeared_ids = [external_id for external_id, active in results if active is False]
    unknown = sum(1 for _, active in results if active is None)
    if disappeared_ids:
        now = datetime.utcnow()
        async with SessionLocal() as session:
            result = await session.execute(select(Listing).where(Listing.external_id.in_(disappeared_ids)))
            found = list(result.scalars().all())
            for row in found:
                if row.is_active:
                    row.is_active = False
                    row.disappeared_at = now
            await session.commit()
    return len(candidates), len(disappeared_ids), unknown


async def fetch_exact_views_v438_compatible(
    parser: KleinanzeigenParser,
    urls: list[str],
    *,
    concurrency: int,
    progress_cb=None,
    traffic_priority: str,
) -> dict[str, ViewCountResult]:
    """Use the remote fleet first and locally recover only missing shards.

    v4.4.0 keeps every exact result already produced by healthy View Workers.
    If one Redis shard times out or gets handed back after refusals, only URLs
    absent from the remote merge run through the proven local view path.
    """
    urls = list(dict.fromkeys(urls))
    if REMOTE_VIEW_WORKER_ENABLED:
        remote_reported = 0

        async def remote_progress(done: int, total: int) -> None:
            nonlocal remote_reported
            remote_reported = max(remote_reported, min(len(urls), max(0, int(done))))
            if progress_cb is not None:
                maybe = progress_cb(remote_reported, len(urls))
                if asyncio.iscoroutine(maybe):
                    await maybe

        remote = await REMOTE_VIEW_MANAGER.fetch(
            urls, progress_cb=remote_progress, traffic_priority=traffic_priority,
        )
        if remote is not None:
            converted = {
                url: ViewCountResult(
                    item.views, item.raw_text, item.source, item.final_url, item.page_title, item.error
                )
                for url, item in remote.items()
            }
            missing = [url for url in urls if url not in converted]
            if not missing:
                return {url: converted[url] for url in urls if url in converted}

            log.warning(
                "Dedicated view partial fallback remote=%s/%s missing=%s; retrying only missing URLs locally",
                len(converted), len(urls), len(missing),
            )

            async def local_progress(done: int, total: int) -> None:
                # Keep Telegram progress monotonic even if a failed remote shard
                # had reported partial work before being retried locally.
                combined = max(remote_reported, min(len(urls), len(converted) + max(0, int(done))))
                if progress_cb is not None:
                    maybe = progress_cb(combined, len(urls))
                    if asyncio.iscoroutine(maybe):
                        await maybe

            local = await parser.fetch_public_view_counts(
                missing, concurrency=concurrency, progress_cb=local_progress,
                traffic_priority=traffic_priority, browser_fallback=True,
                direct_http_only=False, accurate=True,
            )
            converted.update(local)
            if progress_cb is not None:
                maybe = progress_cb(len(urls), len(urls))
                if asyncio.iscoroutine(maybe):
                    await maybe
            return {url: converted[url] for url in urls if url in converted}

        log.warning("Dedicated view worker unavailable; using local v4.3.8 view path")
    return await parser.fetch_public_view_counts(
        urls, concurrency=concurrency, progress_cb=progress_cb,
        traffic_priority=traffic_priority, browser_fallback=True,
        direct_http_only=False, accurate=True,
    )


async def enrich_page_view_counts(
    parser: KleinanzeigenParser,
    items: list[ParsedListing],
    live: CategoryLiveProgress | None = None,
) -> tuple[int, int, int]:
    """Fetch public view counters as part of the category-page pipeline.

    Only missing/stale counters are opened. The same Playwright browser is reused
    by the category parser, and the passive s-vac-inc-get response is preferred.
    Returns (requested, updated, failed).
    """
    if not items:
        return 0, 0, 0

    unique = {item.external_id: item for item in items if item.url}
    if not unique:
        return 0, 0, 0

    async with SessionLocal() as session:
        result = await session.execute(select(Listing).where(
            Listing.external_id.in_(list(unique)),
            Listing.is_promoted.is_(False),
            Listing.is_price_reduced.is_(False),
        ))
        rows = {row.external_id: row for row in result.scalars().all()}
        # Accurate Views Core always verifies every target in the current
        # measurement. A previously stored value is never allowed to satisfy a
        # 50+/100+ filter without a fresh verified read.
        targets = [unique[eid] for eid in rows if eid in unique]

    if not targets:
        return 0, 0, 0

    # v4.2.1: live progress must reflect completed counter requests while the
    # batch is running. v4.2.0 updated ``views_ready`` only after the entire
    # batch returned, so Telegram could sit at 0/575 for several minutes even
    # though Railway logs showed hundreds of successful 200 responses.
    base_ready = int(live.views_ready or 0) if live is not None else 0
    reused_count = 0
    if live is not None:
        live.views_ready = base_ready

    async def live_progress(done: int, total: int) -> None:
        if live is None:
            return
        live.views_ready = base_ready + reused_count + min(max(0, int(done)), len(targets))
        if done == total or (done > 0 and done % 25 == 0):
            log.info(
                "View progress category=%s checked=%s/%s failed_so_far=%s",
                live.category_name, live.views_ready, len(unique), live.views_failed,
            )

    results = await fetch_exact_views_v438_compatible(
        parser, [item.url for item in targets],
        concurrency=VIEW_COUNT_CONCURRENCY,
        progress_cb=live_progress,
        traffic_priority="scan_inline",
    )

    source_counts = Counter(vr.source for vr in results.values())
    log.info(
        "Accurate views batch category=%s total=%s sources=%s",
        (live.category_name if live is not None else "manual"), len(results), dict(source_counts),
    )

    now = datetime.utcnow()
    updated = 0
    failed = 0
    url_to_id = {item.url: item.external_id for item in targets}
    async with db_write_lock:
        async with SessionLocal() as session:
            db_result = await session.execute(select(Listing).where(
                Listing.external_id.in_([item.external_id for item in targets]),
                Listing.is_promoted.is_(False),
                Listing.is_price_reduced.is_(False),
            ))
            db_rows = {row.external_id: row for row in db_result.scalars().all()}
            for item in targets:
                url = item.url
                external_id = url_to_id.get(url)
                row = db_rows.get(external_id) if external_id else None
                if row is None:
                    continue
                vr = results.get(url)
                if vr is None or vr.views is None:
                    failed += 1
                    # Fail closed: an omitted result is the same as an explicit
                    # unknown. A stale number from an earlier measurement must
                    # never survive a failed current verification and pass filters.
                    row.view_count = None
                    row.views_checked_at = now
                    continue
                old = row.view_count
                row.view_count = int(vr.views)
                row.views_checked_at = now
                _apply_organic_view_baseline(row, row.view_count, now)
                if old != row.view_count:
                    session.add(ViewHistory(
                        external_id=row.external_id,
                        view_count=row.view_count,
                        recorded_at=now,
                    ))
                updated += 1
            await session.commit()

    if live is not None:
        # ``views_ready`` is a checked/progress counter in the user-facing card.
        # Failed counters are tracked separately but still count as completed work,
        # otherwise progress could stop forever below 100%.
        live.views_ready = base_ready + len(unique)
        live.views_failed += failed
    return len(targets), updated, failed


async def enrich_autoscan_view_counts(
    parser: KleinanzeigenParser,
    items: list[ParsedListing],
    live: CategoryLiveProgress | None = None,
) -> tuple[int, int, int]:
    """Collect exact Radar counters with the dedicated View Worker fleet first.

    v4.22.6 removes the old "Parser HTTP pass over every ad -> View Worker only for
    misses" bottleneck when the fleet is healthy. The whole exact batch is sharded
    across the dedicated View Workers immediately. The proven local parser remains a
    bounded fallback, and a *small* unresolved tail gets one final targeted exact
    retry. Every unresolved value is persisted as NULL; no approximate/stale counter
    can enter Radar.
    """
    if not items:
        return 0, 0, 0

    unique = {item.external_id: item for item in items if item.url}
    if not unique:
        return 0, 0, 0

    async with SessionLocal() as session:
        result = await session.execute(select(Listing).where(
            Listing.external_id.in_(list(unique)),
            Listing.is_promoted.is_(False),
            Listing.is_price_reduced.is_(False),
        ))
        rows = {row.external_id: row for row in result.scalars().all()}
        targets = [unique[eid] for eid in rows if eid in unique]

    if not targets:
        return 0, 0, 0

    urls = [item.url for item in targets]
    url_to_id = {item.url: item.external_id for item in targets}
    total = len(urls)
    base_ready = int(live.views_ready or 0) if live is not None else 0
    autoscan_view_priority = "scan_inline"
    autoscan_view_concurrency = max(1, min(RADAR_AUTOSCAN_SAFE_VIEW_CONCURRENCY, VIEW_COUNT_CONCURRENCY))

    async def fleet_progress(done: int, _total: int) -> None:
        if live is None:
            return
        live.views_ready = base_ready + min(total, max(0, int(done)))
        if done == _total or (done > 0 and done % 50 == 0):
            log.info(
                "Radar AutoScan views fleet category=%s checked=%s/%s",
                live.category_name, min(total, int(done)), total,
            )

    remote_ready = False
    if REMOTE_VIEW_WORKER_ENABLED:
        try:
            remote_ready = await REMOTE_VIEW_MANAGER.worker_alive()
        except Exception:
            remote_ready = False

    combined: dict[str, ViewCountResult] = {}
    mode = "remote-fleet-first" if remote_ready else "direct-safe-fallback"
    log.info(
        "Radar AutoScan views start category=%s total=%s mode=%s priority=%s concurrency=%s",
        (live.category_name if live is not None else "autoscan"), total, mode,
        autoscan_view_priority, autoscan_view_concurrency,
    )

    try:
        if remote_ready:
            combined = await asyncio.wait_for(
                fetch_exact_views_v438_compatible(
                    parser,
                    urls,
                    concurrency=autoscan_view_concurrency,
                    progress_cb=fleet_progress,
                    traffic_priority=autoscan_view_priority,
                ),
                timeout=RADAR_AUTOSCAN_VIEW_RECOVERY_TIMEOUT_SECONDS,
            )
        else:
            # Do not open hundreds of browser fallbacks inside Parser when the fleet
            # is offline. The verified official counter remains a cheap safe salvage
            # path; any remaining tail stays UNKNOWN and is judged by coverage below.
            combined = await parser.fetch_public_view_counts(
                urls,
                concurrency=autoscan_view_concurrency,
                progress_cb=fleet_progress,
                traffic_priority=autoscan_view_priority,
                browser_fallback=False,
                direct_http_only=True,
                accurate=False,
                batch_size=80,
                batch_pause_seconds=0.05,
            )
    except asyncio.TimeoutError:
        log.warning(
            "Radar AutoScan fleet watchdog category=%s total=%s timeout=%ss",
            (live.category_name if live is not None else "autoscan"), total,
            int(RADAR_AUTOSCAN_VIEW_RECOVERY_TIMEOUT_SECONDS),
        )
        combined = combined or {}

    # Fleet output is authoritative when present, but an explicit UNKNOWN or a
    # timed-out/missing shard may still be recoverable through the cheap verified
    # official counter. Run that salvage only for the unresolved subset, never for
    # URLs the fleet already completed. This preserves partial work and prevents a
    # single bad shard from turning into a complete category replay.
    unresolved = [url for url in urls if combined.get(url) is None or combined[url].views is None]
    direct_salvaged = 0
    if unresolved and remote_ready:
        log.info(
            "Radar AutoScan direct salvage category=%s unresolved=%s/%s",
            (live.category_name if live is not None else "autoscan"), len(unresolved), total,
        )
        try:
            salvage = await parser.fetch_public_view_counts(
                unresolved,
                concurrency=autoscan_view_concurrency,
                progress_cb=None,
                traffic_priority=autoscan_view_priority,
                browser_fallback=False,
                direct_http_only=True,
                accurate=False,
                batch_size=80,
                batch_pause_seconds=0.05,
            )
            for url, result in salvage.items():
                if result is not None and result.views is not None:
                    combined[url] = result
                    direct_salvaged += 1
        except Exception:
            log.warning(
                "Radar AutoScan direct salvage failed category=%s unresolved=%s",
                (live.category_name if live is not None else "autoscan"), len(unresolved),
                exc_info=True,
            )

    unresolved = [url for url in urls if combined.get(url) is None or combined[url].views is None]
    tail_recovered = 0
    # A handful of transient misses should never cause a 20-page category replay.
    # Retry only that exact tail once through the proven accurate path. Large tails
    # are intentionally not browser-flooded here; they remain fail-closed/retryable.
    if unresolved and len(unresolved) <= RADAR_AUTOSCAN_VIEW_TAIL_RETRY_MAX:
        log.info(
            "Radar AutoScan exact tail retry category=%s unresolved=%s/%s",
            (live.category_name if live is not None else "autoscan"), len(unresolved), total,
        )
        try:
            tail = await parser.fetch_public_view_counts(
                unresolved,
                concurrency=min(autoscan_view_concurrency, len(unresolved)),
                progress_cb=None,
                traffic_priority=autoscan_view_priority,
                browser_fallback=True,
                direct_http_only=False,
                accurate=True,
            )
            for url, result in tail.items():
                if result is not None and result.views is not None:
                    combined[url] = result
                    tail_recovered += 1
        except Exception:
            log.warning(
                "Radar AutoScan exact tail retry failed category=%s unresolved=%s",
                (live.category_name if live is not None else "autoscan"), len(unresolved),
                exc_info=True,
            )

    now = datetime.utcnow()
    updated = 0
    failed = 0
    source_counts = Counter()
    async with db_write_lock:
        async with SessionLocal() as session:
            db_result = await session.execute(
                select(Listing).where(
                    Listing.external_id.in_([item.external_id for item in targets]),
                    Listing.is_promoted.is_(False),
                    Listing.is_price_reduced.is_(False),
                )
            )
            db_rows = {row.external_id: row for row in db_result.scalars().all()}
            for url in urls:
                external_id = url_to_id.get(url)
                row = db_rows.get(external_id) if external_id else None
                if row is None:
                    continue
                vr = combined.get(url)
                if vr is None or vr.views is None:
                    failed += 1
                    row.view_count = None
                    row.views_checked_at = now
                    source_counts["unknown"] += 1
                    continue
                source_counts[str(vr.source or "unknown")] += 1
                old = row.view_count
                row.view_count = int(vr.views)
                row.views_checked_at = now
                _apply_organic_view_baseline(row, row.view_count, now)
                if old != row.view_count:
                    session.add(ViewHistory(
                        external_id=row.external_id,
                        view_count=row.view_count,
                        recorded_at=now,
                    ))
                updated += 1
            await session.commit()

    if live is not None:
        # Progress means every target was checked, not that every target succeeded.
        live.views_ready = base_ready + total
        live.views_failed += failed
    log.info(
        "Radar AutoScan views complete category=%s total=%s exact=%s failed=%s coverage=%.2f%% mode=%s direct_salvaged=%s tail_recovered=%s sources=%s",
        (live.category_name if live is not None else "autoscan"), total, updated, failed,
        (updated / max(1, total)) * 100.0, mode, direct_salvaged, tail_recovered, dict(source_counts),
    )
    return total, updated, failed


async def refresh_view_counts(
    rows: list[Listing], message: Message | BotChatAdapter | None = None, *,
    force: bool = False, max_age_seconds: int | None = None,
    traffic_priority: str = "manual",
    progress_message: Message | None = None,
    progress_title: str | None = None,
) -> tuple[int, int, int]:
    """Refresh missing/stale public view counters and persist them.

    Returns (requested, updated, failed). max_age_seconds lets automatic checkpoints
    safely reuse a counter fetched only a few minutes ago by another scan/user.
    """
    if not rows:
        return 0, 0, 0

    if ACCURATE_VIEWS_MODE:
        # Never reuse the historical 30-minute cache. A caller may explicitly
        # request a tiny <=60s coalescing window so an immediate baseline does not
        # reopen hundreds of ads that were verified seconds ago by the same scan.
        effective_ttl = max(0, min(60, int(max_age_seconds or 0)))
    else:
        effective_ttl = VIEW_COUNT_CACHE_TTL_SECONDS if max_age_seconds is None else max(0, int(max_age_seconds))
    cutoff = datetime.utcnow() - timedelta(seconds=effective_ttl)
    eligible = [
        row for row in rows
        if row.url
        and not bool(getattr(row, "is_promoted", False))
        and not bool(getattr(row, "is_price_reduced", False))
    ]
    targets = [
        row for row in eligible
        if force or row.views_checked_at is None or row.views_checked_at < cutoff
    ]
    reused_count = max(0, len(eligible) - len(targets))

    status = progress_message
    status_note = "точный свежий замер" if ACCURATE_VIEWS_MODE else ("свежий контрольный замер" if effective_ttl <= 60 else f"кэш {max(1, effective_ttl // 60)} мин.")
    if status is None and message is not None:
        try:
            status = await message.answer(
                f"👁 <b>Собираю просмотры</b>\n\n"
                f"📦 Объявлений: <b>{len(eligible)}</b>",
                parse_mode=ParseMode.HTML,
            )
        except Exception:
            status = None

    last_progress_edit = 0.0

    async def progress_cb(done: int, total: int):
        nonlocal last_progress_edit
        if status is not None and hasattr(status, "edit_text"):
            now_mono = time.monotonic()
            completed = min(len(eligible), reused_count + done)
            all_total = max(1, len(eligible))
            if completed < all_total and now_mono - last_progress_edit < 1.5:
                return
            last_progress_edit = now_mono
            try:
                pct = round(completed / all_total * 100) if eligible else 100
                title = progress_title or "👁 Собираю просмотры"
                await status.edit_text(
                    f"<b>{html.escape(title)}</b>\n\n"
                    f"{_progress_bar(pct)} <b>{pct}%</b>\n"
                    f"👁 Проверено: <b>{completed}/{len(eligible)}</b>",
                    parse_mode=ParseMode.HTML,
                )
            except Exception:
                pass

    if status is not None:
        await progress_cb(0, len(targets))
    if not targets:
        return 0, 0, 0

    parser = KleinanzeigenParser()
    try:
        results = await fetch_exact_views_v438_compatible(
            parser, [row.url for row in targets],
            concurrency=VIEW_COUNT_CONCURRENCY,
            progress_cb=progress_cb,
            traffic_priority=traffic_priority,
        )
    finally:
        await parser.close()

    source_counts = Counter(vr.source for vr in results.values())
    log.info("Accurate views refresh total=%s sources=%s", len(results), dict(source_counts))

    now = datetime.utcnow()
    updated = 0
    failed = 0
    by_id = {row.external_id: row for row in targets}
    url_to_id = {row.url: row.external_id for row in targets}
    async with db_write_lock:
        async with SessionLocal() as session:
            result = await session.execute(select(Listing).where(
                Listing.external_id.in_(list(by_id)),
                Listing.is_promoted.is_(False),
                Listing.is_price_reduced.is_(False),
            ))
            db_rows = {row.external_id: row for row in result.scalars().all()}
            for item in targets:
                url = item.url
                external_id = url_to_id.get(url)
                row = db_rows.get(external_id) if external_id else None
                if row is None:
                    continue
                vr = results.get(url)
                if vr is None or vr.views is None:
                    failed += 1
                    # Fail closed: an omitted result is the same as an explicit
                    # unknown. A stale number from an earlier measurement must
                    # never survive a failed current verification and pass filters.
                    row.view_count = None
                    row.views_checked_at = now
                    continue
                old = row.view_count
                row.view_count = int(vr.views)
                row.views_checked_at = now
                _apply_organic_view_baseline(row, row.view_count, now)
                if old != row.view_count:
                    session.add(ViewHistory(
                        external_id=row.external_id,
                        view_count=row.view_count,
                        recorded_at=now,
                    ))
                updated += 1
            await session.commit()

    # Update the already-loaded ORM objects so the CSV can be written without a reload.
    for row in targets:
        vr = results.get(row.url)
        if vr and vr.views is not None:
            row.view_count = int(vr.views)
            row.views_checked_at = now
        else:
            row.view_count = None
            row.views_checked_at = now

    if status is not None and hasattr(status, "edit_text"):
        try:
            await status.edit_text(
                f"👁 Точные просмотры готовы: <b>{updated}</b> · не подтверждено: <b>{failed}</b>",
                parse_mode=ParseMode.HTML,
            )
        except Exception:
            pass
    return len(targets), updated, failed


async def claim_due_observation() -> ScanObservation | None:
    """Claim one due automatic checkpoint and quickly discard stale missed ones."""
    for _ in range(100):
        now = datetime.utcnow()
        async with db_write_lock:
            async with SessionLocal() as session:
                result = await session.execute(
                    select(ScanObservation)
                    .where(
                        ScanObservation.status == "pending",
                        ScanObservation.target_hours.in_(OBSERVATION_SCHEDULE_HOURS),
                        ScanObservation.due_at <= now,
                    )
                    .order_by(ScanObservation.due_at.asc())
                    .limit(1)
                )
                obs = result.scalar_one_or_none()
                if obs is None:
                    return None
                if now - obs.due_at > timedelta(minutes=OBSERVATION_LATE_GRACE_MINUTES):
                    obs.status = "missed"
                    obs.completed_at = now
                    obs.error_text = "checkpoint missed while service was offline/busy"
                    await session.commit()
                    continue
                obs.status = "running"
                obs.started_at = now
                await session.commit()
                await session.refresh(obs)
                session.expunge(obs)
                return obs
    return None


async def mark_observation_result(
    observation_id: int, *, status: str, item_count: int = 0, error_text: str | None = None
) -> None:
    async with db_write_lock:
        async with SessionLocal() as session:
            obs = await session.get(ScanObservation, observation_id)
            if obs is None:
                return
            obs.status = status
            obs.completed_at = datetime.utcnow()
            obs.item_count = item_count
            obs.error_text = (error_text or "")[:1000] or None
            await session.commit()


async def recover_running_observations() -> int:
    """Requeue observations left in running state by an interrupted Railway process."""
    cutoff = datetime.utcnow() - timedelta(minutes=15)
    changed = 0
    async with db_write_lock:
        async with SessionLocal() as session:
            result = await session.execute(
                select(ScanObservation).where(
                    ScanObservation.status == "running",
                    ScanObservation.started_at.is_not(None),
                    ScanObservation.started_at < cutoff,
                )
            )
            for obs in result.scalars().all():
                obs.status = "pending"
                obs.started_at = None
                changed += 1
            await session.commit()
    return changed


async def process_observation(bot: Bot, obs: ScanObservation) -> None:
    async with SessionLocal() as session:
        scan = await session.get(UserScan, obs.scan_id)
        scan_settings = await session.get(UserSettings, int(scan.user_id)) if scan is not None else None
    if scan is not None and (scan_settings is None or not bool(getattr(scan_settings, "auto_observations", False))):
        await mark_observation_result(obs.id, status="cancelled", error_text="auto measurements disabled by user")
        return
    if scan is None or scan.status not in {"done", "partial"}:
        await mark_observation_result(obs.id, status="error", error_text="scan not available")
        return

    pairs = await get_scan_rows(scan.id)
    rows = [row for row, _ in pairs]
    if not rows:
        await mark_observation_result(obs.id, status="done", item_count=0)
        return

    try:
        measurement_started = datetime.utcnow() - timedelta(seconds=VIEW_MEASUREMENT_REUSE_SECONDS)
        async with background_view_refresh_lock:
            requested, updated, failed = await refresh_view_counts(
                rows, None, force=False, max_age_seconds=VIEW_MEASUREMENT_REUSE_SECONDS,
                traffic_priority="background"
            )
            recorded = await update_scan_view_refresh(
                scan.id, target_hours=obs.target_hours, fresh_after=measurement_started
            )
        if recorded <= 0:
            await mark_observation_result(
                obs.id, status="error", item_count=0,
                error_text=f"no fresh view values; failures={failed}",
            )
            log.warning("Observation produced no fresh counters scan=%s +%sh", scan.id, obs.target_hours)
            return
        await mark_observation_result(
            obs.id, status="done", item_count=recorded,
            error_text=(f"view failures: {failed}" if failed else None),
        )
        if obs.target_hours > 0:
            try:
                keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔥 Открыть популярное", callback_data="popular_now")],
                    [InlineKeyboardButton(text="📊 Открыть скан", callback_data=f"scan:{scan.id}")],
                ])
                await bot.send_message(
                    scan.user_id,
                    f"✅ <b>Контрольный замер +{obs.target_hours}ч готов</b>\n\n"
                    f"Скан: <b>{html.escape(scan.title)}</b>\n"
                    f"📅 Дата объявлений: <b>{_date_label(scan.target_date)}</b>\n"
                    f"👁 Свежих значений сохранено: <b>{recorded}</b>\n\n"
                    "Теперь в «🔥 Популярное сейчас» доступен TOP роста по каждой категории отдельно.",
                    parse_mode=ParseMode.HTML,
                    reply_markup=keyboard,
                )
            except Exception:
                log.debug("Could not notify user about observation scan=%s +%sh", scan.id, obs.target_hours, exc_info=True)
        log.info(
            "Observation done scan=%s +%sh requested=%s updated=%s recorded=%s failed=%s",
            scan.id, obs.target_hours, requested, updated, recorded, failed,
        )
    except Exception as exc:
        log.exception("Automatic observation failed scan=%s +%sh", scan.id, obs.target_hours)
        await mark_observation_result(obs.id, status="error", error_text=str(exc))


async def observation_scheduler(bot: Bot, worker_id: int = 1) -> None:
    """Persistent +3/+6/+12h view-checkpoint worker."""
    while True:
        try:
            obs = await claim_due_observation()
            if obs is None:
                await asyncio.sleep(OBSERVATION_POLL_SECONDS)
                continue
            await process_observation(bot, obs)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Observation scheduler loop error")
            await asyncio.sleep(OBSERVATION_POLL_SECONDS)


async def scan_archive_scheduler() -> None:
    """Keep the My Scans inbox compact even when the user does not open it."""
    while True:
        try:
            moved = await archive_expired_scans()
            if moved:
                log.info("Auto-archived %s completed scan cards", moved)
            await asyncio.sleep(SCAN_ARCHIVE_SWEEP_SECONDS)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Scan archive scheduler loop error")
            await asyncio.sleep(SCAN_ARCHIVE_SWEEP_SECONDS)


async def _due_high_baseline_rows(limit: int = ORGANIC_HIGH_BATCH_SIZE) -> list[Listing]:
    """Return 400+ baseline listings whose next clean velocity checkpoint is due."""
    cutoff = datetime.utcnow() - timedelta(minutes=ORGANIC_HIGH_CHECKPOINT_MINUTES)
    async with SessionLocal() as session:
        rows = list((await session.execute(
            select(Listing)
            .where(
                Listing.is_active.is_(True),
                Listing.last_seen_at >= datetime.utcnow() - timedelta(hours=24),
                Listing.is_promoted.is_(False),
                Listing.is_price_reduced.is_(False),
                Listing.organic_baseline_views.is_not(None),
                Listing.organic_baseline_views >= int(ORGANIC_HIGH_BASELINE_VIEWS),
                Listing.organic_verified_checkpoints < int(ORGANIC_HIGH_REQUIRED_CHECKPOINTS),
                Listing.organic_history_status.in_(["high_baseline", "high_check_1"]),
                func.coalesce(Listing.organic_last_checkpoint_at, Listing.organic_baseline_at) <= cutoff,
                func.coalesce(Listing.views_checked_at, Listing.organic_baseline_at) <= cutoff,
            )
            .order_by(func.coalesce(Listing.organic_last_checkpoint_at, Listing.organic_baseline_at).asc())
            .limit(max(1, int(limit)))
        )).scalars().all())
        for row in rows:
            session.expunge(row)
        return rows


async def organic_velocity_scheduler() -> None:
    """Collect two low-priority exact checkpoints for first-seen 400+ counters.

    These checks are independent from AI candidacy: a suspiciously large initial
    total cannot create an AI/Radar score first and only be checked afterwards.
    User scans keep priority; the scheduler simply waits while foreground work exists.
    """
    await asyncio.sleep(20)
    while True:
        try:
            running, queued = await _radar_foreground_counts()
            traffic_snapshot = await TRAFFIC.snapshot()
            if running or queued or int(getattr(traffic_snapshot, "scan_jobs_active", 0) or 0) > 0 or int(getattr(traffic_snapshot, "background_pauses", 0) or 0) > 0:
                await asyncio.sleep(ORGANIC_HIGH_POLL_SECONDS)
                continue
            rows = await _due_high_baseline_rows()
            if not rows:
                await asyncio.sleep(ORGANIC_HIGH_POLL_SECONDS)
                continue
            clean_rows: list[Listing] = []
            detail_unknown = 0
            detail_blocked = 0
            foreground_yielded = False
            for row in rows:
                traffic_snapshot = await TRAFFIC.snapshot()
                if int(getattr(traffic_snapshot, "scan_jobs_active", 0) or 0) > 0 or int(getattr(traffic_snapshot, "background_pauses", 0) or 0) > 0:
                    foreground_yielded = True
                    log.info("Verified Organic Velocity yielded to foreground scan after=%s/%s", len(clean_rows), len(rows))
                    break
                allowed, reason, _verified_at = await verify_listing_organic_now(
                    str(row.external_id), traffic_priority="background"
                )
                if allowed:
                    clean_rows.append(row)
                elif "promoted" in str(reason) or "reduced" in str(reason):
                    detail_blocked += 1
                else:
                    detail_unknown += 1
            if foreground_yielded:
                await asyncio.sleep(ORGANIC_HIGH_POLL_SECONDS)
                continue
            if not clean_rows:
                log.info(
                    "Verified Organic Velocity checkpoint gate: due=%s clean=0 blocked=%s unknown=%s",
                    len(rows), detail_blocked, detail_unknown,
                )
                await asyncio.sleep(ORGANIC_HIGH_POLL_SECONDS)
                continue
            ids = [str(row.external_id) for row in clean_rows]
            before: dict[str, int] = {
                str(row.external_id): int(getattr(row, "organic_verified_checkpoints", 0) or 0)
                for row in clean_rows
            }
            async with background_view_refresh_lock:
                requested, updated, failed = await refresh_view_counts(
                    clean_rows, None, force=True, max_age_seconds=0, traffic_priority="background"
                )
            async with SessionLocal() as session:
                refreshed = list((await session.execute(
                    select(Listing).where(Listing.external_id.in_(ids))
                )).scalars().all())
            checkpointed = 0
            newly_verified_ids: list[str] = []
            for row in refreshed:
                current = int(getattr(row, "organic_verified_checkpoints", 0) or 0)
                previous = int(before.get(str(row.external_id), 0))
                if current > previous:
                    checkpointed += 1
                if previous < int(ORGANIC_HIGH_REQUIRED_CHECKPOINTS) <= current and str(getattr(row, "organic_history_status", "") or "") == "observed":
                    newly_verified_ids.append(str(row.external_id))
            radar_saved = 0
            if newly_verified_ids:
                try:
                    radar_saved = await record_verified_velocity_signals(
                        newly_verified_ids, traffic_priority="background"
                    )
                except Exception:
                    log.exception("Verified Organic Velocity Radar merge failed ids=%s", len(newly_verified_ids))
            log.info(
                "Verified Organic Velocity batch due=%s detail_clean=%s detail_blocked=%s detail_unknown=%s requested=%s updated=%s failed=%s checkpointed=%s newly_verified=%s radar_saved=%s threshold=%s",
                len(rows), len(clean_rows), detail_blocked, detail_unknown, requested, updated, failed, checkpointed, len(newly_verified_ids), radar_saved, ORGANIC_HIGH_BASELINE_VIEWS,
            )
            await asyncio.sleep(ORGANIC_HIGH_POLL_SECONDS)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Verified Organic Velocity scheduler error")
            await asyncio.sleep(ORGANIC_HIGH_POLL_SECONDS)


async def radar_maintenance_scheduler() -> None:
    """Radar 3.2 maintenance: preserve evidence; no destructive startup reset/backfill."""
    async def foreground_busy() -> bool:
        running, queued = await _radar_foreground_counts()
        snap = await TRAFFIC.snapshot()
        return bool(running or queued or int(getattr(snap, "scan_jobs_active", 0) or 0) > 0 or int(getattr(snap, "background_pauses", 0) or 0) > 0)
    try:
        await asyncio.sleep(8)
        await prepare_radar_v3_once()
        await repair_radar_v3_historical_scores_once()
        restored_live = await repair_radar_v3_live_retention_once()
        if restored_live:
            log.info("DT Radar 3.2 startup live-retention restore=%s", restored_live)
    except asyncio.CancelledError:
        raise
    except Exception:
        log.exception("DT Radar 3.0 startup reset failed")
    while True:
        try:
            if not await foreground_busy():
                await bump_resurrection_integrity_sweep_once()
            await asyncio.sleep(3600)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("DT Radar 3.0 maintenance loop error")
            await asyncio.sleep(300)


async def radar_v3_observation_scheduler() -> None:
    """Background exact remeasurement for DT-owned Radar 3.0 baselines.

    v4.21.1 claims rows before network work, so multiple Parser replicas cannot
    refresh/write the same RadarObservation batch concurrently.
    """
    owner = f"parser:{os.getenv('RAILWAY_REPLICA_ID', os.getenv('HOSTNAME', 'local'))}:{os.getpid()}:{uuid.uuid4().hex[:8]}"
    await asyncio.sleep(30)
    while True:
        try:
            running, queued = await _radar_foreground_counts()
            # User scans remain absolute priority. AutoScan itself is allowed to run
            # concurrently with one throttled Radar checkpoint lane; otherwise +60m
            # observations can drift by hours during a large 15+15 circle.
            if running or queued:
                await asyncio.sleep(30)
                continue
            expired_obs = await radar_v3_expire_observations()
            expired_products = await radar_v3_expire_stale_products()
            if expired_obs or expired_products:
                log.info("DT Radar 3.0 expiry observations=%s products=%s", expired_obs, expired_products)
            ids = await radar_v3_claim_due_external_ids(owner, limit=250)
            if not ids:
                await asyncio.sleep(30)
                continue
            async with SessionLocal() as session:
                rows = list((await session.execute(select(Listing).where(Listing.external_id.in_(ids)))).scalars().all())
            if not rows:
                released = await radar_v3_release_claims(owner, ids)
                log.warning("DT Radar 3.0 claimed rows missing listings due=%s released=%s", len(ids), released)
                await asyncio.sleep(30)
                continue
            async with radar_v3_view_refresh_lock:
                requested, updated, failed = await refresh_view_counts(rows, None, force=True, max_age_seconds=0, traffic_priority="radar_checkpoint")
            saved = await radar_v3_record_refreshed([str(x.external_id) for x in rows])
            # Successful observations release their own lease while being recorded;
            # failed/unchanged rows are released here for a clean retry next poll.
            released = await radar_v3_release_claims(owner, ids)
            log.info("DT Radar 3.0 observation batch due=%s requested=%s updated=%s failed=%s signals=%s released_claims=%s", len(ids), requested, updated, failed, saved, released)
            await asyncio.sleep(10)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("DT Radar 3.0 observation scheduler error")
            await asyncio.sleep(60)


def _radar_autoscan_category_allowed(cat) -> bool:
    """Use the canonical Radar 3.2 product-market scope for AutoScan."""
    if cat is None:
        return False
    return radar_v3_category_allowed(str(getattr(cat, "key", "")))


def _radar_autoscan_categories() -> list:
    """Product-oriented leaf categories only; normal parser categories are untouched."""
    return [cat for cat in CATEGORIES.values() if _radar_autoscan_category_allowed(cat)]

def _radar_autoscan_now_iso() -> str:
    return datetime.now(MOSCOW).replace(microsecond=0).isoformat()


def _radar_autoscan_default_state() -> dict:
    categories = _radar_autoscan_categories()
    return {
        "policy_version": RADAR_AUTOSCAN_POLICY_VERSION,
        "daily_enabled": False,
        "daily_time": RADAR_AUTOSCAN_DEFAULT_TIME,
        "skip_daily_if_completed_today": True,
        "status": "idle",              # idle | running | paused
        "stop_requested": False,
        "waiting_for_users": False,
        "round_id": "",
        "mode": "",                   # manual | daily
        "target_date": "",
        "category_keys": [cat.key for cat in categories],
        "current_index": 0,
        "current_category_key": "",
        "current_category_name": "",
        "current_stage": "",
        "current_stage_started_at": "",
        "last_watchdog_category": "",
        "total": len(categories),
        "processed": 0,
        "successful": 0,
        "needs_review": 0,
        "system_errors": 0,
        "skipped_nonproduct": 0,
        "failed": 0,  # compatibility total = needs_review + system_errors
        "pages_verified": 0,
        "listings_seen": 0,
        "new_listings": 0,
        "search_promoted_filtered": 0,
        "search_reduced_filtered": 0,
        "radar_saved": 0,
        "views_requested": 0,
        "views_verified": 0,
        "views_failed": 0,
        "view_tail_deferred": 0,
        "view_tail_categories": 0,
        "radar_candidates": 0,
        "radar_high_baseline_pending": 0,
        "radar_high_baseline_verified": 0,
        "radar_detail_checked": 0,
        "radar_organic_passed": 0,
        "radar_promoted_blocked": 0,
        "radar_reduced_blocked": 0,
        "radar_unknown_blocked": 0,
        "radar_unknown_reasons": {},
        "radar_db_blocked": 0,
        "radar_demand_gate_rejected": 0,
        "radar_qualified_candidates": 0,
        "radar_early_admitted": 0,
        "radar_strong_admitted": 0,
        "radar_hot_admitted": 0,
        "radar_already_present": 0,
        "failed_categories": [],
        "retry_parent_total": 0,
        "retry_parent_successful": 0,
        "retry_parent_round_id": "",
        "started_at": "",
        "updated_at": _radar_autoscan_now_iso(),
        "last_completed_date": "",
        "last_daily_date": "",
        "last_context_date": "",
        "context_for_date": "",
        "layer": "fresh",
        "last_summary": {},
        "history": [],
    }


def _radar_autoscan_normalize_state(raw: dict | None) -> dict:
    state = _radar_autoscan_default_state()
    raw_state = raw if isinstance(raw, dict) else {}
    state.update(raw_state)
    if str(state.get("daily_time") or "") not in RADAR_AUTOSCAN_TIME_CHOICES:
        state["daily_time"] = RADAR_AUTOSCAN_DEFAULT_TIME
    if str(state.get("status") or "idle") not in {"idle", "running", "paused"}:
        state["status"] = "idle"
    state["daily_enabled"] = bool(state.get("daily_enabled"))
    state["skip_daily_if_completed_today"] = bool(state.get("skip_daily_if_completed_today", True))
    state["stop_requested"] = bool(state.get("stop_requested"))
    state["waiting_for_users"] = bool(state.get("waiting_for_users"))
    state["last_context_date"] = str(state.get("last_context_date") or "")[:10]
    state["context_for_date"] = str(state.get("context_for_date") or "")[:10]
    state["layer"] = "context" if str(state.get("layer") or "fresh") == "context" else "fresh"
    state["current_stage"] = str(state.get("current_stage") or "")[:80]
    state["current_stage_started_at"] = str(state.get("current_stage_started_at") or "")[:64]
    state["last_watchdog_category"] = str(state.get("last_watchdog_category") or "")[:160]
    state["history"] = list(state.get("history") or [])[:RADAR_AUTOSCAN_HISTORY_LIMIT]
    raw_unknown_reasons = state.get("radar_unknown_reasons") or {}
    state["radar_unknown_reasons"] = {
        str(key)[:64]: max(0, int(value or 0))
        for key, value in (raw_unknown_reasons.items() if isinstance(raw_unknown_reasons, dict) else [])
        if str(key).strip()
    }

    failures = []
    inferred_review = 0
    inferred_system = 0
    for item in list(state.get("failed_categories") or []):
        if not isinstance(item, dict):
            continue
        key = str(item.get("key") or "")
        if key not in CATEGORIES or bool(getattr(CATEGORIES[key], "is_group", False)):
            continue
        kind = str(item.get("kind") or "").strip().lower()
        if kind not in {"partial", "system", "radar_views", "radar_gate_unknown"}:
            # Legacy v4.11.1-v4.11.4 entries did not carry a kind. A category with
            # parser evidence/verified pages is a partial crawl; zero-page generic
            # execution failures are treated as system errors.
            reason = str(item.get("reason") or "partial").lower()
            system_hint = any(token in reason for token in ("traceback", "exception", "ошибка выполнения", "runtimeerror"))
            kind = "system" if system_hint and int(item.get("verified_pages") or 0) <= 0 else "partial"
        if kind == "system":
            inferred_system += 1
        else:
            inferred_review += 1
        failures.append({
            "key": key,
            "name": str(item.get("name") or CATEGORIES[key].name)[:160],
            "reason": str(item.get("reason") or "partial")[:300],
            "verified_pages": max(0, int(item.get("verified_pages") or 0)),
            "depth": max(1, int(item.get("depth") or RADAR_AUTOSCAN_DEPTH)),
            "kind": kind,
        })
    state["failed_categories"] = failures[:500]
    state["retry_parent_total"] = max(0, int(state.get("retry_parent_total") or 0))
    state["retry_parent_successful"] = max(0, int(state.get("retry_parent_successful") or 0))
    state["retry_parent_round_id"] = str(state.get("retry_parent_round_id") or "")[:80]

    # Existing in-progress v4.11.4 rounds keep their exact old key order so a Railway
    # deploy cannot corrupt current_index/counters midway through a persisted round.
    # Every new/retry round uses the current policy and only product-oriented categories.
    stored_policy = max(0, int(raw_state.get("policy_version") or 0))
    if stored_policy < RADAR_AUTOSCAN_POLICY_VERSION:
        # A Radar policy change means old progress/counters/history describe a
        # different market scope and must not leak into the clean dashboard.
        # Preserve only the user's daily schedule preferences.
        daily_enabled = bool(state.get("daily_enabled"))
        daily_time = str(state.get("daily_time") or RADAR_AUTOSCAN_DEFAULT_TIME)
        skip_daily = bool(state.get("skip_daily_if_completed_today", True))
        state = _radar_autoscan_default_state()
        state["daily_enabled"] = daily_enabled
        state["daily_time"] = daily_time if daily_time in RADAR_AUTOSCAN_TIME_CHOICES else RADAR_AUTOSCAN_DEFAULT_TIME
        state["skip_daily_if_completed_today"] = skip_daily
        raw_state = {}
        failures = []
        inferred_review = 0
        inferred_system = 0
    if stored_policy < RADAR_AUTOSCAN_POLICY_VERSION and str(state.get("status") or "idle") in {"running", "paused"}:
        # v4.21.5 retires yesterday/context entirely. Never resume a persisted
        # 15+15 round after deploy; a new run must use 20 pages for today only.
        state["status"] = "idle"
        state["stop_requested"] = False
        state["waiting_for_users"] = False
        state["current_index"] = 0
        state["current_category_key"] = ""
        state["current_category_name"] = ""
        state["mode"] = ""
        state["layer"] = "fresh"
        state["target_date"] = ""
    legacy_active = False
    if legacy_active:
        keys = [str(x) for x in (state.get("category_keys") or []) if str(x) in CATEGORIES and not CATEGORIES[str(x)].is_group]
        state["legacy_policy_round"] = True
        state["policy_version"] = stored_policy
    else:
        keys = [str(x) for x in (state.get("category_keys") or []) if str(x) in CATEGORIES and _radar_autoscan_category_allowed(CATEGORIES[str(x)])]
        state["legacy_policy_round"] = False
        state["policy_version"] = RADAR_AUTOSCAN_POLICY_VERSION
    if not keys:
        keys = [cat.key for cat in _radar_autoscan_categories()]
    state["category_keys"] = keys
    state["total"] = len(keys)
    state["current_index"] = max(0, min(len(keys), int(state.get("current_index") or 0)))

    for key in (
        "processed", "successful", "failed", "skipped_nonproduct", "pages_verified",
        "listings_seen", "new_listings", "radar_saved", "views_requested", "views_verified",
        "views_failed", "view_tail_deferred", "view_tail_categories",
    ):
        state[key] = max(0, int(state.get(key) or 0))
    # New counters are authoritative when present; otherwise infer legacy failures.
    if "needs_review" in raw_state or "system_errors" in raw_state:
        state["needs_review"] = max(0, int(state.get("needs_review") or 0))
        state["system_errors"] = max(0, int(state.get("system_errors") or 0))
    else:
        state["needs_review"] = inferred_review
        state["system_errors"] = inferred_system
    state["failed"] = max(state["failed"], state["needs_review"] + state["system_errors"])
    return state


async def load_radar_autoscan_state() -> dict:
    async with SessionLocal() as session:
        row = await session.get(AppSetting, RADAR_AUTOSCAN_SETTING_KEY)
        if row is None or not str(row.value or "").strip():
            return _radar_autoscan_default_state()
        try:
            raw = json.loads(row.value)
        except Exception:
            log.exception("DT Radar AutoScan state JSON is invalid; using defaults")
            return _radar_autoscan_default_state()
    return _radar_autoscan_normalize_state(raw)


async def save_radar_autoscan_state(state: dict) -> dict:
    state = _radar_autoscan_normalize_state(state)
    state["updated_at"] = _radar_autoscan_now_iso()
    payload = json.dumps(state, ensure_ascii=False, separators=(",", ":"))
    async with SessionLocal() as session:
        row = await session.get(AppSetting, RADAR_AUTOSCAN_SETTING_KEY)
        if row is None:
            row = AppSetting(key=RADAR_AUTOSCAN_SETTING_KEY, value=payload, updated_at=datetime.utcnow())
            session.add(row)
        else:
            row.value = payload
            row.updated_at = datetime.utcnow()
        await session.commit()
    return state


def _radar_autoscan_new_round(state: dict, mode: str) -> dict:
    keys = [cat.key for cat in _radar_autoscan_categories()]
    now = datetime.now(MOSCOW)
    keep = {
        "daily_enabled": bool(state.get("daily_enabled")),
        "daily_time": str(state.get("daily_time") or RADAR_AUTOSCAN_DEFAULT_TIME),
        "skip_daily_if_completed_today": bool(state.get("skip_daily_if_completed_today", True)),
        "last_completed_date": str(state.get("last_completed_date") or ""),
        "last_daily_date": str(state.get("last_daily_date") or ""),
        "last_context_date": str(state.get("last_context_date") or ""),
        "last_summary": dict(state.get("last_summary") or {}),
        "history": list(state.get("history") or [])[:RADAR_AUTOSCAN_HISTORY_LIMIT],
    }
    new_state = _radar_autoscan_default_state()
    new_state.update(keep)
    new_state.update({
        "policy_version": RADAR_AUTOSCAN_POLICY_VERSION,
        "status": "running",
        "stop_requested": False,
        "waiting_for_users": False,
        "round_id": f"{now.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:5]}",
        "mode": "daily" if mode == "daily" else "manual",
        "layer": "fresh",
        "context_for_date": "",
        "target_date": now.date().isoformat(),
        "category_keys": keys,
        "current_index": 0,
        "current_category_key": "",
        "current_category_name": "Запуск первой категории…",
        "current_stage": "starting",
        "current_stage_started_at": now.replace(microsecond=0).isoformat(),
        "total": len(keys),
        "processed": 0,
        "successful": 0,
        "needs_review": 0,
        "system_errors": 0,
        "skipped_nonproduct": 0,
        "failed": 0,
        "pages_verified": 0,
        "listings_seen": 0,
        "new_listings": 0,
        "search_promoted_filtered": 0,
        "search_reduced_filtered": 0,
        "radar_saved": 0,
        "views_requested": 0,
        "views_verified": 0,
        "views_failed": 0,
        "view_tail_deferred": 0,
        "view_tail_categories": 0,
        "radar_candidates": 0,
        "radar_high_baseline_pending": 0,
        "radar_high_baseline_verified": 0,
        "radar_detail_checked": 0,
        "radar_organic_passed": 0,
        "radar_promoted_blocked": 0,
        "radar_reduced_blocked": 0,
        "radar_unknown_blocked": 0,
        "radar_unknown_reasons": {},
        "radar_db_blocked": 0,
        "radar_demand_gate_rejected": 0,
        "radar_qualified_candidates": 0,
        "radar_early_admitted": 0,
        "radar_strong_admitted": 0,
        "radar_hot_admitted": 0,
        "radar_already_present": 0,
        "failed_categories": [],
        "retry_parent_total": 0,
        "retry_parent_successful": 0,
        "retry_parent_round_id": "",
        "started_at": now.replace(microsecond=0).isoformat(),
    })
    return new_state


def _radar_autoscan_retry_round(state: dict) -> dict | None:
    """Create one low-priority round containing only failures from the last report."""
    last = dict(state.get("last_summary") or {})
    failures = [item for item in list(last.get("failed_categories") or []) if isinstance(item, dict)]
    keys: list[str] = []
    for item in failures:
        key = str(item.get("key") or "")
        if key in CATEGORIES and _radar_autoscan_category_allowed(CATEGORIES[key]) and key not in keys:
            keys.append(key)
    if not keys:
        return None
    now = datetime.now(MOSCOW)
    keep = {
        "daily_enabled": bool(state.get("daily_enabled")),
        "daily_time": str(state.get("daily_time") or RADAR_AUTOSCAN_DEFAULT_TIME),
        "skip_daily_if_completed_today": bool(state.get("skip_daily_if_completed_today", True)),
        "last_completed_date": str(state.get("last_completed_date") or ""),
        "last_daily_date": str(state.get("last_daily_date") or ""),
        "last_context_date": str(state.get("last_context_date") or ""),
        "last_summary": last,
        "history": list(state.get("history") or [])[:RADAR_AUTOSCAN_HISTORY_LIMIT],
    }
    new_state = _radar_autoscan_default_state()
    new_state.update(keep)
    coverage_total = int(last.get("coverage_total") or last.get("total") or len(_radar_autoscan_categories()))
    coverage_success = int(last.get("coverage_successful") or last.get("successful") or 0)
    retry_target_date = str(last.get("target_date") or now.date().isoformat())[:10]
    try:
        datetime.strptime(retry_target_date, "%Y-%m-%d")
    except Exception:
        retry_target_date = now.date().isoformat()
    new_state.update({
        "policy_version": RADAR_AUTOSCAN_POLICY_VERSION,
        "status": "running",
        "stop_requested": False,
        "waiting_for_users": False,
        "round_id": f"retry-{now.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:5]}",
        "mode": "retry",
        "layer": "fresh",
        "context_for_date": str(last.get("context_for_date") or ""),
        "target_date": retry_target_date,
        "category_keys": keys,
        "current_index": 0,
        "current_category_key": "",
        "current_category_name": "Запуск первой категории…",
        "current_stage": "starting",
        "current_stage_started_at": now.replace(microsecond=0).isoformat(),
        "total": len(keys),
        "processed": 0,
        "successful": 0,
        "needs_review": 0,
        "system_errors": 0,
        "skipped_nonproduct": 0,
        "failed": 0,
        "pages_verified": 0,
        "listings_seen": 0,
        "new_listings": 0,
        "search_promoted_filtered": 0,
        "search_reduced_filtered": 0,
        "radar_saved": 0,
        "views_requested": 0,
        "views_verified": 0,
        "views_failed": 0,
        "view_tail_deferred": 0,
        "view_tail_categories": 0,
        "radar_candidates": 0,
        "radar_high_baseline_pending": 0,
        "radar_high_baseline_verified": 0,
        "radar_detail_checked": 0,
        "radar_organic_passed": 0,
        "radar_promoted_blocked": 0,
        "radar_reduced_blocked": 0,
        "radar_unknown_blocked": 0,
        "radar_unknown_reasons": {},
        "radar_db_blocked": 0,
        "radar_demand_gate_rejected": 0,
        "radar_qualified_candidates": 0,
        "radar_early_admitted": 0,
        "radar_strong_admitted": 0,
        "radar_hot_admitted": 0,
        "radar_already_present": 0,
        "failed_categories": [],
        "retry_parent_total": max(coverage_total, len(keys)),
        "retry_parent_successful": max(0, coverage_success),
        "retry_parent_round_id": str(last.get("retry_parent_round_id") or last.get("round_id") or ""),
        "started_at": now.replace(microsecond=0).isoformat(),
    })
    return new_state


async def _radar_foreground_counts() -> tuple[int, int]:
    async with job_guard:
        running = sum(1 for job in active_jobs.values() if job.state == "running" and not job.cancel_requested)
        queued = sum(1 for job in active_jobs.values() if job.state == "queued" and not job.cancel_requested)
    return int(running), int(queued)


async def _notify_radar_autoscan_admins(
    bot: Bot, text: str, reply_markup: InlineKeyboardMarkup | None = None
) -> None:
    for admin_id in sorted(ADMIN_IDS):
        try:
            await bot.send_message(int(admin_id), text, parse_mode=ParseMode.HTML, reply_markup=reply_markup)
        except (TelegramForbiddenError, TelegramBadRequest):
            log.debug("Could not notify Radar AutoScan admin=%s", admin_id)
        except Exception:
            log.exception("Radar AutoScan admin notification failed admin=%s", admin_id)


def _radar_autoscan_duration_seconds(started_at: str) -> int:
    try:
        started = datetime.fromisoformat(str(started_at))
        now = datetime.now(MOSCOW)
        if started.tzinfo is None:
            started = started.replace(tzinfo=MOSCOW)
        return max(0, int((now - started.astimezone(MOSCOW)).total_seconds()))
    except Exception:
        return 0


def _radar_unknown_reason_text(reasons: dict | None, *, limit: int = 5) -> str:
    if not isinstance(reasons, dict):
        return ""
    rows = sorted(
        ((str(k), int(v or 0)) for k, v in reasons.items() if int(v or 0) > 0),
        key=lambda item: (-item[1], item[0]),
    )[:max(1, int(limit))]
    return " · ".join(f"{html.escape(key)} {value}" for key, value in rows)


def _radar_autoscan_human_duration(seconds: int) -> str:
    """Exact AutoScan duration including hours; kept separate from queue ETA formatting."""
    seconds = max(0, int(seconds or 0))
    hours, rem = divmod(seconds, 3600)
    minutes, secs = divmod(rem, 60)
    if hours:
        return f"{hours} ч {minutes} мин"
    if minutes:
        return f"{minutes} мин {secs} сек"
    return f"{secs} сек"


def _radar_autoscan_next_run_text(state: dict) -> str:
    if not bool(state.get("daily_enabled")):
        return "—"
    now = datetime.now(MOSCOW)
    try:
        hh, mm = [int(x) for x in str(state.get("daily_time") or RADAR_AUTOSCAN_DEFAULT_TIME).split(":", 1)]
    except Exception:
        hh, mm = 5, 0
    candidate = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
    if state.get("last_daily_date") == now.date().isoformat():
        candidate += timedelta(days=1)
    elif candidate <= now:
        return "сегодня · при первой свободной возможности"
    return candidate.strftime("%d.%m · %H:%M МСК")




async def _radar3_dashboard_snapshot() -> dict:
    """Compact Radar 3.2 control-plane snapshot with bounded DB round-trips."""
    now = datetime.utcnow()
    active_statuses = ["baseline", "candidate", "observed", "confirmed"]
    async with SessionLocal() as session:
        # One aggregate scan replaces the former ~12 sequential COUNT queries.
        # This matters once Radar carries tens of thousands of observations.
        obs = (await session.execute(select(
            func.count(RadarObservation.id).filter(
                RadarObservation.status.in_(active_statuses), RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.checkpoint_count == 0, RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.next_check_at.is_not(None), RadarObservation.next_check_at <= now,
                RadarObservation.expires_at > now, RadarObservation.status.in_(active_statuses)
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.total_delta > 0, RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.status == "candidate", RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.status.in_(["observed", "confirmed"]), RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.status == "confirmed", RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.consecutive_scored >= 2, RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.current_vph >= 30.0, RadarObservation.acceleration_ratio >= 0.20, RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(
                RadarObservation.confidence >= 70, RadarObservation.expires_at > now
            ),
            func.count(RadarObservation.id).filter(RadarObservation.status == "quiet"),
            func.coalesce(func.sum(RadarObservation.total_delta).filter(
                RadarObservation.positive_checkpoints >= 1, RadarObservation.expires_at > now
            ), 0),
        ))).one()
        (active, baseline, due, any_growth, candidate, observed, strong_intervals,
         persistent, accelerating, high_confidence, quiet, total_delta) = [int(x or 0) for x in obs]

        products = (await session.execute(select(
            func.count(RadarProduct.id).filter(
                RadarProduct.latest_source == "radar3_observed", RadarProduct.status == "stable"
            ),
            func.count(RadarProduct.id).filter(
                RadarProduct.latest_source == "radar3_observed", RadarProduct.status == "rising"
            ),
            func.count(RadarProduct.id).filter(
                RadarProduct.latest_source == "radar3_observed", RadarProduct.status == "hot"
            ),
        ))).one()
        early, strong, hot = [int(x or 0) for x in products]

        category_rows = list((await session.execute(
            select(
                RadarObservation.category_key,
                func.count(RadarObservation.id),
                func.coalesce(func.sum(RadarObservation.total_delta), 0),
                func.coalesce(func.avg(RadarObservation.current_vph), 0.0),
                func.coalesce(func.max(RadarObservation.velocity_percentile), 0.0),
                func.coalesce(func.avg(RadarObservation.confidence), 0.0),
            )
            .where(
                RadarObservation.status.in_(["candidate", "observed", "confirmed"]),
                RadarObservation.expires_at > now,
            )
            .group_by(RadarObservation.category_key)
            .order_by(func.coalesce(func.max(RadarObservation.velocity_percentile), 0.0).desc(), func.coalesce(func.avg(RadarObservation.current_vph), 0.0).desc())
            .limit(12)
        )).all())
        signal_rows = list((await session.execute(
            select(RadarProduct.category_key, RadarProduct.status, func.count(RadarProduct.id))
            .where(
                RadarProduct.latest_source == "radar3_observed",
                RadarProduct.status.in_(["stable", "rising", "hot"]),
            )
            .group_by(RadarProduct.category_key, RadarProduct.status)
        )).all())

    signal_map: dict[str, dict[str, int]] = {}
    for key, status, count in signal_rows:
        slot = signal_map.setdefault(str(key or "unknown"), {"stable": 0, "rising": 0, "hot": 0})
        slot[str(status or "")] = int(count or 0)
    category_lines = []
    for key, count, delta, avg_vph, max_pct, avg_conf in category_rows:
        key = str(key or "unknown")
        cat = CATEGORIES.get(key)
        name = str(getattr(cat, "name", None) or key)
        sig = signal_map.get(key, {})
        suffix = []
        if int(sig.get("hot", 0)): suffix.append(f"🔥{int(sig['hot'])}")
        if int(sig.get("rising", 0)): suffix.append(f"📈{int(sig['rising'])}")
        if int(sig.get("stable", 0)): suffix.append(f"🟡{int(sig['stable'])}")
        signal_text = (" · " + " ".join(suffix)) if suffix else ""
        category_lines.append(
            f"• <b>{html.escape(name)}</b>: {float(avg_vph or 0):.1f}/ч avg · P{int(round(float(max_pct or 0)*100))} max · "
            f"conf {int(round(float(avg_conf or 0)))}% · +{int(delta or 0)} · {int(count or 0)} наблюд.{signal_text}"
        )
    return {
        "active": active, "baseline": baseline, "due": due, "any_growth": any_growth,
        "candidate": candidate, "observed": observed, "strong_intervals": strong_intervals,
        "persistent": persistent, "accelerating": accelerating, "high_confidence": high_confidence,
        "quiet": quiet, "total_delta": total_delta,
        "early": early, "strong": strong, "hot": hot, "category_lines": category_lines,
    }


_radar3_dashboard_snapshot_task: asyncio.Task | None = None


def _consume_detached_radar_task(task: asyncio.Task) -> None:
    """Consume a late DB task result without blocking Telegram UI handlers."""
    try:
        task.result()
    except asyncio.CancelledError:
        pass
    except Exception:
        log.exception("Detached DT Radar dashboard task failed")


async def _radar3_dashboard_safe_snapshot(timeout_seconds: float = 2.5) -> dict:
    """Bound dashboard latency without waiting for asyncpg cancellation.

    asyncio.wait_for(coro) cancels the underlying query on timeout and then waits
    for cancellation to finish. Under PostgreSQL contention that cancellation can
    itself take much longer than the requested timeout, leaving Telegram stuck on
    the loading card. Shielding a single shared in-flight snapshot lets the UI
    return immediately while the query finishes in the background.
    """
    global _radar3_dashboard_snapshot_task
    task = _radar3_dashboard_snapshot_task
    if task is None or task.done():
        task = asyncio.create_task(_radar3_dashboard_snapshot(), name="dt-radar-dashboard-snapshot")
        task.add_done_callback(_consume_detached_radar_task)
        _radar3_dashboard_snapshot_task = task
    try:
        return await asyncio.wait_for(asyncio.shield(task), timeout=max(0.5, float(timeout_seconds)))
    except asyncio.TimeoutError:
        log.warning("DT Radar 3.2 dashboard snapshot timed out; UI continues with controls")
        return {"category_lines": ["⚠️ Статистика замеров ещё считается — нажми «Обновить» через несколько секунд"]}
    except Exception:
        log.exception("DT Radar 3.2 dashboard snapshot failed")
        return {"category_lines": ["⚠️ Статистика замеров временно недоступна"]}


async def _radar_autoscan_text() -> tuple[str, dict]:
    """Fast live control panel. Never waits for Radar analytics queries."""
    state = await load_radar_autoscan_state()
    total = max(1, int(state.get("total") or len(_radar_autoscan_categories()) or 1))
    processed = min(total, int(state.get("processed") or 0))
    pct = int(round(processed / total * 100))
    status = str(state.get("status") or "idle")
    if status == "running":
        status_text = "🟡 ждёт пользовательские сканы" if state.get("waiting_for_users") else "🟢 работает"
    elif status == "paused":
        status_text = "⏸ остановлен"
    else:
        status_text = "🔴 не запущен"

    current = str(state.get("current_category_name") or "—")
    live_stage_line = _radar_autoscan_live_stage_line(state) if status == "running" else ""
    depth = _radar_layer_depth(state)
    total_pages_target = total * depth
    pages_verified = int(state.get("pages_verified") or 0)
    listings_seen = int(state.get("listings_seen") or 0)
    new_listings = int(state.get("new_listings") or 0)
    views_requested = int(state.get("views_requested") or 0)
    views_verified = int(state.get("views_verified") or 0)
    baselines = int(state.get("radar_candidates") or 0)
    successful = int(state.get("successful") or 0)
    needs_review = int(state.get("needs_review") or 0)
    system_errors = int(state.get("system_errors") or 0)

    elapsed_text = "—"
    started_at = str(state.get("started_at") or "")
    if started_at and status in {"running", "paused"}:
        try:
            started = datetime.fromisoformat(started_at)
            if started.tzinfo is None:
                started = started.replace(tzinfo=MOSCOW)
            elapsed_text = _radar_autoscan_human_duration(int((datetime.now(MOSCOW) - started.astimezone(MOSCOW)).total_seconds()))
        except Exception:
            pass

    current_number = min(total, processed + 1) if status == "running" else processed
    last = dict(state.get("last_summary") or {})
    last_line = "—"
    if last:
        coverage_total = int(last.get("coverage_total") or last.get("total") or 0)
        coverage_success = int(last.get("coverage_successful") or last.get("successful") or 0)
        last_line = f"{coverage_success}/{coverage_total} категорий · Baseline {int(last.get('radar_candidates') or 0)}"

    text = (
        "<b>📡 DT Radar 3.2 · ADAPTIVE LIVE</b>\n\n"
        f"AutoScan: <b>{status_text}</b>\n"
        f"📂 Категория: <b>{html.escape(current)}</b>\n"
        f"📊 Прогресс категорий: <b>{processed}/{total} · {pct}%</b>"
        + (f" · сейчас {current_number}/{total}" if status == "running" else "") + "\n"
        f"📄 Глубина: <b>{depth} страниц только за сегодня на категорию</b>\n"
        + (live_stage_line + "\n" if live_stage_line else "")
        + f"⏱ Время круга: <b>{elapsed_text}</b>\n\n"
        + "<b>🔎 Текущий круг</b>\n"
        + f"📄 Страниц TODAY собрано: <b>{pages_verified}</b> <i>(макс. {total_pages_target})</i>\n"
        + f"🧾 Обнаружено чистых объявлений: <b>{listings_seen}</b> · новых <b>{new_listings}</b>\n"
        + f"👁 Точные просмотры: <b>{views_verified}/{views_requested}</b>"
        + (f" · ❓ tail <b>{int(state.get('view_tail_deferred') or 0)}</b>" if int(state.get('view_tail_deferred') or 0) else "") + "\n"
        + f"🎯 Baseline создано: <b>{baselines}</b>\n"
        + f"✅ Категорий успешно: <b>{successful}</b>"
        + (f" · 🟡 exact-tail <b>{int(state.get('view_tail_categories') or 0)}</b> кат." if int(state.get('view_tail_categories') or 0) else "")
        + f" · ⚠️ допроверка <b>{needs_review}</b> · ❌ ошибок <b>{system_errors}</b>\n\n"
        + "<b>🧪 Radar-наблюдения</b>\n"
        + "Подробная воронка Candidate / Early / Strong / Hot, Confidence, Acceleration и категории вынесены в <b>📊 Аналитика Radar</b>.\n"
        + "Так тяжёлая статистика больше не может заблокировать управление AutoScan.\n\n"
        + f"Последний круг: <b>{html.escape(last_line)}</b>\n"
        + f"Следующий ежедневный: <b>{html.escape(_radar_autoscan_next_run_text(state))}</b>"
    )
    return text, state


async def _radar3_analytics_text() -> str:
    """Deep Radar analytics. It is intentionally isolated from the live control panel."""
    radar3 = await _radar3_dashboard_safe_snapshot(timeout_seconds=3.0)
    category_lines = list(radar3.get("category_lines") or [])
    category_text = "\n".join(category_lines[:10]) if category_lines else "Пока подтверждённых категорий нет"
    if not any(k in radar3 for k in ("active", "early", "strong", "hot")):
        return (
            "<b>📊 DT Radar 3.2 · ADAPTIVE ANALYTICS</b>\n\n"
            "⚠️ Глубокая статистика сейчас считается или PostgreSQL занят.\n"
            "Live Status и AutoScan при этом продолжают работать независимо.\n\n"
            + category_text
        )
    return (
        "<b>📊 DT Radar 3.2 · ADAPTIVE ANALYTICS</b>\n\n"
        "<b>🔬 Воронка Radar 3.2 · по категориям</b>\n"
        f"Активных наблюдений: <b>{int(radar3.get('active') or 0)}</b>\n"
        f"Ждут первого повторного замера: <b>{int(radar3.get('baseline') or 0)}</b>\n"
        f"Готовы к замеру сейчас: <b>{int(radar3.get('due') or 0)}</b>\n"
        f"Любой DT-observed прирост: <b>{int(radar3.get('any_growth') or 0)}</b>\n"
        f"🟠 Candidate · топ-10% категории: <b>{int(radar3.get('candidate') or 0)}</b>\n"
        f"⭐ Early/Score · топ-5% категории: <b>{int(radar3.get('observed') or 0)}</b>\n"
        f"⚡ Strong interval · топ-2% категории: <b>{int(radar3.get('strong_intervals') or 0)}</b>\n"
        f"🔁 Score подтверждён ≥2 раза: <b>{int(radar3.get('persistent') or 0)}</b>\n"
        f"🚀 Ускоряются ≥20%: <b>{int(radar3.get('accelerating') or 0)}</b>\n"
        f"🛡 Confidence ≥70%: <b>{int(radar3.get('high_confidence') or 0)}</b>\n"
        f"⚫ Noise / Weak: <b>{int(radar3.get('quiet') or 0)}</b>\n"
        f"Суммарный DT-observed прирост: <b>+{int(radar3.get('total_delta') or 0)}</b>\n\n"
        f"🟡 Early: <b>{int(radar3.get('early') or 0)}</b> · 📈 Strong: <b>{int(radar3.get('strong') or 0)}</b> · 🔥 Hot: <b>{int(radar3.get('hot') or 0)}</b>\n\n"
        "<b>🗂 Категории с живым спросом</b>\n"
        + category_text + "\n\n"
        "<i>Radar 3.2: &lt;3/ч — шум. Дальше объявление сравнивается только со своей категорией: "
        "P90 Candidate · P95 Early/Score · P98 Strong · P99 Hot при подтверждении. "
        "DT Score = 50% позиция в категории + 25% устойчивость + 15% ускорение + 10% повторяемость.</i>"
    )


def _radar_autoscan_loading_text() -> str:
    return (
        "<b>📡 DT Radar 3.2 · CONTEXT DEMAND</b>\n\n"
        "Панель открыта ✅\n"
        "⚙️ Управление AutoScan доступно сразу.\n"
        "Статистика Radar загружается… ⏳"
    )


def admin_radar_autoscan_loading_keyboard() -> InlineKeyboardMarkup:
    """Emergency controls if the lightweight live state itself is temporarily unavailable."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="▶️ Запустить AutoScan", callback_data="adminradarauto:start"),
         InlineKeyboardButton(text="⏹ Остановить", callback_data="adminradarauto:stop")],
        [InlineKeyboardButton(text="📊 Аналитика Radar", callback_data="adminradarauto:analytics")],
        [InlineKeyboardButton(text="🔄 Обновить Live", callback_data="adminradarauto")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


async def _radar_autoscan_safe_text(timeout_seconds: float = 5.0) -> tuple[str, dict]:
    """Open Radar even when PostgreSQL is busy or one diagnostic query fails."""
    try:
        return await asyncio.wait_for(_radar_autoscan_text(), timeout=max(1.0, float(timeout_seconds)))
    except Exception as exc:
        log.exception("DT Radar 3.2 control panel failed")
        state = _radar_autoscan_default_state()
        reason = html.escape(type(exc).__name__)
        text = (
            "<b>📡 DT Radar 3.2 · CONTEXT DEMAND</b>\n\n"
            "Панель открыта ✅\n"
            "Radar 3.2 продолжает работать в фоне.\n\n"
            "⚠️ Статистика временно недоступна.\n"
            f"Диагностика: <code>{reason}</code>"
        )
        return text, state


def _radar_autoscan_failure_list(state: dict) -> list[dict]:
    """Return structured category failures from the last AutoScan summary."""
    last = dict(state.get("last_summary") or {})
    return [item for item in list(last.get("failed_categories") or []) if isinstance(item, dict)]


async def _radar_autoscan_errors_text(page: int = 0, per_page: int = 15) -> tuple[str, dict, int, int]:
    state = await load_radar_autoscan_state()
    failures = _radar_autoscan_failure_list(state)
    total = len(failures)
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(0, min(pages - 1, int(page or 0)))
    lines = ["<b>⚠️ DT Radar · категории для допроверки</b>"]
    if not failures:
        legacy_failed = int(dict(state.get("last_summary") or {}).get("failed") or 0)
        if legacy_failed:
            lines += ["", f"В последнем круге было <b>{legacy_failed}</b> ошибок, но он выполнен на v4.11.0, которая ещё не сохраняла список категорий.",
                      "Начиная с v4.11.1 причины и категории сохраняются автоматически."]
        else:
            lines += ["", "Ошибок в последнем круге нет."]
        return "\n".join(lines), state, page, pages
    start = page * per_page
    for idx, item in enumerate(failures[start:start + per_page], start=start + 1):
        reason = str(item.get("reason") or "partial").replace("\n", " ").strip()
        if len(reason) > 120:
            reason = reason[:117] + "…"
        kind = str(item.get("kind") or "partial")
        icon = "❌" if kind == "system" else "⚠️"
        label = "системная" if kind == "system" else "допроверка"
        lines += [
            "",
            f"<b>{icon} {idx}. {html.escape(str(item.get('name') or item.get('key') or '—'))}</b>",
            f"📄 {int(item.get('verified_pages') or 0)}/{int(item.get('depth') or RADAR_AUTOSCAN_DEPTH)} · {label} · {html.escape(reason)}",
        ]
    lines += ["", f"Страница <b>{page + 1}/{pages}</b> · проблемных категорий <b>{total}</b>"]
    return "\n".join(lines), state, page, pages


async def _radar_autoscan_history_text() -> str:
    state = await load_radar_autoscan_state()
    history = list(state.get("history") or [])[:10]
    lines = ["<b>📜 DT Radar · история кругов</b>"]
    if not history:
        lines += ["", "Кругов пока не было."]
        return "\n".join(lines)
    for item in history:
        mode = ("ежедневный" if item.get("mode") == "daily" else ("повтор" if item.get("mode") == "retry" else "ручной"))
        lines += [
            "",
            f"<b>#{html.escape(str(item.get('round_id') or '—'))}</b> · {mode}",
            f"✅ {int(item.get('coverage_successful') or item.get('successful') or 0)}/{int(item.get('coverage_total') or item.get('total') or 0)} · ⚠️ {int(item.get('needs_review') or 0)} · ❌ {int(item.get('system_errors') or 0)}",
            f"📊 {int(item.get('category_coverage_pct') or 0)}% кат. · {int(item.get('page_coverage_pct') or 0)}% стр. от max",
            f"📄 {int(item.get('pages_verified') or 0)} стр. · 🧾 {int(item.get('listings_seen') or 0)} объявлений",
            f"👁 {int(item.get('views_verified') or 0)}/{int(item.get('views_requested') or 0)} exact · 🛡 {int(item.get('radar_organic_passed') or 0)} organic",
            f"📡 Baseline {int(item.get('radar_candidates') or 0)} · ⏱ {_radar_autoscan_human_duration(int(item.get('duration_seconds') or 0))}",
            f"🕒 {html.escape(str(item.get('finished_at_text') or '—'))}",
        ]
    return "\n".join(lines)


async def _radar_autoscan_finish_round(bot: Bot, state: dict) -> dict:
    now = datetime.now(MOSCOW)
    duration = _radar_autoscan_duration_seconds(str(state.get("started_at") or ""))
    run_total = int(state.get("total") or 0)
    run_successful = int(state.get("successful") or 0)
    needs_review = int(state.get("needs_review") or 0)
    system_errors = int(state.get("system_errors") or 0)
    failed = max(int(state.get("failed") or 0), needs_review + system_errors)
    skipped_nonproduct = int(state.get("skipped_nonproduct") or 0)
    mode = str(state.get("mode") or "manual")
    retry_parent_total = int(state.get("retry_parent_total") or 0)
    retry_parent_successful = int(state.get("retry_parent_successful") or 0)
    if mode == "retry" and retry_parent_total:
        coverage_total = retry_parent_total
        coverage_successful = min(coverage_total, retry_parent_successful + run_successful)
    else:
        coverage_total = run_total
        coverage_successful = run_successful
    category_coverage_pct = int(round((coverage_successful / max(1, coverage_total)) * 100))
    if mode == "retry":
        parent = dict(state.get("last_summary") or {})
        parent_failures = [x for x in list(parent.get("failed_categories") or []) if isinstance(x, dict)]
        parent_pages = int(parent.get("coverage_pages_verified") or parent.get("pages_verified") or 0)
        replaced_partial_pages = sum(max(0, int(x.get("verified_pages") or 0)) for x in parent_failures)
        coverage_pages_verified = max(0, parent_pages - replaced_partial_pages + int(state.get("pages_verified") or 0))
    else:
        coverage_pages_verified = int(state.get("pages_verified") or 0)
    page_coverage_pct = int(round((coverage_pages_verified / max(1, coverage_total * _radar_layer_depth(state))) * 100))
    page_coverage_pct = max(0, min(100, page_coverage_pct))
    failed_categories = list(state.get("failed_categories") or [])
    summary = {
        "round_id": str(state.get("round_id") or ""),
        "mode": mode,
        "layer": str(state.get("layer") or ("context" if mode == "context" else "fresh")),
        "context_for_date": str(state.get("context_for_date") or ""),
        "target_date": str(state.get("target_date") or ""),
        "total": run_total,
        "processed": int(state.get("processed") or 0),
        "successful": run_successful,
        "needs_review": needs_review,
        "system_errors": system_errors,
        "skipped_nonproduct": skipped_nonproduct,
        "failed": failed,
        "coverage_total": coverage_total,
        "coverage_successful": coverage_successful,
        "category_coverage_pct": max(0, min(100, category_coverage_pct)),
        "page_coverage_pct": page_coverage_pct,
        "coverage_pages_verified": coverage_pages_verified,
        "pages_verified": int(state.get("pages_verified") or 0),
        "listings_seen": int(state.get("listings_seen") or 0),
        "new_listings": int(state.get("new_listings") or 0),
        "search_promoted_filtered": int(state.get("search_promoted_filtered") or 0),
        "search_reduced_filtered": int(state.get("search_reduced_filtered") or 0),
        "views_requested": int(state.get("views_requested") or 0),
        "views_verified": int(state.get("views_verified") or 0),
        "views_failed": int(state.get("views_failed") or 0),
        "view_tail_deferred": int(state.get("view_tail_deferred") or 0),
        "view_tail_categories": int(state.get("view_tail_categories") or 0),
        "radar_candidates": int(state.get("radar_candidates") or 0),
        "radar_high_baseline_pending": int(state.get("radar_high_baseline_pending") or 0),
        "radar_high_baseline_verified": int(state.get("radar_high_baseline_verified") or 0),
        "radar_detail_checked": int(state.get("radar_detail_checked") or 0),
        "radar_organic_passed": int(state.get("radar_organic_passed") or 0),
        "radar_promoted_blocked": int(state.get("radar_promoted_blocked") or 0),
        "radar_reduced_blocked": int(state.get("radar_reduced_blocked") or 0),
        "radar_unknown_blocked": int(state.get("radar_unknown_blocked") or 0),
        "radar_unknown_reasons": dict(state.get("radar_unknown_reasons") or {}),
        "radar_db_blocked": int(state.get("radar_db_blocked") or 0),
        "radar_demand_gate_rejected": int(state.get("radar_demand_gate_rejected") or 0),
        "radar_qualified_candidates": int(state.get("radar_qualified_candidates") or 0),
        "radar_early_admitted": int(state.get("radar_early_admitted") or 0),
        "radar_strong_admitted": int(state.get("radar_strong_admitted") or 0),
        "radar_hot_admitted": int(state.get("radar_hot_admitted") or 0),
        "radar_already_present": int(state.get("radar_already_present") or 0),
        "radar_saved": int(state.get("radar_saved") or 0),
        "failed_categories": failed_categories,
        "retry_parent_round_id": str(state.get("retry_parent_round_id") or ""),
        "duration_seconds": duration,
        "finished_at": now.replace(microsecond=0).isoformat(),
        "finished_at_text": now.strftime("%d.%m.%Y %H:%M МСК"),
    }
    state["history"] = [summary] + list(state.get("history") or [])
    state["history"] = state["history"][:RADAR_AUTOSCAN_HISTORY_LIMIT]
    state["last_summary"] = summary
    if coverage_successful >= coverage_total and failed == 0 and mode in {"manual", "daily"}:
        state["last_completed_date"] = now.date().isoformat()
    if mode == "daily":
        state["last_daily_date"] = now.date().isoformat()
    if mode == "context":
        state["last_context_date"] = str(state.get("context_for_date") or now.date().isoformat())
    state["status"] = "idle"
    state["stop_requested"] = False
    state["waiting_for_users"] = False
    state["current_category_key"] = ""
    state["current_category_name"] = ""
    state = await save_radar_autoscan_state(state)
    start_context_after_fresh = False
    icon = "✅" if failed == 0 else "⚠️"
    if mode == "retry":
        headline = f"📡 <b>DT Radar — повтор ошибок завершён {icon}</b>"
        scope_line = f"Повторено: <b>{run_successful}/{run_total}</b> успешно"
        if needs_review:
            scope_line += f" · ⚠️ допроверка <b>{needs_review}</b>"
        if system_errors:
            scope_line += f" · ❌ системных <b>{system_errors}</b>"
    elif mode == "context":
        headline = f"📡 <b>DT Radar 3.0 — круг завершён {icon}</b>"
        scope_line = f"Сегодня: <b>{run_successful}/{run_total}</b> категорий подтверждено"
        if needs_review:
            scope_line += f" · ⚠️ допроверка <b>{needs_review}</b>"
        if system_errors:
            scope_line += f" · ❌ системных <b>{system_errors}</b>"
    else:
        headline = f"📡 <b>DT Radar 3.0 — сегодняшний круг завершён {icon}</b>"
        scope_line = f"Сегодня: <b>{run_successful}/{run_total}</b> категорий подтверждено"
        if needs_review:
            scope_line += f" · ⚠️ допроверка <b>{needs_review}</b>"
        if system_errors:
            scope_line += f" · ❌ системных <b>{system_errors}</b>"
    notify_keyboard = None
    if failed_categories:
        notify_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"⚠️ Требуют внимания — {len(failed_categories)}", callback_data="adminradarauto:errors:0")],
            [InlineKeyboardButton(text="🔁 Повторить проблемные", callback_data="adminradarauto:retry")],
        ])
    await _notify_radar_autoscan_admins(
        bot,
        headline + "\n\n"
        + scope_line
        + f"\n📊 Покрытие категорий: <b>{coverage_successful}/{coverage_total} · {category_coverage_pct}%</b>"
        + f"\n📊 Страниц от максимума: <b>{page_coverage_pct}%</b> · <b>{coverage_pages_verified}/{coverage_total * _radar_layer_depth(state)}</b>"
        + (f"\n⏭ Нетоварных пропущено: <b>{skipped_nonproduct}</b>" if skipped_nonproduct else "")
        + f"\n📄 Страниц TODAY собрано в этом запуске: <b>{int(summary['pages_verified'])}</b>"
        + f"\n🧾 Чистых объявлений даты: <b>{int(summary['listings_seen'])}</b> · новых <b>{int(summary['new_listings'])}</b>"
        + f"\n🚫 Сразу исключено: TOP/Promo <b>{int(summary['search_promoted_filtered'])}</b> · снижение <b>{int(summary['search_reduced_filtered'])}</b>"
        + f"\n👁 Точные просмотры: <b>{int(summary['views_verified'])}/{int(summary['views_requested'])}</b>"
        + (f" · ❓ tail <b>{int(summary.get('view_tail_deferred') or 0)}</b>" if int(summary.get('view_tail_deferred') or 0) else "")
        + (
            f"\n🟡 Initial ≥{ORGANIC_HIGH_BASELINE_VIEWS}: <b>{int(summary.get('radar_high_baseline_pending') or 0)}</b> ждут 2 замера · "
            f"✅ delta verified <b>{int(summary.get('radar_high_baseline_verified') or 0)}</b>"
            if int(summary.get('radar_high_baseline_pending') or 0) or int(summary.get('radar_high_baseline_verified') or 0) else ""
        )
        + f"\n📡 Radar 3.0 baseline создано: <b>{int(summary.get('radar_candidates') or 0)}</b>"
        + "\n⏱ Первый счётчик не оценивается; сигналы появятся только после повторных замеров DT"
        + f"\n🛡 Organic: <b>{int(summary['radar_organic_passed'])}</b> · TOP/Promo <b>{int(summary['radar_promoted_blocked'])}</b> · снижение <b>{int(summary['radar_reduced_blocked'])}</b> · unknown <b>{int(summary['radar_unknown_blocked'])}</b>"
        + (f"\n↳ unknown: {_radar_unknown_reason_text(summary.get('radar_unknown_reasons'))}" if int(summary.get('radar_unknown_blocked') or 0) else "")
        + f"\n📍 Немедленных сигналов из baseline: <b>{int(summary['radar_saved'])}</b> (должно быть 0)"
        + (f" · уже были <b>{int(summary.get('radar_already_present') or 0)}</b>" if int(summary.get("radar_already_present") or 0) else "")
        + f"\n⏱ Время: <b>{_radar_autoscan_human_duration(duration)}</b>"
        + "\n\nКруг завершён. Повторные DT-замеры baseline продолжаются автоматически."
        + (
            f" Следующий ежедневный запуск: <b>{html.escape(_radar_autoscan_next_run_text(state))}</b>."
            if state.get("daily_enabled") and not start_context_after_fresh else ""
        ),
        reply_markup=notify_keyboard,
    )
    return state


async def _radar_autoscan_interruptible_sleep(seconds: float) -> bool:
    """Sleep without making Stop wait for a category cooldown/poll interval.

    Returns True when the stop event interrupted the wait.
    """
    seconds = max(0.0, float(seconds or 0.0))
    if _radar_autoscan_stop_event.is_set():
        return True
    if seconds <= 0:
        await asyncio.sleep(0)
        return _radar_autoscan_stop_event.is_set()
    try:
        await asyncio.wait_for(_radar_autoscan_stop_event.wait(), timeout=seconds)
        return True
    except asyncio.TimeoutError:
        return False


async def _radar_autoscan_run_category_controlled(coro, *, category_name: str):
    """Run one category with hard-stop and watchdog semantics.

    The child task is owned exclusively by AutoScan, so cancellation is safe: on
    Stop the current category is NOT advanced and Resume retries it from scratch.
    On watchdog the category is recorded as partial and the round continues.
    """
    task = asyncio.create_task(coro, name=f"dt-radar-autoscan-category-{str(category_name)[:40]}")
    stop_waiter = asyncio.create_task(_radar_autoscan_stop_event.wait(), name="dt-radar-autoscan-hard-stop")
    try:
        done, _pending = await asyncio.wait(
            {task, stop_waiter},
            timeout=RADAR_AUTOSCAN_CATEGORY_TIMEOUT_SECONDS,
            return_when=asyncio.FIRST_COMPLETED,
        )
        if stop_waiter in done and _radar_autoscan_stop_event.is_set():
            if not task.done():
                task.cancel()
            await asyncio.gather(task, return_exceptions=True)
            raise RadarAutoScanStopped()
        if task in done:
            return await task
        if not task.done():
            task.cancel()
        await asyncio.gather(task, return_exceptions=True)
        raise RadarAutoScanCategoryTimeout(
            f"category watchdog {int(RADAR_AUTOSCAN_CATEGORY_TIMEOUT_SECONDS)}s"
        )
    finally:
        if not stop_waiter.done():
            stop_waiter.cancel()
        await asyncio.gather(stop_waiter, return_exceptions=True)


def _radar_autoscan_live_stage_line(state: dict) -> str:
    """Human-readable current category heartbeat for the admin panel."""
    key = str(state.get("current_category_key") or "")
    target = str(state.get("target_date") or "")
    if not key or not target:
        stage = str(state.get("current_stage") or "")
        return f"⚙️ Этап: <b>{html.escape(stage)}</b>" if stage else ""
    try:
        depth = _radar_layer_depth(state)
        live = category_live_progress.get(_progress_key(key, target, depth))
    except Exception:
        live = None
    if live is None:
        stage = str(state.get("current_stage") or "")
        labels = {"starting": "запуск", "scan": "поиск даты / страницы", "organic_gate": "Organic detail-check", "context_gate": "48H Context organic-check"}
        return f"⚙️ Этап: <b>{html.escape(labels.get(stage, stage))}</b>" if stage else ""
    phase = str(getattr(live, "phase", "") or "")
    if phase == "views":
        total_views = max(0, int(getattr(live, "today_seen", 0) or 0))
        ready = min(total_views, max(0, int(getattr(live, "views_ready", 0) or 0)))
        return f"👁 Этап: <b>точные просмотры {ready}/{total_views}</b>"
    if phase in {"collecting", "regional_date"}:
        pages = max(0, min(depth, int(getattr(live, "collection_index", 0) or 0)))
        return f"📄 Этап: <b>страницы {pages}/{depth}</b> · объявлений <b>{int(getattr(live, 'today_seen', 0) or 0)}</b>"
    requests = max(0, int(getattr(live, "network_requests", 0) or 0))
    page = max(0, int(getattr(live, "page", 0) or 0))
    transport = str(getattr(live, "transport_stage", "") or "")
    extra = f" · {html.escape(transport)}" if transport else ""
    return f"🔎 Этап: <b>поиск даты</b> · запросов <b>{requests}</b> · стр. <b>{page or '—'}</b>{extra}"


async def _run_radar_autoscan_round_inner(bot: Bot) -> None:
    """Run/resume one persistent low-priority round until complete or paused.

    v4.11.5 keeps one parser/http/browser session for the whole round instead of
    constructing/closing Chromium context state 80+ times. The session is recycled only
    after an unexpected system error; partial page/date evidence gets a clean browser
    context plus bounded cooldown before the next category.
    """
    state = await load_radar_autoscan_state()
    if state.get("status") != "running":
        return
    keys = list(state.get("category_keys") or [])
    total = len(keys)
    round_depth = _radar_layer_depth(state)
    state["total"] = total
    if not state.get("target_date"):
        state["target_date"] = datetime.now(MOSCOW).date().isoformat()
        state = await save_radar_autoscan_state(state)

    # Pause low-priority Radar sweep / 400+ checkpoints for the whole round.
    # An already-leased background request may finish, but no new background
    # detail/view request can start until AutoScan exits/pauses.
    parser = KleinanzeigenParser()
    background_paused = False
    issue_streak = 0
    try:
        await TRAFFIC.background_pause_started()
        background_paused = True
        while int(state.get("current_index") or 0) < total:
            # Reload on every category boundary so admin Stop/Resume changes are observed.
            state = await load_radar_autoscan_state()
            if state.get("status") != "running":
                return
            if state.get("stop_requested"):
                state["status"] = "paused"
                state["stop_requested"] = False
                state["waiting_for_users"] = False
                state["current_stage"] = "paused"
                state = await save_radar_autoscan_state(state)
                _radar_autoscan_stop_event.set()
                await _notify_radar_autoscan_admins(
                    bot,
                    f"⏸ <b>DT Radar AutoScan остановлен</b>\n\n"
                    f"Прогресс сохранён: <b>{int(state.get('processed') or 0)}/{total}</b>. "
                    "Можно продолжить круг с сохранённого места.",
                )
                return

            # User scans are always first. Do not start a fresh category while any
            # foreground scan is running or queued. A category already in progress is
            # allowed to finish so parser integrity is never sacrificed.
            running, queued = await _radar_foreground_counts()
            if running or queued:
                if not state.get("waiting_for_users"):
                    state["waiting_for_users"] = True
                    state["current_category_name"] = "Ожидание пользовательских сканов"
                    state = await save_radar_autoscan_state(state)
                    log.info("DT Radar AutoScan yielding to users running=%s queued=%s", running, queued)
                if await _radar_autoscan_interruptible_sleep(5):
                    return
                continue
            if state.get("waiting_for_users"):
                state["waiting_for_users"] = False
                state = await save_radar_autoscan_state(state)

            idx = int(state.get("current_index") or 0)
            if idx >= total:
                break
            key = keys[idx]
            cat = CATEGORIES.get(key)
            if cat is None or bool(getattr(cat, "is_group", False)):
                state["current_index"] = idx + 1
                state["processed"] = int(state.get("processed") or 0) + 1
                state["system_errors"] = int(state.get("system_errors") or 0) + 1
                state["failed"] = int(state.get("failed") or 0) + 1
                state.setdefault("failed_categories", []).append({
                    "key": str(key), "name": str(key), "reason": "категория недоступна",
                    "verified_pages": 0, "depth": round_depth, "kind": "system",
                })
                state = await save_radar_autoscan_state(state)
                issue_streak += 1
                invalid_backoff = min(RADAR_AUTOSCAN_MAX_BACKOFF_SECONDS, RADAR_AUTOSCAN_SYSTEM_BACKOFF_BASE_SECONDS * (2 ** min(2, issue_streak - 1)))
                if await _radar_autoscan_interruptible_sleep(invalid_backoff):
                    return
                continue

            # A persisted v4.11.4 round may still contain all 141 categories. Do not
            # waste new traffic on the remaining non-product/service leaves after deploy;
            # skip them cleanly without calling the parser. Fresh rounds contain only the
            # product policy set from the start.
            if not _radar_autoscan_category_allowed(cat):
                state["current_index"] = idx + 1
                state["processed"] = int(state.get("processed") or 0) + 1
                state["skipped_nonproduct"] = int(state.get("skipped_nonproduct") or 0) + 1
                state["current_category_key"] = ""
                state["current_category_name"] = ""
                state = await save_radar_autoscan_state(state)
                log.info(
                    "DT Radar AutoScan skipped non-product category round=%s index=%s/%s category=%s group=%s",
                    state.get("round_id"), idx + 1, total, cat.name, cat.group,
                )
                await asyncio.sleep(0)
                continue

            state["current_category_key"] = cat.key
            state["current_category_name"] = cat.name
            state["current_stage"] = "scan"
            state["current_stage_started_at"] = _radar_autoscan_now_iso()
            state = await save_radar_autoscan_state(state)
            log.info(
                "DT Radar AutoScan category start round=%s index=%s/%s category=%s depth=%s target=%s parser_reused=True",
                state.get("round_id"), idx + 1, total, cat.name, round_depth, state.get("target_date"),
            )

            result = None
            radar_saved = 0
            radar_stats = None
            error_text = ""
            failure_kind = ""
            recycle_parser = False
            await TRAFFIC.scan_job_started()
            parser.prepare_category_scan()
            parser_token = JOB_PARSER.set(parser)
            try:
                async def _category_pipeline():
                    local_result = await scan_one_category(
                        parser, cat, RADAR_AUTOSCAN_USER_ID, round_depth, str(state.get("target_date"))
                    )
                    local_radar_saved = 0
                    local_radar_stats = None
                    local_failure_kind = ""
                    local_error_text = ""
                    if local_result.date_complete and not local_result.radar_views_usable:
                        local_failure_kind = "radar_views"
                        local_error_text = (
                            f"Radar views существенно неполные: {int(local_result.views_verified or 0)}/"
                            f"{int(local_result.views_requested or 0)} точных счётчиков"
                        )
                    elif local_result.date_complete:
                        if not local_result.radar_views_complete and int(local_result.view_tail_count or 0) > 0:
                            log.warning(
                                "DT Radar AutoScan soft exact tail category=%s exact=%s/%s tail=%s budget=%s; "
                                "verified rows seed baseline, UNKNOWN rows stay excluded without full rescan",
                                cat.name, int(local_result.views_verified or 0), int(local_result.views_requested or 0),
                                int(local_result.view_tail_count or 0),
                                _radar_autoscan_view_tail_budget(int(local_result.views_requested or 0)),
                            )
                        live_state = await load_radar_autoscan_state()
                        if live_state.get("status") == "running":
                            live_state["current_stage"] = ("context_gate" if str(state.get("layer") or "fresh") == "context" else "organic_gate")
                            live_state["current_stage_started_at"] = _radar_autoscan_now_iso()
                            await save_radar_autoscan_state(live_state)
                        source_round_id = (
                            str(state.get("retry_parent_round_id") or "").strip()
                            if str(state.get("mode") or "") == "retry"
                            else ""
                        ) or str(state.get("round_id") or "round")
                        local_radar_stats = await record_autoscan_hot_detailed(
                            source_round_id, cat.key, local_result.matched_ids or [],
                            emit_signals=True,
                        )
                        local_radar_saved = int(local_radar_stats.saved or 0)
                        expected_slots = min(RADAR_SCAN_TOP_LIMIT, int(local_radar_stats.qualified_candidates or 0))
                        if int(local_radar_stats.unknown_blocked or 0) > 0:
                            local_failure_kind = "radar_gate_unknown"
                            reason_text = _radar_unknown_reason_text(dict(local_radar_stats.unknown_reasons or ()))
                            local_error_text = (
                                f"Organic detail gate не подтвердил {int(local_radar_stats.unknown_blocked or 0)} "
                                f"вышестоящих кандидатов; подтверждено Radar {int(local_radar_stats.admitted or 0)}/{expected_slots}"
                                + (f"; причины: {reason_text}" if reason_text else "")
                            )
                    else:
                        local_failure_kind = "partial"
                        local_error_text = local_result.reason or "нужно повторно подтвердить категорию"
                    return local_result, local_radar_saved, local_radar_stats, local_failure_kind, local_error_text

                result, radar_saved, radar_stats, failure_kind, error_text = await _radar_autoscan_run_category_controlled(
                    _category_pipeline(), category_name=cat.name
                )
            except RadarAutoScanStopped:
                log.warning(
                    "DT Radar AutoScan hard stop category=%s round=%s index=%s/%s",
                    cat.name, state.get("round_id"), idx + 1, total,
                )
                try:
                    await asyncio.wait_for(parser.reset_scan_browser_context(), timeout=10.0)
                except Exception:
                    log.debug("DT Radar AutoScan hard-stop context reset failed", exc_info=True)
                stopped_state = await load_radar_autoscan_state()
                stopped_state["status"] = "paused"
                stopped_state["stop_requested"] = False
                stopped_state["waiting_for_users"] = False
                stopped_state["current_stage"] = "paused"
                await save_radar_autoscan_state(stopped_state)
                await _notify_radar_autoscan_admins(
                    bot,
                    f"⏸ <b>DT Radar AutoScan остановлен сразу</b>\n\n"
                    f"Категория <b>{html.escape(cat.name)}</b> будет начата заново после продолжения. "
                    f"Сохранено: <b>{int(stopped_state.get('processed') or 0)}/{total}</b>.",
                )
                return
            except RadarAutoScanCategoryTimeout as exc:
                failure_kind = "partial"
                error_text = f"watchdog категории: {int(RADAR_AUTOSCAN_CATEGORY_TIMEOUT_SECONDS)} сек"
                recycle_parser = True
                state["last_watchdog_category"] = cat.name
                log.error(
                    "DT Radar AutoScan category watchdog timeout round=%s category=%s seconds=%s",
                    state.get("round_id"), cat.name, int(RADAR_AUTOSCAN_CATEGORY_TIMEOUT_SECONDS),
                )
            except asyncio.CancelledError:
                raise
            except TemporaryAccessError as exc:
                # Public-site pressure is not a code/system crash. Keep it in the retry
                # list and cool down before touching the next category.
                failure_kind = "partial"
                error_text = str(exc)
                log.warning(
                    "DT Radar AutoScan temporary access category=%s round=%s error=%s",
                    cat.name, state.get("round_id"), error_text,
                )
            except (TimeoutError, asyncio.TimeoutError) as exc:
                failure_kind = "partial"
                error_text = f"временный таймаут: {exc}"[:300]
                log.warning(
                    "DT Radar AutoScan timeout category=%s round=%s error=%s",
                    cat.name, state.get("round_id"), error_text,
                )
            except Exception as exc:
                failure_kind = "system"
                error_text = str(exc) or exc.__class__.__name__
                recycle_parser = True
                log.exception("DT Radar AutoScan system failure round=%s category=%s", state.get("round_id"), cat.name)
            finally:
                JOB_PARSER.reset(parser_token)
                await TRAFFIC.scan_job_finished()

            state = await load_radar_autoscan_state()
            # A Stop pressed during the category is honored by the child waiter above.
            if error_text.startswith("watchdog категории"):
                state["last_watchdog_category"] = cat.name
            state["current_index"] = idx + 1
            state["processed"] = int(state.get("processed") or 0) + 1
            if result is not None:
                state["pages_verified"] = int(state.get("pages_verified") or 0) + min(
                    round_depth, max(0, int(result.collection_pages_confirmed or 0))
                )
                state["listings_seen"] = int(state.get("listings_seen") or 0) + max(0, int(result.today_seen or 0))
                state["new_listings"] = int(state.get("new_listings") or 0) + max(0, int(result.new_count or 0))
                state["search_promoted_filtered"] = int(state.get("search_promoted_filtered") or 0) + max(0, int(result.promoted_filtered or 0))
                state["search_reduced_filtered"] = int(state.get("search_reduced_filtered") or 0) + max(0, int(result.price_reduced_filtered or 0))
                state["views_requested"] = int(state.get("views_requested") or 0) + max(0, int(result.views_requested or 0))
                state["views_verified"] = int(state.get("views_verified") or 0) + max(0, int(result.views_verified or 0))
                state["views_failed"] = int(state.get("views_failed") or 0) + max(0, int(result.view_failures or 0))
                if result.radar_views_usable and not result.radar_views_complete and int(result.view_tail_count or 0) > 0:
                    state["view_tail_deferred"] = int(state.get("view_tail_deferred") or 0) + int(result.view_tail_count or 0)
                    state["view_tail_categories"] = int(state.get("view_tail_categories") or 0) + 1
                state["radar_saved"] = int(state.get("radar_saved") or 0) + max(0, int(radar_saved or 0))
            if radar_stats is not None:
                state["radar_candidates"] = int(state.get("radar_candidates") or 0) + int(radar_stats.eligible_with_views or 0)
                state["radar_high_baseline_pending"] = int(state.get("radar_high_baseline_pending") or 0) + int(radar_stats.high_baseline_pending or 0)
                state["radar_high_baseline_verified"] = int(state.get("radar_high_baseline_verified") or 0) + int(radar_stats.high_baseline_verified or 0)
                state["radar_detail_checked"] = int(state.get("radar_detail_checked") or 0) + int(radar_stats.detail_checked or 0)
                state["radar_organic_passed"] = int(state.get("radar_organic_passed") or 0) + int(radar_stats.organic_passed or 0)
                state["radar_promoted_blocked"] = int(state.get("radar_promoted_blocked") or 0) + int(radar_stats.promoted_blocked or 0)
                state["radar_reduced_blocked"] = int(state.get("radar_reduced_blocked") or 0) + int(radar_stats.reduced_blocked or 0)
                state["radar_unknown_blocked"] = int(state.get("radar_unknown_blocked") or 0) + int(radar_stats.unknown_blocked or 0)
                unknown_reasons = dict(state.get("radar_unknown_reasons") or {})
                for reason, count in tuple(radar_stats.unknown_reasons or ()):
                    key = str(reason or "detail_unknown")[:64]
                    unknown_reasons[key] = int(unknown_reasons.get(key) or 0) + int(count or 0)
                state["radar_unknown_reasons"] = unknown_reasons
                state["radar_db_blocked"] = int(state.get("radar_db_blocked") or 0) + int(radar_stats.db_blocked or 0)
                state["radar_demand_gate_rejected"] = int(state.get("radar_demand_gate_rejected") or 0) + int(radar_stats.demand_gate_rejected or 0)
                state["radar_qualified_candidates"] = int(state.get("radar_qualified_candidates") or 0) + int(radar_stats.qualified_candidates or 0)
                state["radar_early_admitted"] = int(state.get("radar_early_admitted") or 0) + int(radar_stats.early_admitted or 0)
                state["radar_strong_admitted"] = int(state.get("radar_strong_admitted") or 0) + int(radar_stats.strong_admitted or 0)
                state["radar_hot_admitted"] = int(state.get("radar_hot_admitted") or 0) + int(radar_stats.hot_admitted or 0)
                state["radar_already_present"] = int(state.get("radar_already_present") or 0) + int(radar_stats.already_present or 0)

            if result is not None and result.date_complete and not failure_kind:
                state["successful"] = int(state.get("successful") or 0) + 1
                issue_streak = 0
                try:
                    rollover_retired = await radar_v3_rollover_successful_category(
                        cat.key, result.matched_ids or []
                    )
                    if rollover_retired:
                        log.info(
                            "DT Radar AutoScan category freshness rollover round=%s category=%s retired=%s",
                            state.get("round_id"), cat.name, rollover_retired,
                        )
                except Exception:
                    # Category scanning and evidence collection already succeeded;
                    # a DB-only catalogue rollover must never downgrade that scan to
                    # a parser failure. The 24h hard cap remains the safe fallback.
                    log.exception(
                        "DT Radar AutoScan category freshness rollover failed round=%s category=%s",
                        state.get("round_id"), cat.name,
                    )
            else:
                state["failed"] = int(state.get("failed") or 0) + 1
                if failure_kind == "system":
                    state["system_errors"] = int(state.get("system_errors") or 0) + 1
                else:
                    failure_kind = failure_kind or "partial"
                    state["needs_review"] = int(state.get("needs_review") or 0) + 1
                state.setdefault("failed_categories", []).append({
                    "key": cat.key,
                    "name": cat.name,
                    "reason": error_text or "нужно повторно подтвердить категорию",
                    "verified_pages": min(
                        round_depth,
                        max(0, int(result.collection_pages_confirmed or 0)) if result is not None else 0,
                    ),
                    "depth": round_depth,
                    "kind": failure_kind,
                })
                issue_streak += 1

            log.info(
                "DT Radar AutoScan category finish round=%s index=%s/%s category=%s complete=%s radar_saved=%s kind=%s error=%s",
                state.get("round_id"), idx + 1, total, cat.name,
                bool(result and result.date_complete), radar_saved, failure_kind or "ok", error_text[:240],
            )
            state["current_category_key"] = ""
            state["current_category_name"] = ""
            state["current_stage"] = ""
            state["current_stage_started_at"] = ""
            state = await save_radar_autoscan_state(state)

            if result is not None and not result.date_complete and not recycle_parser:
                # Keep the long-lived HTTP client/cookies, but drop any bad browser page
                # verdict before the next category. The parser itself stays reusable.
                try:
                    await parser.reset_scan_browser_context()
                except Exception:
                    log.debug("DT Radar AutoScan browser context reset failed", exc_info=True)

            if recycle_parser:
                try:
                    await asyncio.wait_for(parser.close(), timeout=15.0)
                except Exception:
                    log.debug("DT Radar AutoScan parser close during recycle failed", exc_info=True)
                parser = KleinanzeigenParser()
                log.warning("DT Radar AutoScan parser recycled after system error category=%s", cat.name)

            if failure_kind:
                base = RADAR_AUTOSCAN_SYSTEM_BACKOFF_BASE_SECONDS if failure_kind == "system" else RADAR_AUTOSCAN_PARTIAL_BACKOFF_BASE_SECONDS
                cooldown = min(RADAR_AUTOSCAN_MAX_BACKOFF_SECONDS, base * (2 ** min(3, max(0, issue_streak - 1))))
                log.info(
                    "DT Radar AutoScan cooldown kind=%s streak=%s seconds=%.1f category=%s",
                    failure_kind, issue_streak, cooldown, cat.name,
                )
                if await _radar_autoscan_interruptible_sleep(cooldown):
                    return
            else:
                if await _radar_autoscan_interruptible_sleep(RADAR_AUTOSCAN_SUCCESS_GAP_SECONDS):
                    return

        state = await load_radar_autoscan_state()
        if state.get("status") == "running" and int(state.get("current_index") or 0) >= total:
            await _radar_autoscan_finish_round(bot, state)
    finally:
        try:
            await asyncio.wait_for(parser.close(), timeout=15.0)
        except Exception:
            log.debug("DT Radar AutoScan persistent parser close failed", exc_info=True)
        if background_paused:
            await TRAFFIC.background_pause_finished()


async def _run_radar_autoscan_round(bot: Bot) -> None:
    """Single-flight wrapper for all AutoScan launch paths."""
    async with _radar_autoscan_run_guard:
        state = await load_radar_autoscan_state()
        if state.get("status") != "running":
            return
        log.info(
            "DT Radar AutoScan runner entered round=%s index=%s/%s mode=%s",
            state.get("round_id"), int(state.get("current_index") or 0),
            int(state.get("total") or 0), state.get("mode"),
        )
        await _run_radar_autoscan_round_inner(bot)


def _kick_radar_autoscan(bot: Bot, reason: str) -> None:
    """Start the runner immediately without depending only on scheduler wake-up."""
    global _radar_autoscan_kick_task
    if _radar_autoscan_kick_task is not None and not _radar_autoscan_kick_task.done():
        log.info("DT Radar AutoScan kick already active reason=%s", reason)
        return

    async def _runner() -> None:
        try:
            log.info("DT Radar AutoScan immediate kick reason=%s", reason)
            await _run_radar_autoscan_round(bot)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("DT Radar AutoScan immediate kick failed reason=%s", reason)

    _radar_autoscan_kick_task = asyncio.create_task(
        _runner(), name=f"dt-radar-autoscan-kick-{reason}"
    )


def _schedule_radar_autoscan_launch_watchdog(bot: Bot, round_id: str) -> None:
    """Re-kick a manual round if it is still untouched after the launch grace period."""
    async def _watch() -> None:
        await asyncio.sleep(RADAR_AUTOSCAN_LAUNCH_WATCHDOG_SECONDS)
        try:
            state = await load_radar_autoscan_state()
            if str(state.get("round_id") or "") != str(round_id or ""):
                return
            if state.get("status") != "running" or state.get("waiting_for_users"):
                return
            untouched = (
                int(state.get("current_index") or 0) == 0
                and not str(state.get("current_category_key") or "").strip()
                and int(state.get("processed") or 0) == 0
            )
            if untouched:
                log.warning(
                    "DT Radar AutoScan launch watchdog re-kick round=%s after=%ss",
                    round_id, RADAR_AUTOSCAN_LAUNCH_WATCHDOG_SECONDS,
                )
                _kick_radar_autoscan(bot, "watchdog")
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("DT Radar AutoScan launch watchdog failed round=%s", round_id)

    asyncio.create_task(_watch(), name=f"dt-radar-autoscan-watchdog-{round_id}")


async def radar_autoscan_scheduler(bot: Bot) -> None:
    """Resume a persisted round or start one configured daily round in Moscow time."""
    await asyncio.sleep(12)
    while True:
        try:
            state = await load_radar_autoscan_state()
            if state.get("status") == "running":
                await _run_radar_autoscan_round(bot)
                continue

            now = datetime.now(MOSCOW)
            today = now.date().isoformat()
            if state.get("daily_enabled") and state.get("status") == "idle":
                try:
                    hh, mm = [int(x) for x in str(state.get("daily_time") or RADAR_AUTOSCAN_DEFAULT_TIME).split(":", 1)]
                except Exception:
                    hh, mm = 5, 0
                due = now >= now.replace(hour=hh, minute=mm, second=0, microsecond=0)
                if due:
                    if state.get("last_daily_date") != today:
                        if state.get("skip_daily_if_completed_today", True) and state.get("last_completed_date") == today:
                            state["last_daily_date"] = today
                            state = await save_radar_autoscan_state(state)
                            await _notify_radar_autoscan_admins(
                                bot,
                                "📡 <b>DT Radar 3.0 — сегодняшний круг уже готов</b>\n\n"
                                "20 страниц за сегодня уже собраны ручным запуском. Повторные DT-замеры продолжаются автоматически.",
                            )
                        else:
                            async with _radar_autoscan_guard:
                                current = await load_radar_autoscan_state()
                                if current.get("status") == "idle":
                                    current = _radar_autoscan_new_round(current, "daily")
                                    await save_radar_autoscan_state(current)
                                    _radar_autoscan_stop_event.clear()
                                    await _notify_radar_autoscan_admins(
                                        bot,
                                        f"📡 <b>DT Radar 3.0 — сегодняшний круг запущен</b>\n\n"
                                        f"{len(current.get('category_keys') or [])} товарных категорий × {RADAR_AUTOSCAN_DEPTH} страниц только за сегодня. "
                                        "Пользовательские сканы имеют приоритет и тоже добавляют сегодняшние baseline.",
                                    )
                                    continue

            try:
                await asyncio.wait_for(_radar_autoscan_wakeup.wait(), timeout=RADAR_AUTOSCAN_POLL_SECONDS)
                _radar_autoscan_wakeup.clear()
            except asyncio.TimeoutError:
                pass
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("DT Radar AutoScan scheduler error")
            await asyncio.sleep(RADAR_AUTOSCAN_POLL_SECONDS)


async def send_smart_export(
    message: Message | BotChatAdapter,
    user_id: int,
    selected_count: int,
    *,
    category_keys_override: set[str] | None = None,
    rows_override: list[Listing] | None = None,
    price_filter_override: str | None = None,
) -> int:
    s = await get_settings(user_id)
    mode = s.output_mode
    effective_price_filter = (price_filter_override or "any").strip()
    all_rows = (
        list(rows_override)
        if rows_override is not None
        else await today_rows(include_price_reduced=(mode == "price_drop"))
    )
    selected_keys = category_keys_override if category_keys_override is not None else await get_selected(user_id)
    if selected_keys:
        all_rows = [row for row in all_rows if row.category_key in selected_keys]
    raw_base = base_filter(
        all_rows, period=None, price_filter=effective_price_filter, clean_noise=s.clean_noise,
        include_words=s.include_words or "", exclude_words=s.exclude_words or "",
    )

    # Frequency intentionally sees all distinct IDs; otherwise smart de-duplication
    # would hide the very repetitions this mode is meant to measure. Unique is
    # also evaluated on raw filtered rows so duplicates cannot turn into a fake
    # unique product after being collapsed.
    if mode == "unique":
        base = unique_rows(raw_base)
    else:
        base = raw_base
        if s.smart_dedupe and mode != "frequent":
            base = dedupe_rows(base)

    # v2.7.0: view counters are collected during the category scan itself.
    # Export is intentionally read-only so the result file is available immediately.

    if mode == "frequent":
        result = frequent_rows(base, min_count=3)
        if not result:
            await message.answer("🔥 Пока нет групп минимум с 3 публикациями по текущим фильтрам.", reply_markup=main_keyboard(selected_count))
            return 0
        result = _group_export_rows(result)
        path = write_frequent_xlsx(result)
        caption = f"🔥 Часто публикуемые группы: {len(result)} · 📊 XLSX"

    elif mode == "below_market":
        result = below_market_rows(base)
        if not result:
            await message.answer("💰 Нужны минимум 5 цен в одной уверенной группе; сейчас нет позиций ≥20% ниже медианы похожих объявлений.", reply_markup=main_keyboard(selected_count))
            return 0
        result = _group_export_rows(result)
        path = write_market_xlsx(result)
        caption = f"💰 Потенциально ниже рынка: {len(result)} · 📊 XLSX"

    elif mode == "fast_disappearing":
        status = await message.answer(
            f"⚡ Проверяю доступность до <b>{AVAILABILITY_CHECK_LIMIT}</b> сегодняшних объявлений…",
            parse_mode=ParseMode.HTML,
        )
        checked, newly_disappeared, unknown = await refresh_availability(base)
        # Reload rows because availability status may have changed.
        all_rows = await today_rows()
        refreshed = base_filter(
            all_rows, period=None, price_filter=effective_price_filter, clean_noise=s.clean_noise,
            include_words=s.include_words or "", exclude_words=s.exclude_words or "",
        )
        if s.smart_dedupe:
            refreshed = dedupe_rows(refreshed)
        result = disappearing_rows(refreshed, max_lifespan_hours=12)
        await status.edit_text(
            f"⚡ Проверено: <b>{checked}</b> · новых исчезнувших: <b>{newly_disappeared}</b> · неопределённых: <b>{unknown}</b>",
            parse_mode=ParseMode.HTML,
        )
        if not result:
            await message.answer(
                "⚡ Пока нет объявлений, исчезнувших примерно за ≤12 часов. Точность растёт при регулярных проверках в течение дня.",
                reply_markup=main_keyboard(selected_count),
            )
            return 0
        result = _group_export_rows(result)
        path = write_disappearing_xlsx(result)
        caption = f"⚡ Быстро исчезающие: {len(result)} · 📊 XLSX"

    elif mode == "price_drop":
        histories = await histories_for(base)
        result = price_drop_rows(base, histories, min_drop_pct=5, min_drop_eur=5)
        if not result:
            await message.answer(
                "📉 Пока нет подтверждённых снижений минимум на 5 € и 5%. Нужен повторный парсинг после изменения цены объявления.",
                reply_markup=main_keyboard(selected_count),
            )
            return 0
        result = _group_export_rows(result)
        path = write_price_drop_xlsx(result)
        caption = f"📉 Снижение цены: {len(result)} · 📊 XLSX"

    else:
        result = sort_rows(base, s.sort_mode)
        if not result:
            await message.answer("📦 По текущим фильтрам ничего не найдено.", reply_markup=main_keyboard(selected_count))
            return 0
        result = _group_export_rows(result)
        path = write_listing_xlsx(result, mode)
        caption = f"📦 {MODE_LABELS.get(mode, mode)}: {len(result)} · 📊 XLSX"

    try:
        await message.answer_document(FSInputFile(path), caption=caption, reply_markup=main_keyboard(selected_count))
    finally:
        shutil.rmtree(path.parent, ignore_errors=True)
    return len(result)


def settings_text(s: UserSettings) -> str:
    include = html.escape(s.include_words) if s.include_words else "—"
    exclude = html.escape(s.exclude_words) if s.exclude_words else "—"
    return (
        "<b>⚙️ Настройки результата</b>\n\n"
        "Эти параметры определяют, <b>какие объявления попадут в XLSX и TOP</b> после сбора.\n\n"
        f"Режим: <b>{MODE_LABELS.get(s.output_mode, s.output_mode)}</b>\n"
        f"👁 Просмотры: <b>{min_views_label(getattr(s, 'min_views', 0))}</b>\n"
        f"🧠 Умные дубли: <b>{'Вкл' if s.smart_dedupe else 'Выкл'}</b>\n"
        f"🧹 Очистка шума: <b>{'Вкл' if s.clean_noise else 'Выкл'}</b>\n"
        f"↕️ Сортировка: <b>{SORT_LABELS.get(s.sort_mode, s.sort_mode)}</b>\n"
        f"🔎 Ключевые слова: <b>{include}</b>\n"
        f"🚫 Исключения: <b>{exclude}</b>\n\n"
        "💡 Не знаешь, что выбрать? Нажми <b>«ℹ️ Что выбрать?»</b>.\n"
        "<i>Дата, цена и глубина 15/25/50 страниц выбираются отдельно при каждом новом скане.</i>"
    )


async def stats_text() -> str:
    start_utc, end_utc = berlin_today_utc_bounds()
    day_key = berlin_date_key()
    async with SessionLocal() as session:
        total = (await session.execute(select(func.count(Listing.id)))).scalar_one()
        today = (await session.execute(select(func.count(Listing.id)).where(
            Listing.first_seen_at >= start_utc, Listing.first_seen_at < end_utc,
        ))).scalar_one()
        priced = (await session.execute(select(func.count(Listing.id)).where(
            Listing.first_seen_at >= start_utc, Listing.first_seen_at < end_utc, Listing.price_text.is_not(None),
        ))).scalar_one()
        viewed = (await session.execute(select(func.count(Listing.id)).where(
            Listing.first_seen_at >= start_utc, Listing.first_seen_at < end_utc, Listing.view_count.is_not(None),
        ))).scalar_one()
        drops = (await session.execute(select(func.count(PriceHistory.id)))).scalar_one()
        runs = (await session.execute(select(func.count(ParserRun.id)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc, ParserRun.success.is_(True),
        ))).scalar_one()
        fast_runs = (await session.execute(select(func.count(ParserRun.id)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc,
            ParserRun.success.is_(True), ParserRun.mode == "fast",
        ))).scalar_one()
        pages = (await session.execute(select(func.coalesce(func.sum(ParserRun.pages_scanned), 0)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc, ParserRun.success.is_(True),
        ))).scalar_one()
        scan_new = (await session.execute(select(func.coalesce(func.sum(ParserRun.new_count), 0)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc, ParserRun.success.is_(True),
        ))).scalar_one()
        avg_quality = (await session.execute(select(func.coalesce(func.avg(ParserRun.quality_score), 0)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc, ParserRun.success.is_(True),
            ParserRun.quality_score > 0,
        ))).scalar_one()
        missing_dates = (await session.execute(select(func.coalesce(func.sum(ParserRun.missing_date_count), 0)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc, ParserRun.success.is_(True),
        ))).scalar_one()
        invalid_pages = (await session.execute(select(func.coalesce(func.sum(ParserRun.invalid_pages), 0)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc, ParserRun.success.is_(True),
        ))).scalar_one()
        repeated_pages = (await session.execute(select(func.coalesce(func.sum(ParserRun.repeated_pages), 0)).where(
            ParserRun.started_at >= start_utc, ParserRun.started_at < end_utc, ParserRun.success.is_(True),
        ))).scalar_one()
        fast_ready = (await session.execute(select(func.count(CategoryScanState.category_key)).where(
            CategoryScanState.scan_date == day_key, CategoryScanState.day_seed_complete.is_(True),
        ))).scalar_one()
        db_running_jobs = (await session.execute(select(func.count(UserScan.id)).where(
            UserScan.status == "running", UserScan.finished_at.is_(None),
        ))).scalar_one()
        db_queued_jobs = (await session.execute(select(func.count(UserScan.id)).where(
            UserScan.status == "queued", UserScan.finished_at.is_(None),
        ))).scalar_one()
        stable_jobs_done = (await session.execute(select(func.count(StableCategoryJob.id)).where(
            StableCategoryJob.status == "done", StableCategoryJob.updated_at >= start_utc,
        ))).scalar_one()
        stable_jobs_partial = (await session.execute(select(func.count(StableCategoryJob.id)).where(
            StableCategoryJob.status == "partial", StableCategoryJob.updated_at >= start_utc,
        ))).scalar_one()
        stable_checkpoints = (await session.execute(select(func.count(StablePageCheckpoint.id)).where(
            StablePageCheckpoint.status == "verified", StablePageCheckpoint.checked_at >= start_utc,
        ))).scalar_one()
    storage = DATABASE_BACKEND
    warning = ""
    coverage = round(priced / today * 100) if today else 0
    view_coverage = round(viewed / today * 100) if today else 0
    async with job_guard:
        running_jobs_count = sum(1 for j in active_jobs.values() if j.state == "running")
        queued_jobs_count = sum(1 for j in active_jobs.values() if j.state == "queued" and not j.cancel_requested)
        inflight_categories_count = len(category_inflight)
    multiuser_text = ""
    if DISTRIBUTED_WORKERS:
        parser_workers = views_workers = 0
        try:
            parser_workers = await COORDINATOR.worker_count("parser")
            views_workers = await COORDINATOR.worker_count("views")
        except Exception:
            pass
        multiuser_text = (
            "<b>⚙️ Multi-user Core</b>\n"
            "Режим: <b>Redis / distributed</b>\n"
            f"Parser-worker: <b>{parser_workers}</b> · Views-worker: <b>{views_workers}</b>\n"
            f"Сейчас сканируют: <b>{int(db_running_jobs or 0)}</b> · В очереди: <b>{int(db_queued_jobs or 0)}</b>\n\n"
        )
    else:
        multiuser_text = (
            "<b>⚙️ Multi-user Core</b>\n"
            f"Режим: <b>локальный</b> · Запущено: <b>{running_jobs_count}</b> · Очередь: <b>{queued_jobs_count}</b>\n\n"
        )
    return (
        f"<b>📊 База и парсинг</b>\n\n"
        f"Сегодня собрано: <b>{today}</b>\n"
        f"С ценой: <b>{priced}</b> ({coverage}%)\n"
        f"С просмотрами: <b>{viewed}</b> ({view_coverage}%)\n"
        f"Всего сохранено: <b>{total}</b>\n"
        f"Записей истории цен: <b>{drops}</b>\n\n"
        f"{multiuser_text}"
        f"<b>🧱 Stable Scan Engine</b>\n"
        f"Общих category/date jobs завершено: <b>{stable_jobs_done}</b>\n"
        f"Частичных после автоповторов: <b>{stable_jobs_partial}</b>\n"
        f"PostgreSQL checkpoints сегодня: <b>{stable_checkpoints}</b>\n\n"
        f"<b>🛡 Качество сегодня</b>\n"
        f"Запусков категорий: <b>{runs}</b>\n"
        f"Среднее качество: <b>{round(float(avg_quality or 0))}/100</b>\n"
        f"Дат не распознано: <b>{missing_dates}</b>\n"
        f"Невалидных страниц: <b>{invalid_pages}</b>\n"
        f"Повторов страниц: <b>{repeated_pages}</b>\n"
        f"Сетевых страниц: <b>{pages}</b>\n"
        f"Найдено новых за запуски: <b>{scan_new}</b>\n\n"
        f"База: <b>{storage}</b>{warning}"
    )


@dataclass
class ScanResult:
    new_count: int
    pages_scanned: int
    today_seen: int
    known_count: int
    enriched_count: int
    hit_limit: bool
    reason: str
    mode: str
    avoided_pages: int = 0
    date_complete: bool = False
    oldest_date_seen: str = ""
    max_page_reached: int = 0
    matched_ids: list[str] | None = None
    # v3.1 quality telemetry. These fields are intentionally part of the shared
    # ScanResult so cached/shared scans preserve the same reliability verdict.
    cards_seen: int = 0
    listings_parsed: int = 0
    missing_date_count: int = 0
    missing_price_count: int = 0
    promoted_filtered: int = 0
    price_reduced_filtered: int = 0
    duplicate_count: int = 0
    invalid_pages: int = 0
    repeated_pages: int = 0
    low_quality_pages: int = 0
    verified_pages: int = 0
    # Exact number of verified pages that actually contained the requested day.
    # `verified_pages` above is transport/quality telemetry and may include date probes.
    collection_pages_confirmed: int = 0
    view_failures: int = 0
    views_requested: int = 0
    views_verified: int = 0
    radar_views_complete: bool = True
    radar_views_usable: bool = True
    view_tail_count: int = 0
    quality_score: int = 0
    quality_note: str = ""


def _calculate_scan_quality(
    *,
    listings_parsed: int,
    missing_dates: int,
    missing_prices: int,
    invalid_pages: int,
    repeated_pages: int,
    low_quality_pages: int,
    verified_pages: int,
    pages_scanned: int,
    view_failures: int,
    date_complete: bool,
) -> tuple[int, str]:
    """Return a conservative 0-100 quality score plus one compact reason."""
    score = 100.0
    notes: list[str] = []
    if listings_parsed > 0:
        date_cov = max(0.0, min(1.0, (listings_parsed - missing_dates) / listings_parsed))
        price_cov = max(0.0, min(1.0, (listings_parsed - missing_prices) / listings_parsed))
        score -= (1.0 - date_cov) * 35.0
        score -= (1.0 - price_cov) * 8.0
        if date_cov < 0.80:
            notes.append(f"дат распознано {round(date_cov * 100)}%")
    elif date_complete:
        # A verified empty day can still be a valid scan.
        notes.append("объявлений за дату не найдено")

    # `invalid_pages` and `repeated_pages` are mutually exclusive unique-page
    # counters in v4.8.5. Repeated content is one defect, not a repeated+invalid
    # double penalty for the same physical page.
    score -= min(30.0, invalid_pages * 10.0)
    score -= min(25.0, repeated_pages * 12.0)
    score -= min(20.0, low_quality_pages * 4.0)
    if pages_scanned and verified_pages == 0:
        score -= 10.0
        notes.append("страницы слабо подтверждены")
    if invalid_pages:
        notes.append(f"невалидных страниц {invalid_pages}")
    if repeated_pages:
        notes.append(f"повторов страниц {repeated_pages}")
    if view_failures:
        # Views are secondary data: a few failures should not make the category
        # parser look broken, but a large number is still worth surfacing.
        score -= min(8.0, view_failures * 0.15)
    if not date_complete:
        score = min(score, 69.0)
        if listings_parsed == 0:
            score = min(score, 45.0)
        notes.append("охват даты не подтверждён")

    final = max(0, min(100, int(round(score))))
    if not notes:
        notes.append("проверки пройдены")
    return final, "; ".join(notes[:3])


@dataclass
class CategoryDispatchResult:
    source: str  # scan | shared | cache
    result: ScanResult | None = None
    cache_age_seconds: int = 0


@dataclass
class ScanJob:
    job_id: str
    user_id: int
    chat_id: int
    status_message_id: int
    category_keys: list[str]
    created_at: datetime
    state: str = "queued"
    cancel_requested: bool = False
    worker_id: int | None = None
    current_category: str = ""
    completed_categories: int = 0
    total_new: int = 0
    total_pages: int = 0
    total_avoided: int = 0
    cache_hits: int = 0
    shared_hits: int = 0
    scanned_categories: int = 0
    fast_categories: int = 0
    full_categories: int = 0
    warnings: list[str] | None = None
    last_status_update: float = 0.0
    current_category_key: str = ""
    current_category_index: int = 0
    started_running_monotonic: float = 0.0
    page_limit: int = 50
    current_progress_key: str = ""
    scan_id: int | None = None
    target_date: str = ""
    price_filter: str = "any"
    is_trial: bool = False
    incomplete_categories: int = 0
    scan_notes: list[str] | None = None
    matched_ids: set[str] | None = None
    quality_scores: list[int] | None = None
    quality_notes: list[str] | None = None
    incomplete_category_keys: set[str] | None = None
    retry_note: str = ""
    recovery_note: str = ""
    recovery_attempt: int = 0
    recovery_total: int = 0
    auto_recovered_categories: int = 0
    recovered: bool = False
    stop_event: asyncio.Event = field(default_factory=asyncio.Event, repr=False, compare=False)


@dataclass
class CategoryLiveProgress:
    category_key: str
    category_name: str
    mode: str
    page: int = 0
    today_seen: int = 0
    new_count: int = 0
    known_count: int = 0
    views_ready: int = 0
    views_failed: int = 0
    estimated_pages: int = 10
    started_monotonic: float = 0.0
    page_limit: int = 50
    oldest_date_seen: str = ""
    current_page_date: str = ""
    phase: str = "seeking"
    segment_name: str = ""
    segments_done: int = 0
    segments_total: int = 0
    collection_index: int = 0
    collection_start_page: int = 0
    date_coverage_pct: int = 0
    quality_score: int = 100
    quality_warning: str = ""
    # Number of real category-page HTTP responses already processed.  This gives
    # the UI a truthful heartbeat during date-location before collection starts.
    network_requests: int = 0
    checkpoint_hits: int = 0
    request_started_at_ts: float = 0.0
    current_request_page: int = 0
    last_request_ms: int = 0
    request_timeouts: int = 0
    transport_stage: str = ""


category_live_progress: dict[str, CategoryLiveProgress] = {}


def _date_scan_limit(target_date: str) -> int:
    """Verified public page window for one Kleinanzeigen result feed."""
    return PUBLIC_SEARCH_PAGE_CAP if target_date else MAX_PAGES_PER_CATEGORY


def _progress_key(category_key: str, target_date: str, page_limit: int | None = None) -> str:
    depth = int(page_limit or 0)
    return f"v410:{category_key}:date:{target_date}:depth:{depth}"


scan_queue: asyncio.Queue[ScanJob] = asyncio.Queue()
active_jobs: dict[int, ScanJob] = {}
queued_job_ids: list[str] = []
job_guard = asyncio.Lock()
category_inflight: dict[str, asyncio.Task[ScanResult]] = {}
category_inflight_waiters: dict[str, int] = {}
category_inflight_guard = asyncio.Lock()
# Exact-date cache must preserve the exact 15/25/50-page result set, so v3.0.6
# caches ScanResult (including matched IDs) in memory instead of reconstructing a
# result from every listing ever seen for that date.
category_result_cache: dict[str, tuple[float, ScanResult]] = {}
db_write_lock = asyncio.Lock()


async def _lock_listing_integrity_ids(session, external_ids) -> None:
    """Share v4.15.3 integrity locks with the Radar admission transaction."""
    bind = session.get_bind()
    if bind is None or bind.dialect.name != "postgresql":
        return
    for external_id in sorted({str(x).strip() for x in external_ids if str(x).strip()}):
        await session.execute(
            text("SELECT pg_advisory_xact_lock(CAST(hashtext(:integrity_key) AS bigint))"),
            {"integrity_key": f"organic-integrity:{external_id}"},
        )

# v3.1.6 manual view refreshes are true background jobs. A user can navigate
# anywhere in the bot while the refresh continues, and duplicate refreshes of the
# same scan are coalesced into one task.
manual_view_tasks: dict[int, asyncio.Task] = {}
manual_view_tasks_guard = asyncio.Lock()
# One lightweight background collector at a time. This makes DB cache reuse deterministic
# across users/scans and prevents automatic checkpoints from multiplying the same ID requests.
background_view_refresh_lock = asyncio.Lock()
radar_v3_view_refresh_lock = asyncio.Lock()


def scan_job_from_record(scan: UserScan, *, recovered: bool = False) -> ScanJob:
    keys = [k for k in _scan_category_keys(scan) if k in CATEGORIES]
    job = ScanJob(
        job_id=str(scan.job_uid),
        user_id=int(scan.user_id),
        chat_id=int(scan.chat_id or scan.user_id),
        status_message_id=int(scan.status_message_id or 0),
        category_keys=keys,
        created_at=scan.created_at or datetime.utcnow(),
        warnings=[],
        page_limit=int(scan.page_limit or 25),
        scan_id=int(scan.id),
        target_date=scan.target_date or _moscow_today_iso(),
        price_filter=(getattr(scan, "price_filter", "any") or "any"),
        is_trial=bool(getattr(scan, "is_trial", False)),
        recovered=recovered,
    )
    job.completed_categories = int(scan.completed_categories or 0)
    job.incomplete_categories = int(getattr(scan, "incomplete_categories", 0) or 0)
    job.incomplete_category_keys = set(
        x for x in (getattr(scan, "incomplete_category_keys", "") or "").split(",") if x
    )
    if scan.status == "running":
        job.state = "running"
    elif scan.status == "cancelling":
        job.state = "queued"
        job.cancel_requested = True
        job.stop_event.set()
    else:
        job.state = "queued"
    return job


async def load_scan_job_by_uid(job_uid: str) -> ScanJob | None:
    async with SessionLocal() as session:
        result = await session.execute(select(UserScan).where(UserScan.job_uid == str(job_uid)).limit(1))
        scan = result.scalar_one_or_none()
        if scan is None:
            return None
        return scan_job_from_record(scan, recovered=scan.status == "running")


async def _distributed_cancel_watcher(job: ScanJob) -> None:
    """Mirror cross-process cancel state into the worker-local asyncio.Event."""
    while not job.stop_event.is_set() and job.state in {"queued", "running"}:
        redis_cancelled = False
        try:
            redis_cancelled = await COORDINATOR.is_cancel_requested(job.job_id)
        except Exception:
            log.debug("Distributed Redis cancel check failed job=%s", job.job_id, exc_info=True)
        if redis_cancelled:
            job.cancel_requested = True
            job.stop_event.set()
            return

        # PostgreSQL is an independent cancellation channel. It keeps Stop working
        # even during a temporary Redis connectivity issue after the job started.
        if job.scan_id is not None:
            try:
                async with SessionLocal() as session:
                    scan = await session.get(UserScan, int(job.scan_id))
                    if scan is not None and scan.status == "cancelling":
                        job.cancel_requested = True
                        job.stop_event.set()
                        return
            except Exception:
                log.debug("Distributed DB cancel check failed job=%s", job.job_id, exc_info=True)
        await asyncio.sleep(0.8)


async def get_category_scan_state(category_key: str) -> CategoryScanState | None:
    async with SessionLocal() as session:
        return await session.get(CategoryScanState, category_key)


async def save_category_scan_state(
    category_key: str,
    *,
    target_date: str,
    mode: str,
    pages_scanned: int,
    new_count: int,
    today_seen: int,
    reason: str,
    head_ids: list[str],
    seed_complete: bool,
    seed_capped: bool,
    coverage_pages: int | None = None,
) -> CategoryScanState:
    day_key = berlin_date_key()
    async with SessionLocal() as session:
        state = await session.get(CategoryScanState, category_key)
        if state is None:
            state = CategoryScanState(category_key=category_key, scan_date=day_key)
            session.add(state)
        new_day = state.scan_date != day_key or (state.target_date or "") != target_date
        if new_day:
            state.scan_date = day_key
            state.target_date = target_date
            state.total_runs = 0
            state.day_seed_complete = False
            state.day_seed_capped = False
            state.day_full_pages = 0
            state.head_ids = ""

        if head_ids:
            state.head_ids = ",".join(head_ids[:INCREMENTAL_HEAD_SIZE])
        state.target_date = target_date
        state.last_scan_at = datetime.utcnow()
        state.last_mode = mode
        state.last_pages = pages_scanned
        state.last_new = new_count
        state.last_today_seen = today_seen
        state.last_stop_reason = reason[:255]
        state.total_runs = (state.total_runs or 0) + 1
        if mode in {"full", "date", "stable-date"}:
            # Keep the deepest seeded window for the day. A 25-page seed enables
            # later 25-page fast scans, while a later 100-page request can deepen it.
            state.day_full_pages = max(state.day_full_pages or 0, int(coverage_pages or pages_scanned))
            if seed_complete:
                state.day_seed_complete = True
                state.day_seed_capped = False
            elif seed_capped and not state.day_seed_complete:
                state.day_seed_capped = True
        await session.commit()
        await session.refresh(state)
        return state


async def record_parser_run(
    user_id: int,
    cat,
    result: ScanResult,
    started_at: datetime,
    *,
    success: bool = True,
    error_text: str | None = None,
) -> None:
    async with SessionLocal() as session:
        session.add(ParserRun(
            user_id=user_id,
            category_key=cat.key,
            category_name=cat.name,
            mode=result.mode,
            started_at=started_at,
            finished_at=datetime.utcnow(),
            pages_scanned=result.pages_scanned,
            today_seen=result.today_seen,
            new_count=result.new_count,
            known_count=result.known_count,
            enriched_count=result.enriched_count,
            cards_seen=result.cards_seen,
            listings_parsed=result.listings_parsed,
            missing_date_count=result.missing_date_count,
            missing_price_count=result.missing_price_count,
            promoted_filtered=result.promoted_filtered,
            duplicate_count=result.duplicate_count,
            invalid_pages=result.invalid_pages,
            repeated_pages=result.repeated_pages,
            low_quality_pages=result.low_quality_pages,
            view_failures=result.view_failures,
            quality_score=result.quality_score,
            stop_reason=result.reason[:255],
            success=success,
            error_text=(error_text[:1000] if error_text else None),
        ))
        await session.commit()


async def scan_one_category(parser: KleinanzeigenParser, cat, user_id: int, page_limit: int, target_date: str) -> ScanResult:
    """Reliably locate the selected Moscow date and collect 15/25/50-page depth.

    v3.1 treats page identity and publication-date coverage as data-quality signals.
    A weak/normalized/repeated page may contribute diagnostics, but it is never used
    as proof that a date is absent. This prevents a silent parser degradation from
    turning into a believable zero-result scan.
    """
    depth = page_limit if page_limit in PAGE_LIMIT_CHOICES else 50
    target_day = datetime.strptime(target_date, "%Y-%m-%d").date()
    # v4.1.0 Universal Date Stream. Every date uses the same newest-sorted
    # chronology algorithm. There are no special "today/yesterday" branches:
    # page 1 -> sequential evidence -> target window -> older boundary. If the
    # target is genuinely deeper than the public window, the same stream is run
    # over independent location shards.
    moscow_today = datetime.now(MOSCOW).date()
    universal_date_stream = bool(STABLE_SCAN_ENGINE)
    recent_fast_path = 0 <= (moscow_today - target_day).days <= 2  # UI/telemetry only
    today_fast_path = target_day == moscow_today  # UI/telemetry only
    progress_key = _progress_key(cat.key, target_date, depth)
    mode = "stable-date" if STABLE_SCAN_ENGINE else "date"
    scan_settings = await get_settings(user_id)
    # v4.1.6: view-threshold scans still need fresh counters before final filtering,
    # but fetching 20-30 counters after *every* page made a 25-page crawl look
    # frozen.  Collect the category pages first, then run one concurrent views
    # phase for all matched target-day listings.
    need_view_counts = PRIMARY_SCAN_INLINE_VIEWS or int(getattr(scan_settings, "min_views", 0) or 0) > 0
    deferred_view_items: dict[str, ParsedListing] = {}
    if STABLE_SCAN_ENGINE:
        try:
            await mark_category_job(cat.key, target_date, depth, status="running")
        except Exception:
            log.debug("Could not persist stable category job start", exc_info=True)

    category_live_progress[progress_key] = CategoryLiveProgress(
        category_key=cat.key,
        category_name=cat.name,
        mode=mode,
        estimated_pages=depth,
        started_monotonic=time.monotonic(),
        page_limit=depth,
        phase="stable_scan" if universal_date_stream else "jumping",
    )

    new_count = 0
    today_seen = 0
    known_total = 0
    enriched_total = 0
    target_seen_any = False
    request_complete = False
    oldest_date_seen = ""
    first_page_head_ids: list[str] = []
    processed_target_ids: set[str] = set()
    started_at = datetime.utcnow()
    reason = ""
    hit_limit = False
    collection_start_page = 0
    direct_pages_collected = 0
    collection_pages_confirmed = 0
    network_requests = 0
    max_page_reached = 0

    # v3.1 quality telemetry. Counters increase only for actual network responses,
    # never when a page is reused from the locator's in-memory cache.
    cards_seen = 0
    listings_parsed = 0
    missing_date_count = 0
    missing_price_count = 0
    promoted_filtered = 0
    price_reduced_filtered = 0
    duplicate_count = 0
    invalid_pages = 0
    repeated_pages = 0
    low_quality_pages = 0
    verified_pages = 0
    view_failures = 0
    views_requested = 0
    views_verified = 0
    # v4.8.5 Quality Integrity: quality is about unique logical pages, not
    # network attempts. A page retried four times must still count as one page.
    # Snapshots also let a later successful retry remove the earlier penalty.
    page_quality_metrics: dict[tuple[str, int], dict[str, int]] = {}
    invalid_page_keys: set[tuple[str, int]] = set()
    repeated_page_keys: set[tuple[str, int]] = set()
    low_quality_page_keys: set[tuple[str, int]] = set()

    def classify(items):
        profile = profile_page_dates(items, target_day)
        return profile.relation, profile.pairs, profile.days, profile

    def update_quality_live(note: str = "") -> None:
        live = category_live_progress.get(progress_key)
        if live is None:
            return
        if listings_parsed:
            coverage = max(0.0, min(1.0, (listings_parsed - missing_date_count) / listings_parsed))
            live.date_coverage_pct = round(coverage * 100)
        rough, rough_note = _calculate_scan_quality(
            listings_parsed=listings_parsed,
            missing_dates=missing_date_count,
            missing_prices=missing_price_count,
            invalid_pages=invalid_pages,
            repeated_pages=repeated_pages,
            low_quality_pages=low_quality_pages,
            verified_pages=verified_pages,
            pages_scanned=network_requests,
            view_failures=view_failures,
            date_complete=True,
        )
        live.quality_score = rough
        live.quality_warning = note or (rough_note if rough < 85 else "")

    def update_live(page: int, days: list, phase: str, collection_index: int | None = None) -> None:
        nonlocal oldest_date_seen, max_page_reached
        max_page_reached = max(max_page_reached, int(page or 0))
        page_date_hint = ""
        if days:
            page_oldest = min(days)
            page_newest = max(days)
            page_date_hint = page_oldest.isoformat() if page_oldest == page_newest else f"{page_newest.isoformat()}..{page_oldest.isoformat()}"
            if not oldest_date_seen or page_oldest.isoformat() < oldest_date_seen:
                oldest_date_seen = page_oldest.isoformat()
        live = category_live_progress.get(progress_key)
        if live is not None:
            live.network_requests = max(live.network_requests, network_requests)
            live.page = page
            live.oldest_date_seen = oldest_date_seen
            live.current_page_date = page_date_hint
            live.phase = phase
            if collection_index is not None:
                live.collection_index = min(depth, max(0, collection_index))
                live.page_limit = depth
        update_quality_live()

    async def process_target_items(items, pairs, limit: int | None = None) -> int:
        nonlocal new_count, today_seen, known_total, enriched_total, target_seen_any, first_page_head_ids, view_failures
        target_items = [
            item for item, item_day in pairs
            if item_day == target_day and item.external_id not in processed_target_ids
        ]
        if limit is not None:
            target_items = target_items[:max(0, int(limit))]
        if not target_items:
            return 0
        processed_target_ids.update(item.external_id for item in target_items)
        new_items, known_count, enriched_count, nonorganic_ids = await upsert_page_items(
            cat.key, cat.name, target_items
        )
        if nonorganic_ids:
            target_items = [item for item in target_items if item.external_id not in nonorganic_ids]
        if not target_items:
            update_quality_live()
            return 0
        target_seen_any = True
        today_seen += len(target_items)
        if not first_page_head_ids:
            first_page_head_ids = [item.external_id for item in target_items[:INCREMENTAL_HEAD_SIZE]]
        live = category_live_progress.get(progress_key)
        if need_view_counts:
            for item in target_items:
                deferred_view_items[item.external_id] = item
        new_count += len(new_items)
        known_total += known_count
        enriched_total += enriched_count
        live = category_live_progress.get(progress_key)
        if live is not None:
            live.today_seen = today_seen
            live.new_count = new_count
            live.known_count = known_total
        update_quality_live()
        return len(target_items)

    _REMOTE_HINT_UNSET = object()

    async def locate_feed(base_url: str, feed_name: str, remote_hint_override=_REMOTE_HINT_UNSET):
        # v3.1.8: use Kleinanzeigen's official Anbieter=Privat filter at the
        # search-feed level. Commercial/store listings therefore never consume
        # scan depth and never enter snapshots, views or TOP analytics.
        base_url = private_provider_url(base_url)
        """Locate the first target-date page inside one verified <=50-page feed."""
        nonlocal network_requests, cards_seen, listings_parsed, missing_date_count
        nonlocal missing_price_count, promoted_filtered, price_reduced_filtered, duplicate_count, invalid_pages
        nonlocal repeated_pages, low_quality_pages, verified_pages
        cache: dict[int, object] = {}
        fingerprints: dict[str, int] = {}
        # v4.10.2: if a Page Worker cache entry duplicates a page already seen
        # in this exact category scan, never replay that Redis value through
        # stable_fetch retries. The page is forced through the local stable
        # parser for the remainder of this locator so retry/reset can actually
        # obtain fresh content instead of reading the same poisoned cache.
        remote_repeat_bypass_pages: set[int] = set()
        effective_limit = PUBLIC_SEARCH_PAGE_CAP
        site_max_page: int | None = None
        discovered_shards: list[tuple[str, int | None]] = []
        invalid_note = ""
        last_invalid_kind = ""

        def locator_result(status: str, reason_text: str = "", candidate_page: int | None = None):
            # v4.0.4: the Stable Engine must actually use its per-page retry wrapper
            # during collection. Older builds returned raw `fetch` here, so a single
            # weak/invalid page bypassed STABLE_PAGE_RETRIES and could make the whole
            # category partial even though stable_fetch() existed. Name resolution is
            # intentionally late: locator_result is called only after stable_fetch has
            # been defined.
            page_fetch = stable_fetch if STABLE_SCAN_ENGINE else fetch
            return {
                "status": status, "reason": reason_text, "fetch": page_fetch,
                "limit": effective_limit, "site_max_page": site_max_page,
                "candidate": candidate_page, "shards": list(discovered_shards),
                "base_url": base_url,
                # Expose the final stable-fetch defect class without widening the
                # fetch() tuple used by the date locator and regional collector.
                "last_invalid_kind": lambda: last_invalid_kind,
            }

        async def fetch(page: int, phase: str):
            nonlocal network_requests, effective_limit, site_max_page, discovered_shards, invalid_note, last_invalid_kind
            nonlocal cards_seen, listings_parsed, missing_date_count, missing_price_count
            nonlocal promoted_filtered, price_reduced_filtered, duplicate_count, invalid_pages, repeated_pages
            nonlocal low_quality_pages, verified_pages
            page = max(1, min(effective_limit, int(page)))
            fresh = page not in cache
            from_checkpoint = False
            remote_page = False
            local_page_fetched = False
            if not fresh:
                info = cache[page]
            else:
                live_req = category_live_progress.get(progress_key)
                info = None
                if STABLE_SCAN_ENGINE:
                    try:
                        info = await load_page_checkpoint(cat.key, target_date, base_url, page)
                    except Exception:
                        log.debug("Stable checkpoint read failed category=%s page=%s", cat.key, page, exc_info=True)
                    if info is not None:
                        from_checkpoint = True
                        if live_req is not None:
                            live_req.checkpoint_hits += 1
                            live_req.transport_stage = "postgres-checkpoint"

                if info is None:
                    requested_url = page_url(base_url, page)
                    # v4.3.21 Page Worker. Date-location probes stay on the proven
                    # local parser. Only the post-locator collection phase may consume
                    # the 180-second Redis page cache warmed by dedicated workers.
                    if (
                        phase == "collecting"
                        and REMOTE_PAGE_WORKER_ENABLED
                        and page not in remote_repeat_bypass_pages
                    ):
                        try:
                            info = await REMOTE_PAGE_MANAGER.get_cached_wait(requested_url, page)
                            remote_page = info is not None
                            if remote_page:
                                # v4.3.23 strict remote gate. Stable retries must never
                                # replay one weak Redis response. Validate both page
                                # identity and target-date chronology before accepting
                                # Page Worker output; otherwise discard it and use the
                                # original local stable parser immediately.
                                remote_relation, _rpairs, _rdays, _rprofile = classify(info.items)
                                remote_dated = max(
                                    0,
                                    len(info.items) - int(getattr(info, "missing_date_count", 0) or 0),
                                )
                                remote_safe = (
                                    bool(getattr(info, "request_matches_page", True))
                                    and bool(getattr(info, "page_verified", False))
                                    and not bool(getattr(info, "suspicious", False))
                                    and remote_relation not in {"unknown", "invalid"}
                                    and (
                                        not info.items
                                        or remote_dated >= 2
                                        or remote_relation == "target"
                                    )
                                )
                                if not remote_safe:
                                    log.warning(
                                        "Rejected weak Page Worker cache category=%s page=%s relation=%s verified=%s matches=%s suspicious=%s",
                                        cat.key, page, remote_relation,
                                        bool(getattr(info, "page_verified", False)),
                                        bool(getattr(info, "request_matches_page", True)),
                                        bool(getattr(info, "suspicious", False)),
                                    )
                                    try:
                                        await REMOTE_PAGE_MANAGER.invalidate_cached(requested_url, page)
                                    except Exception:
                                        pass
                                    info = None
                                    remote_page = False
                                elif live_req is not None:
                                    live_req.transport_stage = "page-worker-cache"
                        except Exception:
                            info = None
                            remote_page = False
                            log.debug(
                                "Remote Page Worker cache failed category=%s page=%s",
                                cat.key, page, exc_info=True,
                            )

                    if info is None:
                        checkpoint_before = int(getattr(parser, "scan_page_checkpoint_hits", 0) or 0)
                        req_started = time.monotonic()
                        if live_req is not None:
                            live_req.request_started_at_ts = time.time()
                            live_req.current_request_page = page
                            try:
                                live_req.transport_stage = parser.scan_transport_status()
                            except Exception:
                                live_req.transport_stage = getattr(parser, "scan_transport", "http")
                        try:
                            info = await parser.parse_category_page_info(requested_url, page)
                        except Exception as exc:
                            if STABLE_SCAN_ENGINE:
                                try:
                                    await record_page_failure(cat.key, target_date, base_url, page, f"{type(exc).__name__}: {exc}")
                                except Exception:
                                    log.debug("Could not persist failed page checkpoint", exc_info=True)
                            if live_req is not None and (
                                isinstance(exc, (asyncio.TimeoutError, TimeoutError)) or "timeout" in str(exc).lower()
                            ):
                                live_req.request_timeouts += 1
                            raise
                        finally:
                            if live_req is not None:
                                live_req.last_request_ms = max(0, int((time.monotonic() - req_started) * 1000))
                                live_req.request_started_at_ts = 0.0
                                try:
                                    live_req.transport_stage = parser.scan_transport_status()
                                except Exception:
                                    pass
                        checkpoint_after = int(getattr(parser, "scan_page_checkpoint_hits", 0) or 0)
                        from_checkpoint = checkpoint_after > checkpoint_before
                        local_page_fetched = not from_checkpoint
                        if live_req is not None and from_checkpoint:
                            live_req.checkpoint_hits += 1
                    elif remote_page and live_req is not None:
                        # Keep normal quality/accounting below; the only difference is
                        # that another Railway service performed the network navigation.
                        live_req.request_started_at_ts = 0.0

                cache[page] = info
                promoted_ids = list(getattr(info, "promoted_ids", None) or [])
                if promoted_ids:
                    await mark_promoted_listings(promoted_ids, reason="search_promotion_marker")
                price_reduced_ids = list(getattr(info, "price_reduced_ids", None) or [])
                if price_reduced_ids:
                    await mark_price_reduced_listings(price_reduced_ids)
                if not from_checkpoint:
                    network_requests += 1
                if getattr(info, "max_page", None):
                    site_max_page = max(1, int(info.max_page))
                    effective_limit = max(1, min(PUBLIC_SEARCH_PAGE_CAP, site_max_page))
                if page == 1 and getattr(info, "location_shards", None):
                    discovered_shards = list(info.location_shards or [])
                if phase == "jumping" and DATE_JUMP_PROBE_DELAY_SECONDS and not STABLE_SCAN_ENGINE:
                    await asyncio.sleep(DATE_JUMP_PROBE_DELAY_SECONDS)

            items = info.items
            last_invalid_kind = ""
            relation, pairs, days, profile = classify(items)
            valid = bool(getattr(info, "request_matches_page", True)) and not bool(getattr(info, "suspicious", False))
            fp = getattr(info, "fingerprint", "") or ""
            repeated = False
            if fp:
                previous = fingerprints.get(fp)
                if previous is None:
                    fingerprints[fp] = page
                elif previous != page and len(items) >= 5:
                    valid = False
                    repeated = True
                    invalid_note = f"страница {page} повторила содержимое страницы {previous}"

            # v4.10.2 Page Worker poison recovery. v4.10.1 correctly detected
            # repeated-content, but stable_fetch() only popped the in-process
            # cache. A bad prefetched Redis page was therefore replayed on every
            # retry and even after BrowserContext reset. Once a remote page is
            # proven to duplicate an earlier page, delete that shared cache entry
            # and pin this page to the local parser for all subsequent retries.
            # A healthy local retry will clear the repeated-page quality marker
            # below; a genuinely repeated local page still follows the bounded
            # repeated-content recovery path.
            if repeated and remote_page:
                remote_repeat_bypass_pages.add(int(page))
                try:
                    await REMOTE_PAGE_MANAGER.invalidate_cached(page_url(base_url, page), page)
                except Exception:
                    log.debug(
                        "Could not invalidate repeated Page Worker cache category=%s page=%s",
                        cat.key, page, exc_info=True,
                    )
                log.warning(
                    "Repeated Page Worker cache invalidated category=%s page=%s previous=%s; forcing local retry",
                    cat.name, page, previous,
                )

            if fresh:
                metric_key = (str(base_url), int(page))
                current_metrics = {
                    "cards": int(getattr(info, "raw_candidates", 0) or 0),
                    "parsed": len(info.items),
                    "missing_date": int(getattr(info, "missing_date_count", 0) or 0),
                    "missing_price": int(getattr(info, "missing_price_count", 0) or 0),
                    "promoted": int(getattr(info, "promoted_filtered", 0) or 0),
                    "price_reduced": int(getattr(info, "price_reduced_filtered", 0) or 0),
                    "duplicates": int(getattr(info, "duplicate_cards", 0) or 0),
                    "verified": 1 if bool(getattr(info, "page_verified", False)) else 0,
                }
                previous_metrics = page_quality_metrics.get(metric_key, {})
                cards_seen += current_metrics["cards"] - int(previous_metrics.get("cards", 0))
                listings_parsed += current_metrics["parsed"] - int(previous_metrics.get("parsed", 0))
                missing_date_count += current_metrics["missing_date"] - int(previous_metrics.get("missing_date", 0))
                missing_price_count += current_metrics["missing_price"] - int(previous_metrics.get("missing_price", 0))
                promoted_filtered += current_metrics["promoted"] - int(previous_metrics.get("promoted", 0))
                price_reduced_filtered += current_metrics["price_reduced"] - int(previous_metrics.get("price_reduced", 0))
                duplicate_count += current_metrics["duplicates"] - int(previous_metrics.get("duplicates", 0))
                verified_pages += current_metrics["verified"] - int(previous_metrics.get("verified", 0))
                page_quality_metrics[metric_key] = current_metrics

                # A repeated-content page is a dedicated defect class; do not also
                # count it as a generic invalid page. A later healthy retry clears it.
                if repeated:
                    repeated_page_keys.add(metric_key)
                    invalid_page_keys.discard(metric_key)
                elif not valid:
                    invalid_page_keys.add(metric_key)
                    repeated_page_keys.discard(metric_key)
                else:
                    invalid_page_keys.discard(metric_key)
                    repeated_page_keys.discard(metric_key)

                if items and relation == "unknown" and valid:
                    low_quality_page_keys.add(metric_key)
                else:
                    low_quality_page_keys.discard(metric_key)

                invalid_pages = len(invalid_page_keys)
                repeated_pages = len(repeated_page_keys)
                low_quality_pages = len(low_quality_page_keys)

            if not valid:
                relation, pairs, days = "invalid", [], []
                invalid_note = invalid_note or f"страница {page} была нормализована/не подтверждена сайтом"

            dated_count = max(0, len(items) - int(getattr(info, "missing_date_count", 0) or 0))
            strong_page = (
                valid
                and bool(getattr(info, "page_verified", False))
                and relation not in {"unknown", "invalid"}
                and (not items or dated_count >= 2 or relation == "target")
            )
            if STABLE_SCAN_ENGINE and strong_page and not from_checkpoint:
                try:
                    await save_page_checkpoint(
                        cat.key, target_date, base_url, page, info, relation=relation
                    )
                except Exception:
                    log.debug("Stable checkpoint write failed category=%s page=%s", cat.key, page, exc_info=True)

            # v4.3.22 cooperative fallback: publish only a page that passed the same
            # stable quality gate used for PostgreSQL checkpoints. Weak/challenge
            # responses are never allowed into the shared 180-second cache.
            if (
                phase == "collecting"
                and REMOTE_PAGE_WORKER_ENABLED
                and local_page_fetched
                and strong_page
            ):
                try:
                    await REMOTE_PAGE_MANAGER.store_cached(requested_url, page, info)
                except Exception:
                    log.debug(
                        "Could not publish local collecting page category=%s page=%s",
                        cat.key, page, exc_info=True,
                    )

            update_live(page, days, phase)
            if relation == "unknown":
                update_quality_live(f"не хватает дат на странице {page}")
            elif relation == "invalid":
                update_quality_live(invalid_note)
            invalid_reason = ""
            if not valid:
                if bool(getattr(info, "suspicious", False)):
                    invalid_reason = "challenge"
                elif not bool(getattr(info, "request_matches_page", True)):
                    invalid_reason = "page-identity"
                elif repeated:
                    invalid_reason = "repeated-content"
                else:
                    invalid_reason = invalid_note or "unknown"
            last_invalid_kind = invalid_reason
            log.info(
                "category=%s feed=%s phase=%s page=%s relation=%s actual=%s max=%s verified=%s "
                "date_cov=%.0f%% parsed=%s missing_date=%s raw=%s promoted=%s reduced_price=%s duplicates=%s valid=%s "
                "invalid_reason=%s requests=%s checkpoint=%s",
                cat.name, feed_name, phase, page, relation,
                getattr(info, "actual_page", None), getattr(info, "max_page", None),
                getattr(info, "page_verified", False), float(getattr(info, "date_coverage", 0.0) or 0.0) * 100,
                len(items), getattr(info, "missing_date_count", 0), getattr(info, "raw_candidates", 0),
                getattr(info, "promoted_filtered", 0), getattr(info, "price_reduced_filtered", 0),
                getattr(info, "duplicate_cards", 0), valid, invalid_reason,
                network_requests, from_checkpoint,
            )
            return items, relation, pairs, days

        async def stable_fetch(page: int, phase: str):
            """Retry only the weak page; never restart a whole category for one bad response."""
            last = None
            for attempt in range(1, STABLE_PAGE_RETRIES + 1):
                try:
                    last = await fetch(page, phase)
                except TemporaryAccessError:
                    raise
                except Exception as exc:
                    if attempt >= STABLE_PAGE_RETRIES:
                        raise
                    cache.pop(page, None)
                    await asyncio.sleep(STABLE_PAGE_RETRY_SECONDS * attempt)
                    continue
                if last[1] not in {"unknown", "invalid"}:
                    return last
                if attempt < STABLE_PAGE_RETRIES:
                    cache.pop(page, None)
                    await asyncio.sleep(STABLE_PAGE_RETRY_SECONDS * attempt)
            # v4.15.5: if the exact page is still invalid OR chronologically unknown
            # after normal retries, recycle only this job's browser context and make
            # one final request. A stale browser/session template must not make retry
            # rounds reproduce the exact same UNKNOWN forever.
            # Do not restart the category and do not replay already verified pages.
            if (
                STABLE_SINGLE_SERVICE_MODE and last is not None and last[1] in {"unknown", "invalid"}
                and getattr(parser, "scan_transport", "") == "browser"
            ):
                try:
                    await parser.reset_scan_browser_context()
                    cache.pop(page, None)
                    await asyncio.sleep(STABLE_PAGE_RETRY_SECONDS)
                    final = await fetch(page, phase)
                    if final[1] not in {"unknown", "invalid"}:
                        log.info(
                            "Stable Reset recovered page after browser recycle category=%s page=%s",
                            cat.name, page,
                        )
                        return final
                    last = final
                except TemporaryAccessError:
                    raise
                except Exception as exc:
                    log.warning(
                        "Stable Reset final page retry failed category=%s page=%s: %s",
                        cat.name, page, exc,
                    )
            if STABLE_SCAN_ENGINE and last is not None and last[1] in {"unknown", "invalid"}:
                try:
                    await record_page_failure(
                        cat.key, target_date, base_url, page,
                        f"weak chronology after page retries/context recycle: {last[1]}",
                    )
                except Exception:
                    log.debug("Could not persist weak page failure", exc_info=True)
            return last

        async def remember_confirmed_date_hint(page: int) -> None:
            # v4.3.27 predictor still learns only from pages the foreground stable
            # parser has already confirmed. Remote Date Worker guesses never
            # become training data by themselves.
            if not REMOTE_DATE_WORKER_ENABLED:
                return
            try:
                await REMOTE_DATE_MANAGER.record_confirmed_hint(base_url, target_date, int(page))
            except Exception:
                log.debug("Date predictor confirmation write failed", exc_info=True)

        if STABLE_SCAN_ENGINE:
            # v4.22.6 TODAY FAST PATH. AutoScan is today-only, so the first verified
            # nationwide page is the cheapest authoritative chronology probe. If it
            # already contains today's listings, skip Date Worker prediction entirely
            # and start collection at page 1. If the newest verified page is already
            # older/empty, today's category is proven empty. Weak/mixed evidence falls
            # through to the unchanged universal Date Worker/local locator.
            if user_id == RADAR_AUTOSCAN_USER_ID and today_fast_path and str(feed_name) == "nationwide":
                try:
                    _tf_items, tf_relation, _tf_pairs, _tf_days = await stable_fetch(1, "today_fast")
                    tf_info = cache.get(1)
                    tf_verified = bool(
                        tf_info is not None
                        and getattr(tf_info, "page_verified", False)
                        and getattr(tf_info, "request_matches_page", True)
                        and not getattr(tf_info, "suspicious", False)
                    )
                    if tf_relation == "target" and tf_verified:
                        try:
                            await save_date_index(
                                cat.key, target_date, base_url, status="found",
                                candidate_page=1, max_page=site_max_page,
                            )
                        except Exception:
                            log.debug("Today fast-path date-index write failed", exc_info=True)
                        await remember_confirmed_date_hint(1)
                        log.info(
                            "Today fast path HIT category=%s target=%s page=1 requests=%s",
                            cat.name, target_date, network_requests,
                        )
                        return locator_result("found", candidate_page=1)
                    if tf_relation in {"older", "empty"} and tf_verified:
                        try:
                            await save_date_index(
                                cat.key, target_date, base_url, status="absent", max_page=site_max_page,
                            )
                        except Exception:
                            log.debug("Today fast-path absent write failed", exc_info=True)
                        log.info(
                            "Today fast path EMPTY category=%s target=%s relation=%s requests=%s",
                            cat.name, target_date, tf_relation, network_requests,
                        )
                        return locator_result("absent", "самая новая подтверждённая страница уже не содержит сегодняшних объявлений")
                    log.info(
                        "Today fast path fallback category=%s target=%s relation=%s verified=%s; universal locator continues",
                        cat.name, target_date, tf_relation, tf_verified,
                    )
                except TemporaryAccessError:
                    raise
                except Exception:
                    log.debug("Today fast path failed; universal locator continues", exc_info=True)

            # v4.1.0 kept one deterministic newest-sorted locator for every date/feed.
            # Sparse/hidden timestamp templates are not a broken page: any weak
            # probe falls back to the deterministic sequential stream below.
            # v4.2.4 Fast Date Search. Keep the same verified stable_fetch() and
            # sequential collection, but locate the first target-date page with
            # exponential probes + a binary boundary search. If chronology is weak
            # at any probe, fall back to the old deterministic sequential locator.
            indexed = None
            try:
                indexed = await load_date_index(cat.key, target_date, base_url)
            except Exception:
                log.debug("Stable date-index read failed", exc_info=True)
            if indexed and indexed.get("status") == "found" and indexed.get("candidate_page"):
                candidate_page = max(1, int(indexed["candidate_page"]))
                try:
                    _items, relation, _pairs, _days = await stable_fetch(candidate_page, "stable_scan")
                    if relation == "target":
                        await remember_confirmed_date_hint(candidate_page)
                        return locator_result("found", candidate_page=candidate_page)
                except TemporaryAccessError:
                    raise
                except Exception:
                    log.debug("Stored date index could not be revalidated", exc_info=True)

            # v4.3.24 Date Worker PRO. Dedicated Railway replicas probe the
            # exponential chronology checkpoints in parallel and return only a
            # boundary hint. The stable foreground parser ALWAYS revalidates that
            # boundary locally before accepting the selected date. If anything is
            # weak/inconsistent/offline, execution falls through to the unchanged
            # v4.2.4 local exponential + binary locator below.
            if REMOTE_DATE_WORKER_ENABLED:
                live_date = category_live_progress.get(progress_key)
                previous_stage = getattr(live_date, "transport_stage", "") if live_date is not None else ""
                previous_phase = getattr(live_date, "phase", "") if live_date is not None else ""
                active_date_phase = "regional_date" if str(feed_name).startswith("hidden:") else "date_worker"
                if live_date is not None:
                    live_date.transport_stage = "date-worker-probes"
                    live_date.phase = active_date_phase
                try:
                    # v4.3.33: hidden_fill may already have a Date Worker hint
                    # running in parallel for this regional feed.  Reuse that exact
                    # hint instead of starting a duplicate Redis/date-probe search.
                    # The stable foreground parser below still verifies the page.
                    if remote_hint_override is _REMOTE_HINT_UNSET:
                        hint = await REMOTE_DATE_MANAGER.locate_hint(base_url, target_date)
                    else:
                        hint = remote_hint_override
                    if hint and hint.get("boundary"):
                        boundary_hint = max(1, min(effective_limit, int(hint["boundary"])))
                        # v4.3.28 Cold Date Turbo can prove remotely that even page
                        # 50 is still newer than the requested day. Verify that one
                        # exact page locally and jump straight to regional sharding;
                        # do not repeat the entire local 1/2/4/8/16/32/50 ladder.
                        if bool(hint.get("beyond_public")):
                            _bi, beyond_relation, _bp, _bd = await stable_fetch(effective_limit, "date_verify")
                            if beyond_relation == "newer":
                                # v4.3.30: a hidden/regional feed that is itself deeper
                                # than the public 50-page window MUST expose its child
                                # location shards before we return `too_deep`. Those
                                # shards are discovered from page 1. v4.3.28/29 skipped
                                # page 1 on the Cold Turbo shortcut, so hidden_fill()
                                # could not recurse and incorrectly marked the category
                                # partial even though the chronology was valid.
                                if str(feed_name).startswith("hidden:") and not discovered_shards:
                                    _si, shard_relation, _sp, _sd = await stable_fetch(1, "date_verify")
                                    if shard_relation in {"unknown", "invalid"}:
                                        log.warning(
                                            "Cold Date Turbo shard discovery weak; falling back to local locator category=%s feed=%s target=%s relation=%s",
                                            cat.name, feed_name, target_date, shard_relation,
                                        )
                                    else:
                                        log.info(
                                            "Cold Date Turbo discovered %s child shards before regional too_deep category=%s feed=%s",
                                            len(discovered_shards), cat.name, feed_name,
                                        )
                                # If page 1 itself was weak, do not accept the shortcut:
                                # let the proven local locator continue and recover it.
                                if not (
                                    str(feed_name).startswith("hidden:")
                                    and not discovered_shards
                                    and shard_relation in {"unknown", "invalid"}
                                ):
                                    try:
                                        await save_date_index(
                                            cat.key, target_date, base_url, status="too_deep",
                                            max_page=site_max_page,
                                        )
                                    except Exception:
                                        log.debug("Stable date-index too_deep write failed", exc_info=True)
                                    log.info(
                                        "Cold Date Turbo verified target beyond public window category=%s target=%s page=%s workers=%s shards=%s",
                                        cat.name, target_date, effective_limit, int(hint.get("workers", 0) or 0),
                                        len(discovered_shards),
                                    )
                                    return locator_result(
                                        "too_deep",
                                        "дата глубже публичного окна; региональная структура подтверждена",
                                    )
                        # v4.8.5 Smart Hint Rejection. Verify the hinted page first,
                        # then move ONE page toward the target based on chronology.
                        # If both pages say the same direction, the remote hint is far
                        # off and the proven local locator is cheaper than checking four
                        # more neighbours. Example: hint=17 -> older, page16 -> older: 
                        # immediately fall back instead of probing 18/15/19/20 too.
                        remote_confirmed = None
                        remote_weak = False
                        first_relation = None
                        _vi, first_relation, _vp, _vd = await stable_fetch(boundary_hint, "date_verify")
                        if first_relation == "target":
                            remote_confirmed = boundary_hint
                        elif first_relation in {"unknown", "invalid"}:
                            remote_weak = True
                        elif first_relation in {"newer", "older"}:
                            direction = 1 if first_relation == "newer" else -1
                            directional_page = boundary_hint + direction
                            if 1 <= directional_page <= effective_limit:
                                _vi2, second_relation, _vp2, _vd2 = await stable_fetch(directional_page, "date_verify")
                                if second_relation == "target":
                                    remote_confirmed = directional_page
                                elif second_relation in {"unknown", "invalid"}:
                                    remote_weak = True
                                elif second_relation == first_relation:
                                    log.info(
                                        "Date Worker directional miss category=%s target=%s hint=%s relation=%s next=%s; local locator fallback",
                                        cat.name, target_date, boundary_hint, first_relation, directional_page,
                                    )
                                # A direction flip without an exact target is a bracket,
                                # not proof of the first target page. Let the local
                                # exponential/binary locator resolve it safely.
                        else:
                            # Mixed/empty hints are acceleration misses, not errors.
                            remote_weak = False
                        if remote_confirmed is not None:
                            candidate = remote_confirmed
                            walkback_steps = 0
                            while candidate > 1:
                                prev = candidate - 1
                                _i2, prev_relation, _p2, _d2 = await stable_fetch(prev, "date_verify")
                                if prev_relation == "target":
                                    candidate = prev
                                    walkback_steps += 1
                                    if walkback_steps >= REMOTE_DATE_MAX_LINEAR_WALKBACK:
                                        # A remote hint landed deep inside a long target-day
                                        # run.  Do not keep walking page-by-page: reject the
                                        # acceleration hint and let the local exponential /
                                        # binary locator find the exact first target page.
                                        remote_weak = True
                                        log.warning(
                                            "Date Worker wide target hint capped category=%s target=%s hint=%s walkback=%s; local locator fallback",
                                            cat.name, target_date, boundary_hint, walkback_steps,
                                        )
                                        break
                                    continue
                                if prev_relation == "newer":
                                    break
                                if prev_relation == "unknown":
                                    # v4.15.5: do not accept a remote hint across a
                                    # persistently weak predecessor. Let the local
                                    # locator/sequential recovery prove the boundary.
                                    remote_weak = True
                                    log.warning(
                                        "Date Worker walkback predecessor unknown category=%s target=%s page=%s; local recovery",
                                        cat.name, target_date, prev,
                                    )
                                    break
                                if prev_relation == "invalid":
                                    remote_weak = True
                                break
                            if not remote_weak:
                                try:
                                    await save_date_index(
                                        cat.key, target_date, base_url, status="found",
                                        candidate_page=candidate, max_page=site_max_page,
                                    )
                                except Exception:
                                    log.debug("Stable date-index write failed", exc_info=True)
                                await remember_confirmed_date_hint(candidate)
                                log.info(
                                    "Date Worker confirmed category=%s target=%s hint=%s page=%s workers=%s local_requests=%s predictor=%s/%s",
                                    cat.name, target_date, boundary_hint, candidate,
                                    int(hint.get("workers", 0) or 0), network_requests,
                                    str(hint.get("predictor_source") or "cold"),
                                    int(hint.get("predictor_page", 0) or 0) or "—",
                                )
                                return locator_result("found", candidate_page=candidate)
                        log.info(
                            "Date Worker hint not locally confirmed category=%s target=%s hint=%s weak=%s; local locator fallback",
                            cat.name, target_date, boundary_hint, remote_weak,
                        )
                except TemporaryAccessError:
                    raise
                except Exception:
                    log.debug("Date Worker acceleration failed; local locator fallback", exc_info=True)
                finally:
                    if live_date is not None and getattr(live_date, "transport_stage", "") == "date-worker-probes":
                        live_date.transport_stage = previous_stage or "browser"
                    if live_date is not None and getattr(live_date, "phase", "") == active_date_phase:
                        live_date.phase = previous_phase or ("regional_date" if str(feed_name).startswith("hidden:") else "jumping")

            async def sequential_locator(start_page: int = 1):
                saw_newer = False
                saw_chronology = False
                weak_total = 0
                page = max(1, int(start_page))
                while page <= effective_limit:
                    items, relation, pairs, days = await stable_fetch(page, "stable_scan")
                    if relation == "target":
                        saw_chronology = True
                        if weak_total:
                            return locator_result(
                                "unknown",
                                f"последовательный recovery встретил {weak_total} слабых страниц до найденной даты",
                            )
                        try:
                            await save_date_index(
                                cat.key, target_date, base_url, status="found",
                                candidate_page=page, max_page=site_max_page,
                            )
                        except Exception:
                            log.debug("Stable date-index write failed", exc_info=True)
                        await remember_confirmed_date_hint(page)
                        return locator_result("found", candidate_page=page)
                    if relation == "newer":
                        saw_chronology = True
                        saw_newer = True
                        page += 1
                        continue
                    if relation == "empty":
                        if weak_total:
                            return locator_result(
                                "unknown",
                                f"последовательный recovery завершился после {weak_total} слабых страниц",
                            )
                        try:
                            await save_date_index(cat.key, target_date, base_url, status="absent", max_page=site_max_page)
                        except Exception:
                            pass
                        return locator_result("absent", "последовательный проход завершил выдачу")
                    if relation == "older":
                        saw_chronology = True
                        if weak_total:
                            return locator_result(
                                "unknown",
                                f"последовательный recovery пересёк дату после {weak_total} слабых страниц",
                            )
                        try:
                            await save_date_index(cat.key, target_date, base_url, status="absent", max_page=site_max_page)
                        except Exception:
                            pass
                        reason_text = (
                            "последовательный проход пересёк выбранную дату без объявлений"
                            if saw_newer else "самые новые объявления уже старше выбранной даты"
                        )
                        return locator_result("absent", reason_text)
                    if relation == "mixed":
                        saw_chronology = True
                        if weak_total:
                            return locator_result(
                                "unknown",
                                f"последовательный recovery встретил границу после {weak_total} слабых страниц",
                            )
                        try:
                            await save_date_index(cat.key, target_date, base_url, status="absent", max_page=site_max_page)
                        except Exception:
                            pass
                        return locator_result("absent", "граница выбранной даты пройдена без точных совпадений")
                    if relation == "invalid":
                        return locator_result("invalid", invalid_note or f"не удалось получить страницу {page}")
                    if relation == "unknown":
                        weak_total += 1
                        page += 1
                        continue
                    page += 1

                if not saw_chronology and weak_total:
                    return locator_result(
                        "unknown",
                        f"на {weak_total} страницах не удалось извлечь ни одной надёжной даты",
                    )
                try:
                    await save_date_index(cat.key, target_date, base_url, status="too_deep", max_page=site_max_page)
                except Exception:
                    pass
                return locator_result("too_deep", "дата глубже публичного окна; перехожу к независимым регионам")

            async def sequential_recovery(reason: str, start_page: int = 1):
                start_page = max(1, min(effective_limit, int(start_page or 1)))
                log.warning(
                    "Date sequential recovery category=%s target=%s feed=%s start=%s reason=%s",
                    cat.name, target_date, feed_name, start_page, reason,
                )
                result = await sequential_locator(start_page)
                log.info(
                    "Date sequential recovery result category=%s target=%s feed=%s status=%s candidate=%s requests=%s",
                    cat.name, target_date, feed_name, result.get("status"), result.get("candidate_page"), network_requests,
                )
                return result

            async def recover_weak_probe(page: int, low_bound: int = 1, high_bound: int | None = None):
                """Resolve one weak chronology probe locally, never with a full linear rewind.

                v4.2.4 could turn a fast 1/2/4/8/... lookup into a 1..50 crawl as
                soon as one page hid its timestamps.  We now inspect at most four
                adjacent pages (distance 1-2).  If none has trustworthy chronology,
                the scan reports an unknown boundary instead of opening dozens of
                extra browser navigations.
                """
                upper = effective_limit if high_bound is None else min(effective_limit, int(high_bound))
                lower = max(1, int(low_bound))
                checked = {int(page)}
                for radius in (1, 2):
                    for candidate in (int(page) - radius, int(page) + radius):
                        if candidate in checked or candidate < lower or candidate > upper:
                            continue
                        checked.add(candidate)
                        _ri, recovered_relation, _rp, _rd = await stable_fetch(candidate, "date_recover")
                        if recovered_relation == "target":
                            return candidate, recovered_relation
                        if recovered_relation not in {"unknown", "invalid"}:
                            return candidate, recovered_relation
                return int(page), "unknown"

            # Page 1 is always checked first. For today's date this normally ends
            # immediately; yesterday/older dates then jump 2,4,8,16,32,50.
            low_newer = 0
            high: int | None = None
            probe = 1
            while True:
                _items, relation, _pairs, _days = await stable_fetch(probe, "date_probe")
                if relation == "invalid":
                    return locator_result("invalid", invalid_note or f"не удалось получить страницу {probe}")
                if relation == "unknown":
                    recovered_page, recovered_relation = await recover_weak_probe(
                        probe, max(1, low_newer + 1), effective_limit
                    )
                    log.info(
                        "Fast date local recovery category=%s target=%s weak_probe=%s recovered_page=%s relation=%s",
                        cat.name, target_date, probe, recovered_page, recovered_relation,
                    )
                    if recovered_relation == "unknown":
                        return await sequential_recovery(
                            f"weak exponential probe page={probe}", max(1, low_newer + 1)
                        )
                    probe = recovered_page
                    relation = recovered_relation
                if relation == "newer":
                    low_newer = probe
                    if probe >= effective_limit:
                        try:
                            await save_date_index(cat.key, target_date, base_url, status="too_deep", max_page=site_max_page)
                        except Exception:
                            pass
                        return locator_result("too_deep", "дата глубже публичного окна; перехожу к независимым регионам")
                    next_probe = min(effective_limit, 2 if probe == 1 else probe * 2)
                    if next_probe == probe:
                        return locator_result("too_deep", "дата глубже публичного окна")
                    probe = next_probe
                    continue
                # target / older / mixed / empty all prove that the first target
                # page, if present, is no deeper than this probe.
                high = probe
                break

            lo = max(1, low_newer + 1)
            hi = max(lo, int(high or lo))
            while lo < hi:
                mid = (lo + hi) // 2
                _items, relation, _pairs, _days = await stable_fetch(mid, "date_probe")
                if relation == "invalid":
                    return locator_result("invalid", invalid_note or f"не удалось получить страницу {mid}")
                if relation == "unknown":
                    recovered_page, recovered_relation = await recover_weak_probe(mid, lo, hi)
                    log.info(
                        "Fast date binary local recovery category=%s target=%s weak_mid=%s recovered_page=%s relation=%s",
                        cat.name, target_date, mid, recovered_page, recovered_relation,
                    )
                    if recovered_relation == "unknown":
                        return await sequential_recovery(
                            f"weak binary boundary page={mid}", max(1, low_newer + 1)
                        )
                    mid = recovered_page
                    relation = recovered_relation
                if relation == "newer":
                    lo = min(hi, mid + 1)
                else:
                    hi = max(lo, mid)

            boundary = lo
            # Verify a tight neighborhood. This protects against a page where a few
            # cards hide dates without turning date discovery back into a 25-page walk.
            start_verify = max(1, boundary - 2)
            end_verify = min(effective_limit, boundary + 3)
            saw_newer = False
            saw_older = False
            saw_unknown = False
            for page in range(start_verify, end_verify + 1):
                _items, relation, _pairs, _days = await stable_fetch(page, "date_verify")
                if relation == "target":
                    candidate = page
                    # Walk back only while the immediately previous page is also
                    # target. Binary search already placed us at the first
                    # non-newer boundary, so this normally costs zero/one request.
                    while candidate > 1:
                        prev = candidate - 1
                        _i2, prev_relation, _p2, _d2 = await stable_fetch(prev, "date_verify")
                        if prev_relation == "target":
                            candidate = prev
                            continue
                        if prev_relation == "newer":
                            break
                        if prev_relation == "unknown":
                            return await sequential_recovery(
                                f"weak predecessor immediately before target page={candidate}",
                                max(1, low_newer + 1),
                            )
                        if prev_relation == "invalid":
                            return locator_result("invalid", invalid_note or f"не удалось получить страницу {prev}")
                        break
                    try:
                        await save_date_index(
                            cat.key, target_date, base_url, status="found",
                            candidate_page=candidate, max_page=site_max_page,
                        )
                    except Exception:
                        log.debug("Stable date-index write failed", exc_info=True)
                    await remember_confirmed_date_hint(candidate)
                    log.info(
                        "Fast date locator found category=%s target=%s page=%s requests=%s",
                        cat.name, target_date, candidate, network_requests,
                    )
                    return locator_result("found", candidate_page=candidate)
                if relation == "newer":
                    saw_newer = True
                elif relation in {"older", "mixed", "empty"}:
                    saw_older = True
                elif relation == "invalid":
                    return locator_result("invalid", invalid_note or f"не удалось получить страницу {page}")
                elif relation == "unknown":
                    saw_unknown = True
                    continue

            full_feed_visible = site_max_page is not None and site_max_page <= effective_limit
            if saw_newer and saw_older and full_feed_visible and not saw_unknown:
                try:
                    await save_date_index(cat.key, target_date, base_url, status="absent", max_page=site_max_page)
                except Exception:
                    pass
                return locator_result("absent", "быстрый поиск подтвердил пересечение даты без объявлений")
            if saw_older and low_newer == 0 and full_feed_visible and not saw_unknown:
                return locator_result("absent", "самые новые объявления уже старше выбранной даты")
            if saw_unknown:
                return await sequential_recovery(
                    "weak pages around verified boundary", max(1, low_newer + 1)
                )
            # Large feeds need the existing hidden/regional logic rather than a
            # false zero when the public 50-page window cannot prove absence.
            if not full_feed_visible and saw_older:
                return locator_result("ambiguous_absent", "large feed requires independent sub-feed verification")
            return await sequential_recovery(
                "fast locator could not prove date boundary", max(1, low_newer + 1)
            )

        low_newer = 0
        high: int | None = None
        probe = 1
        while True:
            items, relation, pairs, days = await fetch(probe, "jumping")
            if relation in {"target", "older", "mixed", "empty"}:
                high = probe
                break
            if relation == "invalid":
                return locator_result("invalid", invalid_note)
            if relation == "unknown":
                # Unknown chronology is not a valid jump signal. Nearby pages may be
                # healthier, but if the current probe is page 1 we cannot safely infer
                # a direction and therefore return a partial result instead of zero.
                return locator_result("unknown", "publication dates could not be verified")
            if relation == "newer":
                low_newer = probe
            if probe >= effective_limit:
                return locator_result("too_deep", "target beyond public page window")
            next_probe = min(effective_limit, 2 if probe == 1 else probe * 2)
            if next_probe == probe:
                return locator_result("too_deep", "target beyond public page window")
            probe = next_probe

        lo = max(1, low_newer + 1)
        hi = high
        while lo < hi:
            mid = (lo + hi) // 2
            items, relation, pairs, days = await fetch(mid, "jumping")
            if relation == "invalid":
                return locator_result("invalid", invalid_note)
            if relation == "unknown":
                return locator_result("unknown", "weak date coverage near boundary")
            if relation == "newer":
                lo = mid + 1
            else:
                hi = mid
        boundary = lo

        candidate = None
        saw_newer = saw_older = saw_unknown = False
        for page in range(max(1, boundary - 3), min(effective_limit, boundary + 5) + 1):
            items, relation, pairs, days = await fetch(page, "jumping")
            if relation == "target":
                candidate = page
                break
            if relation == "newer":
                saw_newer = True
            elif relation in {"older", "mixed", "empty"}:
                saw_older = True
            elif relation == "unknown":
                saw_unknown = True
            elif relation == "invalid":
                return locator_result("invalid", invalid_note)

        if candidate is not None:
            # Walk back a few pages so the first boundary card cannot be missed.
            for back in range(candidate - 1, max(0, candidate - 4), -1):
                items, relation, pairs, days = await fetch(back, "jumping")
                if relation == "target":
                    candidate = back
                elif relation == "newer":
                    break
                elif relation in {"unknown", "invalid"}:
                    return locator_result("unknown", "could not verify page immediately before target")
                else:
                    break
            await remember_confirmed_date_hint(candidate)
            return locator_result("found", candidate_page=candidate)

        # A zero is allowed only when the *whole* feed is visible inside the public
        # page window. For a large feed (>50 pages), a local crossing is not enough
        # evidence to conclude that an entire category has zero listings on that day;
        # we must split this category into smaller official location feeds first.
        full_feed_visible = site_max_page is not None and site_max_page <= effective_limit
        if saw_newer and saw_older and not saw_unknown:
            if full_feed_visible:
                return locator_result("absent", "verified date crossing without target listings")
            return locator_result("ambiguous_absent", "large feed requires independent sub-feed verification")
        if saw_older and low_newer == 0 and not saw_unknown:
            if full_feed_visible:
                return locator_result("absent", "feed starts after the selected calendar day")
            return locator_result("ambiguous_absent", "large feed requires independent sub-feed verification")
        return locator_result("unknown", "could not verify date boundary")

    async def prefetch_collection_range(base_url: str, start_page: int, end_page: int) -> None:
        """Warm post-locator pages on dedicated Railway Page Worker replicas.

        This helper is acceleration-only: any worker outage, timeout or partial
        batch leaves missing pages to the original local fetch() path. The date
        locator itself is never delegated.
        """
        if not REMOTE_PAGE_WORKER_ENABLED or end_page < start_page:
            return
        requests = [
            (page, page_url(base_url, page))
            for page in range(max(1, int(start_page)), max(1, int(end_page)) + 1)
        ]
        if not requests:
            return
        live = category_live_progress.get(progress_key)
        previous_stage = getattr(live, "transport_stage", "") if live is not None else ""
        if live is not None:
            live.transport_stage = "page-worker-prefetch"
        try:
            # v4.3.22 streaming dispatch: enqueue the range and immediately start
            # foreground collection. The previous v4.3.21 path awaited the entire
            # remote batch here, which was the visible 0/N pause before progress.
            await REMOTE_PAGE_MANAGER.prefetch(requests, wait_for_results=False)
        except Exception:
            log.debug(
                "Page Worker prefetch failed category=%s pages=%s-%s; local fallback stays active",
                cat.key, start_page, end_page, exc_info=True,
            )
        finally:
            if live is not None and getattr(live, "transport_stage", "") == "page-worker-prefetch":
                live.transport_stage = previous_stage or "browser"

    async def collect_direct(locator) -> tuple[str, int]:
        """Collect nationwide pages without turning isolated weak pages into deep fallback.

        v4.0.1 uses the sorted feed as a stream of chronology evidence. A weak
        card template is retried and recorded, but only a sustained chronology
        failure can stop the direct pass. Regional hidden-fill is reserved for a
        real public-window/depth problem.
        """
        nonlocal direct_pages_collected, collection_pages_confirmed, collection_start_page, request_complete, reason, hit_limit
        candidate = int(locator["candidate"])
        limit = int(locator["limit"])
        fetch = locator["fetch"]
        collection_start_page = candidate
        live = category_live_progress.get(progress_key)
        if live is not None:
            live.phase = "collecting"
            live.collection_start_page = candidate
            live.collection_index = 0

        # v4.4.0 rolling Page Worker prefetch. Warm only a short window and top
        # it up while the target day continues, instead of scheduling the whole
        # 15/25/50-page request before chronology proves those pages are needed.
        direct_base_url = str(locator.get("base_url") or cat.url)
        direct_prefetch_last = candidate

        async def top_up_direct_prefetch(current_page: int) -> None:
            nonlocal direct_prefetch_last
            next_range = rolling_prefetch_range(
                current_page, direct_prefetch_last, limit,
                window_pages=PAGE_PREFETCH_WINDOW_PAGES + PAGE_PREFETCH_EXTRA_PAGES,
                low_water_pages=PAGE_PREFETCH_LOW_WATER_PAGES,
            )
            if next_range is None:
                return
            start_page, end_page = next_range
            await prefetch_collection_range(direct_base_url, start_page, end_page)
            direct_prefetch_last = max(direct_prefetch_last, end_page)

        await top_up_direct_prefetch(candidate)

        page = candidate
        weak_streak = 0
        weak_pages = 0
        repeated_recovery_skips = 0
        # Small look-ahead compensates for weak card-template pages without
        # silently reducing the requested 15/25/50-page depth.
        # Recent dates may start many pages below page 1. For them the requested
        # 15/25/50 depth is measured from the first target-date page, not from page 1,
        # so allow the deterministic stream to walk the whole public window.
        # Universal stream may need to skip many newer pages before reaching an
        # arbitrary historical date, so the search/collection walk may use the
        # entire verified public window. Only confirmed target pages count toward
        # the user's requested 15/25/50 depth.
        hard_stop = limit
        while page <= hard_stop:
            items, relation, pairs, days = await fetch(page, "collecting")
            target_on_page = any(d == target_day for d in days)

            if relation == "invalid":
                # v4.10.1 Repeated Page Recovery. A stable repeated-content verdict
                # means this exact page duplicated another page even after the normal
                # retries and a clean BrowserContext recycle. It contributes NO rows
                # and NO confirmed depth, but one isolated duplicate must not abort
                # the whole category. Walk to the next nationwide page and replace
                # the missing verified page later. Other invalid classes (challenge,
                # page identity, transport/normalization) remain strict failures.
                invalid_kind_getter = locator.get("last_invalid_kind")
                invalid_kind = invalid_kind_getter() if callable(invalid_kind_getter) else ""
                if invalid_kind == "repeated-content":
                    repeated_recovery_skips += 1
                    log.warning(
                        "Repeated page recovery skip category=%s page=%s skip=%s/%s confirmed=%s/%s",
                        cat.name, page, repeated_recovery_skips, DIRECT_REPEATED_RECOVERY_LIMIT,
                        direct_pages_collected, depth,
                    )
                    if repeated_recovery_skips > DIRECT_REPEATED_RECOVERY_LIMIT:
                        hit_limit = True
                        reason = (
                            f"слишком много повторяющихся страниц выдачи: "
                            f"{repeated_recovery_skips} за проход"
                        )
                        return "invalid_stop", direct_pages_collected
                    await top_up_direct_prefetch(page)
                    page += 1
                    if PAGE_DELAY_SECONDS:
                        await asyncio.sleep(min(PAGE_DELAY_SECONDS, 0.25))
                    continue

                # A persistent non-repeat invalid page is a genuine page/transport
                # failure. The per-page retry wrapper has already exhausted attempts.
                hit_limit = True
                reason = f"страница {page} не была корректно получена после повторов"
                return "invalid_stop", direct_pages_collected

            if relation == "unknown":
                weak_streak += 1
                weak_pages += 1
                # Unknown chronology is not a failed page. Preserve exact target-day
                # cards if any were parsed, then keep walking until a later page gives
                # us a trustworthy target/older boundary. Never make the whole scan
                # partial merely because several cards hide their timestamp.
                if target_on_page:
                    await process_target_items(items, pairs)
                    direct_pages_collected += 1
                    collection_pages_confirmed += 1
                    update_live(page, days, "collecting", direct_pages_collected)
                    if direct_pages_collected >= depth:
                        request_complete = True
                        reason = f"собрано {depth} подтверждённых страниц выбранной даты"
                        return "done", direct_pages_collected
                await top_up_direct_prefetch(page)
                page += 1
                continue

            weak_streak = 0
            if relation == "empty":
                request_complete = True
                reason = "выдача закончилась раньше выбранной глубины"
                return "done", direct_pages_collected
            if relation == "older" or (relation == "mixed" and not target_on_page):
                request_complete = True
                reason = (
                    "выбранная дата закончилась раньше выбранной глубины"
                    if target_seen_any else "выбранная дата пройдена; объявлений за неё не найдено"
                )
                return "done", direct_pages_collected

            if target_on_page or relation == "target":
                direct_pages_collected += 1
                collection_pages_confirmed += 1
                update_live(page, days, "collecting", direct_pages_collected)
                await process_target_items(items, pairs)
                if direct_pages_collected >= depth:
                    request_complete = True
                    reason = f"собрано {depth} страниц от начала выбранной даты"
                    return "done", direct_pages_collected

            await top_up_direct_prefetch(page)
            page += 1
            if PAGE_DELAY_SECONDS:
                await asyncio.sleep(min(PAGE_DELAY_SECONDS, 0.25))

        if page > limit:
            hit_limit = True
            return "needs_hidden", direct_pages_collected

        # Reaching the public window without an older boundary means the target
        # may continue deeper. Ask the shard layer for the remaining confirmed
        # target pages; sparse timestamp pages never cause a fake partial by themselves.
        return "needs_hidden", direct_pages_collected

    # v4.3.28 rolling regional Date Worker prewarm. Only Date Worker probes run
    # concurrently here; foreground local verification/collection stays on the
    # proven single parser path. This makes the speed-up low-risk.
    hidden_prewarm_tasks: dict[str, asyncio.Task] = {}
    _regional_prewarm_enabled = bool(
        REMOTE_DATE_WORKER_ENABLED
        and (HIDDEN_DATE_PREWARM_ENABLED or REGIONAL_DATE_PIPELINE_ENABLED)
    )
    _regional_prewarm_window = max(
        HIDDEN_DATE_PREWARM_WINDOW if HIDDEN_DATE_PREWARM_ENABLED else 0,
        REGIONAL_DATE_PIPELINE_WINDOW if REGIONAL_DATE_PIPELINE_ENABLED else 0,
        1,
    )
    _regional_prewarm_concurrency = max(
        HIDDEN_DATE_PREWARM_CONCURRENCY if HIDDEN_DATE_PREWARM_ENABLED else 0,
        REGIONAL_DATE_PIPELINE_CONCURRENCY if REGIONAL_DATE_PIPELINE_ENABLED else 0,
        1,
    )
    hidden_prewarm_sem = asyncio.Semaphore(_regional_prewarm_concurrency)

    async def _prewarm_hidden_date(feed_url: str):
        if not _regional_prewarm_enabled:
            return None
        async with hidden_prewarm_sem:
            try:
                return await REMOTE_DATE_MANAGER.locate_hint(feed_url, target_date)
            except asyncio.CancelledError:
                raise
            except Exception:
                log.debug("Regional Date Worker pipeline hint failed url=%s", feed_url, exc_info=True)
                return None

    def schedule_hidden_date_prewarm(entries: list[tuple[str, str, int]]) -> None:
        if not _regional_prewarm_enabled:
            return
        scheduled = sum(1 for task in hidden_prewarm_tasks.values() if not task.done())
        for _name, feed_url, _level in entries:
            if feed_url in hidden_prewarm_tasks:
                continue
            if scheduled >= _regional_prewarm_window:
                break
            hidden_prewarm_tasks[feed_url] = asyncio.create_task(
                _prewarm_hidden_date(feed_url),
                name=f"date-prewarm-{cat.key}-{len(hidden_prewarm_tasks)+1}",
            )
            scheduled += 1

    async def await_hidden_date_prewarm(feed_url: str):
        task = hidden_prewarm_tasks.get(feed_url)
        if task is None:
            return _REMOTE_HINT_UNSET
        try:
            return await task
        except asyncio.CancelledError:
            raise
        except Exception:
            return None

    async def cancel_hidden_date_prewarm() -> None:
        # v4.3.29: never leave speculative regional date jobs running after the
        # category has finished. Orphan prewarm traffic can collide with Page/View
        # phases and turn a healthy scan into a transient partial.
        pending = [task for task in hidden_prewarm_tasks.values() if not task.done()]
        for task in pending:
            task.cancel()
        if pending:
            await asyncio.gather(*pending, return_exceptions=True)

    async def hidden_fill(remaining_virtual_pages: int) -> tuple[bool, bool]:
        """Fill remaining depth from independent location feeds.

        v3.1.3 keeps the multi-category false-zero fix and adds resilient 403 recovery. Every category owns its own
        locator state. If a state feed is itself larger than Kleinanzeigen's public
        50-page window, it is recursively split into smaller official location feeds
        discovered from that category page. Nothing from the previous selected
        category is reused.
        """
        nonlocal request_complete, reason, hit_limit, collection_pages_confirmed
        if remaining_virtual_pages <= 0:
            return True, False

        # v3.3.1: hidden/regional fallback measures depth by real verified result
        # pages, not by how many listings survive seller/promotion/dedupe filters.
        # A Kleinanzeigen page can contain far fewer than 25 usable private ads;
        # treating 25 surviving rows as one page caused false "partial" warnings.
        hidden_pages_collected = 0
        goal_pages = max(0, int(remaining_virtual_pages))
        unresolved = False
        visited: set[str] = set()
        max_hidden_feeds = 180
        max_shard_depth = 2
        feeds_processed = 0

        live = category_live_progress.get(progress_key)
        if live is not None:
            live.phase = "regional_date"
            live.transport_stage = "date-worker-regional-prewarm"
            live.collection_start_page = 0
            live.segment_name = ""
            live.segments_done = 0
            live.segments_total = 0

        queue: list[tuple[str, str, int]] = [
            (state_name, _regional_category_url(cat.url, slug, location_id), 0)
            for state_name, slug, location_id in GERMAN_STATE_SEGMENTS
        ]
        schedule_hidden_date_prewarm(queue)

        def add_children(parent_name: str, loc: dict, level: int) -> bool:
            if level >= max_shard_depth:
                return False
            children = list(loc.get("shards") or [])
            if not children:
                return False
            added = 0
            # Prefer smaller counted feeds; they are more likely to expose the
            # requested historical date within the public 50-page window.
            children.sort(key=lambda item: (item[1] is None, item[1] or 10**12, item[0]))
            for child_url, child_count in children:
                if child_url in visited or any(existing[1] == child_url for existing in queue):
                    continue
                label = f"{parent_name}/{added + 1}"
                queue.append((label, child_url, level + 1))
                added += 1
                if added >= 60:
                    break
            if added:
                schedule_hidden_date_prewarm(queue)
            return added > 0

        regional_locator_wait_seconds = 0.0
        regional_collect_seconds = 0.0

        while queue and hidden_pages_collected < goal_pages and feeds_processed < max_hidden_feeds:
            # v4.3.33: prefer a region whose remote Date Worker hint is already
            # ready.  This prevents the foreground parser from idling behind the
            # first state in list order while another state has already finished.
            selected_index = 0
            if _regional_prewarm_enabled:
                for idx, (_n, queued_url, _l) in enumerate(queue):
                    task = hidden_prewarm_tasks.get(queued_url)
                    if task is not None and task.done():
                        selected_index = idx
                        break
            state_name, feed_url, level = queue.pop(selected_index)
            if feed_url in visited:
                continue
            # Keep the next regional Date Worker window busy while this feed is
            # locally verified and then collected through Page Worker/cache.
            schedule_hidden_date_prewarm(queue)
            visited.add(feed_url)
            feeds_processed += 1
            if live is not None:
                live.phase = "regional_date"
                live.transport_stage = "date-worker-regional-pipeline"
                live.segment_name = state_name
                live.segments_done = feeds_processed - 1
                live.segments_total = max(feeds_processed, feeds_processed + len(queue))

            try:
                locator_wait_started = time.monotonic()
                precomputed_hint = await await_hidden_date_prewarm(feed_url)
                regional_locator_wait_seconds += max(0.0, time.monotonic() - locator_wait_started)
                loc = await locate_feed(
                    feed_url, f"hidden:{state_name}",
                    remote_hint_override=precomputed_hint,
                )
            except TemporaryAccessError as exc:
                unresolved = True
                log.warning("hidden date shard temporary limit category=%s state=%s http=%s", cat.name, state_name, exc.status_code)
                continue
            except Exception as exc:
                unresolved = True
                log.warning("hidden date shard failed category=%s state=%s: %s", cat.name, state_name, exc)
                continue

            status = loc["status"]
            if status in {"too_deep", "ambiguous_absent"}:
                if not add_children(state_name, loc, level):
                    unresolved = True
                continue
            if status in {"invalid", "unknown"}:
                unresolved = True
                continue
            if status == "absent":
                # This is a trustworthy zero only because locate_feed returns
                # `absent` exclusively for a fully visible feed.
                continue

            collect_started = time.monotonic()
            candidate = int(loc["candidate"])
            feed_limit = int(loc["limit"])
            fetch = loc["fetch"]
            page = candidate
            state_exhausted = False
            remaining_goal = max(1, goal_pages - hidden_pages_collected)
            hidden_base_url = str(loc.get("base_url") or feed_url)
            hidden_prefetch_last = candidate

            async def top_up_hidden_prefetch(current_page: int) -> None:
                nonlocal hidden_prefetch_last
                next_range = rolling_prefetch_range(
                    current_page, hidden_prefetch_last, feed_limit,
                    window_pages=min(
                        PAGE_PREFETCH_WINDOW_PAGES + PAGE_PREFETCH_EXTRA_PAGES,
                        max(1, remaining_goal + PAGE_PREFETCH_EXTRA_PAGES),
                    ),
                    low_water_pages=PAGE_PREFETCH_LOW_WATER_PAGES,
                )
                if next_range is None:
                    return
                start_page, end_page = next_range
                await prefetch_collection_range(hidden_base_url, start_page, end_page)
                hidden_prefetch_last = max(hidden_prefetch_last, end_page)

            await top_up_hidden_prefetch(candidate)
            while page <= feed_limit and hidden_pages_collected < goal_pages:
                try:
                    items, relation, pairs, days = await fetch(page, "collecting")
                except TemporaryAccessError:
                    unresolved = True
                    break
                if relation == "invalid":
                    unresolved = True
                    break
                if relation == "unknown":
                    # Same universal rule as the nationwide stream: sparse dates are
                    # not a transport failure. Continue until target/older evidence.
                    await top_up_hidden_prefetch(page)
                    page += 1
                    continue
                if relation == "empty":
                    state_exhausted = True
                    break
                if relation == "older" or (relation == "mixed" and not any(d == target_day for d in days)):
                    state_exhausted = True
                    break
                # Count the actual verified page even when business/promoted cards
                # were filtered out and only a handful of usable listings remain.
                target_on_page = any(d == target_day for d in days)
                if target_on_page:
                    await process_target_items(items, pairs)
                    hidden_pages_collected += 1
                    collection_pages_confirmed += 1
                    update_live(
                        page, days, "collecting",
                        direct_pages_collected + hidden_pages_collected,
                    )
                if hidden_pages_collected < goal_pages:
                    await top_up_hidden_prefetch(page)
                page += 1
                if PAGE_DELAY_SECONDS:
                    await asyncio.sleep(min(PAGE_DELAY_SECONDS, 0.15))

            if page > feed_limit and not state_exhausted and hidden_pages_collected < goal_pages:
                # The target day continues beyond this feed's visible window. Drill
                # down again instead of declaring the category empty/skipped.
                if not add_children(state_name, loc, level):
                    unresolved = True
            regional_collect_seconds += max(0.0, time.monotonic() - collect_started)

        log.info(
            "Regional pipeline category=%s target=%s feeds=%s pages=%s locator_wait=%.2fs collect=%.2fs prewarm_window=%s prewarm_concurrency=%s",
            cat.name, target_date, feeds_processed, hidden_pages_collected,
            regional_locator_wait_seconds, regional_collect_seconds,
            _regional_prewarm_window, _regional_prewarm_concurrency,
        )

        if hidden_pages_collected >= goal_pages:
            # v4.8.6 Coverage Integrity: the user's depth is a count of VERIFIED
            # target-date pages, not a requirement that every regional feed remain
            # flawless. If one regional page repeats/normalizes but another
            # independent region supplies a verified replacement, 50/50 really is
            # 50 confirmed pages and the crawl is complete. Only a verified-page
            # shortfall may produce a partial scan.
            request_complete = True
            hit_limit = False
            if unresolved:
                reason = (
                    f"проверено {depth} подтверждённых страниц выбранной даты; "
                    "один слабый региональный участок заменён другой подтверждённой страницей"
                )
            else:
                reason = f"проверено {depth} реальных страниц выбранной даты"
            return True, False

        if queue and feeds_processed >= max_hidden_feeds:
            unresolved = True

        if not unresolved and not queue:
            # All independent feeds were fully verified and the selected date ended
            # before the requested depth. This may be a real zero for a tiny category.
            request_complete = True
            reason = "выбранная дата закончилась раньше выбранной глубины"
            return False, False

        hit_limit = True
        reason = (
            f"частичный результат: подтверждено {collection_pages_confirmed}/{depth} страниц; "
            f"собрано {today_seen} объявлений выбранной даты после локальных повторов"
        )
        return False, True

    try:
        try:
            nationwide = await locate_feed(cat.url, "nationwide")
        except TemporaryAccessError as exc:
            reason = f"временный лимит Kleinanzeigen (HTTP {exc.status_code}) во время поиска даты"
            nationwide = None

        if nationwide is not None and not reason:
            if nationwide["status"] == "found":
                # v4.3.31: 15/25/50 is a maximum depth inside the nationwide feed.
                # Do not turn a 50-page old-date request into a second multi-minute
                # regional crawl just because the public nationwide window ends.
                # The existing regional engine remains available behind an explicit
                # rollback flag, but the default path is nationwide-only.
                if REGIONAL_HIDDEN_FILL_ENABLED:
                    try:
                        likely_hidden = (
                            int(nationwide.get("candidate") or 1) + int(depth) - 1
                            > int(nationwide.get("limit") or PUBLIC_SEARCH_PAGE_CAP)
                        )
                    except Exception:
                        likely_hidden = False
                    if likely_hidden:
                        initial_hidden = [
                            (state_name, _regional_category_url(cat.url, slug, location_id), 0)
                            for state_name, slug, location_id in GERMAN_STATE_SEGMENTS
                        ]
                        schedule_hidden_date_prewarm(initial_hidden)

                outcome, direct_pages_collected = await collect_direct(nationwide)
                if outcome == "needs_hidden" and not request_complete:
                    remaining = max(0, depth - direct_pages_collected)
                    # v4.10.1: if the nationwide pass lost verified depth specifically
                    # to repeated-content, recover only that shortfall from independent
                    # regional feeds. This is integrity replacement, not the old broad
                    # historical hidden-fill mode. Ordinary clean nationwide scans keep
                    # REGIONAL_HIDDEN_FILL_ENABLED=0 and therefore keep their fast path.
                    repeated_shortfall = max(0, int(repeated_pages or 0))
                    if REGIONAL_HIDDEN_FILL_ENABLED or (remaining > 0 and repeated_shortfall > 0):
                        if remaining > 0 and repeated_shortfall > 0 and not REGIONAL_HIDDEN_FILL_ENABLED:
                            log.info(
                                "Repeated page recovery regional replacement category=%s remaining=%s repeats=%s",
                                cat.name, remaining, repeated_shortfall,
                            )
                        await hidden_fill(remaining)
                    else:
                        # The selected date is still present at the end of the public
                        # nationwide window. Those are all target-date pages we can
                        # verify in nationwide mode, so finish successfully instead
                        # of fabricating the requested depth from regional feeds.
                        request_complete = True
                        hit_limit = False
                        if direct_pages_collected:
                            reason = (
                                f"общая выдача проверена до публичного лимита; "
                                f"собрано {direct_pages_collected} страниц выбранной даты"
                            )
                        else:
                            reason = "общая выдача проверена до публичного лимита"
                elif outcome == "invalid_stop" and not request_complete:
                    # A genuine page/transport identity failure still stays partial.
                    reason = reason or "не удалось корректно получить один из участков выдачи"
            elif nationwide["status"] == "absent":
                request_complete = True
                reason = "выбранная дата надёжно пройдена; объявлений за неё не найдено"
            elif nationwide["status"] == "too_deep":
                # v4.3.32: a far historical date can be completely outside the
                # nationwide 50-page public window. In that case nationwide-only
                # mode cannot possibly return correct rows, so automatically fall
                # back to regional shards even though ordinary hidden-fill remains
                # disabled. This avoids the v4.3.31 false 0-result regression.
                if REGIONAL_HIDDEN_FILL_ENABLED or AUTO_REGIONAL_FALLBACK_TOO_DEEP:
                    initial_hidden = [
                        (state_name, _regional_category_url(cat.url, slug, location_id), 0)
                        for state_name, slug, location_id in GERMAN_STATE_SEGMENTS
                    ]
                    # Do not speculative-prewarm here unless explicitly enabled;
                    # hidden_fill() will use the existing safe Date Worker path.
                    schedule_hidden_date_prewarm(initial_hidden)
                    await hidden_fill(depth)
                else:
                    # Explicit diagnostic-only mode: never turn an unreachable date
                    # into a successful zero. Mark it incomplete so UI/recovery logic
                    # cannot present "nothing found" as a verified result.
                    request_complete = False
                    hit_limit = True
                    reason = (
                        f"выбранная дата находится глубже публичных "
                        f"{int(nationwide.get('limit') or PUBLIC_SEARCH_PAGE_CAP)} страниц общей выдачи; "
                        "для точного результата нужен региональный добор"
                    )
            else:
                # Unknown chronology is a parser-quality issue, not proof that the
                # date is deep. Avoid multiplying one weak page into hundreds of
                # regional requests; automatic recovery will retry the weak area.
                reason = nationwide.get("reason") or "не удалось подтвердить хронологию прямой выдачи"

        if not reason:
            reason = "завершено"

        # v4.1.6: the category crawl is complete before public view counters are
        # collected.  This keeps page traversal fast and gives Telegram a distinct
        # "Собираю просмотры" phase instead of spending 10-15 seconds on every page.
        # v4.8.4 Scan Integrity: never start the expensive views phase from a
        # structurally incomplete crawl. A partial category must first pass the
        # bounded automatic recovery pipeline. If recovery succeeds, that complete
        # pass will collect views once; if it stays partial, we preserve confirmed
        # page data without pretending the scan has entered its final phase.
        if request_complete and need_view_counts and deferred_view_items:
            live = category_live_progress.get(progress_key)
            if live is not None:
                live.phase = "views"
                live.today_seen = today_seen
                live.page_limit = depth
            try:
                if user_id == RADAR_AUTOSCAN_USER_ID:
                    requested_views, verified_views, failed_views = await enrich_autoscan_view_counts(
                        parser, list(deferred_view_items.values()), live
                    )
                else:
                    requested_views, verified_views, failed_views = await enrich_page_view_counts(
                        parser, list(deferred_view_items.values()), live
                    )
                views_requested += int(requested_views or 0)
                views_verified += int(verified_views or 0)
                view_failures += int(failed_views or 0)
            except Exception:
                # A view-threshold result will naturally exclude rows without a
                # counter.  Do not convert a successfully crawled date into a
                # partial category merely because the optional counter endpoint
                # had a transient problem.
                views_requested += len(deferred_view_items)
                view_failures += len(deferred_view_items)
                log.exception(
                    "Deferred view-count phase failed category=%s target=%s",
                    cat.name, target_date,
                )
        elif need_view_counts and deferred_view_items and not request_complete:
            log.warning(
                "Deferred views skipped for incomplete crawl category=%s target=%s confirmed_pages=%s/%s reason=%s",
                cat.name, target_date, collection_pages_confirmed, depth, reason,
            )

        quality_score, quality_note = _calculate_scan_quality(
            listings_parsed=listings_parsed,
            missing_dates=missing_date_count,
            missing_prices=missing_price_count,
            invalid_pages=invalid_pages,
            repeated_pages=repeated_pages,
            low_quality_pages=low_quality_pages,
            verified_pages=verified_pages,
            pages_scanned=network_requests,
            view_failures=view_failures,
            date_complete=request_complete,
        )
        update_quality_live(quality_note if quality_score < 85 else "")

        interrupted = reason.startswith("временный лимит Kleinanzeigen")
        seed_complete = bool(request_complete and not interrupted)
        seed_capped = bool(hit_limit and not interrupted)
        pages_scanned = network_requests

        # CategoryScanState is an optimization/checkpoint summary, not the scan
        # result itself.  Never turn an already successful parse into a partial
        # result merely because this bookkeeping write failed.
        try:
            await save_category_scan_state(
                cat.key,
                target_date=target_date,
                mode=mode,
                pages_scanned=pages_scanned,
                new_count=new_count,
                today_seen=today_seen,
                reason=reason,
                head_ids=first_page_head_ids,
                seed_complete=seed_complete,
                seed_capped=seed_capped,
                coverage_pages=depth if request_complete else 0,
            )
        except Exception:
            log.exception(
                "Non-fatal CategoryScanState save failed category=%s target=%s",
                cat.key, target_date,
            )

        result = ScanResult(
            new_count=new_count,
            pages_scanned=pages_scanned,
            today_seen=today_seen,
            known_count=known_total,
            enriched_count=enriched_total,
            hit_limit=hit_limit,
            reason=reason,
            mode=mode,
            avoided_pages=0,
            date_complete=request_complete,
            oldest_date_seen=oldest_date_seen,
            max_page_reached=max_page_reached,
            matched_ids=sorted(processed_target_ids),
            cards_seen=cards_seen,
            listings_parsed=listings_parsed,
            missing_date_count=missing_date_count,
            missing_price_count=missing_price_count,
            promoted_filtered=promoted_filtered,
            price_reduced_filtered=price_reduced_filtered,
            duplicate_count=duplicate_count,
            invalid_pages=invalid_pages,
            repeated_pages=repeated_pages,
            low_quality_pages=low_quality_pages,
            verified_pages=verified_pages,
            collection_pages_confirmed=collection_pages_confirmed,
            view_failures=view_failures,
            views_requested=views_requested,
            views_verified=views_verified,
            radar_views_complete=bool(user_id != RADAR_AUTOSCAN_USER_ID or views_requested == views_verified),
            radar_views_usable=bool(
                user_id != RADAR_AUTOSCAN_USER_ID
                or _radar_autoscan_views_usable(views_requested, views_verified)
            ),
            view_tail_count=max(0, int(views_requested or 0) - int(views_verified or 0)),
            quality_score=quality_score,
            quality_note=quality_note,
        )
        await record_parser_run(user_id, cat, result, started_at)
        if STABLE_SCAN_ENGINE:
            try:
                await mark_category_job(
                    cat.key, target_date, depth,
                    status="done" if request_complete else "partial",
                    verified_pages=verified_pages,
                    network_requests=network_requests,
                    matched_count=len(processed_target_ids),
                    error_text="" if request_complete else reason,
                )
            except Exception:
                log.debug("Could not persist stable category job summary", exc_info=True)
        log.info(
            "category=%s v3.1-quality target=%s depth=%s requests=%s matched=%s complete=%s quality=%s "
            "cards=%s parsed=%s missing_date=%s promoted=%s reduced_price=%s duplicates=%s invalid_pages=%s repeated_pages=%s low_quality=%s views_failed=%s reason=%s",
            cat.name, target_date, depth, pages_scanned, today_seen, request_complete, quality_score,
            cards_seen, listings_parsed, missing_date_count, promoted_filtered, price_reduced_filtered, duplicate_count,
            invalid_pages, repeated_pages, low_quality_pages, view_failures, reason,
        )
        await cancel_hidden_date_prewarm()
        return result

    except Exception as exc:
        quality_score, quality_note = _calculate_scan_quality(
            listings_parsed=listings_parsed,
            missing_dates=missing_date_count,
            missing_prices=missing_price_count,
            invalid_pages=invalid_pages,
            repeated_pages=repeated_pages,
            low_quality_pages=low_quality_pages,
            verified_pages=verified_pages,
            pages_scanned=network_requests,
            view_failures=view_failures,
            date_complete=False,
        )
        failed = ScanResult(
            new_count=new_count,
            pages_scanned=network_requests,
            today_seen=today_seen,
            known_count=known_total,
            enriched_count=enriched_total,
            hit_limit=False,
            reason="ошибка",
            mode=mode,
            avoided_pages=0,
            date_complete=False,
            oldest_date_seen=oldest_date_seen,
            max_page_reached=max_page_reached,
            matched_ids=sorted(processed_target_ids),
            cards_seen=cards_seen,
            listings_parsed=listings_parsed,
            missing_date_count=missing_date_count,
            missing_price_count=missing_price_count,
            promoted_filtered=promoted_filtered,
            duplicate_count=duplicate_count,
            invalid_pages=invalid_pages,
            repeated_pages=repeated_pages,
            low_quality_pages=low_quality_pages,
            verified_pages=verified_pages,
            collection_pages_confirmed=collection_pages_confirmed,
            view_failures=view_failures,
            views_requested=views_requested,
            views_verified=views_verified,
            radar_views_complete=bool(user_id != RADAR_AUTOSCAN_USER_ID or views_requested == views_verified),
            radar_views_usable=bool(
                user_id != RADAR_AUTOSCAN_USER_ID
                or _radar_autoscan_views_usable(views_requested, views_verified)
            ),
            view_tail_count=max(0, int(views_requested or 0) - int(views_verified or 0)),
            quality_score=quality_score,
            quality_note=quality_note,
        )
        try:
            await record_parser_run(user_id, cat, failed, started_at, success=False, error_text=str(exc))
        except Exception:
            log.exception("Could not record failed parser run")
        if STABLE_SCAN_ENGINE:
            try:
                await mark_category_job(
                    cat.key, target_date, depth, status="failed",
                    verified_pages=verified_pages, network_requests=network_requests,
                    matched_count=len(processed_target_ids), error_text=str(exc),
                )
            except Exception:
                log.debug("Could not persist failed stable category job", exc_info=True)
        await cancel_hidden_date_prewarm()
        raise

def job_keyboard(job_id: str, *, queued: bool = False) -> InlineKeyboardMarkup:
    if queued:
        return InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отменить очередь", callback_data=f"cancel_scan:{job_id}")],
        ])
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⏹ Остановить парсер", callback_data=f"cancel_scan:{job_id}")],
    ])


def stopped_job_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗂 Выбрать категории", callback_data="groups")],
        [InlineKeyboardButton(text="▶️ Новый скан", callback_data="start_scan")],
        [InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ])


def trial_result_keyboard(scan_id: int | None, remaining: int) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if scan_id is not None:
        rows.append([
            InlineKeyboardButton(text="🔥 TOP-12", callback_data=f"scantop:{scan_id}"),
            InlineKeyboardButton(text="📋 TOP-50", callback_data=f"scantop50:{scan_id}:0"),
        ])
        rows.append([InlineKeyboardButton(text="📊 Открыть скан", callback_data=f"scan:{scan_id}")])
    if remaining > 0:
        rows.append([InlineKeyboardButton(text="🎁 Использовать ещё 1 бесплатный скан", callback_data="start_scan")])
    rows.append([InlineKeyboardButton(text="💎 Полный доступ", callback_data="subscription")])
    rows.append([InlineKeyboardButton(text="🏠 Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def trial_conversion_keyboard(remaining: int) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if remaining > 0:
        rows.append([InlineKeyboardButton(text="🎁 Второй бесплатный скан", callback_data="start_scan")])
    rows.append([InlineKeyboardButton(text="💎 Получить полный доступ", callback_data="subscription")])
    rows.append([InlineKeyboardButton(text="📊 Мои сканы", callback_data="my_scans")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def failed_job_keyboard(scan_id: int | None) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if scan_id is not None:
        rows.append([InlineKeyboardButton(text="🔄 Повторить этот скан", callback_data=f"scanrepeat:{scan_id}")])
    rows.append([InlineKeyboardButton(text="📊 Мои сканы", callback_data="my_scans")])
    rows.append([InlineKeyboardButton(text="🏠 Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _cacheable_category_result(result: ScanResult | None) -> bool:
    """Only fully verified category scans may poison/share the short result cache.

    v4.0/v4.0.1 cached partial results for five minutes, so one weak pass could be
    replayed to the same user and to other Railway replicas. Partial/failed work is
    recovery input, never a reusable final result.
    """
    return bool(result is not None and result.date_complete)


async def fresh_category_cache_age(category_key: str, page_limit: int, target_date: str) -> int | None:
    """Legacy DB cache hook. Exact-depth scans use the in-memory ScanResult cache.

    Reconstructing a 25-page result from every listing ever stored for the same
    category/date would silently turn it into a 50/100-page result, so v3.0.6
    intentionally does not use the old DB-only cache for exact-date scans.
    """
    return None


async def _scan_category_task(cat, user_id: int, page_limit: int, target_date: str) -> ScanResult:
    # A browser worker binds one parser/browser session to the whole user job.
    # asyncio tasks inherit ContextVar values, so every sequential category in the
    # job reuses the same independent Chromium context instead of launching a new
    # browser for every category. Local/legacy calls still get an owned parser.
    parser = JOB_PARSER.get()
    owned = parser is None
    if parser is None:
        parser = KleinanzeigenParser()
    try:
        return await scan_one_category(parser, cat, user_id, page_limit, target_date)
    finally:
        if owned:
            await parser.close()


async def _publish_distributed_category_progress(inflight_key: str, lock_token: str) -> None:
    last_refresh = 0.0
    while True:
        try:
            live = category_live_progress.get(inflight_key)
            if live is not None:
                await COORDINATOR.set_category_progress(inflight_key, live)
            now = time.monotonic()
            if now - last_refresh >= 20.0:
                if not await COORDINATOR.refresh_category_lock(inflight_key, lock_token):
                    return
                last_refresh = now
        except asyncio.CancelledError:
            raise
        except Exception:
            log.debug("Could not publish distributed category progress key=%s", inflight_key, exc_info=True)
        await asyncio.sleep(1.0)


async def _run_distributed_category_owner(
    inflight_key: str, lock_token: str, cat, user_id: int, page_limit: int, target_date: str
) -> ScanResult:
    publisher = asyncio.create_task(
        _publish_distributed_category_progress(inflight_key, lock_token),
        name=f"category-progress-publisher:{inflight_key}",
    )
    try:
        result = await _scan_category_task(cat, user_id, page_limit, target_date)
        ttl = max(60, CATEGORY_CACHE_TTL_SECONDS or 60)
        try:
            if _cacheable_category_result(result):
                await COORDINATOR.set_category_result(inflight_key, result, ttl)
            else:
                await COORDINATOR.delete_category_result(inflight_key)
        except Exception:
            # Redis cache is an optimization. The actual listing/result data has
            # already been written to PostgreSQL and must still be returned.
            log.warning("Could not publish distributed category result key=%s", inflight_key, exc_info=True)
        return result
    finally:
        publisher.cancel()
        await asyncio.gather(publisher, return_exceptions=True)
        try:
            await COORDINATOR.release_category_lock(inflight_key, lock_token)
            await COORDINATOR.clear_category_progress(inflight_key)
        except Exception:
            log.debug("Could not release distributed category lock key=%s", inflight_key, exc_info=True)


async def _wait_distributed_category_or_take_over(
    inflight_key: str, cat, user_id: int, page_limit: int, target_date: str
) -> ScanResult:
    """Wait for another replica's scan, mirror progress, and take over on owner loss."""
    redis_failures = 0
    while True:
        try:
            cached = await COORDINATOR.get_category_result(inflight_key)
            if cached is not None:
                cached_result = ScanResult(**cached)
                if _cacheable_category_result(cached_result):
                    return cached_result
                await COORDINATOR.delete_category_result(inflight_key)

            remote = await COORDINATOR.get_category_progress(inflight_key)
            if remote:
                fields = CategoryLiveProgress.__dataclass_fields__
                safe = {k: v for k, v in remote.items() if k in fields}
                try:
                    category_live_progress[inflight_key] = CategoryLiveProgress(**safe)
                except Exception:
                    pass

            if not await COORDINATOR.category_lock_exists(inflight_key):
                token = await COORDINATOR.acquire_category_lock(inflight_key)
                if token:
                    return await _run_distributed_category_owner(
                        inflight_key, token, cat, user_id, page_limit, target_date
                    )
            redis_failures = 0
        except asyncio.CancelledError:
            raise
        except Exception:
            redis_failures += 1
            log.warning(
                "Distributed shared-scan coordination unavailable key=%s failure=%s",
                inflight_key, redis_failures, exc_info=True,
            )
            if redis_failures >= 5:
                # PostgreSQL persistence + per-process traffic controls still make a
                # local scan safe. This prevents a short Redis cache outage from
                # turning an already-consumed user job into a hard failure.
                return await _scan_category_task(cat, user_id, page_limit, target_date)
        await asyncio.sleep(0.7)


async def dispatch_category(
    cat,
    user_id: int,
    page_limit: int,
    target_date: str,
    *,
    stop_event: asyncio.Event | None = None,
    force_refresh: bool = False,
) -> CategoryDispatchResult:
    """Share exact category/date/depth work locally and across Redis worker replicas."""
    inflight_key = _progress_key(cat.key, target_date, page_limit)

    if stop_event is not None and stop_event.is_set():
        raise ScanStopRequested()

    if force_refresh:
        category_result_cache.pop(inflight_key, None)
        if DISTRIBUTED_WORKERS:
            try:
                await COORDINATOR.delete_category_result(inflight_key)
            except Exception:
                log.debug("Could not clear distributed result cache for force refresh key=%s", inflight_key, exc_info=True)

    # Fast process-local cache.
    if not force_refresh and CATEGORY_CACHE_TTL_SECONDS > 0:
        cached = category_result_cache.get(inflight_key)
        if cached is not None:
            cached_at, cached_result = cached
            age = max(0, int(time.monotonic() - cached_at))
            if age <= CATEGORY_CACHE_TTL_SECONDS and _cacheable_category_result(cached_result):
                if stop_event is not None and stop_event.is_set():
                    raise ScanStopRequested()
                return CategoryDispatchResult(source="cache", result=cached_result, cache_age_seconds=age)
            category_result_cache.pop(inflight_key, None)

    # Cross-replica cache. PostgreSQL contains the listings themselves; Redis only
    # stores the small ScanResult membership/quality summary for a few minutes.
    if not force_refresh and DISTRIBUTED_WORKERS:
        try:
            remote_cached = await COORDINATOR.get_category_result(inflight_key)
            if remote_cached is not None:
                result = ScanResult(**remote_cached)
                if _cacheable_category_result(result):
                    if CATEGORY_CACHE_TTL_SECONDS > 0:
                        category_result_cache[inflight_key] = (time.monotonic(), result)
                    return CategoryDispatchResult(source="cache", result=result, cache_age_seconds=0)
                await COORDINATOR.delete_category_result(inflight_key)
        except Exception:
            log.exception("Redis category cache read failed key=%s", inflight_key)

    # Browser-isolated mode intentionally does not subscribe a second active user
    # to another user's in-flight category. Each user keeps moving in their own
    # Chromium session. Completed-result caching is still kept above.
    if not SHARE_ACTIVE_CATEGORY_SCANS:
        task = asyncio.create_task(
            _scan_category_task(cat, user_id, page_limit, target_date),
            name=f"category-scan-isolated:{user_id}:{inflight_key}",
        )
        try:
            result = await wait_for_task_or_stop(task, stop_event)
        except ScanStopRequested:
            task.cancel()
            await asyncio.gather(task, return_exceptions=True)
            raise
        if CATEGORY_CACHE_TTL_SECONDS > 0 and _cacheable_category_result(result):
            category_result_cache[inflight_key] = (time.monotonic(), result)
        if DISTRIBUTED_WORKERS:
            try:
                if _cacheable_category_result(result):
                    await COORDINATOR.set_category_result(
                        inflight_key, result, max(60, CATEGORY_CACHE_TTL_SECONDS or 60)
                    )
                else:
                    await COORDINATOR.delete_category_result(inflight_key)
            except Exception:
                log.debug("Could not publish isolated category cache key=%s", inflight_key, exc_info=True)
        return CategoryDispatchResult(source="scan", result=result)

    async with category_inflight_guard:
        task = category_inflight.get(inflight_key)
        if task is None:
            if DISTRIBUTED_WORKERS:
                redis_lock_failed = False
                try:
                    token = await COORDINATOR.acquire_category_lock(inflight_key)
                except Exception:
                    log.exception("Redis category lock failed key=%s; falling back locally", inflight_key)
                    token = None
                    redis_lock_failed = True
                if token:
                    task = asyncio.create_task(
                        _run_distributed_category_owner(
                            inflight_key, token, cat, user_id, page_limit, target_date
                        ),
                        name=f"category-scan-owner:{inflight_key}",
                    )
                    source = "scan"
                elif redis_lock_failed:
                    task = asyncio.create_task(
                        _scan_category_task(cat, user_id, page_limit, target_date),
                        name=f"category-scan-local-fallback:{inflight_key}",
                    )
                    source = "scan"
                else:
                    task = asyncio.create_task(
                        _wait_distributed_category_or_take_over(
                            inflight_key, cat, user_id, page_limit, target_date
                        ),
                        name=f"category-scan-subscriber:{inflight_key}",
                    )
                    source = "shared"
            else:
                task = asyncio.create_task(
                    _scan_category_task(cat, user_id, page_limit, target_date),
                    name=f"category-scan:{inflight_key}",
                )
                source = "scan"
            category_inflight[inflight_key] = task
        else:
            source = "shared"
        category_inflight_waiters[inflight_key] = category_inflight_waiters.get(inflight_key, 0) + 1

    stopped = False
    cancel_underlying = False
    try:
        result = await wait_for_task_or_stop(task, stop_event)
        if CATEGORY_CACHE_TTL_SECONDS > 0 and _cacheable_category_result(result):
            category_result_cache[inflight_key] = (time.monotonic(), result)
        return CategoryDispatchResult(source=source, result=result)
    except ScanStopRequested:
        stopped = True
        raise
    finally:
        async with category_inflight_guard:
            remaining = max(0, category_inflight_waiters.get(inflight_key, 1) - 1)
            if remaining:
                category_inflight_waiters[inflight_key] = remaining
            else:
                category_inflight_waiters.pop(inflight_key, None)
                # In local mode, or when this process owns/waits on the task with no
                # remaining local subscribers, cancel only our task. A remote Redis
                # owner is never cancelled merely because one subscriber stops.
                if stopped and category_inflight.get(inflight_key) is task and not task.done():
                    cancel_underlying = True
                    task.cancel()

            if task.done() and category_inflight.get(inflight_key) is task:
                category_inflight.pop(inflight_key, None)

        if cancel_underlying:
            await asyncio.gather(task, return_exceptions=True)
            async with category_inflight_guard:
                if category_inflight.get(inflight_key) is task:
                    category_inflight.pop(inflight_key, None)

        if task.done() or cancel_underlying:
            category_live_progress.pop(inflight_key, None)


async def queue_status_text(user_id: int) -> str:
    async with job_guard:
        running = [j for j in active_jobs.values() if j.state == "running"]
        queued = [j for j in active_jobs.values() if j.state == "queued" and not j.cancel_requested]
        mine = active_jobs.get(user_id)
        position = None
        if mine and mine.state == "queued" and mine.job_id in queued_job_ids:
            position = queued_job_ids.index(mine.job_id) + 1

    free_slots = max(0, MAX_CONCURRENT_JOBS - len(running))
    lines = [
        "<b>📥 Очередь парсинга</b>",
        "",
        f"⚙️ Активно: <b>{len(running)}/{MAX_CONCURRENT_JOBS}</b>",
        f"🟢 Свободно: <b>{free_slots}</b>",
        f"⏳ Ждут: <b>{len(queued)}</b>",
    ]
    if mine:
        lines += ["", "<b>Твоя задача</b>"]
        if mine.state == "queued":
            lines.append(f"⏳ В очереди" + (f" · позиция <b>{position}/{len(queued)}</b>" if position else ""))
            if position:
                lines.append(f"👥 Перед тобой: <b>{max(0, position - 1)}</b>")
            lines.append("Позиция обновляется автоматически.")
        elif mine.state == "running":
            lines.append(f"⚙️ Выполняется воркером <b>#{mine.worker_id}</b>")
            if mine.current_category:
                lines.append(f"Сейчас: <b>{html.escape(mine.current_category)}</b>")
            lines.append(f"Готово категорий: <b>{mine.completed_categories}/{len(mine.category_keys)}</b>")
        elif mine.cancel_requested:
            lines.append("❌ Ожидает отмены")
    else:
        lines += ["", "У тебя сейчас нет активного запуска."]
    return "\n".join(lines)


def _human_duration(seconds: float | int | None) -> str:
    if seconds is None:
        return "считаю…"
    seconds = max(0, int(seconds))
    if seconds < 60:
        return f"{seconds} сек"
    minutes, secs = divmod(seconds, 60)
    if minutes < 10 and secs >= 30:
        minutes += 1
    return f"{minutes} мин"


def _human_eta(seconds: float | int | None) -> str:
    """Conservative ETA: avoid misleading 10–15 second guesses."""
    if seconds is None:
        return "считаю…"
    seconds = max(0, float(seconds))
    if seconds < 45:
        return "меньше 1 мин"
    import math
    return f"≈ {max(1, math.ceil(seconds / 60))} мин"


def _base_category_eta_seconds(page_limit: int) -> int:
    limit = min(PAGE_LIMIT_CHOICES, key=lambda x: abs(x - int(page_limit)))
    return PAGE_LIMIT_BASE_ETA_SECONDS[limit]


def _progress_bar(percent: int) -> str:
    percent = max(0, min(100, percent))
    filled = min(10, percent // 10)
    return "█" * filled + "░" * (10 - filled)


def render_user_job_status(job: ScanJob) -> str:
    """Compact user-facing scan progress. Technical diagnostics stay in logs/results."""
    total = max(1, len(job.category_keys))
    depth = job.page_limit if job.page_limit in PAGE_LIMIT_CHOICES else 50

    if job.state == "queued":
        waited = max(0, int((datetime.utcnow() - job.created_at).total_seconds()))
        if STABLE_SINGLE_SERVICE_MODE and not job.recovered:
            try:
                position = queued_job_ids.index(job.job_id) + 1
            except ValueError:
                position = 1
            running_count = sum(1 for item in active_jobs.values() if item.state == "running" and not item.cancel_requested)
            waiting_count = sum(1 for item in active_jobs.values() if item.state == "queued" and not item.cancel_requested)
            ahead = max(0, position - 1)
            free_slots = max(0, MAX_CONCURRENT_JOBS - running_count)
            slot_line = (
                f"⚙️ Занято слотов: <b>{running_count}/{MAX_CONCURRENT_JOBS}</b>\n"
                if free_slots == 0
                else f"🟢 Свободных слотов: <b>{free_slots}</b> · запуск начинается\n"
            )
            return (
                "⏳ <b>Скан в очереди</b>\n\n"
                f"📍 Твоя позиция: <b>{position}/{max(position, waiting_count)}</b>\n"
                + slot_line
                + f"👥 Перед тобой: <b>{ahead}</b>\n"
                + f"🗂 Категорий: <b>{total}</b>\n"
                + f"📅 <b>{_date_label(job.target_date)}</b> · 📄 <b>{depth} стр.</b>\n"
                + f"💶 <b>{html.escape(price_filter_label(job.price_filter))}</b>\n"
                + f"⏱ Ждёшь: <b>{_human_duration(waited)}</b>\n\n"
                + "Позиция обновляется автоматически. Как только освободится слот, скан запустится сам."
            )
        headline = "♻️ <b>Восстанавливаю скан…</b>\n\n" if job.recovered else "⏳ <b>Подготавливаю скан</b>\n\n"
        return (
            headline
            + f"🗂 Категорий: <b>{total}</b>\n"
            + f"📅 <b>{_date_label(job.target_date)}</b> · 📄 <b>{depth} стр.</b>\n"
            + f"💶 <b>{html.escape(price_filter_label(job.price_filter))}</b>\n"
            + f"⏱ <b>{_human_duration(waited)}</b>"
        )

    live = category_live_progress.get(job.current_progress_key) if job.current_progress_key else None
    current_today = live.today_seen if live is not None else 0
    live_views_ready = live.views_ready if live is not None else 0

    # v3.4.3: progress must move from the first network request, not only after
    # the date locator has finished.  Date discovery is not linear, so it owns a
    # conservative first 18% of the current category.  Collection owns the next
    # 77%; the final 5% is reserved for persistence/export.  This is a UI progress
    # estimate, not a fake page count.
    current_fraction = 0.02
    if live is not None:
        if live.phase == "collecting" and depth > 0:
            collected = min(1.0, max(0.0, live.collection_index / depth))
            current_fraction = 0.18 + 0.77 * collected
        elif live.phase == "stable_scan":
            request_steps = min(50, max(int(live.page or 0), int(live.network_requests or 0)))
            current_fraction = 0.03 + 0.15 * (request_steps / 50.0)
        elif live.phase in {"jumping", "seeking"}:
            request_steps = min(12, max(0, int(live.network_requests or 0)))
            current_fraction = 0.03 + 0.15 * (request_steps / 12.0)
        elif live.phase == "regional_date" and depth > 0:
            # Do not visually reset progress when nationwide collection hands off
            # to regional depth. Keep the already confirmed page count visible.
            collected = min(1.0, max(0.0, float(live.collection_index or 0) / depth))
            current_fraction = 0.18 + 0.77 * collected
        elif live.phase == "views":
            total_views = max(1, int(live.today_seen or 0))
            view_ratio = min(1.0, max(0.0, float(live.views_ready) / total_views))
            current_fraction = 0.95 + 0.04 * view_ratio
    percent = int(max(0.0, min(0.99, (job.completed_categories + current_fraction) / total)) * 100)
    if job.completed_categories >= total:
        percent = 100

    elapsed = 0
    if job.started_running_monotonic:
        elapsed = max(0, int(time.monotonic() - job.started_running_monotonic))

    category_line = html.escape(job.current_category) if job.current_category else "Подготовка…"
    category_index = max(1, job.current_category_index)

    # v4.5.1: everyday users see one simple scan state. Date search, regional
    # fallback, HTTP/Chromium transport, retries and worker routing stay in logs/admin.
    if live is None or live.phase not in {"collecting", "views"}:
        pages_done = max(0, min(depth, int(live.collection_index or 0))) if live is not None else 0
        return (
            f"🔎 <b>Сканирование · {percent}%</b>\n"
            f"{_progress_bar(percent)}\n\n"
            f"🗂 <b>{category_line}</b> · {category_index}/{total}\n"
            f"📄 <b>{pages_done}/{depth}</b> страниц\n"
            f"📦 <b>{current_today}</b> объявлений\n"
            f"⏱ <b>{_human_duration(elapsed)}</b>"
        )

    if live.phase == "views":
        total_views = max(1, int(live.today_seen or 0))
        ready = min(total_views, int(live.views_ready or 0))
        return (
            f"👁 <b>Собираю просмотры · {percent}%</b>\n"
            f"{_progress_bar(percent)}\n\n"
            f"🗂 <b>{category_line}</b> · {category_index}/{total}\n"
            f"📦 Объявлений: <b>{live.today_seen}</b>\n"
            f"👁 Проверено: <b>{ready}/{total_views}</b>\n"
            f"⏱ <b>{_human_duration(elapsed)}</b>"
        )

    pages_done = max(0, min(depth, int(live.collection_index or 0)))
    views_text = ""
    if current_today:
        views_text = f" · 👁 <b>{min(live_views_ready, current_today)}</b>"

    return (
        f"🔎 <b>Сканирование · {percent}%</b>\n"
        f"{_progress_bar(percent)}\n\n"
        f"🗂 <b>{category_line}</b> · {category_index}/{total}\n"
        f"📄 <b>{pages_done}/{depth}</b> страниц\n"
        f"📦 <b>{current_today}</b> объявлений{views_text}\n"
        f"⏱ <b>{_human_duration(elapsed)}</b>"
    )


async def progress_ticker(bot: Bot) -> None:
    """Continuously refresh user-facing progress without exposing internal scheduling."""
    while True:
        await asyncio.sleep(max(2.0, STATUS_UPDATE_INTERVAL_SECONDS))
        async with job_guard:
            jobs = list(active_jobs.values())
        for job in jobs:
            if job.state not in {"queued", "running"} or job.cancel_requested:
                continue
            try:
                await edit_job_status(bot, job, render_user_job_status(job))
            except Exception:
                log.debug("Could not refresh live progress for job=%s", job.job_id, exc_info=True)


async def edit_job_status(bot: Bot, job: ScanJob, text: str, *, force: bool = False) -> None:
    now = time.monotonic()
    if not force and now - job.last_status_update < STATUS_UPDATE_INTERVAL_SECONDS:
        return
    job.last_status_update = now
    try:
        await bot.edit_message_text(
            chat_id=job.chat_id,
            message_id=job.status_message_id,
            text=text,
            parse_mode=ParseMode.HTML,
            reply_markup=job_keyboard(job.job_id, queued=(job.state == "queued")) if job.state in {"queued", "running"} else None,
        )
    except Exception as exc:
        # Telegram may reject an identical edit; this must never stop parsing.
        log.debug("Could not edit job status %s: %s", job.job_id, exc)


async def finish_job(bot: Bot, job: ScanJob, *, cancelled: bool = False) -> None:
    await finalize_user_scan(job, cancelled=cancelled)
    elapsed_seconds = max(0, int((datetime.utcnow() - job.created_at).total_seconds()))
    mins, secs = divmod(elapsed_seconds, 60)
    elapsed_text = f"{mins} мин {secs} сек" if mins else f"{secs} сек"

    job.state = "cancelled" if cancelled else ("partial" if job.incomplete_categories else "done")
    if cancelled:
        text = (
            "⏹ <b>Скан остановлен</b>\n\n"
            f"🗂 Обработано: <b>{job.completed_categories}/{len(job.category_keys)}</b> категорий\n"
            f"📦 Найдено до остановки: <b>{job.total_new}</b>\n"
            f"⏱ Время: <b>{elapsed_text}</b>"
        )
        await edit_job_status(bot, job, text, force=True)
        try:
            await bot.edit_message_reply_markup(
                chat_id=job.chat_id,
                message_id=job.status_message_id,
                reply_markup=stopped_job_keyboard(),
            )
        except Exception:
            log.debug("Could not attach stopped-job actions job=%s", job.job_id, exc_info=True)
        return

    # Keep the live card clean while the CSV is being built, then replace it with
    # the final product-style summary after export returns the actual result count.
    await edit_job_status(
        bot,
        job,
        "✅ <b>Сканирование завершено</b>\n\n📄 Готовлю результат…",
        force=True,
    )

    result_count: int | None = None
    export_ok = False
    snapshot_rows: list[Listing] = []
    if job.scan_id is not None:
        try:
            snapshot_rows = [row for row, _ in await get_scan_rows(job.scan_id)]
        except Exception:
            log.exception("Could not load snapshot rows for job=%s", job.job_id)

    if job.incomplete_categories:
        # v4.8.6: partial data is provisional. Do not run the final smart export
        # (especially min-views filtering) before the views phase has legally run.
        # The confirmed rows are saved and remain available from the scan card.
        result_count = len(job.matched_ids or set())
    else:
        try:
            result_prefix = (
                "📄 <b>Результат скана</b>\n"
                f"📅 {_date_label(job.target_date)} · 🗂 {job.completed_categories}/{len(job.category_keys)} категорий"
            )
            trial_status = await get_trial_status(job.user_id) if job.is_trial else None
            trial_remaining = (trial_status.remaining if trial_status is not None and trial_status.eligible else 0)
            adapter = BotChatAdapter(
                bot,
                job.chat_id,
                prefix=result_prefix,
                reply_markup=(
                    trial_result_keyboard(job.scan_id, trial_remaining)
                    if job.is_trial and trial_status is not None
                    else post_scan_keyboard(job.scan_id, recheck=False)
                ),
            )
            result_count = await send_smart_export(
                adapter,
                job.user_id,
                len(job.category_keys),
                category_keys_override=set(job.category_keys),
                rows_override=snapshot_rows,
                price_filter_override=job.price_filter,
            )
            export_ok = True
        except Exception:
            log.exception("Could not auto-export result for job=%s", job.job_id)
            try:
                await bot.send_message(
                    job.chat_id,
                    "⚠️ <b>Результат сохранён, но XLSX не отправился.</b>\n\n"
                    "Открой скан — файл можно скачать повторно.",
                    parse_mode=ParseMode.HTML,
                    reply_markup=post_scan_keyboard(job.scan_id, recheck=False),
                )
            except Exception:
                pass

    quality_values = [int(x) for x in (job.quality_scores or []) if x is not None]
    quality_avg = round(sum(quality_values) / len(quality_values)) if quality_values else 0
    headline = "⚠️ <b>Скан завершён частично</b>" if job.incomplete_categories else "✅ <b>Скан завершён</b>"
    lines = [
        headline,
        "",
        f"📅 Дата: <b>{_date_label(job.target_date)}</b>",
        f"🗂 Категории: <b>{job.completed_categories}/{len(job.category_keys)}</b>",
    ]
    if result_count is not None:
        if job.incomplete_categories:
            lines.append(f"📦 Подтверждено объявлений: <b>{result_count}</b>")
        else:
            lines.append(f"📦 В результате: <b>{result_count}</b>")
    if job.auto_recovered_categories:
        lines.append(f"🔧 Автовосстановлено: <b>{job.auto_recovered_categories}</b> категорий")
    if job.incomplete_categories:
        label = "Непроверенные участки после автоповторов" if STABLE_SCAN_ENGINE else "Требуют ручной проверки"
        lines.append(f"⚠️ {label}: <b>{job.incomplete_categories}</b> категорий")
    elif quality_avg and quality_avg < 90:
        lines.append(f"🛡 Качество: <b>{quality_avg}/100</b>")
    lines.append(f"⏱ Время: <b>{elapsed_text}</b>")
    if not job.incomplete_categories:
        if job.is_trial:
            lines.append("🎁 Пробный скан · <b>автозамеры доступны по подписке</b>")
        else:
            auto_enabled = await auto_observations_enabled(job.user_id)
            if auto_enabled:
                lines.append("🔔 Автозамеры: <b>✅ ВКЛ · 3 · 6 · 12 ч</b>")
            else:
                lines.append("🔔 Автозамеры: <b>⛔ ВЫКЛ</b> · ручное обновление доступно всегда")
    if job.incomplete_categories:
        result_tail = (
            "Подтверждённые объявления сохранены. Финальные фильтры по просмотрам и итоговый XLSX "
            "применяются только после полного подтверждения глубины."
        )
    elif export_ok and result_count:
        result_tail = "📊 XLSX отправлен ниже."
    elif export_ok:
        result_tail = "По текущим фильтрам подходящих объявлений нет."
    else:
        result_tail = "Данные сохранены в скане."
    lines += ["", result_tail]
    await edit_job_status(bot, job, "\n".join(lines), force=True)

    if job.is_trial and not cancelled and not allowed(job.user_id):
        try:
            trial = await get_trial_status(job.user_id)
            usable_remaining = trial.remaining if trial.eligible else 0
            if usable_remaining > 0:
                trial_text = (
                    "🎁 <b>Бесплатный скан готов</b>\n\n"
                    f"У тебя остался ещё <b>{usable_remaining}</b> бесплатный скан. "
                    "Попробуй другую категорию или открой полный доступ без ограничений пробного режима."
                )
            else:
                trial_text = (
                    f"🔥 <b>{FREE_TRIAL_SCAN_LIMIT} бесплатных скана использованы</b>\n\n"
                    "Твои результаты сохраняются в «Мои сканы». "
                    "Чтобы продолжить сканирование, открой полный доступ."
                )
            await bot.send_message(
                job.chat_id, trial_text, parse_mode=ParseMode.HTML,
                reply_markup=trial_conversion_keyboard(usable_remaining),
            )
        except Exception:
            log.exception("Could not send free-trial conversion card job=%s", job.job_id)

    if job.incomplete_categories:
        try:
            if STABLE_SCAN_ENGINE:
                if STABLE_SINGLE_SERVICE_MODE:
                    text = (
                        "<b>⚠️ Не все страницы удалось подтвердить</b>\n\n"
                        f"{job.incomplete_categories} из {len(job.category_keys)} категорий завершены не полностью. "
                        "Парсер уже повторил проблемную страницу и один раз обновил браузерный контекст. "
                        "Успешные страницы сохранены в PostgreSQL, поэтому следующий запуск продолжит с проверенных данных."
                    )
                else:
                    text = (
                        "<b>⚠️ Не все участки удалось подтвердить</b>\n\n"
                        f"Сервер автоматически выполнил до {SCAN_AUTO_RECOVERY_PASSES} повторных проходов. "
                        f"{job.incomplete_categories} из {len(job.category_keys)} категорий всё ещё имеют непроверенные участки. "
                        "Все подтверждённые страницы сохранены в PostgreSQL и будут переиспользованы при следующем запуске — "
                        "повторный скан не начнётся с нуля."
                    )
                markup = InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔄 Повторить скан", callback_data=f"scanrepeat:{job.scan_id}")],
                    [InlineKeyboardButton(text="📊 Открыть сохранённый скан", callback_data=f"scan:{job.scan_id}")],
                ]) if job.scan_id is not None else None
            else:
                text = (
                    "<b>⚠️ Автовосстановление не завершило все участки</b>\n\n"
                    f"{job.incomplete_categories} из {len(job.category_keys)} категорий всё ещё проверены не полностью. "
                    "Найденные данные сохранены. Ручная допроверка доступна как резервный вариант."
                )
                markup = partial_recheck_keyboard(job.scan_id) if job.scan_id is not None else None
            await bot.send_message(job.chat_id, text, parse_mode=ParseMode.HTML, reply_markup=markup)
        except Exception:
            log.exception("Could not send partial-scan notice for job=%s", job.job_id)
    elif job.warnings:
        log.info("scan job=%s warnings=%s", job.job_id, job.warnings[:20])


async def dispatch_category_with_retry(
    bot: Bot, job: ScanJob, cat, *, force_refresh: bool = False
) -> CategoryDispatchResult:
    """Run one category with a small job-level retry safety net.

    The parser itself already retries HTTP/403/429/5xx responses. This outer layer
    protects the user from a one-off transport/runtime failure and keeps a partial
    scan recoverable instead of immediately turning the category into an error.
    """
    last_exc: Exception | None = None
    for attempt in range(1, SCAN_CATEGORY_ATTEMPTS + 1):
        try:
            job.retry_note = ""
            return await dispatch_category(
                cat, job.user_id, job.page_limit, job.target_date,
                stop_event=job.stop_event, force_refresh=force_refresh,
            )
        except ScanStopRequested:
            raise
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            last_exc = exc
            await record_user_scan_retry(job.scan_id, f"{cat.name}: {type(exc).__name__}: {exc}")
            if attempt >= SCAN_CATEGORY_ATTEMPTS:
                raise
            job.retry_note = (
                f"Попытка {attempt + 1}/{SCAN_CATEGORY_ATTEMPTS} через "
                f"{int(SCAN_CATEGORY_RETRY_SECONDS * attempt)} сек."
            )
            await edit_job_status(bot, job, render_user_job_status(job), force=True)
            await asyncio.sleep(SCAN_CATEGORY_RETRY_SECONDS * attempt)
            if job.stop_event.is_set() or job.cancel_requested:
                raise ScanStopRequested()
    assert last_exc is not None
    raise last_exc


async def auto_recover_partial_category(
    bot: Bot, job: ScanJob, cat, dispatched: CategoryDispatchResult
) -> CategoryDispatchResult:
    """Automatically recheck only weak/missing page areas before exposing partial UI.

    The per-job KleinanzeigenParser keeps verified CategoryPageInfo checkpoints. A
    forced category pass therefore reuses strong pages and performs network work for
    pages that were weak, missing or never reached. This is intentionally bounded.
    """
    result = dispatched.result
    if result is None or result.date_complete or SCAN_AUTO_RECOVERY_PASSES <= 0:
        return dispatched

    merged_ids = set(result.matched_ids or [])
    best = result
    best_dispatch = dispatched
    job.recovery_total = SCAN_AUTO_RECOVERY_PASSES

    for attempt in range(1, SCAN_AUTO_RECOVERY_PASSES + 1):
        if job.stop_event.is_set() or job.cancel_requested:
            raise ScanStopRequested()
        job.recovery_attempt = attempt
        job.recovery_note = (
            "Основной проход получился неполным. Уже подтверждённые страницы сохранены; "
            "перепроверяю только слабые или недостающие участки."
        )
        await edit_job_status(bot, job, render_user_job_status(job), force=True)
        if attempt > 1 or SCAN_AUTO_RECOVERY_DELAY_SECONDS:
            await asyncio.sleep(SCAN_AUTO_RECOVERY_DELAY_SECONDS * attempt)
        try:
            candidate_dispatch = await dispatch_category_with_retry(
                bot, job, cat, force_refresh=True
            )
            candidate = candidate_dispatch.result
        except ScanStopRequested:
            raise
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            await record_user_scan_retry(
                job.scan_id, f"auto-recovery {cat.name} {attempt}: {type(exc).__name__}: {exc}"
            )
            log.warning(
                "Automatic partial recovery failed job=%s category=%s attempt=%s/%s: %s",
                job.job_id, cat.name, attempt, SCAN_AUTO_RECOVERY_PASSES, exc,
            )
            continue

        if candidate is None:
            continue
        merged_ids.update(candidate.matched_ids or [])
        candidate.matched_ids = sorted(merged_ids)
        # The recovery pass revisits rows already inserted by the first pass, so
        # its DB-level `new_count` can legitimately be zero. Preserve the strongest
        # user-facing counters from either pass while using the recovery verdict.
        candidate.new_count = max(int(result.new_count or 0), int(candidate.new_count or 0))
        candidate.today_seen = max(int(result.today_seen or 0), int(candidate.today_seen or 0))
        candidate.enriched_count = max(int(result.enriched_count or 0), int(candidate.enriched_count or 0))
        # Prefer a complete pass; otherwise keep the strongest partial verdict.
        if candidate.date_complete:
            candidate.reason = f"{candidate.reason}; автодопроверка успешна"
            job.auto_recovered_categories += 1
            job.recovery_note = ""
            job.retry_note = ""
            job.recovery_attempt = 0
            job.recovery_total = 0
            return candidate_dispatch
        if (candidate.quality_score, candidate.verified_pages, candidate.today_seen) > (
            best.quality_score, best.verified_pages, best.today_seen
        ):
            best = candidate
            best_dispatch = candidate_dispatch

    best.matched_ids = sorted(merged_ids)
    job.recovery_note = ""
    job.retry_note = ""
    job.recovery_attempt = 0
    job.recovery_total = 0
    return CategoryDispatchResult(
        source=best_dispatch.source, result=best, cache_age_seconds=best_dispatch.cache_age_seconds
    )


async def process_scan_job(bot: Bot, job: ScanJob, worker_id: int) -> None:
    job.state = "running"
    if job.scan_id is not None:
        async with SessionLocal() as session:
            scan = await session.get(UserScan, job.scan_id)
            if scan is not None:
                scan.status = "running"
                await session.commit()
    job.worker_id = worker_id
    job.started_running_monotonic = time.monotonic()
    job.warnings = job.warnings or []
    job.scan_notes = job.scan_notes or []
    job.matched_ids = job.matched_ids or set()
    job.quality_scores = job.quality_scores or []
    job.quality_notes = job.quality_notes or []
    job.incomplete_category_keys = job.incomplete_category_keys or set()
    waited_before_start = max(0, int((datetime.utcnow() - job.created_at).total_seconds()))
    await edit_job_status(bot, job, render_user_job_status(job), force=True)
    if waited_before_start >= QUEUE_START_NOTIFY_AFTER_SECONDS:
        try:
            await bot.send_message(
                job.chat_id,
                "🚀 <b>Твой скан вышел из очереди и начался.</b>\n\n"
                f"⏱ Ожидание: <b>{_human_duration(waited_before_start)}</b>\n"
                "Прогресс дальше обновляется в основной карточке.",
                parse_mode=ParseMode.HTML,
            )
        except Exception:
            log.debug("Could not send queue-start notification job=%s", job.job_id, exc_info=True)

    for idx, key in enumerate(job.category_keys, start=1):
        if job.cancel_requested:
            break
        cat = CATEGORIES.get(key)
        if cat is None:
            job.warnings.append(f"Неизвестная категория: {key}")
            job.incomplete_categories += 1
            job.incomplete_category_keys = job.incomplete_category_keys or set()
            job.incomplete_category_keys.add(key)
            job.completed_categories += 1
            job.quality_scores = job.quality_scores or []
            job.quality_scores.append(0)
            continue
        job.current_category = cat.name
        job.current_category_key = cat.key
        job.current_progress_key = _progress_key(cat.key, job.target_date, job.page_limit)
        job.current_category_index = idx
        # Multi-category isolation: every selected category starts with a clean
        # live-progress slot and performs its own date location cycle. No boundary
        # or progress state from the previous category may leak into this one.
        category_live_progress.pop(job.current_progress_key, None)
        log.info(
            "multi-category start job=%s category=%s index=%s/%s target=%s depth=%s",
            job.job_id, cat.name, idx, len(job.category_keys), job.target_date, job.page_limit,
        )
        try:
            await edit_job_status(bot, job, render_user_job_status(job), force=True)

            async def _run_category_pipeline() -> CategoryDispatchResult:
                dispatched_local = await dispatch_category_with_retry(bot, job, cat)
                if job.cancel_requested or job.stop_event.is_set():
                    raise ScanStopRequested()
                dispatched_local = await auto_recover_partial_category(bot, job, cat, dispatched_local)
                if job.cancel_requested or job.stop_event.is_set():
                    raise ScanStopRequested()
                return dispatched_local

            # Hard safety net only. Normal browser/page requests have their own much
            # shorter timeouts. This prevents a genuinely wedged category from
            # occupying one of the three user lanes forever.
            dispatched = await asyncio.wait_for(
                _run_category_pipeline(), timeout=SCAN_CATEGORY_HARD_TIMEOUT_SECONDS
            )
            result = dispatched.result
            source_label = "♻️ кэш"
            if dispatched.source == "cache":
                job.cache_hits += 1
                source_label = f"♻️ кэш ({dispatched.cache_age_seconds} сек.)"
            elif dispatched.source == "shared":
                job.shared_hits += 1
                source_label = "🤝 общий скан"
            else:
                job.scanned_categories += 1
                source_label = "🌐 новый скан"

            if result is not None:
                job.matched_ids = job.matched_ids or set()
                job.matched_ids.update(result.matched_ids or [])
                job.quality_scores = job.quality_scores or []
                job.quality_notes = job.quality_notes or []
                job.quality_scores.append(int(result.quality_score or 0))
                job.quality_notes.append(f"{cat.name}: {result.quality_score}/100 — {result.quality_note}")
                # New_count is a global DB fact from this shared scan. Pages are counted
                # only for the job that actually started the network scan.
                job.total_new += result.new_count
                if dispatched.source == "scan":
                    job.total_pages += result.pages_scanned
                    job.total_avoided += result.avoided_pages
                    if result.mode == "fast":
                        job.fast_categories += 1
                    else:
                        job.full_categories += 1
                if not result.date_complete:
                    job.incomplete_categories += 1
                    job.incomplete_category_keys = job.incomplete_category_keys or set()
                    job.incomplete_category_keys.add(cat.key)
                    reached = _date_label(result.oldest_date_seen) if result.oldest_date_seen else "не определена"
                    note = (
                        f"{cat.name}: охват {_date_label(job.target_date)} не подтверждён полностью; "
                        f"самая старая распознанная дата — {reached}; "
                        f"сетевых запросов {result.pages_scanned}; качество {result.quality_score}/100; "
                        f"причина: {result.reason}"
                    )
                    job.warnings.append(note)
                    job.scan_notes = job.scan_notes or []
                    job.scan_notes.append(note)
                elif result.reason.startswith("временный лимит Kleinanzeigen"):
                    job.warnings.append(
                        f"{cat.name}: Kleinanzeigen временно ограничил запросы; "
                        f"успели сделать {result.pages_scanned} запросов (до стр. {result.max_page_reached or '?'}), можно повторить позже"
                    )

            log.info(
                "multi-category finish job=%s category=%s matched=%s complete=%s reason=%s",
                job.job_id, cat.name, (result.today_seen if result is not None else 0),
                (result.date_complete if result is not None else False),
                (result.reason if result is not None else "no result"),
            )
            job.completed_categories += 1

            # If an interactive category already spent the full recovery window and
            # Kleinanzeigen still refuses the process, immediately trying the next
            # selected category only extends the block. Stop this job gracefully;
            # completed categories remain saved and no false zeros are produced.
            if result is not None and "временный лимит Kleinanzeigen" in (result.reason or ""):
                remaining_categories = max(0, len(job.category_keys) - idx)
                if remaining_categories:
                    note = (
                        f"Kleinanzeigen всё ещё ограничивает доступ после автоматического ожидания. "
                        f"Оставшиеся категории ({remaining_categories}) не запускались, чтобы не усиливать лимит."
                    )
                    job.warnings.append(note)
                    job.scan_notes = job.scan_notes or []
                    job.scan_notes.append(note)
                    job.incomplete_categories += remaining_categories
                    job.incomplete_category_keys = job.incomplete_category_keys or set()
                    job.incomplete_category_keys.update(job.category_keys[idx:])
                    job.completed_categories += remaining_categories
                break
            # User sees only useful progress; cache/shared/worker details stay internal.
            await edit_job_status(bot, job, render_user_job_status(job), force=True)
        except asyncio.TimeoutError:
            # Cancelled wait_for() work can leave a Playwright page mid-navigation.
            # Recycle ONLY this job's BrowserContext; the two other user contexts
            # inside the shared Chromium process remain untouched.
            parser = JOB_PARSER.get()
            if parser is not None:
                try:
                    await parser.reset_scan_browser_context()
                except Exception:
                    log.debug("Could not recycle timed-out browser context job=%s", job.job_id, exc_info=True)
            minutes = max(1, int(round(SCAN_CATEGORY_HARD_TIMEOUT_SECONDS / 60.0)))
            note = f"{cat.name}: превышен защитный лимит {minutes} мин.; категория остановлена, остальные продолжаются"
            log.error("Category watchdog timeout job=%s category=%s seconds=%s", job.job_id, cat.name, SCAN_CATEGORY_HARD_TIMEOUT_SECONDS)
            job.warnings.append(note)
            job.scan_notes = job.scan_notes or []
            job.scan_notes.append(note)
            job.quality_scores = job.quality_scores or []
            job.quality_notes = job.quality_notes or []
            job.quality_scores.append(0)
            job.quality_notes.append(f"{cat.name}: 0/100 — защитный таймаут")
            job.incomplete_categories += 1
            job.incomplete_category_keys = job.incomplete_category_keys or set()
            job.incomplete_category_keys.add(cat.key)
            job.completed_categories += 1
        except ScanStopRequested:
            job.cancel_requested = True
            log.info("User stopped scan job=%s category=%s", job.job_id, cat.name)
            break
        except Exception as exc:
            log.exception("Queue scan error job=%s category=%s", job.job_id, cat.name)
            note = f"{cat.name}: ошибка скана — {str(exc)[:160]}"
            job.warnings.append(note)
            job.scan_notes = job.scan_notes or []
            job.scan_notes.append(note)
            job.quality_scores = job.quality_scores or []
            job.quality_notes = job.quality_notes or []
            job.quality_scores.append(0)
            job.quality_notes.append(f"{cat.name}: 0/100 — ошибка скана")
            job.incomplete_categories += 1
            job.incomplete_category_keys = job.incomplete_category_keys or set()
            job.incomplete_category_keys.add(cat.key)
            job.completed_categories += 1

    await finish_job(bot, job, cancelled=job.cancel_requested)


async def _mark_local_job_running_on_claim(job: ScanJob, worker_id: int) -> None:
    """Persist the local worker claim before parser/browser setup begins.

    v4.9.0 only changed PostgreSQL to ``running`` inside process_scan_job(). If
    browser/parser setup ever took noticeable time, admin diagnostics could show a
    worker-owned job as still queued. v4.9.1 makes the claim atomic from the UI's
    point of view: once one of the four consumers takes a job, PostgreSQL immediately
    says ``running``.
    """
    job.state = "running"
    job.worker_id = worker_id
    if job.scan_id is None:
        return
    async with SessionLocal() as session:
        scan = await session.get(UserScan, int(job.scan_id))
        if scan is not None and scan.status == "queued":
            scan.status = "running"
            await session.commit()


async def scan_worker(bot: Bot, worker_id: int) -> None:
    log.info("Scan worker #%s started", worker_id)
    while True:
        job = await scan_queue.get()
        try:
            async with job_guard:
                if job.job_id in queued_job_ids:
                    queued_job_ids.remove(job.job_id)
                if job.cancel_requested:
                    job.state = "cancelled"
                else:
                    job.state = "running"
                    job.worker_id = worker_id

            if job.cancel_requested:
                await finish_job(bot, job, cancelled=True)
            else:
                # Persist the slot claim before any parser/browser setup. This keeps
                # admin status truthful and makes all four local lanes visible at once.
                await _mark_local_job_running_on_claim(job, worker_id)
                await TRAFFIC.scan_job_started()
                parser = KleinanzeigenParser()
                parser_token = JOB_PARSER.set(parser)
                try:
                    await process_scan_job(bot, job, worker_id)
                finally:
                    JOB_PARSER.reset(parser_token)
                    await parser.close()
                    await TRAFFIC.scan_job_finished()
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Unhandled scan worker error worker=%s job=%s", worker_id, job.job_id)
            job.state = "failed"
            if job.scan_id is not None:
                try:
                    async with SessionLocal() as session:
                        scan = await session.get(UserScan, job.scan_id)
                        if scan is not None:
                            scan.status = "failed"
                            scan.finished_at = datetime.utcnow()
                            scan.last_error = "Необработанная ошибка воркера; см. Railway logs"
                            await session.commit()
                except Exception:
                    log.exception("Could not mark user scan failed scan_id=%s", job.scan_id)
            try:
                await bot.edit_message_text(
                    chat_id=job.chat_id,
                    message_id=job.status_message_id,
                    text=(
                        "❌ <b>Парсер не смог завершить этот запуск</b>\n\n"
                        "Уже собранные данные в PostgreSQL не потеряны. Можно повторить этот же скан одной кнопкой."
                    ),
                    parse_mode=ParseMode.HTML,
                    reply_markup=failed_job_keyboard(job.scan_id),
                )
            except Exception:
                pass
        finally:
            async with job_guard:
                if active_jobs.get(job.user_id) is job:
                    active_jobs.pop(job.user_id, None)
                if job.job_id in queued_job_ids:
                    queued_job_ids.remove(job.job_id)
            scan_queue.task_done()



async def cleanup_stale_distributed_queue_rows() -> int:
    """Retire abandoned queued DB cards from pre-fleet test deployments.

    Redis messages for these rows are harmless: when a worker encounters them it
    sees the finished DB row and ACKs them immediately.  This prevents a newly
    bootstrapped single worker from spending minutes on ancient user jobs before
    it reaches the scan the user just launched.
    """
    if not DISTRIBUTED_WORKERS:
        return 0
    now = datetime.utcnow()
    cutoff = now - timedelta(seconds=DISTRIBUTED_STALE_QUEUE_SECONDS)
    async with SessionLocal() as session:
        rows = list((await session.execute(
            select(UserScan).where(
                UserScan.status == "queued",
                UserScan.finished_at.is_(None),
                UserScan.created_at < cutoff,
            )
        )).scalars().all())
        retired_trial_ids = [int(scan.id) for scan in rows if bool(getattr(scan, "is_trial", False))]
        for scan in rows:
            scan.status = "failed"
            scan.finished_at = now
            scan.last_error = "Старая очередь очищена после запуска Browser Fleet; повтори скан"
        if rows:
            await session.commit()
    for scan_id in retired_trial_ids:
        try:
            await _refund_trial_credit_for_scan(scan_id)
        except Exception:
            log.exception("Could not refund stale queued trial scan=%s", scan_id)
    if rows:
        log.warning("Retired stale distributed queued scans: %s", len(rows))
    return len(rows)


async def distributed_queue_ui_ticker(bot: Bot) -> None:
    """Keep queued distributed cards alive while all worker lanes are busy.

    Previously a job waiting behind another Redis job stayed forever at
    `Подготавливаю скан · 0 сек`, making a healthy queue look broken.
    """
    while True:
        await asyncio.sleep(DISTRIBUTED_QUEUE_UI_SECONDS)
        try:
            async with SessionLocal() as session:
                rows = list((await session.execute(
                    select(UserScan)
                    .where(UserScan.status == "queued", UserScan.finished_at.is_(None))
                    .order_by(UserScan.created_at.asc(), UserScan.id.asc())
                )).scalars().all())
            if not rows:
                continue
            try:
                workers = await COORDINATOR.worker_count(prefix="parser")
            except Exception:
                workers = 0
            total_waiting = len(rows)
            for position, scan in enumerate(rows, start=1):
                if not scan.chat_id or not scan.status_message_id:
                    continue
                waited = max(0, int((datetime.utcnow() - scan.created_at).total_seconds()))
                text = (
                    "⏳ <b>Скан в очереди</b>\n\n"
                    f"📍 Твоя позиция: <b>{position}/{total_waiting}</b>\n"
                    f"⚙️ Активных parser-worker: <b>{workers}</b>\n"
                    f"👥 Перед тобой: <b>{max(0, position - 1)}</b>\n"
                    f"📅 <b>{_date_label(scan.target_date)}</b> · 📄 <b>{scan.page_limit} стр.</b>\n"
                    f"⏱ Ждёшь: <b>{_human_duration(waited)}</b>\n\n"
                    "Позиция обновляется автоматически. Как только освободится worker, скан начнётся сам."
                )
                try:
                    await bot.edit_message_text(
                        chat_id=int(scan.chat_id),
                        message_id=int(scan.status_message_id),
                        text=text,
                        parse_mode=ParseMode.HTML,
                        reply_markup=job_keyboard(str(scan.job_uid), queued=True),
                    )
                except Exception:
                    # Identical edits / deleted messages are non-fatal.
                    pass
        except asyncio.CancelledError:
            raise
        except Exception:
            log.debug("Distributed queue UI ticker failed", exc_info=True)


async def recover_distributed_unfinished_scans() -> int:
    """Ensure PostgreSQL unfinished scans have a Redis Stream item.

    enqueue_scan() is idempotent through a Redis marker, so normal restarts do not
    duplicate work. If Redis itself was recreated, the marker is gone and the DB
    record is safely re-enqueued.
    """
    if not DISTRIBUTED_WORKERS:
        return 0
    async with SessionLocal() as session:
        rows = list((await session.execute(
            select(UserScan)
            .where(UserScan.status.in_(ACTIVE_SCAN_STATUSES), UserScan.finished_at.is_(None))
            .order_by(UserScan.created_at.asc())
        )).scalars().all())
    added = 0
    for scan in rows:
        try:
            if await COORDINATOR.enqueue_scan(str(scan.job_uid)):
                added += 1
        except Exception:
            log.exception("Could not recover distributed scan job=%s", scan.job_uid)
    return added


async def distributed_worker_heartbeat(worker_id: str, kind: str = "parser") -> None:
    while True:
        try:
            await COORDINATOR.heartbeat(worker_id, kind)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.debug("Worker heartbeat failed worker=%s", worker_id, exc_info=True)
        await asyncio.sleep(6.0)


async def _distributed_job_lock_refresher(job_uid: str, token: str) -> None:
    while True:
        await asyncio.sleep(12.0)
        try:
            if not await COORDINATOR.refresh_job_lock(job_uid, token):
                return
        except asyncio.CancelledError:
            raise
        except Exception:
            log.debug("Could not refresh distributed job lock job=%s", job_uid, exc_info=True)


async def distributed_scan_worker(bot: Bot, worker_id: str) -> None:
    """Consume persistent Redis Stream jobs; safe to run in many Railway replicas."""
    consumer = worker_id.replace(":", "-")
    log.info("Distributed scan worker started consumer=%s", consumer)
    while True:
        message_id: str | None = None
        job_uid: str | None = None
        job_lock_token: str | None = None
        lock_refresher: asyncio.Task | None = None
        execution_started = False
        try:
            item = await COORDINATOR.consume_scan(consumer)
            if item is None:
                continue
            message_id, job_uid = item

            # Prevent accidental duplicate Stream entries/recovery races from ever
            # running the same UserScan on two worker replicas simultaneously.
            try:
                job_lock_token = await COORDINATOR.acquire_job_lock(job_uid)
            except Exception:
                log.warning("Redis job lock unavailable job=%s; leaving delivery pending", job_uid, exc_info=True)
                await asyncio.sleep(1.0)
                continue
            if not job_lock_token:
                # Keep this delivery pending. If the real owner crashes, its short
                # lock expires and XAUTOCLAIM makes the job recoverable automatically.
                await asyncio.sleep(0.5)
                continue
            lock_refresher = asyncio.create_task(
                _distributed_job_lock_refresher(job_uid, job_lock_token),
                name=f"job-lock-refresh:{job_uid}",
            )

            job = await load_scan_job_by_uid(job_uid)
            if job is not None:
                log.info(
                    "Distributed job claimed job=%s consumer=%s scan_id=%s user=%s chat=%s message=%s",
                    job_uid, consumer, job.scan_id, job.user_id, job.chat_id, job.status_message_id,
                )
            if job is None:
                await COORDINATOR.ack_scan(message_id)
                await COORDINATOR.mark_job_complete(job_uid)
                continue

            # Another delivery of an already-finished job is harmless.
            async with SessionLocal() as session:
                db_scan = await session.get(UserScan, int(job.scan_id or 0))
                if db_scan is None or db_scan.finished_at is not None or db_scan.status in ARCHIVABLE_SCAN_STATUSES:
                    await COORDINATOR.ack_scan(message_id)
                    await COORDINATOR.mark_job_complete(job_uid)
                    continue
                current_status = db_scan.status

            redis_cancelled = False
            try:
                redis_cancelled = await COORDINATOR.is_cancel_requested(job_uid)
            except Exception:
                # The PostgreSQL status is also checked by the worker-side watcher;
                # a short Redis hiccup must not fail the user scan.
                log.debug("Could not read Redis cancel flag job=%s", job_uid, exc_info=True)
            if current_status == "cancelling" or redis_cancelled:
                execution_started = True
                job.cancel_requested = True
                job.stop_event.set()
                await finish_job(bot, job, cancelled=True)
                await COORDINATOR.ack_scan(message_id)
                await COORDINATOR.mark_job_complete(job_uid)
                continue

            # From this point a caught execution error is terminal for this delivery.
            # Errors before this line leave the Stream message pending for recovery.
            execution_started = True
            # Local active_jobs exists only so this worker's progress ticker can
            # render the current CategoryLiveProgress. PostgreSQL is authoritative.
            async with job_guard:
                active_jobs[job.user_id] = job
                if job.job_id not in queued_job_ids:
                    queued_job_ids.append(job.job_id)

            cancel_watcher = asyncio.create_task(
                _distributed_cancel_watcher(job), name=f"cancel-watch:{job.job_id}"
            )
            await TRAFFIC.scan_job_started()
            parser = KleinanzeigenParser()
            parser_token = JOB_PARSER.set(parser)
            try:
                await process_scan_job(bot, job, 1)
            finally:
                JOB_PARSER.reset(parser_token)
                await parser.close()
                await TRAFFIC.scan_job_finished()
                cancel_watcher.cancel()
                await asyncio.gather(cancel_watcher, return_exceptions=True)
                async with job_guard:
                    if active_jobs.get(job.user_id) is job:
                        active_jobs.pop(job.user_id, None)
                    if job.job_id in queued_job_ids:
                        queued_job_ids.remove(job.job_id)

            await COORDINATOR.ack_scan(message_id)
            await COORDINATOR.mark_job_complete(job_uid)
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            log.exception("Distributed scan worker failure job=%s message=%s", job_uid, message_id)
            # Only failures after actual execution starts are terminal. Queue/Redis/DB
            # coordination errors before execution keep the message pending so another
            # worker can reclaim it instead of showing the user a false failure.
            if execution_started:
                if job_uid:
                    try:
                        async with SessionLocal() as session:
                            result = await session.execute(select(UserScan).where(UserScan.job_uid == job_uid).limit(1))
                            scan = result.scalar_one_or_none()
                            if scan is not None and scan.finished_at is None:
                                scan.status = "failed"
                                scan.finished_at = datetime.utcnow()
                                scan.last_error = f"Worker error: {type(exc).__name__}: {exc}"[:1000]
                                await session.commit()
                    except Exception:
                        log.exception("Could not persist distributed worker failure job=%s", job_uid)
                if message_id:
                    try:
                        await COORDINATOR.ack_scan(message_id)
                    except Exception:
                        pass
                if job_uid:
                    try:
                        await COORDINATOR.mark_job_complete(job_uid)
                    except Exception:
                        pass
            else:
                log.warning("Job coordination failed before execution; Redis delivery will be reclaimed job=%s", job_uid)
            await asyncio.sleep(1.0)
        finally:
            if lock_refresher is not None:
                lock_refresher.cancel()
                await asyncio.gather(lock_refresher, return_exceptions=True)
            if job_uid and job_lock_token:
                try:
                    await COORDINATOR.release_job_lock(job_uid, job_lock_token)
                except Exception:
                    log.debug("Could not release distributed job lock job=%s", job_uid, exc_info=True)


async def enqueue_user_scan(
    message: Message, user_id: int, category_keys: list[str], page_limit: int, target_date: str,
    *, price_filter: str = "any", is_trial: bool = False,
) -> ScanJob | None:
    """Create a persistent scan card and queue the network job.

    v4.1.4: in distributed mode do not create a user scan until at least one
    live parser/fleet worker heartbeat exists. This prevents the UI from sitting
    forever on "Подготавливаю скан" when Redis is healthy but the Browser Fleet
    service is absent, misconfigured, or still booting.
    """
    category_keys = _validate_scan_category_count(category_keys)
    if is_trial:
        if len(category_keys) != FREE_TRIAL_MAX_CATEGORIES:
            await message.answer("🎁 Бесплатный скан доступен только для одной категории.")
            return None
        if int(page_limit) > FREE_TRIAL_MAX_PAGES:
            await message.answer(
                f"🔒 В бесплатном режиме доступно до {FREE_TRIAL_MAX_PAGES} страниц. "
                "50 страниц открываются по подписке."
            )
            return None
    try:
        requested_day = datetime.strptime(target_date, "%Y-%m-%d").date()
    except Exception:
        await message.answer("⚠️ Некорректная дата скана.")
        return None
    today_msk = datetime.now(MOSCOW).date()
    oldest_allowed = today_msk - timedelta(days=DATE_MAX_AGE_DAYS)
    if requested_day > today_msk or requested_day < oldest_allowed:
        await message.answer(
            f"⚠️ <b>Эта дата уже недоступна для нового скана.</b>\n\n"
            f"Можно сканировать только последние <b>{DATE_MAX_AGE_DAYS + 1} дней</b>: "
            f"с <b>{oldest_allowed:%d.%m.%Y}</b> по <b>{today_msk:%d.%m.%Y}</b>.",
            parse_mode=ParseMode.HTML,
        )
        return None
    if STABLE_SINGLE_SERVICE_MODE:
        existing = await get_persisted_active_scan(user_id)
        if existing is not None:
            await message.answer(
                "⏳ <b>У тебя уже есть активный скан.</b>\n\n"
                "Дождись завершения или останови его. Второй scan_id одновременно не создаётся.",
                parse_mode=ParseMode.HTML,
            )
            return None
    if DISTRIBUTED_WORKERS:
        ready_wait = 8
        try:
            ready_wait = max(0, min(30, int(os.getenv("DISTRIBUTED_WORKER_READY_WAIT_SECONDS", "8"))))
        except Exception:
            ready_wait = 8
        probe = await message.answer(
            "⏳ <b>Проверяю Browser Fleet…</b>", parse_mode=ParseMode.HTML
        )
        workers = 0
        deadline = asyncio.get_running_loop().time() + ready_wait
        while True:
            try:
                workers = await COORDINATOR.worker_count(prefix="parser")
            except Exception:
                workers = 0
            if workers > 0 or asyncio.get_running_loop().time() >= deadline:
                break
            await asyncio.sleep(1.0)
        if workers <= 0:
            log.error(
                "Distributed scan rejected: no live parser/fleet workers user=%s target=%s depth=%s",
                user_id, target_date, page_limit,
            )
            await probe.edit_text(
                "⚠️ <b>Browser Fleet не запущен</b>\n\n"
                "Redis работает, но сейчас нет ни одного активного parser/fleet-worker. "
                "Скан не был запущен и не будет висеть в очереди.",
                parse_mode=ParseMode.HTML,
            )
            return None
        await probe.edit_text(
            f"✅ <b>Browser Fleet готов</b> · активных worker: {workers}\n\nСоздаю скан…",
            parse_mode=ParseMode.HTML,
        )

    job_uid = uuid.uuid4().hex[:12]
    scan = await create_user_scan(
        user_id, job_uid, category_keys, page_limit, target_date,
        price_filter=price_filter, is_trial=is_trial,
    )
    if is_trial:
        reserved = await _consume_trial_credit(user_id)
        if reserved is None:
            async with SessionLocal() as session:
                doomed = await session.get(UserScan, int(scan.id))
                if doomed is not None:
                    await session.delete(doomed)
                    await session.commit()
            await message.answer(
                "🎁 Бесплатные сканы уже использованы или акция выключена. Полный доступ можно открыть по подписке.",
                reply_markup=await subscription_keyboard(user_id),
            )
            return None
    status = await message.answer("⏳ <b>Подготавливаю скан…</b>", parse_mode=ParseMode.HTML)
    await attach_user_scan_message(scan.id, message.chat.id, status.message_id)
    job = ScanJob(
        job_id=job_uid,
        user_id=user_id,
        chat_id=message.chat.id,
        status_message_id=status.message_id,
        category_keys=category_keys,
        created_at=datetime.utcnow(),
        warnings=[],
        page_limit=page_limit,
        scan_id=scan.id,
        target_date=target_date,
        price_filter=price_filter or "any",
        is_trial=bool(is_trial),
    )
    if DISTRIBUTED_WORKERS:
        try:
            await COORDINATOR.enqueue_scan(job.job_id)
        except Exception as exc:
            log.exception("Could not enqueue distributed scan job=%s", job.job_id)
            async with SessionLocal() as session:
                db_scan = await session.get(UserScan, int(scan.id))
                if db_scan is not None:
                    db_scan.status = "failed"
                    db_scan.finished_at = datetime.utcnow()
                    db_scan.last_error = f"Redis queue unavailable: {type(exc).__name__}"[:1000]
                    await session.commit()
            if is_trial:
                await _refund_trial_credit_for_scan(scan.id)
            queue_error_text = (
                "⚠️ <b>Не удалось поставить скан в очередь</b>\n\n"
                + (
                    "Сервис очереди временно недоступен. Бесплатный запуск не списан — попробуй ещё раз через несколько секунд."
                    if is_trial else
                    "Сервис очереди временно недоступен. Попробуй ещё раз через несколько секунд."
                )
            )
            await status.edit_text(queue_error_text, parse_mode=ParseMode.HTML)
            job.state = "failed"
            return job
    else:
        async with job_guard:
            active_jobs[job.user_id] = job
            queued_job_ids.append(job.job_id)
            scan_queue.put_nowait(job)
    await status.edit_text(
        render_user_job_status(job),
        parse_mode=ParseMode.HTML,
        reply_markup=job_keyboard(job.job_id, queued=True),
    )
    return job



async def recover_interrupted_user_scans(bot: Bot) -> int:
    """Close unfinished pre-restart jobs; never resurrect them automatically.

    v4.2.2 Hard Stable Reset keeps one scan = one explicit user launch. A Railway
    deploy/restart may interrupt a job, but it must not create a second Telegram
    card or silently start an old date hours later. Verified page checkpoints stay
    in PostgreSQL and can be reused when the user explicitly repeats the scan.
    """
    now = datetime.utcnow()
    async with SessionLocal() as session:
        rows = list((await session.execute(
            select(UserScan).where(
                UserScan.status.in_(["queued", "running", "cancelling"]),
                UserScan.finished_at.is_(None),
            )
        )).scalars().all())
        queued_trial_ids = {
            int(scan.id) for scan in rows
            if scan.status == "queued" and bool(getattr(scan, "is_trial", False))
        }
        for scan in rows:
            if scan.status == "cancelling":
                scan.status = "cancelled"
                scan.last_error = "Остановлен перед перезапуском сервиса"
            else:
                scan.status = "failed"
                scan.last_error = "Скан прерван перезапуском Railway; повтори его вручную"
            scan.finished_at = now
        await session.commit()

    # Jobs that were only waiting in the queue never consumed parser/network work.
    # Return their launch credit after a Railway restart; the helper is idempotent.
    for scan_id in queued_trial_ids:
        try:
            await _refund_trial_credit_for_scan(scan_id)
        except Exception:
            log.exception("Could not refund interrupted queued trial scan=%s", scan_id)

    for scan in rows:
        if not scan.chat_id or not scan.status_message_id:
            continue
        try:
            await bot.edit_message_text(
                chat_id=int(scan.chat_id),
                message_id=int(scan.status_message_id),
                text=(
                    "⚠️ <b>Скан был прерван перезапуском сервиса</b>\n\n"
                    "Он не запускается повторно автоматически. Уже подтверждённые страницы "
                    "сохранены; при необходимости нажми «Повторить скан» вручную."
                ),
                parse_mode=ParseMode.HTML,
            )
        except Exception:
            log.debug("Could not mark interrupted Telegram card scan=%s", scan.id, exc_info=True)
    return len(rows)





def _utc_to_msk_text(value: datetime | None) -> str:
    if value is None:
        return "—"
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(MOSCOW).strftime("%d.%m.%Y %H:%M")


def _access_mode_label(mode: str | None = None) -> str:
    mode = mode or current_access_mode()
    return {
        "admin_only": "🔒 Только админы",
        "subscription": "💎 По подписке",
        "open": "🌍 Открытый доступ",
    }.get(mode, mode)


def _provider_label(provider: str) -> str:
    return {"cryptobot": "🤖 CryptoBot", "xrocket": "🚀 xRocket"}.get(provider, provider)


def _payment_status_label(status: str) -> str:
    return {
        "pending": "⏳ ожидает",
        "paid": "✅ оплачено",
        "expired": "⌛ истекло",
        "failed": "❌ ошибка",
        "cancelled": "❌ отменено",
    }.get(status, status)


async def subscription_text(user_id: int) -> str:
    mode = current_access_mode()
    user = await get_commerce_user(user_id)
    until = user.access_until if user else cached_access_until(user_id)
    payments = await user_payments(user_id, 10)
    pending_count = sum(1 for p in payments if p.status == "pending")

    if user_id in ADMIN_IDS:
        status = "👑 <b>Администратор — доступ без ограничений</b>"
    elif is_banned_cached(user_id):
        status = "⛔ <b>Доступ заблокирован администратором</b>"
    elif mode == "open":
        status = "🌍 <b>Сервис сейчас открыт для всех</b>"
    elif until and until > datetime.utcnow():
        left = until - datetime.utcnow()
        hours = max(0, int(left.total_seconds() // 3600))
        days = hours // 24
        rem_hours = hours % 24
        status = (
            f"✅ <b>Подписка активна до {_utc_to_msk_text(until)} МСК</b>\n"
            f"Осталось: <b>{days} дн. {rem_hours} ч.</b>\n"
            "Можно продлить заранее — новые дни прибавятся к текущему сроку."
        )
    elif mode == "admin_only":
        status = "🔒 <b>Сервис пока работает в закрытом тестовом режиме.</b>"
    else:
        status = "❌ <b>Подписка не активна</b>\nВыбери тариф ниже, чтобы открыть доступ."

    providers = providers_status()
    methods = []
    if providers["cryptobot"]:
        methods.append("CryptoBot")
    if providers["xrocket"]:
        methods.append("xRocket")
    methods_text = " · ".join(methods) if methods else "временно недоступна"
    pending_text = f"\n⏳ Ожидающих счетов: <b>{pending_count}</b>" if pending_count else ""
    admin_mode = f"\nРежим: <b>{_access_mode_label(mode)}</b>" if user_id in ADMIN_IDS else ""
    trial_note = ""
    if user_id not in ADMIN_IDS and mode == "subscription" and not allowed(user_id):
        trial = await get_trial_status(user_id)
        if trial.eligible:
            trial_note = (
                f"\n\n🎁 <b>Стартовая акция:</b> осталось <b>{trial.remaining}</b> "
                f"бесплатных скан(а) · 1 категория · до {FREE_TRIAL_MAX_PAGES} страниц."
            )
    return (
        "<b>💎 Подписка</b>\n\n"
        f"{status}{trial_note}\n\n"
        f"Оплата: <b>{methods_text}</b>{pending_text}{admin_mode}"
    )


async def subscription_keyboard(user_id: int) -> InlineKeyboardMarkup:
    plans = await get_plans(active_only=True)
    user = await get_commerce_user(user_id)
    active = bool(user and user.access_until and user.access_until > datetime.utcnow())
    rows: list[list[InlineKeyboardButton]] = []
    if current_access_mode() == "subscription" and not is_banned_cached(user_id):
        for plan in plans:
            prefix = "🔄 Продлить · " if active else ""
            rows.append([InlineKeyboardButton(
                text=f"{prefix}{plan.title} · {plan.price_usdt:g} USDT",
                callback_data=f"buyplan:{plan.key}",
            )])
    rows.append([InlineKeyboardButton(text="💳 Мои платежи", callback_data="mypayments")])
    rows.append([InlineKeyboardButton(text="🌐 Язык", callback_data="language_settings")])
    if read_only_history_allowed(user_id):
        rows.append([InlineKeyboardButton(text="🏠 Меню", callback_data="home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def user_payments_keyboard(payments: list[SubscriptionPayment]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for p in payments[:5]:
        if p.status == "pending":
            if p.pay_url:
                rows.append([InlineKeyboardButton(text=f"💳 Открыть счёт #{p.id}", url=p.pay_url)])
            rows.append([InlineKeyboardButton(text=f"✅ Проверить счёт #{p.id}", callback_data=f"paycheck:{p.id}")])
    rows.append([InlineKeyboardButton(text="⬅️ Подписка", callback_data="subscription")])
    return InlineKeyboardMarkup(inline_keyboard=rows)



def payment_provider_keyboard(plan_key: str) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if provider_enabled("cryptobot"):
        rows.append([InlineKeyboardButton(text="🤖 Оплатить через CryptoBot", callback_data=f"payprovider:cryptobot:{plan_key}")])
    if provider_enabled("xrocket"):
        rows.append([InlineKeyboardButton(text="🚀 Оплатить через xRocket", callback_data=f"payprovider:xrocket:{plan_key}")])
    rows.append([InlineKeyboardButton(text="⬅️ К подписке", callback_data="subscription")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def payment_invoice_keyboard(payment: SubscriptionPayment) -> InlineKeyboardMarkup:
    rows = []
    if payment.pay_url:
        rows.append([InlineKeyboardButton(text="💳 Открыть оплату", url=payment.pay_url)])
    rows.append([InlineKeyboardButton(text="✅ Проверить оплату", callback_data=f"paycheck:{payment.id}")])
    rows.append([InlineKeyboardButton(text="⬅️ Подписка", callback_data="subscription")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_keyboard(ai_unread: int = 0, active_scans: int = 0) -> InlineKeyboardMarkup:
    active = max(0, int(active_scans or 0))
    parsing_label = "👀 Кто сейчас парсит" if active <= 0 else f"👀 Сейчас парсят · {active}"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Статистика", callback_data="adminstats")],
        [InlineKeyboardButton(text="⚙️ Воркеры", callback_data="adminworkers"),
         InlineKeyboardButton(text=parsing_label, callback_data="adminactive")],
        [InlineKeyboardButton(text="👥 Пользователи", callback_data="adminusers"),
         InlineKeyboardButton(text="💳 Платежи", callback_data="adminpayments")],
        [InlineKeyboardButton(text="🎟 Тарифы", callback_data="adminplans"),
         InlineKeyboardButton(text="🔐 Режим доступа", callback_data="adminmode")],
        [InlineKeyboardButton(text="🎁 Бесплатные сканы", callback_data="admintrial"),
         InlineKeyboardButton(text="👥 Рефералы", callback_data="adminreferral")],
        [InlineKeyboardButton(text="📡 DT Radar 3.0", callback_data="adminradarauto")],
        [InlineKeyboardButton(text="🟣 Vinted Lab", callback_data="av:home")],
        [InlineKeyboardButton(text="🔎 Найти пользователя", callback_data="adminusersearch")],
        [InlineKeyboardButton(text="📨 Daily Radar", callback_data="admindailyradar"),
         InlineKeyboardButton(text="📣 Рассылка", callback_data="adminbroadcast")],
        [InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ])


def admin_back_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")]
    ])


def admin_trial_keyboard(enabled: bool) -> InlineKeyboardMarkup:
    toggle = "⏸ Выключить акцию" if enabled else "▶️ Включить акцию"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=toggle, callback_data="admintrial:toggle")],
        [InlineKeyboardButton(text="📡 Воронка бесплатного Radar", callback_data="adminradarfunnel:0")],
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="admintrial")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


def admin_referral_keyboard(enabled: bool) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text="⏸ Выключить акцию" if enabled else "▶️ Включить акцию",
            callback_data="adminreferral:toggle",
        )],
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="adminreferral")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


def admin_radar_funnel_keyboard(visitors: list[dict], page: int, pages: int) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for visitor in visitors:
        uid = int(visitor.get("user_id") or 0)
        username = str(visitor.get("username") or "").strip()
        first_name = str(visitor.get("first_name") or "").strip()
        label = f"@{username}" if username else (first_name or str(uid))
        if len(label) > 30:
            label = label[:29].rstrip() + "…"
        rows.append([InlineKeyboardButton(text=f"👤 {label}", callback_data=f"adminuser:{uid}")])
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"adminradarfunnel:{page - 1}"))
    if page + 1 < pages:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"adminradarfunnel:{page + 1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton(text="🔄 Обновить", callback_data=f"adminradarfunnel:{page}")])
    rows.append([InlineKeyboardButton(text="⬅️ Бесплатные сканы", callback_data="admintrial")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_radar_autoscan_keyboard(state: dict) -> InlineKeyboardMarkup:
    status = str(state.get("status") or "idle")
    daily_enabled = bool(state.get("daily_enabled"))
    skip_today = bool(state.get("skip_daily_if_completed_today", True))
    rows: list[list[InlineKeyboardButton]] = []
    if status == "running":
        rows.append([InlineKeyboardButton(text="⏹ Остановить сейчас", callback_data="adminradarauto:stop")])
    elif status == "paused":
        rows.append([InlineKeyboardButton(text="▶️ Продолжить круг", callback_data="adminradarauto:resume"),
                     InlineKeyboardButton(text="🔄 Новый круг", callback_data="adminradarauto:start")])
    else:
        rows.append([InlineKeyboardButton(text="▶️ Запустить 1 круг", callback_data="adminradarauto:start")])
        last = dict(state.get("last_summary") or {})
        last_failed = int(last.get("failed") or 0)
        failure_details = [x for x in list(last.get("failed_categories") or []) if isinstance(x, dict)]
        if last_failed:
            rows.append([InlineKeyboardButton(text=f"⚠️ Требуют внимания: {last_failed}", callback_data="adminradarauto:errors:0")])
        if failure_details:
            rows.append([InlineKeyboardButton(text="🔁 Повторить проблемные", callback_data="adminradarauto:retry")])
    rows.append([InlineKeyboardButton(
        text=f"🔄 Ежедневный автокруг: {'ВКЛ' if daily_enabled else 'ВЫКЛ'}",
        callback_data="adminradarauto:daily",
    )])
    rows.append([InlineKeyboardButton(
        text=f"🕐 Время: {state.get('daily_time') or RADAR_AUTOSCAN_DEFAULT_TIME}",
        callback_data="adminradarauto:time",
    )])
    rows.append([InlineKeyboardButton(
        text=f"✅ Пропускать автокруг после ручного: {'ДА' if skip_today else 'НЕТ'}",
        callback_data="adminradarauto:skipday",
    )])
    rows.append([InlineKeyboardButton(text="📊 Аналитика Radar", callback_data="adminradarauto:analytics")])
    rows.append([InlineKeyboardButton(text="📜 История кругов", callback_data="adminradarauto:history"),
                 InlineKeyboardButton(text="🔄 Обновить Live", callback_data="adminradarauto")])
    rows.append([InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_radar_autoscan_errors_keyboard(state: dict, page: int, pages: int) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"adminradarauto:errors:{page - 1}"))
    if page + 1 < pages:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"adminradarauto:errors:{page + 1}"))
    if nav:
        rows.append(nav)
    if _radar_autoscan_failure_list(state) and str(state.get("status") or "idle") == "idle":
        rows.append([InlineKeyboardButton(text="🔁 Повторить проблемные", callback_data="adminradarauto:retry")])
    rows.append([InlineKeyboardButton(text="⬅️ Radar AutoScan", callback_data="adminradarauto")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_radar_autoscan_time_keyboard(current: str) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    row: list[InlineKeyboardButton] = []
    for value in RADAR_AUTOSCAN_TIME_CHOICES:
        label = ("✅ " if value == current else "") + value
        row.append(InlineKeyboardButton(text=label, callback_data=f"adminradarauto:settime:{value}"))
        if len(row) == 3:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="⬅️ Radar AutoScan", callback_data="adminradarauto")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_broadcast_preview_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🚀 Отправить всем", callback_data="adminbroadcast:send")],
        [InlineKeyboardButton(text="✏️ Заменить пост", callback_data="adminbroadcast:replace")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="adminbroadcast:cancel")],
    ])


def admin_daily_radar_keyboard(state: dict) -> InlineKeyboardMarkup:
    enabled = bool(state.get("enabled", True))
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text="⏸ Выключить Daily Radar" if enabled else "▶️ Включить Daily Radar",
            callback_data="admindailyradar:toggle",
        )],
        [InlineKeyboardButton(text=f"🕐 Время: {state.get('time') or RADAR_DAILY_DIGEST_DEFAULT_TIME}", callback_data="admindailyradar:time")],
        [InlineKeyboardButton(text="📣 Отправить сейчас", callback_data="admindailyradar:sendnow")],
        [InlineKeyboardButton(text="🧪 Тест только мне", callback_data="admindailyradar:test")],
        [InlineKeyboardButton(text="🔄 Обновить цифры", callback_data="admindailyradar")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


def admin_daily_radar_time_keyboard(current: str = "") -> InlineKeyboardMarkup:
    rows = []
    row: list[InlineKeyboardButton] = []
    for value in RADAR_DAILY_DIGEST_TIME_CHOICES:
        label = ("✅ " if value == current else "") + value
        # Encode HHMM so callback parsing cannot lose the hour/minute separator.
        row.append(InlineKeyboardButton(text=label, callback_data=f"admindailyradar:settime:{value.replace(':', '')}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="✍️ Ввести своё время", callback_data="admindailyradar:customtime")])
    rows.append([InlineKeyboardButton(text="⬅️ Daily Radar", callback_data="admindailyradar")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_daily_radar_send_confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🚀 Да, отправить всем", callback_data="admindailyradar:sendconfirm")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="admindailyradar")],
    ])


def admin_broadcast_back_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


def admin_workers_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 Date", callback_data="admindates"),
         InlineKeyboardButton(text="📄 Page", callback_data="adminpages"),
         InlineKeyboardButton(text="👁 View", callback_data="adminviews")],
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="adminworkers")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


def admin_active_scans_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="adminactive")],
        [InlineKeyboardButton(text="⚙️ Воркеры", callback_data="adminworkers")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


def admin_view_worker_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="adminviews")],
        [InlineKeyboardButton(text="⬅️ Воркеры", callback_data="adminworkers")],
    ])


def admin_page_worker_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="adminpages")],
        [InlineKeyboardButton(text="⬅️ Воркеры", callback_data="adminworkers")],
    ])


def admin_date_worker_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="admindates")],
        [InlineKeyboardButton(text="⬅️ Воркеры", callback_data="adminworkers")],
    ])


def admin_mode_keyboard() -> InlineKeyboardMarkup:
    current = current_access_mode()
    def b(mode: str, label: str) -> InlineKeyboardButton:
        prefix = "✅ " if current == mode else ""
        return InlineKeyboardButton(text=prefix + label, callback_data=f"adminsetmode:{mode}")
    return InlineKeyboardMarkup(inline_keyboard=[
        [b("admin_only", "🔒 Только админы")],
        [b("subscription", "💎 По подписке")],
        [b("open", "🌍 Открытый доступ")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


def admin_users_keyboard(users: list[BotUser]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    now = datetime.utcnow()
    for user in users[:20]:
        if user.is_banned:
            icon = "⛔"
        elif user.access_until and user.access_until > now:
            icon = "✅"
        else:
            icon = "▫️"
        name = f"@{user.username}" if user.username else (user.first_name or str(user.user_id))
        rows.append([InlineKeyboardButton(text=f"{icon} {name[:28]} · {user.user_id}", callback_data=f"adminuser:{user.user_id}")])
    rows.append([InlineKeyboardButton(text="🔎 Поиск", callback_data="adminusersearch")])
    rows.append([InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_user_keyboard(user: BotUser) -> InlineKeyboardMarkup:
    ban_label = "✅ Разблокировать" if user.is_banned else "⛔ Заблокировать"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="+1 день", callback_data=f"admingrant:{user.user_id}:1"),
         InlineKeyboardButton(text="+3 дня", callback_data=f"admingrant:{user.user_id}:3")],
        [InlineKeyboardButton(text="+7 дней", callback_data=f"admingrant:{user.user_id}:7"),
         InlineKeyboardButton(text="+30 дней", callback_data=f"admingrant:{user.user_id}:30")],
        [InlineKeyboardButton(text="➕ Свой срок", callback_data=f"admincustom:{user.user_id}")],
        [InlineKeyboardButton(text="💳 Платежи", callback_data=f"adminuserpayments:{user.user_id}"),
         InlineKeyboardButton(text="📊 Сканы", callback_data=f"adminuserscans:{user.user_id}")],
        [InlineKeyboardButton(text="⚠️ Ошибки", callback_data=f"adminusererrors:{user.user_id}")],
        [InlineKeyboardButton(text="🗑 Забрать доступ", callback_data=f"adminrevoke:{user.user_id}"),
         InlineKeyboardButton(text=ban_label, callback_data=f"adminban:{user.user_id}")],
        [InlineKeyboardButton(text="⬅️ Пользователи", callback_data="adminusers")],
    ])


def admin_user_back_keyboard(user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ К пользователю", callback_data=f"adminuser:{user_id}")],
        [InlineKeyboardButton(text="👥 Пользователи", callback_data="adminusers")],
    ])



def admin_plans_keyboard(plans: list[SubscriptionPlan]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for plan in plans:
        icon = "✅" if plan.is_active else "▫️"
        rows.append([InlineKeyboardButton(
            text=f"{icon} {plan.title} · {plan.price_usdt:g} USDT",
            callback_data=f"adminplan:{plan.key}",
        )])
    rows.append([InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_plan_keyboard(plan: SubscriptionPlan) -> InlineKeyboardMarkup:
    toggle_label = "⏸ Выключить тариф" if plan.is_active else "▶️ Включить тариф"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💰 Изменить цену", callback_data=f"adminplanprice:{plan.key}")],
        [InlineKeyboardButton(text=toggle_label, callback_data=f"adminplantoggle:{plan.key}")],
        [InlineKeyboardButton(text="⬅️ Тарифы", callback_data="adminplans")],
    ])


async def render_admin_user(user_id: int) -> tuple[str, InlineKeyboardMarkup] | None:
    user = await get_commerce_user(user_id)
    if user is None:
        return None
    now = datetime.utcnow()
    active = bool(user.access_until and user.access_until > now and not user.is_banned)
    async with SessionLocal() as session:
        scans = (await session.execute(select(func.count(UserScan.id)).where(UserScan.user_id == user.user_id))).scalar_one()
        paid = (await session.execute(select(func.count(SubscriptionPayment.id)).where(
            SubscriptionPayment.user_id == user.user_id, SubscriptionPayment.status == "paid"
        ))).scalar_one()
        pending = (await session.execute(select(func.count(SubscriptionPayment.id)).where(
            SubscriptionPayment.user_id == user.user_id, SubscriptionPayment.status == "pending"
        ))).scalar_one()
        errors = (await session.execute(select(func.count(ParserRun.id)).where(
            ParserRun.user_id == user.user_id, ParserRun.success.is_(False)
        ))).scalar_one()
        last_scan = (await session.execute(
            select(UserScan).where(UserScan.user_id == user.user_id).order_by(UserScan.created_at.desc()).limit(1)
        )).scalar_one_or_none()
    name = f"@{html.escape(user.username)}" if user.username else html.escape(user.first_name or "без username")
    scan_line = "—"
    if last_scan is not None:
        scan_line = f"{_date_label(last_scan.target_date)} · {last_scan.status} · {last_scan.result_count} результатов"
    text = (
        f"<b>👤 Пользователь</b>\n\n"
        f"{name}\n"
        f"ID: <code>{user.user_id}</code>\n"
        f"Статус: <b>{'⛔ заблокирован' if user.is_banned else ('✅ активен' if active else '▫️ без доступа')}</b>\n"
        f"Доступ до: <b>{_utc_to_msk_text(user.access_until)} МСК</b>\n"
        f"Первый вход: <b>{_utc_to_msk_text(user.joined_at)} МСК</b>\n"
        f"Последняя активность: <b>{_utc_to_msk_text(user.last_seen_at)} МСК</b>\n\n"
        f"📊 Сканов: <b>{int(scans or 0)}</b>\n"
        f"🎁 Пробные сканы: <b>{max(0, int(getattr(user, 'trial_scans_used', 0) or 0))}/{FREE_TRIAL_SCAN_LIMIT}</b>\n"
        f"Последний: <b>{html.escape(scan_line)}</b>\n"
        f"⚠️ Ошибок парсера: <b>{int(errors or 0)}</b>\n"
        f"💳 Оплат: <b>{int(paid or 0)}</b> · ожидают: <b>{int(pending or 0)}</b>\n"
        f"💰 Оплачено всего: <b>{float(user.paid_total_usdt or 0):g} USDT</b>"
    )
    return text, admin_user_keyboard(user)


async def send_access_screen(message: Message, user_id: int) -> None:
    if is_banned_cached(user_id):
        await message.answer("⛔ <b>Доступ к сервису заблокирован.</b>", parse_mode=ParseMode.HTML)
        return
    await message.answer(
        await subscription_text(user_id),
        parse_mode=ParseMode.HTML,
        reply_markup=await subscription_keyboard(user_id),
    )


def language_keyboard(current: str | None = None) -> InlineKeyboardMarkup:
    ru = "✅ 🇷🇺 Русский" if current == LANG_RU else "🇷🇺 Русский"
    en = "✅ 🇬🇧 English" if current == LANG_EN else "🇬🇧 English"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=ru, callback_data="language:ru"),
         InlineKeyboardButton(text=en, callback_data="language:en")],
    ])


def language_prompt_text(current: str | None = None) -> str:
    if current in (LANG_RU, LANG_EN):
        return (
            "<b>🌐 Язык интерфейса / Interface language</b>\n\n"
            f"Текущий язык / Current language: <b>{language_name(current)}</b>\n\n"
            "Выберите язык / Choose language:"
        )
    return (
        "<b>🌐 Выберите язык / Choose language</b>\n\n"
        "🇷🇺 Русский\n🇬🇧 English\n\n"
        "Язык можно изменить позже в настройках.\n"
        "You can change the language later in Settings."
    )


async def _send_language_picker(message: Message, current: str | None = None) -> None:
    # Deliberately bilingual and not run through localization: this is the one
    # screen shown before the user has chosen a language.
    token = _UI_LANGUAGE.set(LANG_RU)
    try:
        await message.answer(
            language_prompt_text(current),
            parse_mode=ParseMode.HTML,
            reply_markup=language_keyboard(current),
        )
    finally:
        _UI_LANGUAGE.reset(token)


async def _admin_surface_requires_russian(event, data) -> bool:
    """Return True only for the actual admin surface, not the admin user's normal UI."""
    if isinstance(event, Message):
        text = (event.text or "").strip().lower()
        if text.startswith("/admin"):
            return True
    elif isinstance(event, CallbackQuery):
        callback_data = (event.data or "").strip().lower()
        if callback_data.startswith("admin") or callback_data.startswith("aic:"):
            return True

    # Admin text-entry flows (user search, custom days, plan price) do not carry
    # an `admin*` callback while the user is typing, so preserve Russian there too.
    state = data.get("state") if isinstance(data, dict) else None
    if state is not None:
        try:
            raw_state = await state.get_state()
        except Exception:
            raw_state = None
        if raw_state and str(raw_state).startswith(f"{AdminInput.__name__}:"):
            return True
    return False


class LanguageContextMiddleware(BaseMiddleware):
    """Load per-user UI language; keep only the actual admin panel in Russian."""

    async def __call__(self, handler, event, data):
        tg_user = getattr(event, "from_user", None)
        if tg_user is None:
            return await handler(event, data)
        uid = int(tg_user.id)
        force_admin_ru = uid in ADMIN_IDS and await _admin_surface_requires_russian(event, data)
        try:
            language = LANG_RU if force_admin_ru else await get_user_language(uid)
        except Exception:
            log.exception("Could not load UI language user=%s", tg_user.id)
            language = LANG_RU if force_admin_ru else None
        token = _UI_LANGUAGE.set(language)
        try:
            # Existing users from pre-v4.6.5 are gated once, too, so nobody gets
            # a half-Russian/half-English experience after upgrade.
            if language is None:
                if isinstance(event, Message):
                    text = (event.text or "").strip()
                    if text.startswith("/start") or text.startswith("/language"):
                        return await handler(event, data)
                    await _send_language_picker(event)
                    return None
                if isinstance(event, CallbackQuery):
                    callback_data = event.data or ""
                    if callback_data.startswith("language:") or callback_data == "language_settings":
                        return await handler(event, data)
                    await event.answer()
                    if event.message:
                        await _send_language_picker(event.message)
                    return None
            return await handler(event, data)
        finally:
            _UI_LANGUAGE.reset(token)


class ActivityAccessMiddleware(BaseMiddleware):
    """Track users and keep commercial access checks outside parser handlers."""

    async def __call__(self, handler, event, data):
        tg_user = getattr(event, "from_user", None)
        if tg_user is None:
            return await handler(event, data)
        referral_referrer_id: int | None = None
        if isinstance(event, Message):
            raw_text = (event.text or "").strip()
            match = re.match(r"^/start(?:@[A-Za-z0-9_]+)?\s+ref_(\d+)\s*$", raw_text, flags=re.IGNORECASE)
            if match:
                try:
                    referral_referrer_id = int(match.group(1))
                except Exception:
                    referral_referrer_id = None
        try:
            await touch_user(tg_user, referral_referrer_id=referral_referrer_id)
        except Exception:
            log.exception("Could not update user activity user=%s", tg_user.id)

        uid = int(tg_user.id)
        if uid in ADMIN_IDS or allowed(uid):
            return await handler(event, data)

        trial = await get_trial_status(uid)

        # /start, the read-only home/history, and subscription/payment callbacks
        # stay reachable after a subscription expires. Trial users additionally get
        # only the setup flow required for their two launch scans; repeat/update/
        # auto-observation actions remain subscription-only.
        if isinstance(event, Message):
            text = (event.text or "").strip()
            public_commands = ("/start", "/menu", "/my_scans", "/admin", "/subscription", "/help", "/language")
            if text.startswith(public_commands):
                return await handler(event, data)
            if text.startswith("/radar") and free_radar_preview_allowed(uid):
                return await handler(event, data)
            if trial.eligible and text.startswith(("/new_scan", "/categories", "/settings")):
                return await handler(event, data)
            if text.startswith("/stop") and read_only_history_allowed(uid):
                return await handler(event, data)
            state = data.get("state") if isinstance(data, dict) else None
            raw_state = None
            if state is not None:
                try:
                    raw_state = await state.get_state()
                except Exception:
                    raw_state = None
            if trial.eligible and raw_state and (
                str(raw_state).startswith(f"{ScanInput.__name__}:")
                or str(raw_state) in {
                    f"{SettingsInput.__name__}:min_views",
                    f"{SettingsInput.__name__}:include_words",
                    f"{SettingsInput.__name__}:exclude_words",
                }
            ):
                return await handler(event, data)
            await send_access_screen(event, uid)
            return None

        if isinstance(event, CallbackQuery):
            callback_data = event.data or ""
            public_prefixes = ("buyplan:", "payprovider:", "paycheck:", "mypayments", "language:")
            if callback_data in {"subscription", "language_settings", "referral"} or callback_data.startswith(public_prefixes):
                return await handler(event, data)
            if is_banned_cached(uid):
                await event.answer("Доступ заблокирован", show_alert=True)
                return None

            if free_radar_preview_allowed(uid):
                radar_public_exact = {"radar_home", "radarbest", "radar_locked", "radardaily_open"}
                radar_public_prefixes = ("radar_locked:", "radarlist:", "radarpreviewitem:", "radar_upgrade:")
                if callback_data in radar_public_exact or callback_data.startswith(radar_public_prefixes):
                    return await handler(event, data)

            readonly_exact = {
                "home", "post_home", "my_scans", "archive_my_scans", "archive_noop", "queue_status",
                "radar_locked",
            }
            readonly_prefixes = (
                "scan_archive:", "scan:", "scanproducts:", "scantop:", "scantop50:",
                "scangrowth:", "scangrowthexport:", "scanhistory:", "scanexport:", "cancel_scan:",
            )
            if read_only_history_allowed(uid) and (
                callback_data in readonly_exact or callback_data.startswith(readonly_prefixes)
            ):
                return await handler(event, data)

            if trial.eligible:
                # Explicitly keep every network follow-up outside the free trial.
                if callback_data in {"auto_obs_menu", "toggle_auto_obs", "view_test"} or callback_data.startswith(
                    ("scanrepeat:", "scanrecheck:", "scanviews:")
                ):
                    await event.answer("Эта функция доступна по подписке", show_alert=True)
                    if event.message:
                        await send_access_screen(event.message, uid)
                    return None
                trial_exact = {
                    "start_scan", "groups", "clear_all", "selected", "settings", "post_settings",
                    "mode_help", "set_mode", "set_period", "set_price", "set_min_views", "set_sort",
                    "set_include", "set_exclude", "reset_settings", "toggle_dedupe", "toggle_noise",
                    "scanprice_menu",
                }
                trial_prefixes = (
                    "onboard:", "grp:", "cat:", "grpall:", "quickmode:", "mode:", "period:",
                    "price:", "minviews:", "sort:", "scan_date:", "scanprice:", "scanpages:",
                )
                if callback_data in trial_exact or callback_data.startswith(trial_prefixes):
                    return await handler(event, data)

            if callback_data == "start_scan" or callback_data.startswith(("scanrepeat:", "scanrecheck:", "scanviews:")):
                alert = "Для этого действия нужна активная подписка"
            else:
                alert = "Нужна активная подписка"
            await event.answer(alert, show_alert=True)
            if event.message:
                await send_access_screen(event.message, uid)
            return None
        return None


dp = Dispatcher()

# Language selection is resolved before commercial access so first-run users and
# expired subscribers can always choose/change the interface language.
_language_middleware = LanguageContextMiddleware()
dp.message.outer_middleware(_language_middleware)
dp.callback_query.outer_middleware(_language_middleware)

# Commercial access/user tracking runs before regular handlers. It never performs
# parser work, so Telegram navigation stays responsive while scans run in background.
_access_middleware = ActivityAccessMiddleware()
dp.message.outer_middleware(_access_middleware)
dp.callback_query.outer_middleware(_access_middleware)


def _is_admin(user_id: int) -> bool:
    return int(user_id) in ADMIN_IDS


async def _admin_dashboard_text() -> str:
    stats = await admin_stats()
    providers = providers_status()
    traffic = await TRAFFIC.snapshot()
    return (
        "<b>🛠 Админ-панель</b>\n\n"
        f"Доступ: <b>{_access_mode_label()}</b>\n"
        f"Оплата: CryptoBot {'✅' if providers['cryptobot'] else '▫️'} · "
        f"xRocket {'✅' if providers['xrocket'] else '▫️'}\n\n"
        "<b>👥 Пользователи</b>\n"
        f"Всего: <b>{stats['total_users']}</b> · за 24ч активны: <b>{stats['active_24h']}</b>\n"
        f"Новых за 24ч: <b>{stats['new_24h']}</b> · активных подписок: <b>{stats['active_users']}</b>\n\n"
        "<b>📊 Использование</b>\n"
        f"Сканов: <b>{stats['total_scans']}</b> · за 24ч: <b>{stats['scans_24h']}</b>\n"
        f"Активные сетевые лимиты: scan <b>{traffic.scan_limit}</b> · views <b>{traffic.view_limit}</b> · "
        f"global <b>{traffic.global_limit}</b>\n\n"
        "<b>💳 Платежи</b>\n"
        f"Успешных: <b>{stats['paid_count']}</b> · ожидают: <b>{stats['pending_payments']}</b>\n"
        f"За 24ч: <b>{stats['paid_24h']:g} USDT</b> · всего: <b>{stats['paid_total']:g} USDT</b>"
    )


async def _admin_running_scan_count() -> int:
    async with SessionLocal() as session:
        value = (await session.execute(
            select(func.count(UserScan.id)).where(
                UserScan.status == "running",
                UserScan.finished_at.is_(None),
            )
        )).scalar_one()
        return int(value or 0)


async def _admin_live_keyboard() -> InlineKeyboardMarkup:
    active_scans = await _admin_running_scan_count()
    return admin_keyboard(0, active_scans)


def _worker_health_label(status: dict, *, active_key: str = "active_total") -> str:
    if not status.get("enabled"):
        return "▫️ выключен"
    if not status.get("alive"):
        return "🔴 offline"
    workers = len(list(status.get("workers") or []))
    active = int(status.get(active_key, 0) or 0)
    queue = int(status.get("queue_depth", 0) or 0)
    return f"🟢 {workers} · активных {active} · очередь {queue}"


async def _admin_workers_text() -> str:
    date_status, page_status, view_status = await asyncio.gather(
        REMOTE_DATE_MANAGER.status(),
        REMOTE_PAGE_MANAGER.status(),
        REMOTE_VIEW_MANAGER.status(),
    )
    total_active = (
        int(date_status.get("active_total", 0) or 0)
        + int(page_status.get("active_total", 0) or 0)
        + int(view_status.get("active_jobs", 0) or 0)
    )
    total_queue = (
        int(date_status.get("queue_depth", 0) or 0)
        + int(page_status.get("queue_depth", 0) or 0)
        + int(view_status.get("queue_depth", 0) or 0)
    )
    return "\n".join([
        "<b>⚙️ ВОРКЕРЫ DT PARSER</b>",
        "",
        f"📅 Date Worker: <b>{_worker_health_label(date_status)}</b>",
        f"📄 Page Worker: <b>{_worker_health_label(page_status)}</b>",
        f"👁 View Worker: <b>{_worker_health_label(view_status, active_key='active_jobs')}</b>",
        "",
        f"Всего активных worker-задач: <b>{total_active}</b>",
        f"Всего в worker-очередях: <b>{total_queue}</b>",
        "",
        "<i>Нажми Date / Page / View ниже, чтобы открыть подробную диагностику конкретного воркера.</i>",
    ])


def _admin_scan_user_label(user: BotUser | None, user_id: int) -> str:
    if user is not None and user.username:
        return f"@{html.escape(user.username)}"
    if user is not None and user.first_name:
        return html.escape(user.first_name)
    return f"ID {int(user_id)}"


async def _admin_active_scans_text(limit: int = 20) -> str:
    async with SessionLocal() as session:
        result = await session.execute(
            select(UserScan, BotUser)
            .outerjoin(BotUser, BotUser.user_id == UserScan.user_id)
            .where(
                UserScan.status.in_(ACTIVE_SCAN_STATUSES),
                UserScan.finished_at.is_(None),
            )
            .order_by(UserScan.created_at.asc(), UserScan.id.asc())
            .limit(max(1, int(limit)))
        )
        rows = list(result.all())

    running = [(scan, user) for scan, user in rows if scan.status == "running"]
    queued = [(scan, user) for scan, user in rows if scan.status == "queued"]
    cancelling = [(scan, user) for scan, user in rows if scan.status == "cancelling"]
    lines = [
        "<b>👀 КТО СЕЙЧАС ПАРСИТ</b>",
        "",
        f"🟢 Парсят сейчас: <b>{len(running)}/{GUARANTEED_LOCAL_PARSER_LANES}</b> · 🟡 в очереди: <b>{len(queued)}</b>",
    ]
    if cancelling:
        lines[-1] += f" · 🟠 останавливаются: <b>{len(cancelling)}</b>"

    if not rows:
        lines.extend(["", "Сейчас активных сканов нет."] )
        return "\n".join(lines)

    now = datetime.utcnow()
    if running:
        lines.extend(["", "<b>🟢 Активные сканы</b>"] )
        for scan, user in running[:12]:
            elapsed = max(0, int((now - scan.created_at).total_seconds()))
            lines.extend([
                f"• <b>{_admin_scan_user_label(user, scan.user_id)}</b> · <code>{scan.user_id}</code>",
                f"  {html.escape(scan.title or 'Скан')} · {_date_label(scan.target_date)} · {int(scan.page_limit or 0)} стр.",
                f"  Статус: <b>парсит</b> · идёт <b>{_human_duration(elapsed)}</b>",
            ])

    if queued:
        lines.extend(["", "<b>🟡 Очередь</b>"] )
        for pos, (scan, user) in enumerate(queued[:8], start=1):
            waited = max(0, int((now - scan.created_at).total_seconds()))
            lines.append(
                f"{pos}. <b>{_admin_scan_user_label(user, scan.user_id)}</b> · "
                f"{html.escape(scan.title or 'Скан')} · {_date_label(scan.target_date)} · "
                f"{int(scan.page_limit or 0)} стр. · ждёт {_human_duration(waited)}"
            )

    if cancelling:
        lines.extend(["", "<b>🟠 Останавливаются</b>"] )
        for scan, user in cancelling[:5]:
            lines.append(f"• <b>{_admin_scan_user_label(user, scan.user_id)}</b> · {html.escape(scan.title or 'Скан')}")
    return "\n".join(lines)


async def _admin_view_worker_text() -> str:
    status = await REMOTE_VIEW_MANAGER.status()
    if not status.get("enabled"):
        return (
            "<b>👁 View Worker</b>\n\n"
            "Статус: <b>▫️ выключен</b>\n"
            "REMOTE_VIEW_WORKER_ENABLED=1 не активирован или REDIS_URL не задан."
        )
    if not status.get("alive"):
        err = html.escape(str(status.get("error") or "heartbeat не найден"))
        return (
            "<b>👁 View Worker</b>\n\n"
            "Статус: <b>🔴 offline</b>\n"
            f"Redis queue: <b>{int(status.get('queue_depth', 0) or 0)}</b>\n"
            f"Причина: <code>{err[:300]}</code>\n\n"
            "Основной бот автоматически использует локальный v4.3.8 fallback."
        )

    workers = list(status.get("workers") or [])
    queue_depth = int(status.get("queue_depth", 0) or 0)
    active_jobs = int(status.get("active_jobs", 0) or 0)
    pool_total = int(status.get("pool_total", 0) or 0)
    browser_total = int(status.get("browser_total", 0) or 0)
    rate_total = float(status.get("rate_total", 0.0) or 0.0)
    sharding_enabled = bool(status.get("sharding_enabled"))
    shard_size = int(status.get("shard_size", 0) or 0)
    last_shard_count = int(status.get("last_shard_count", 0) or 0)
    last_shard_total = int(status.get("last_shard_total", 0) or 0)
    last_shard_workers = int(status.get("last_shard_workers", 0) or 0)
    last_shard_failed = int(status.get("last_shard_failed", 0) or 0)
    expected_replicas = int(status.get("expected_replicas", 0) or 0)
    lines = [
        "<b>👁 VIEW MANAGER PRO</b>",
        "",
        f"Статус: <b>🟢 online</b> · workers: <b>{len(workers)}</b> / ожидается <b>{expected_replicas or len(workers)}</b>",
        f"Redis queue: <b>{queue_depth}</b> · активных jobs: <b>{active_jobs}</b>",
        f"Общий HTTP pool: <b>{pool_total}</b> · Browser: <b>{browser_total}</b>",
        f"Скорость: <b>{rate_total:.1f} views/sec</b>",
        f"View Sharding: <b>{'✅ ВКЛ' if sharding_enabled else '⛔ ВЫКЛ'}</b> · цель ≈ <b>{shard_size}</b> URL/job",
    ]
    if last_shard_count:
        lines.append(
            f"Последний batch: <b>{last_shard_total}</b> URL → <b>{last_shard_count}</b> jobs "
            f"· live при старте: <b>{last_shard_workers}</b> · failed shards: <b>{last_shard_failed}</b>"
        )
    for idx, worker in enumerate(workers[:4], start=1):
        processed = int(worker.get("processed_total", 0) or 0)
        exact_pct = float(worker.get("exact_pct", 0.0) or 0.0)
        fallback_pct = float(worker.get("fallback_pct", 0.0) or 0.0)
        pool = int(worker.get("view_pool", 0) or 0)
        pool_min = int(worker.get("pool_min", pool) or pool)
        pool_max = int(worker.get("pool_max", pool) or pool)
        traffic_limit = int(worker.get("traffic_view_limit", pool) or pool)
        browser_pool = int(worker.get("browser_pool", 0) or 0)
        rate = float(worker.get("rate_ema", 0.0) or 0.0)
        item_ms = float(worker.get("item_ms_ema", 0.0) or 0.0)
        r403 = int(worker.get("http_403", 0) or 0)
        r429 = int(worker.get("http_429", 0) or 0)
        refusals_60 = int(worker.get("refusals_60s", 0) or 0)
        penalty = int(worker.get("penalty", 0) or 0)
        cooldown = float(worker.get("cooldown_seconds", 0.0) or 0.0)
        requeues = int(worker.get("requeues_total", 0) or 0)
        failures = int(worker.get("rounds_failed", 0) or 0)
        reason = html.escape(str(worker.get("adaptive_reason") or "—"))[:180]
        consumer = html.escape(str(worker.get("consumer") or f"worker-{idx}"))[:45]
        lines.extend([
            "",
            f"<b>Worker {idx}</b> · v<b>{html.escape(str(worker.get('version') or '—'))}</b> · <code>{consumer}</code>",
            f"Fleet: <b>{html.escape(str(worker.get('fleet_bucket') or '—'))}</b> · view/global <b>{int(worker.get('fleet_view_limit', 0) or 0)}/{int(worker.get('fleet_global_limit', 0) or 0)}</b>",
            f"Pool: <b>{pool}</b> [{pool_min}–{pool_max}] · effective: <b>{traffic_limit}</b> · browser: <b>{browser_pool}</b>",
            f"Rate: <b>{rate:.1f}/s</b> · ~<b>{item_ms:.0f} ms/item</b>",
            f"Exact: <b>{exact_pct:.1f}%</b> · browser fallback: <b>{fallback_pct:.1f}%</b> · processed: <b>{processed}</b>",
            f"403: <b>{r403}</b> · 429: <b>{r429}</b> · refusals/60s: <b>{refusals_60}</b>",
            f"Penalty: <b>{penalty}</b> · cooldown: <b>{cooldown:.1f}s</b> · requeue: <b>{requeues}</b> · errors: <b>{failures}</b>",
            f"Ожидание: traffic <b>{float(worker.get('traffic_wait_ms_avg', 0.0) or 0.0):.0f} ms</b> · Redis limiter <b>{float(worker.get('redis_wait_ms_avg', 0.0) or 0.0):.0f} ms</b>",
            f"Adaptive: <code>{reason}</code>",
        ])
    return "\n".join(lines)


AI_BADGE_EVENT_TYPES = ("winner", "confirmed")


async def _edit_or_answer(target: Message, text: str, *, reply_markup=None) -> None:
    """Open a UI screen with one Telegram request whenever possible.

    The home menu is a photo message. Calling edit_text() on a photo always
    produces a Telegram Bad Request, which previously added a full network
    round-trip before the fallback answer() and made every tab feel 1-2s slow.
    Media messages now go straight to answer(); text messages still edit in
    place and keep the old fallback for genuinely stale/non-editable messages.
    """
    if not getattr(target, "text", None):
        await target.answer(text, parse_mode=ParseMode.HTML, reply_markup=reply_markup)
        return
    try:
        await target.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=reply_markup)
    except Exception:
        await target.answer(text, parse_mode=ParseMode.HTML, reply_markup=reply_markup)


@dp.callback_query(F.data.startswith("radar_upgrade:"))
async def radar_upgrade_handler(callback: CallbackQuery, state: FSMContext) -> None:
    source = str(callback.data or "").split(":", 1)[1] if ":" in str(callback.data or "") else "radar"
    if free_radar_preview_allowed(callback.from_user.id):
        await record_free_radar_event(callback.from_user.id, "upgrade_click", feature=source)
    await state.clear()
    await callback.answer()
    text, markup = await asyncio.gather(
        subscription_text(callback.from_user.id),
        subscription_keyboard(callback.from_user.id),
    )
    await _edit_or_answer(callback.message, text, reply_markup=markup)


async def _referral_screen(user_id: int, bot: Bot) -> tuple[str, InlineKeyboardMarkup]:
    enabled, stats = await asyncio.gather(
        referral_promo_enabled(), referral_user_stats(user_id)
    )
    me = await bot.get_me()
    username = (me.username or "").strip()
    referral_link = f"https://t.me/{username}?start=ref_{int(user_id)}" if username else f"ref_{int(user_id)}"
    total = int(stats.get("total", 0) or 0)
    eligible = int(stats.get("eligible", 0) or 0)
    progress = int(stats.get("progress", 0) or 0)
    days_earned = int(stats.get("days_earned", 0) or 0)
    left = 2 - progress if progress else 2

    if enabled:
        status = (
            "🟢 <b>Акция активна</b>\n\n"
            "Приведи <b>2 новых пользователей</b> в DT Parser и получи "
            "<b>+1 день полной подписки</b>.\n"
            "Каждые следующие 2 новых пользователя дают ещё +1 день."
        )
        progress_text = (
            f"🎯 Прогресс: <b>{progress}/2</b>\n"
            f"До следующего дня: <b>{left} чел.</b>"
        )
        share_text = (
            "Попробуй DT Parser для Kleinanzeigen по моей ссылке 👇"
        )
    else:
        status = (
            "⏸ <b>Реферальная акция сейчас на паузе.</b>\n\n"
            "Твоя персональная ссылка продолжает учитывать новых пользователей, "
            "но входы во время паузы не дают бонусные дни."
        )
        progress_text = "🎯 Бонусный зачёт временно остановлен."
        share_text = "DT Parser для Kleinanzeigen 👇"

    text = (
        "<b>🎁 Получить день бесплатно</b>\n\n"
        f"{status}\n\n"
        f"👥 Всего новых по твоей ссылке: <b>{total}</b>\n"
        f"✅ В зачёте акции: <b>{eligible}</b>\n"
        f"💎 Получено: <b>+{days_earned} дн.</b>\n"
        f"{progress_text}\n\n"
        "<b>Твоя персональная ссылка:</b>\n"
        f"<code>{html.escape(referral_link)}</code>"
    )

    share_url = (
        "https://t.me/share/url?url=" + quote(referral_link, safe="") +
        "&text=" + quote(share_text, safe="")
    )
    rows = [
        [InlineKeyboardButton(text="📤 Поделиться ссылкой", url=share_url)],
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="referral")],
        [InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ]
    return text, InlineKeyboardMarkup(inline_keyboard=rows)


@dp.callback_query(F.data == "referral")
async def referral_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await callback.answer()
    text, markup = await _referral_screen(callback.from_user.id, callback.bot)
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data == "subscription")
async def subscription_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await callback.answer()
    text, markup = await asyncio.gather(
        subscription_text(callback.from_user.id),
        subscription_keyboard(callback.from_user.id),
    )
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data == "mypayments")
async def my_payments_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    payments = await user_payments(callback.from_user.id, 15)
    lines = ["<b>💳 Мои платежи</b>", ""]
    if not payments:
        lines.append("Платежей пока нет.")
    else:
        for p in payments[:10]:
            plan = await get_plan(p.plan_key)
            title = plan.title if plan else p.plan_key
            date_text = _utc_to_msk_text(p.paid_at or p.created_at)
            lines.append(
                f"{_payment_status_label(p.status)} · <b>{html.escape(title)}</b> · "
                f"{p.amount_usdt:g} USDT · {_provider_label(p.provider)}\n"
                f"#{p.id} · {date_text} МСК"
            )
    await callback.answer()
    await _edit_or_answer(callback.message, "\n\n".join(lines), reply_markup=user_payments_keyboard(payments))


@dp.callback_query(F.data.startswith("buyplan:"))
async def buy_plan_handler(callback: CallbackQuery) -> None:
    if is_banned_cached(callback.from_user.id):
        await callback.answer("Доступ заблокирован", show_alert=True)
        return
    if current_access_mode() != "subscription":
        await callback.answer("Продажа подписок сейчас выключена", show_alert=True)
        return
    key = callback.data.split(":", 1)[1]
    plan = await get_plan(key)
    if plan is None or not plan.is_active:
        await callback.answer("Этот тариф сейчас недоступен", show_alert=True)
        return
    providers = providers_status()
    if not any(providers.values()):
        await callback.answer("Оплата пока не настроена", show_alert=True)
        return
    await callback.answer()
    text = (
        "<b>💎 Оформление подписки</b>\n\n"
        f"Тариф: <b>{html.escape(plan.title)}</b>\n"
        f"Срок: <b>{plan.days} дн.</b>\n"
        f"Стоимость: <b>{plan.price_usdt:g} USDT</b>\n\n"
        "Выбери способ оплаты. Сумма и срок задаются ботом автоматически."
    )
    await _edit_or_answer(callback.message, text, reply_markup=payment_provider_keyboard(plan.key))


@dp.callback_query(F.data.startswith("payprovider:"))
async def create_payment_handler(callback: CallbackQuery) -> None:
    if is_banned_cached(callback.from_user.id):
        await callback.answer("Доступ заблокирован", show_alert=True)
        return
    try:
        _, provider, plan_key = callback.data.split(":", 2)
    except ValueError:
        await callback.answer("Некорректный способ оплаты", show_alert=True)
        return
    await callback.answer("Создаю счёт…")
    try:
        payment = await create_subscription_payment(callback.from_user.id, plan_key, provider)
    except PaymentProviderError as exc:
        await callback.message.answer(
            f"⚠️ <b>Не удалось создать счёт</b>\n{html.escape(str(exc))}",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⬅️ Подписка", callback_data="subscription")]
            ]),
        )
        return
    except Exception:
        log.exception("Could not create payment invoice")
        await callback.message.answer("⚠️ Не удалось создать счёт. Попробуй чуть позже.")
        return

    plan = await get_plan(payment.plan_key)
    title = plan.title if plan else payment.plan_key
    text = (
        "<b>💳 Счёт создан</b>\n\n"
        f"Способ: <b>{_provider_label(payment.provider)}</b>\n"
        f"Тариф: <b>{html.escape(title)}</b>\n"
        f"Сумма: <b>{payment.amount_usdt:g} USDT</b>\n"
        f"Действует до: <b>{_utc_to_msk_text(payment.expires_at)} МСК</b>\n\n"
        "Нажми «Открыть оплату». После оплаты бот проверит счёт автоматически; "
        "кнопка «Проверить оплату» нужна только если хочешь проверить сразу."
    )
    await callback.message.answer(text, parse_mode=ParseMode.HTML, reply_markup=payment_invoice_keyboard(payment))


@dp.callback_query(F.data.startswith("paycheck:"))
async def check_payment_handler(callback: CallbackQuery) -> None:
    try:
        payment_id = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Некорректный счёт", show_alert=True)
        return
    payment = await get_payment(payment_id)
    if payment is None or (payment.user_id != callback.from_user.id and not _is_admin(callback.from_user.id)):
        await callback.answer("Счёт не найден", show_alert=True)
        return
    await callback.answer("Проверяю…")
    refreshed, just_activated = await refresh_payment(payment_id)
    if refreshed is None:
        await callback.message.answer("⚠️ Счёт не найден.")
        return
    if refreshed.status == "paid":
        user = await get_commerce_user(refreshed.user_id)
        text = (
            "✅ <b>Оплата подтверждена</b>\n\n"
            f"Подписка активна до <b>{_utc_to_msk_text(user.access_until if user else None)} МСК</b>."
        )
        markup = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🏠 Перейти в сервис", callback_data="home")],
            [InlineKeyboardButton(text="💎 Подписка", callback_data="subscription")],
        ])
        await callback.message.answer(text, parse_mode=ParseMode.HTML, reply_markup=markup)
        return
    if refreshed.status in {"expired", "failed", "cancelled"}:
        title = "⌛ <b>Срок счёта истёк.</b>" if refreshed.status == "expired" else "❌ <b>Этот счёт нельзя подтвердить.</b>"
        await callback.message.answer(
            title + " Создай новый счёт — старый повторно использовать не нужно.",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔄 Создать новый счёт", callback_data=f"buyplan:{refreshed.plan_key}")],
                [InlineKeyboardButton(text="💳 Мои платежи", callback_data="mypayments")],
                [InlineKeyboardButton(text="💎 Подписка", callback_data="subscription")]
            ]),
        )
        return
    await callback.message.answer(
        "⏳ <b>Оплата пока не подтверждена.</b>\nЕсли ты только что оплатил, подожди несколько секунд — бот проверяет счета автоматически.",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Проверить ещё раз", callback_data=f"paycheck:{refreshed.id}")],
            [InlineKeyboardButton(text="💳 Мои платежи", callback_data="mypayments")],
        ]),
    )


async def _admin_date_worker_text() -> str:
    status = await REMOTE_DATE_MANAGER.status()
    if not status.get("enabled"):
        return (
            "<b>📅 Date Worker</b>\n\n"
            "Статус: <b>▫️ выключен</b>\n"
            "REDIS_URL не задан или REMOTE_DATE_WORKER_ENABLED=0."
        )
    if not status.get("alive"):
        err = html.escape(str(status.get("error") or "heartbeat не найден"))
        return (
            "<b>📅 Date Worker</b>\n\n"
            "Статус: <b>🔴 offline</b>\n"
            f"Redis queue: <b>{int(status.get('queue_depth', 0) or 0)}</b>\n"
            f"Date cache: <b>{int(status.get('cache_ttl', 180) or 180)} сек.</b>\n"
            f"Причина: <code>{err[:300]}</code>\n\n"
            "Основной бот автоматически использует стабильный локальный поиск даты."
        )

    workers = list(status.get("workers") or [])
    expected_replicas = int(status.get("expected_replicas", 0) or 0)
    lines = [
        "<b>📅 DATE MANAGER</b>",
        "",
        f"Статус: <b>🟢 online</b> · workers: <b>{len(workers)}</b> / ожидается <b>{expected_replicas or len(workers)}</b>",
        f"Redis queue: <b>{int(status.get('queue_depth', 0) or 0)}</b> · активных: <b>{int(status.get('active_total', 0) or 0)}</b>",
        f"Date cache: <b>{int(status.get('cache_ttl', 180) or 180)} сек.</b> · Predictor: <b>{int(status.get('predictor_ttl', 3600) or 3600) // 60} мин.</b> · окно дат: <b>{int(status.get('max_age_days', DATE_MAX_AGE_DAYS) or DATE_MAX_AGE_DAYS) + 1} дней</b>",
        f"Общая скорость: <b>{float(status.get('rate_total', 0.0) or 0.0):.2f} probes/sec</b>",
        f"Проб отправлено: <b>{int(status.get('probes_queued_total', 0) or 0)}</b> · cache hits: <b>{int(status.get('cache_hits_total', 0) or 0)}</b>",
        f"Predictor hits/miss: <b>{int(status.get('predictor_hits_total', 0) or 0)}/{int(status.get('predictor_misses_total', 0) or 0)}</b> · подтверждений: <b>{int(status.get('predictor_writes_total', 0) or 0)}</b>",
        f"Continue search: <b>{int(status.get('predictor_continue_success_total', 0) or 0)}/{int(status.get('predictor_continue_total', 0) or 0)}</b> успешных продолжений от hint",
        f"Cold Date Turbo: <b>{'ВКЛ' if status.get('cold_turbo_enabled') else 'ВЫКЛ'}</b> · запусков: <b>{int(status.get('cold_turbo_total', 0) or 0)}</b> · сразу глубже 50: <b>{int(status.get('cold_turbo_public_beyond_total', 0) or 0)}</b>",
    ]
    last_probes = int(status.get("last_batch_probes", 0) or 0)
    if last_probes:
        lines.extend([
            "",
            f"Последний поиск: <b>{html.escape(str(status.get('last_target_date') or '—'))}</b> · boundary ≈ <b>{int(status.get('last_boundary', 0) or 0) or '—'}</b>",
            f"Проб: <b>{last_probes}</b> · из кэша: <b>{int(status.get('last_batch_cached', 0) or 0)}</b> · отправлено: <b>{int(status.get('last_batch_queued', 0) or 0)}</b>",
            f"Время remote-поиска: <b>{float(status.get('last_batch_seconds', 0.0) or 0.0):.2f} сек.</b>",
            f"Predictor: <b>{html.escape(str(status.get('last_predictor_source') or 'cold'))}</b> → page <b>{int(status.get('last_predictor_page', 0) or 0) or '—'}</b> · learned points: <b>{int(status.get('last_predictor_points', 0) or 0)}</b>",
            f"Продолжение от hint: rounds <b>{int(status.get('last_predictor_continue_rounds', 0) or 0)}</b> · новых pages <b>{int(status.get('last_predictor_continue_pages', 0) or 0)}</b> · полный fallback <b>{'ДА' if status.get('last_predictor_fallback') else 'нет'}</b>",
            f"Cold turbo last: <b>{'ДА' if status.get('last_cold_turbo') else 'нет'}</b> · age <b>{int(status.get('last_cold_age_days', 0) or 0)}</b> · grid <code>{html.escape(str(status.get('last_cold_probe_pages') or []))}</code>",
        ])
    for idx, worker in enumerate(workers[:4], start=1):
        lines.extend([
            "",
            f"<b>Worker {idx}</b> · v<b>{html.escape(str(worker.get('version') or '—'))}</b> · concurrency <b>{int(worker.get('concurrency', 0) or 0)}</b> · active <b>{int(worker.get('active', 0) or 0)}</b>",
            f"Fleet: <b>{html.escape(str(worker.get('fleet_bucket') or '—'))}</b> · scan/browser/global <b>{int(worker.get('fleet_scan_limit', 0) or 0)}/{int(worker.get('fleet_browser_limit', 0) or 0)}/{int(worker.get('fleet_global_limit', 0) or 0)}</b>",
            f"probes: <b>{int(worker.get('processed', 0) or 0)}</b> · speed <b>{float(worker.get('rate_ema', 0.0) or 0.0):.2f}/s</b>",
            f"HTTP fast: <b>{int(worker.get('http_fast_ok', 0) or 0)}</b> · browser confirm: <b>{int(worker.get('browser_confirm_ok', 0) or 0)}/{int(worker.get('browser_confirms', 0) or 0)}</b>",
            f"403/429: <b>{int(worker.get('http_403', 0) or 0)}/{int(worker.get('http_429', 0) or 0)}</b> · conflicts: <b>{int(worker.get('transport_conflicts', 0) or 0)}</b> · errors <b>{int(worker.get('errors', 0) or 0)}</b>",
            f"Ожидание: traffic <b>{float(worker.get('traffic_wait_ms_avg', 0.0) or 0.0):.0f} ms</b> · Redis limiter <b>{float(worker.get('redis_wait_ms_avg', 0.0) or 0.0):.0f} ms</b>",
        ])
    lines.extend([
        "",
        "<i>Predictor хранит только локально подтверждённые границы. Если граница уехала, поиск расширяется от hint, а не начинается заново с page 1. Финальная граница всегда перепроверяется стабильным локальным parser.</i>",
    ])
    return "\n".join(lines)


async def _admin_page_worker_text() -> str:
    status = await REMOTE_PAGE_MANAGER.status()
    if not status.get("enabled"):
        return (
            "<b>📄 Page Worker</b>\n\n"
            "Статус: <b>▫️ выключен</b>\n"
            "REDIS_URL не задан или REMOTE_PAGE_WORKER_ENABLED=0."
        )
    if not status.get("alive"):
        err = html.escape(str(status.get("error") or "heartbeat не найден"))
        return (
            "<b>📄 Page Worker</b>\n\n"
            "Статус: <b>🔴 offline</b>\n"
            f"Redis queue: <b>{int(status.get('queue_depth', 0) or 0)}</b>\n"
            f"Кэш страниц: <b>{int(status.get('cache_ttl', 180) or 180)} сек.</b>\n"
            f"Причина: <code>{err[:300]}</code>\n\n"
            "Основной бот автоматически использует локальный сбор страниц."
        )

    workers = list(status.get("workers") or [])
    expected_replicas = int(status.get("expected_replicas", 0) or 0)
    lines = [
        "<b>📄 PAGE MANAGER</b>",
        "",
        f"Статус: <b>🟢 online</b> · workers: <b>{len(workers)}</b> / ожидается <b>{expected_replicas or len(workers)}</b>",
        f"Redis queue: <b>{int(status.get('queue_depth', 0) or 0)}</b> · активных: <b>{int(status.get('active_total', 0) or 0)}</b>",
        f"Кэш страниц: <b>{int(status.get('cache_ttl', 180) or 180)} сек.</b>",
        f"Streaming: <b>✅ ВКЛ</b> · Rolling: <b>{'✅' if status.get('rolling_prefetch') else '—'}</b> "
        f"окно <b>{int(status.get('prefetch_window_pages', 0) or 0)}</b> / low-water <b>{int(status.get('prefetch_low_water_pages', 0) or 0)}</b> · "
        f"ожидание следующей страницы ≤ <b>{int(status.get('cache_wait_ms', 0) or 0)} мс</b>",
        f"Общая скорость: <b>{float(status.get('rate_total', 0.0) or 0.0):.2f} pages/sec</b>",
        f"Обработано worker'ами: <b>{int(status.get('processed_total', 0) or 0)}</b> · ошибок: <b>{int(status.get('errors_total', 0) or 0)}</b>",
    ]
    last_pages = int(status.get("last_batch_pages", 0) or 0)
    if last_pages:
        lines.extend([
            "",
            f"Последний prefetch: <b>{last_pages}</b> страниц · workers: <b>{int(status.get('last_batch_workers', 0) or 0)}</b>",
            f"Из кэша: <b>{int(status.get('last_batch_cached', 0) or 0)}</b> · отправлено: <b>{int(status.get('last_batch_remote', 0) or 0)}</b> · fallback: <b>{int(status.get('last_batch_failed', 0) or 0)}</b>",
            f"Постановка batch в Redis: <b>{float(status.get('last_batch_seconds', 0.0) or 0.0):.2f} сек.</b>",
        ])
    for idx, worker in enumerate(workers[:4], start=1):
        lines.extend([
            "",
            f"<b>Worker {idx}</b> · v<b>{html.escape(str(worker.get('version') or '—'))}</b> · concurrency <b>{int(worker.get('concurrency', 0) or 0)}</b> · active <b>{int(worker.get('active', 0) or 0)}</b>",
            f"Fleet: <b>{html.escape(str(worker.get('fleet_bucket') or '—'))}</b> · scan/browser/global <b>{int(worker.get('fleet_scan_limit', 0) or 0)}/{int(worker.get('fleet_browser_limit', 0) or 0)}/{int(worker.get('fleet_global_limit', 0) or 0)}</b>",
            f"pages: <b>{int(worker.get('processed', 0) or 0)}</b> · cache-hit worker: <b>{int(worker.get('cache_served', 0) or 0)}</b> · speed <b>{float(worker.get('rate_ema', 0.0) or 0.0):.2f}/s</b>",
            f"403/429: <b>{int(worker.get('http_403', 0) or 0)}/{int(worker.get('http_429', 0) or 0)}</b> · penalty <b>{int(worker.get('penalty', 0) or 0)}</b>",
            f"Ожидание: traffic <b>{float(worker.get('traffic_wait_ms_avg', 0.0) or 0.0):.0f} ms</b> · Redis limiter <b>{float(worker.get('redis_wait_ms_avg', 0.0) or 0.0):.0f} ms</b>",
        ])
    lines.extend([
        "",
        "<i>Page Worker ускоряет сбор страниц после того, как Date Worker/локальный fallback подтвердил границу даты.</i>",
    ])
    return "\n".join(lines)


# ---- v4.22.4 Admin-only Vinted Lab -------------------------------------------------
# Vinted is intentionally isolated from the Kleinanzeigen user parser. The Telegram
# bot only orchestrates Vinted-specific Redis streams/tables and renders diagnostics.
VINTED_ADMIN_MAX_CATEGORIES = 24
_VINTED_ADMIN_CFG: dict[int, dict[str, Any]] = {}
_VINTED_ADMIN_WATCHERS: dict[int, asyncio.Task] = {}


def _vinted_cfg(user_id: int) -> dict[str, Any]:
    cfg = _VINTED_ADMIN_CFG.get(int(user_id))
    if cfg is None:
        cfg = {"selected": {}, "pages": 3, "mode": "manual", "node": 0}
        _VINTED_ADMIN_CFG[int(user_id)] = cfg
    return cfg


def _vinted_status_label(status: str) -> str:
    return {
        "queued": "⏳ В очереди",
        "running": "🟣 Сканирование",
        "metrics": "👁 Метрики",
        "cancel_requested": "⏹ Останавливается",
        "cancelled": "⏹ Остановлен",
        "completed": "✅ Завершён",
        "partial": "⚠️ Завершён частично",
        "failed": "❌ Ошибка",
    }.get(str(status or ""), html.escape(str(status or "—")))


def _vinted_mode_label(mode: str) -> str:
    return "📡 Vinted Radar 1.0" if str(mode) == "radar" else "🔎 Vinted Parser"


def _vinted_terminal(status: str) -> bool:
    return str(status or "") in {"completed", "partial", "failed", "cancelled"}


def _vinted_home_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📡 Vinted Radar 1.0", callback_data="av:radar:all:0")],
        [InlineKeyboardButton(text="⚙️ Настроить Radar", callback_data="av:new:r")],
        [InlineKeyboardButton(text="🔎 Новый Vinted-скан", callback_data="av:new:m")],
        [InlineKeyboardButton(text="📂 История сканов", callback_data="av:history")],
        [InlineKeyboardButton(text="🔐 Vinted Session", callback_data="av:session")],
        [InlineKeyboardButton(text="⚙️ Vinted Workers", callback_data="av:workers")],
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="av:home")],
        [InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")],
    ])


async def _vinted_home_text() -> str:
    worker_status, scans, radar_snapshot = await asyncio.gather(
        VINTED_QUEUE.worker_status(), list_vinted_scans(1), build_vinted_radar_snapshot()
    )
    scan_workers = len(list(worker_status.get("scan_workers") or []))
    metrics_workers = len(list(worker_status.get("metrics_workers") or []))
    lines = [
        "<b>🟣 Vinted Lab</b>",
        "<i>Закрытый тестовый контур. Kleinanzeigen Page/Date/View Worker здесь не используются.</i>",
        "",
        f"Scan Worker: <b>{'🟢' if scan_workers else '🔴'} {scan_workers}/2</b>",
        f"Metrics Worker: <b>{'🟢' if metrics_workers else '🔴'} {metrics_workers}/2</b>",
        f"Очередь: scan <b>{int(worker_status.get('scan_queue', 0) or 0)}</b> · metrics <b>{int(worker_status.get('metrics_queue', 0) or 0)}</b>",
        "",
        f"📡 Radar AutoScan: <b>{'🟢 ВКЛ' if radar_snapshot.auto_enabled else '⏸ ВЫКЛ'}</b> · Live <b>{radar_snapshot.live_total}</b> · HOT <b>{radar_snapshot.hot}</b> · Rising <b>{radar_snapshot.rising}</b>",
        "",
        "<b>Режимы</b>",
        "🔎 Parser — разовый проход выбранных категорий.",
        f"📡 Radar 1.0 — Catalog Likes + повторные замеры каждые {VINTED_RADAR_INTERVAL_MINUTES} мин.; Live-окно {VINTED_RADAR_LIVE_HOURS} ч.",
    ]
    if worker_status.get("error"):
        lines.extend(["", f"⚠️ Redis: <code>{html.escape(str(worker_status['error'])[:180])}</code>"])
    if scans:
        scan = scans[0]
        latest_snapshot = await vinted_scan_progress(scan.id)
        like_stats = dict((latest_snapshot or {}).get("catalog_likes") or {})
        lines.extend([
            "",
            "<b>Последний запуск</b>",
            f"{_vinted_mode_label(scan.mode)} · {_vinted_status_label(scan.status)}",
            f"Категории: <b>{int(scan.completed_categories or 0)}/{int(scan.total_categories or 0)}</b> · товаров: <b>{int(scan.total_items or 0)}</b>",
            f"❤️ Catalog likes: <b>{int(like_stats.get('total', 0) or 0)}</b> · с лайками: <b>{int(like_stats.get('nonzero', 0) or 0)}</b>/{int(like_stats.get('known', 0) or 0)}",
        ])
        if scan.mode != "radar":
            lines.append(f"Exact views: <b>{int(scan.exact_views or 0)}</b> · chronology: <b>{int(scan.chronology_count or 0)}</b>")
    return "\n".join(lines)


async def _vinted_tree_context() -> tuple[list[dict[str, Any]], dict[int, dict[str, Any]], str]:
    roots, source = await fetch_vinted_catalog_tree()
    return roots, flatten_vinted_catalog_tree(roots), source


def _vinted_setup_header(cfg: dict[str, Any], source: str) -> str:
    selected = dict(cfg.get("selected") or {})
    mode = str(cfg.get("mode") or "manual")
    selected_names = [html.escape(str(name)) for name in selected.values()]
    selected_line = ""
    if selected_names:
        preview = ", ".join(selected_names[:5])
        if len(selected_names) > 5:
            preview += f" … +{len(selected_names) - 5}"
        selected_line = f"\nВыбрано: <i>{preview}</i>"
    radar_note = (
        f"\n\n<i>Radar 1.0 после запуска будет повторять эти категории каждые {VINTED_RADAR_INTERVAL_MINUTES} мин. "
        f"Товар участвует в Live {VINTED_RADAR_LIVE_HOURS} ч от первого обнаружения.</i>"
        if mode == "radar" else ""
    )
    return (
        f"<b>{_vinted_mode_label(mode)}</b>\n\n"
        f"Категории: <b>{len(selected)}/{VINTED_ADMIN_MAX_CATEGORIES}</b> · глубина: <b>{int(cfg.get('pages') or 3)} стр.</b>\n"
        f"Каталог Vinted: <b>{'live' if str(source).startswith('live') else 'fallback'}</b>{selected_line}\n\n"
        "Выбирай разделы. Можно зайти внутрь дерева или выбрать весь текущий раздел."
        + radar_note
    )


def _vinted_pages_row(cfg: dict[str, Any]) -> list[InlineKeyboardButton]:
    current = int(cfg.get("pages") or 3)
    return [InlineKeyboardButton(text=("✅ " if current == value else "") + f"{value} стр.", callback_data=f"av:pg:{value}") for value in (1, 3, 5, 10)]


async def _vinted_category_screen(user_id: int, node_id: int = 0, page: int = 0) -> tuple[str, InlineKeyboardMarkup]:
    roots, flat, source = await _vinted_tree_context()
    cfg = _vinted_cfg(user_id)
    cfg["node"] = int(node_id or 0)
    selected: dict[int, str] = dict(cfg.get("selected") or {})
    page = max(0, int(page or 0))

    if node_id and node_id in flat:
        node = flat[node_id]
        children = list(node.get("catalogs") or [])
        title = str(node.get("title") or f"Catalog {node_id}")
        parent_id = int(node.get("parent_id") or 0)
        text = _vinted_setup_header(cfg, source) + f"\n\n<b>📁 {html.escape(title)}</b>"
    else:
        children = roots
        title = "Категории Vinted"
        parent_id = 0
        node_id = 0
        text = _vinted_setup_header(cfg, source) + "\n\n<b>📂 Категории Vinted</b>"

    rows: list[list[InlineKeyboardButton]] = []
    if node_id:
        is_selected = int(node_id) in selected
        rows.append([InlineKeyboardButton(
            text=("✅ Убрать этот раздел" if is_selected else "➕ Выбрать этот раздел"),
            callback_data=f"av:pick:{int(node_id)}",
        )])

    per_page = 12
    start = page * per_page
    chunk = children[start:start + per_page]
    for child in chunk:
        cid = int(child.get("id") or 0)
        if cid <= 0:
            continue
        name = str(child.get("title") or f"Catalog {cid}")
        has_children = bool(child.get("catalogs"))
        mark = "✅ " if cid in selected else ""
        icon = "📁" if has_children else "▫️"
        callback_data = f"av:cat:{cid}:0" if has_children else f"av:pick:{cid}"
        rows.append([InlineKeyboardButton(text=f"{mark}{icon} {name}"[:60], callback_data=callback_data)])

    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"av:cat:{int(node_id)}:{page - 1}"))
    if start + per_page < len(children):
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"av:cat:{int(node_id)}:{page + 1}"))
    if nav:
        rows.append(nav)

    rows.append(_vinted_pages_row(cfg))
    if selected:
        start_label = (f"📡 Запустить Radar · {len(selected)} кат." if str(cfg.get("mode") or "manual") == "radar" else f"▶️ Запустить · {len(selected)} кат.")
        rows.append([InlineKeyboardButton(text=start_label, callback_data="av:start")])
        rows.append([InlineKeyboardButton(text="🧹 Очистить выбор", callback_data="av:clear")])
    if node_id:
        rows.append([InlineKeyboardButton(text="⬅️ Уровень выше", callback_data=f"av:cat:{parent_id}:0")])
    else:
        rows.append([InlineKeyboardButton(text="⬅️ Vinted Lab", callback_data="av:home")])
    return text, InlineKeyboardMarkup(inline_keyboard=rows)


async def _vinted_toggle_catalog(user_id: int, catalog_id: int) -> str | None:
    _roots, flat, _source = await _vinted_tree_context()
    node = flat.get(int(catalog_id))
    if not node:
        return "Категория больше не найдена в текущем каталоге Vinted."
    cfg = _vinted_cfg(user_id)
    selected: dict[int, str] = dict(cfg.get("selected") or {})
    if int(catalog_id) in selected:
        selected.pop(int(catalog_id), None)
    else:
        if len(selected) >= VINTED_ADMIN_MAX_CATEGORIES:
            return f"Для теста максимум {VINTED_ADMIN_MAX_CATEGORIES} категории за один запуск."
        selected[int(catalog_id)] = str(node.get("title") or f"Catalog {catalog_id}")
    cfg["selected"] = selected
    return None


def _vinted_scan_keyboard(scan: VintedScan) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = [
        [InlineKeyboardButton(text="🔄 Обновить", callback_data=f"av:scan:{scan.id}"),
         InlineKeyboardButton(text="📦 Результаты", callback_data=f"av:res:{scan.id}:0")],
    ]
    if not _vinted_terminal(scan.status) and scan.status != "cancel_requested":
        rows.append([InlineKeyboardButton(text="⏹ Остановить", callback_data=f"av:stop:{scan.id}")])
    rows.append([InlineKeyboardButton(text="⬅️ Vinted Lab", callback_data="av:home")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _vinted_scan_text(scan_id: int) -> tuple[str, VintedScan | None]:
    snapshot, worker_status = await asyncio.gather(vinted_scan_progress(scan_id), VINTED_QUEUE.worker_status())
    if not snapshot:
        return "<b>Vinted scan не найден.</b>", None
    scan: VintedScan = snapshot["scan"]
    cats: list[VintedScanCategory] = list(snapshot["categories"])
    scan_pct = float(snapshot.get("scan_percent") or 0.0)
    metric_pct = float(snapshot.get("metrics_percent") or 0.0)
    metrics_workers = list(worker_status.get("metrics_workers") or [])
    provider_states = [str(w.get("provider_status") or "") for w in metrics_workers]
    if any(state == "ready" for state in provider_states):
        provider_line = "🟢 browser-session ready"
    elif any(state in {"blocked", "challenge", "expired", "session_expired", "circuit_open"} for state in provider_states):
        provider_line = "🔴 session blocked/expired"
    elif metrics_workers and all(state == "session_missing" for state in provider_states):
        provider_line = "🟠 Vinted Session не задана"
    elif metrics_workers:
        provider_line = "🟡 " + html.escape(", ".join(sorted(set(filter(None, provider_states))) or ["starting"])[:80])
    else:
        provider_line = "🔴 Metrics Worker offline"
    unknown_views = max(0, int(scan.metrics_done or 0) - int(scan.exact_views or 0))
    like_stats = dict(snapshot.get("catalog_likes") or {})
    lines = [
        f"<b>{_vinted_mode_label(scan.mode)}</b>",
        f"Статус: <b>{_vinted_status_label(scan.status)}</b>",
        "",
        "<b>📄 Каталоги</b>",
        f"{_progress_bar(int(scan_pct))} <b>{scan_pct:.1f}%</b>",
        f"Категории: <b>{int(scan.completed_categories or 0)}/{int(scan.total_categories or 0)}</b>",
        f"Найдено уникальных: <b>{int(scan.total_items or 0)}</b>",
        "",
        "<b>❤️ Likes из каталога</b>",
        f"Считано: <b>{int(like_stats.get('known', 0) or 0)}/{int(like_stats.get('items', 0) or 0)}</b> · с лайками: <b>{int(like_stats.get('nonzero', 0) or 0)}</b>",
        f"Всего ❤️: <b>{int(like_stats.get('total', 0) or 0)}</b> · максимум у товара: <b>{int(like_stats.get('max', 0) or 0)}</b>",
    ]
    if scan.mode == "radar":
        lines.extend([
            "",
            "<b>📡 Radar 1.0</b>",
            f"Этот круг — один snapshot. Повтор: каждые <b>{VINTED_RADAR_INTERVAL_MINUTES} мин.</b>",
            f"Live-окно товара: <b>{VINTED_RADAR_LIVE_HOURS} ч</b> от первого обнаружения.",
            "<i>Views/detail API для Radar не используются. Первый замер ❤️ — baseline; Rising/Hot появляются только после подтверждённого роста.</i>",
        ])
    else:
        lines.extend([
            "",
            "<b>👁 Exact Metrics</b>",
            f"{_progress_bar(int(metric_pct))} <b>{metric_pct:.1f}%</b>",
            f"Проверено: <b>{int(scan.metrics_done or 0)}/{int(scan.metrics_total or 0)}</b>",
            f"✅ Exact views: <b>{int(scan.exact_views or 0)}</b> · ❓ views UNKNOWN: <b>{unknown_views}</b>",
            f"🕒 Chronology: <b>{int(scan.chronology_count or 0)}</b> · detail likes: <b>{int(scan.exact_favourites or 0)}</b>",
            f"Provider: <b>{provider_line}</b>",
        ])
    if cats:
        lines.extend(["", "<b>Категории</b>"])
        for row in cats[:8]:
            icon = {"queued": "▫️", "running": "🟣", "completed": "✅", "partial": "⚠️", "failed": "❌", "cancelled": "⏹"}.get(row.status, "▫️")
            lines.append(
                f"{icon} {html.escape(row.category_name[:28])} · стр. <b>{int(row.pages_fetched or 0)}/{int(row.pages_target or 0)}</b> · <b>{int(row.unique_items or 0)}</b>"
            )
        if len(cats) > 8:
            lines.append(f"… ещё {len(cats) - 8}")
    if scan.error_text:
        lines.extend(["", f"⚠️ <code>{html.escape(scan.error_text[:300])}</code>"])
    return "\n".join(lines), scan


async def _vinted_workers_text() -> str:
    status = await VINTED_QUEUE.worker_status()
    scan_workers = list(status.get("scan_workers") or [])
    metrics_workers = list(status.get("metrics_workers") or [])
    lines = [
        "<b>🟣 Vinted Workers</b>",
        "<i>Полностью отдельные очереди от Kleinanzeigen.</i>",
        "",
        f"Scan Worker: <b>{len(scan_workers)}/2</b> · очередь <b>{int(status.get('scan_queue', 0) or 0)}</b>",
        f"Metrics Worker: <b>{len(metrics_workers)}/2</b> · очередь <b>{int(status.get('metrics_queue', 0) or 0)}</b>",
    ]
    for label, workers in (("Scan", scan_workers), ("Metrics", metrics_workers)):
        for idx, worker in enumerate(sorted(workers, key=lambda x: str(x.get("worker_id") or ""))[:4], start=1):
            active = int(worker.get("active", 0) or 0)
            lines.extend([
                "",
                f"<b>{label} {idx}</b> · {'🟢 busy' if active else '🟢 idle'} · v{html.escape(str(worker.get('version') or '—'))}",
                f"ID: <code>{html.escape(str(worker.get('worker_id') or '—')[:42])}</code>",
            ])
            if label == "Scan":
                lines.append(f"Категорий: <b>{int(worker.get('processed_categories', 0) or 0)}</b> · сейчас: <b>{html.escape(str(worker.get('category') or '—')[:28])}</b> · page <b>{int(worker.get('page', 0) or 0)}</b>")
            else:
                lines.append(f"Items: <b>{int(worker.get('processed', 0) or 0)}</b> · exact <b>{int(worker.get('exact_total', 0) or 0)}</b> · unknown <b>{int(worker.get('unknown_total', 0) or 0)}</b> · errors <b>{int(worker.get('errors', 0) or 0)}</b>")
                lines.append(
                    f"Provider: <b>{html.escape(str(worker.get('provider_status') or '—'))}</b> · "
                    f"pool <b>{int(worker.get('concurrency', worker.get('provider_concurrency', 0)) or 0)}</b> · "
                    f"last <b>{int(worker.get('last_ms', 0) or 0)} ms</b>"
                )
                detail = str(worker.get('provider_detail') or '').strip()
                if detail:
                    lines.append(f"Detail: <code>{html.escape(detail[:120])}</code>")
    if status.get("error"):
        lines.extend(["", f"⚠️ <code>{html.escape(str(status['error'])[:250])}</code>"])
    return "\n".join(lines)


def _vinted_workers_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Обновить", callback_data="av:workers")],
        [InlineKeyboardButton(text="⬅️ Vinted Lab", callback_data="av:home")],
    ])


async def _vinted_session_screen() -> tuple[str, InlineKeyboardMarkup]:
    session_raw, meta, service, worker_status = await asyncio.gather(
        load_vinted_session_json(),
        load_vinted_session_meta(),
        get_session_service(),
        VINTED_QUEUE.worker_status(),
    )
    metrics_workers = list(worker_status.get("metrics_workers") or [])
    provider_states = sorted({str(w.get("provider_status") or "") for w in metrics_workers if w.get("provider_status")})
    sources = sorted({str(w.get("session_source") or "") for w in metrics_workers if w.get("session_source")})
    service_online = bool(service.get("online"))
    public_url = str(service.get("public_url") or "").strip()
    captured = str(meta.get("captured_at") or meta.get("updated_at") or "")
    if captured:
        captured = captured.replace("T", " ")[:19] + " UTC"
    lines = [
        "<b>🔐 Vinted Session</b>",
        "<i>Вход выполняется в твоём обычном Chrome с твоего интернет-соединения. Railway больше не открывает страницу Vinted.</i>",
        "",
        f"Session: <b>{'🟢 сохранена' if session_raw else '🟠 не настроена'}</b>",
    ]
    if session_raw:
        lines.append(
            f"Cookies: <b>{int(meta.get('cookie_count', 0) or 0)}</b> · access/refresh: "
            f"<b>{'✅' if meta.get('has_access_token_web') else '—'}/{'✅' if meta.get('has_refresh_token_web') else '—'}</b>"
        )
        if captured:
            lines.append(f"Сохранена: <b>{html.escape(captured)}</b>")
    lines.extend([
        "",
        f"Session Worker: <b>{'🟢 online' if service_online else '🔴 offline'}</b>",
    ])
    if service_online:
        lines.append(f"v<b>{html.escape(str(service.get('version') or '—'))}</b> · HTTPS: <b>{'✅' if public_url.startswith('https://') else '⚠️'}</b>")
        if not public_url.startswith("https://"):
            lines.append("⚠️ Railway → <b>Vinted Session Worker → Networking → Generate Domain</b>.")
    else:
        lines.append("Для входа нужен отдельный Railway-сервис <b>Vinted Session Worker</b> с публичным Domain.")
    if metrics_workers:
        lines.extend([
            "",
            f"Metrics Worker: <b>{len(metrics_workers)}/2</b> · provider: <b>{html.escape(', '.join(provider_states) or 'starting')}</b>",
            f"Источник сессии: <b>{html.escape(', '.join(sources) or '—')}</b>",
        ])
    lines.extend([
        "",
        "После локального входа Helper передаст только cookies авторизованной Vinted-сессии. Metrics Worker подхватят её автоматически примерно за 10–15 секунд — redeploy не нужен.",
    ])
    rows: list[list[InlineKeyboardButton]] = []
    if service_online and public_url.startswith("https://"):
        rows.append([InlineKeyboardButton(text="🌐 Войти через мой Chrome", callback_data="av:sessionnew")])
    rows.append([InlineKeyboardButton(text="🔄 Проверить", callback_data="av:session")])
    if session_raw:
        rows.append([InlineKeyboardButton(text="🗑 Удалить сессию", callback_data="av:sessionclearask")])
    rows.append([InlineKeyboardButton(text="⬅️ Vinted Lab", callback_data="av:home")])
    return "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=rows)


def _vinted_session_open_keyboard(url: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🌐 Открыть локальный вход", url=url)],
        [InlineKeyboardButton(text="🔄 Я вошёл · проверить", callback_data="av:session")],
        [InlineKeyboardButton(text="⬅️ Vinted Session", callback_data="av:session")],
    ])


async def _vinted_history_screen() -> tuple[str, InlineKeyboardMarkup]:
    scans = await list_vinted_scans(12)
    rows: list[list[InlineKeyboardButton]] = []
    lines = ["<b>📂 Vinted · история сканов</b>"]
    if not scans:
        lines.extend(["", "Запусков пока нет."])
    for scan in scans:
        created = scan.created_at.strftime("%d.%m %H:%M") if scan.created_at else "—"
        label = f"{'📡' if scan.mode == 'radar' else '🔎'} {created} · {int(scan.total_items or 0)} · {scan.status}"
        rows.append([InlineKeyboardButton(text=label[:60], callback_data=f"av:scan:{scan.id}")])
    rows.append([InlineKeyboardButton(text="⬅️ Vinted Lab", callback_data="av:home")])
    return "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=rows)


async def _vinted_result_screen(scan_id: int, offset: int) -> tuple[str, InlineKeyboardMarkup]:
    offset = max(0, int(offset or 0))
    items, total = await list_vinted_scan_items(scan_id, offset=offset, limit=1)
    if not items:
        return (
            "<b>📦 Результаты Vinted</b>\n\nПока нет сохранённых объявлений.",
            InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ К скану", callback_data=f"av:scan:{scan_id}")]]),
        )
    item: VintedScanItem = items[0]
    views = str(item.view_count) if item.view_count is not None else "UNKNOWN"
    catalog_fav = item.catalog_favourite_count
    detail_fav = item.favourite_count
    fav_value = detail_fav if detail_fav is not None else catalog_fav
    fav = str(fav_value) if fav_value is not None else "UNKNOWN"
    fav_source = "detail" if detail_fav is not None else ("catalog" if catalog_fav is not None else "UNKNOWN")
    like_delta = await vinted_catalog_like_delta(scan_id, item.item_id)
    delta_value = like_delta.get("delta")
    delta_text = "—" if delta_value is None else (f"+{int(delta_value)}" if int(delta_value) > 0 else str(int(delta_value)))
    price = "—" if item.price_amount is None else f"{float(item.price_amount):g} {html.escape(item.currency or 'EUR')}"
    lines = [
        f"<b>📦 Vinted · {offset + 1}/{total}</b>",
        "",
        f"<b>{html.escape(item.title or f'Item {item.item_id}')}</b>",
        f"Категория: {html.escape(item.category_name or '—')}",
        f"Цена: <b>{price}</b>",
        f"Бренд: <b>{html.escape(item.brand or '—')}</b> · размер: <b>{html.escape(item.size or '—')}</b>",
        "",
        f"👁 Exact views: <b>{views}</b>",
        f"❤️ Likes: <b>{fav}</b> · source: <b>{html.escape(fav_source)}</b> · Δ к прошлому скану: <b>{delta_text}</b>",
        f"🕒 Chronology: <b>{html.escape(str(item.upload_raw or 'UNKNOWN'))}</b>",
        f"Identity: <b>{'✅' if item.identity_ok else '▫️ UNKNOWN'}</b> · metric: <code>{html.escape(item.metric_outcome or item.metric_status or '—')}</code>",
    ]
    if item.promoted:
        lines.append("⚠️ Promoted: <b>да</b>")
    rows: list[list[InlineKeyboardButton]] = []
    if item.url:
        rows.append([InlineKeyboardButton(text="🔗 Открыть Vinted", url=item.url)])
    nav: list[InlineKeyboardButton] = []
    if offset > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"av:res:{scan_id}:{offset - 1}"))
    if offset + 1 < total:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"av:res:{scan_id}:{offset + 1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton(text="⬅️ К скану", callback_data=f"av:scan:{scan_id}")])
    return "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=rows)


def _vinted_radar_status_icon(status: str) -> str:
    return {
        "hot": "🔥 HOT",
        "rising": "📈 RISING",
        "deal": "💎 DEAL",
        "candidate": "👀 CANDIDATE",
        "baseline": "▫️ BASELINE",
    }.get(str(status or ""), "▫️ BASELINE")


def _vinted_radar_age(hours: float) -> str:
    minutes = max(0, int(float(hours or 0.0) * 60))
    if minutes < 60:
        return f"{minutes}м"
    h, m = divmod(minutes, 60)
    return f"{h}ч {m}м" if m else f"{h}ч"


def _vinted_radar_filter_name(value: str) -> str:
    return {
        "all": "Все сигналы",
        "hot": "HOT",
        "rising": "Rising",
        "deal": "Deals",
        "candidate": "Candidates",
    }.get(str(value or "all"), "Все сигналы")


async def _vinted_radar_screen(filter_name: str = "all", page: int = 0) -> tuple[str, InlineKeyboardMarkup]:
    filter_name = filter_name if filter_name in {"all", "hot", "rising", "deal", "candidate"} else "all"
    page = max(0, int(page or 0))
    snapshot = await build_vinted_radar_snapshot()
    entries = [e for e in snapshot.entries if e.status != "baseline"]
    if filter_name != "all":
        entries = [e for e in entries if e.status == filter_name]
    per_page = 5
    pages = max(1, (len(entries) + per_page - 1) // per_page)
    page = min(page, pages - 1)
    chunk = entries[page * per_page:(page + 1) * per_page]

    if snapshot.auto_enabled:
        auto_line = f"🟢 AutoScan: каждые <b>{int(snapshot.auto_interval_minutes)} мин.</b>"
        if snapshot.next_scan_at:
            seconds = max(0, int((snapshot.next_scan_at - datetime.utcnow()).total_seconds()))
            auto_line += f" · следующий примерно через <b>{max(0, seconds // 60)} мин.</b>"
    else:
        auto_line = "⏸ AutoScan: <b>выключен</b>"

    lines = [
        "<b>📡 Vinted Radar 1.0</b>",
        f"<i>Like Momentum · Live {VINTED_RADAR_LIVE_HOURS}ч · обучение {VINTED_RADAR_HISTORY_DAYS} дней</i>",
        "",
        auto_line,
        f"Категории: <b>{len(snapshot.categories)}</b> · глубина: <b>{int(snapshot.pages)} стр.</b>",
        "",
        f"🔥 HOT: <b>{snapshot.hot}</b> · 📈 Rising: <b>{snapshot.rising}</b> · 💎 Deals: <b>{snapshot.deals}</b>",
        f"👀 Candidates: <b>{snapshot.candidates}</b> · baseline: <b>{snapshot.baselines}</b>",
        f"Live товаров: <b>{snapshot.live_total}</b> · история: <b>{snapshot.history_items}</b>",
        "",
        f"<b>{html.escape(_vinted_radar_filter_name(filter_name))}</b> · стр. {page + 1}/{pages}",
    ]
    if not chunk:
        lines.extend([
            "",
            "Пока подтверждённых сигналов нет.",
            "<i>Первый замер ❤️ — только baseline. Для Rising/HOT нужен следующий Radar-круг и реальный рост лайков.</i>",
        ])
    else:
        for entry in chunk:
            delta = "—" if entry.like_delta is None else (f"+{entry.like_delta}" if entry.like_delta > 0 else str(entry.like_delta))
            velocity = "—" if entry.like_velocity is None else f"{entry.like_velocity:.2f}/ч"
            price = "—" if entry.price_amount is None else f"{entry.price_amount:g} {html.escape(entry.currency or 'EUR')}"
            edge = "—" if entry.price_edge_pct is None else f"{entry.price_edge_pct:+.0f}%"
            likes = "UNKNOWN" if entry.likes is None else str(entry.likes)
            lines.extend([
                "",
                f"{_vinted_radar_status_icon(entry.status)} · <b>{entry.score}/100</b>",
                f"<b>{html.escape((entry.title or f'Item {entry.item_id}')[:62])}</b>",
                f"💶 {price} · 💸 к рынку <b>{edge}</b>",
                f"❤️ <b>{likes}</b> · Δ <b>{delta}</b> · скорость <b>{velocity}</b>",
                f"⏱ В Radar: <b>{_vinted_radar_age(entry.age_hours)}</b> · бренд: <b>{html.escape(entry.brand or '—')}</b>",
            ])

    rows: list[list[InlineKeyboardButton]] = [
        [InlineKeyboardButton(text="🔥 HOT", callback_data="av:radar:hot:0"), InlineKeyboardButton(text="📈 Rising", callback_data="av:radar:rising:0")],
        [InlineKeyboardButton(text="💎 Deals", callback_data="av:radar:deal:0"), InlineKeyboardButton(text="👀 Candidates", callback_data="av:radar:candidate:0")],
        [InlineKeyboardButton(text="📡 Все", callback_data="av:radar:all:0"), InlineKeyboardButton(text="🔄 Обновить", callback_data=f"av:radar:{filter_name}:{page}")],
    ]
    for entry in chunk:
        rows.append([InlineKeyboardButton(text=f"{_vinted_radar_status_icon(entry.status).split()[0]} {entry.score} · {(entry.title or str(entry.item_id))[:42]}", callback_data=f"av:ri:{entry.item_id}")])
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"av:radar:{filter_name}:{page - 1}"))
    if page + 1 < pages:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"av:radar:{filter_name}:{page + 1}"))
    if nav:
        rows.append(nav)
    if snapshot.auto_enabled:
        rows.append([InlineKeyboardButton(text="⏸ Остановить Radar AutoScan", callback_data="av:radarstop")])
    else:
        rows.append([InlineKeyboardButton(text="⚙️ Настроить / запустить Radar", callback_data="av:new:r")])
    rows.append([InlineKeyboardButton(text="⬅️ Vinted Lab", callback_data="av:home")])
    return "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=rows)


async def _vinted_radar_item_screen(item_id: int) -> tuple[str, InlineKeyboardMarkup]:
    entry = await get_vinted_radar_entry(item_id)
    if entry is None:
        return (
            "<b>📡 Vinted Radar</b>\n\nТовар уже вышел из Live-окна или ещё не имеет Radar snapshot.",
            InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Radar", callback_data="av:radar:all:0")]]),
        )
    delta = "—" if entry.like_delta is None else (f"+{entry.like_delta}" if entry.like_delta > 0 else str(entry.like_delta))
    velocity = "—" if entry.like_velocity is None else f"{entry.like_velocity:.2f} ❤️/ч"
    accel = "—" if entry.acceleration is None else f"{entry.acceleration:+.2f} ❤️/ч²"
    price = "—" if entry.price_amount is None else f"{entry.price_amount:g} {html.escape(entry.currency or 'EUR')}"
    median = "—" if entry.price_median is None else f"{entry.price_median:g} {html.escape(entry.currency or 'EUR')}"
    edge = "—" if entry.price_edge_pct is None else f"{entry.price_edge_pct:+.1f}%"
    likes = "UNKNOWN" if entry.likes is None else str(entry.likes)
    pct = int(round(entry.like_percentile * 100)) if entry.like_percentile > 0 else 0
    c = entry.components
    lines = [
        f"<b>{_vinted_radar_status_icon(entry.status)} · {entry.score}/100</b>",
        "",
        f"<b>{html.escape(entry.title or f'Item {entry.item_id}')}</b>",
        f"Категория: <b>{html.escape(entry.category_name or '—')}</b>",
        f"Бренд: <b>{html.escape(entry.brand or '—')}</b> · размер: <b>{html.escape(entry.size or '—')}</b>",
        f"Состояние: <b>{html.escape(entry.condition or '—')}</b>",
        "",
        f"❤️ Likes: <b>{likes}</b> · Δ <b>{delta}</b>",
        f"🚀 Like Velocity: <b>{velocity}</b> · percentile <b>P{pct}</b>",
        f"⚡ Acceleration: <b>{accel}</b>",
        f"⏱ В Radar: <b>{_vinted_radar_age(entry.age_hours)}</b> · замеров: <b>{entry.sample_count}</b>",
        "",
        f"💶 Цена: <b>{price}</b> · медиана похожих: <b>{median}</b>",
        f"💸 Price Edge: <b>{edge}</b>",
        f"💎 Похожих по бренду в Live: <b>{entry.scarcity_count}</b>",
        f"👤 Объявлений этого продавца в Live: <b>{entry.seller_active_count}</b>",
        "",
        "<b>Score</b>",
        f"❤️ Velocity <b>{c.get('like_velocity', 0)}/35</b> · 🚀 accel <b>{c.get('acceleration', 0)}/15</b>",
        f"💸 price <b>{c.get('price_edge', 0)}/20</b> · ❤️ peers <b>{c.get('likes_vs_peers', 0)}/10</b>",
        f"💎 scarcity <b>{c.get('scarcity', 0)}/10</b> · 👤 seller <b>{c.get('seller', 0)}/5</b> · 🔥 brand <b>{c.get('brand_momentum', 0)}/5</b>",
    ]
    if entry.sample_count < 2:
        lines.extend(["", "<i>Первый замер — baseline. Demand-статус HOT/RISING появится только после подтверждённого роста ❤️.</i>"])
    rows: list[list[InlineKeyboardButton]] = []
    if entry.url:
        rows.append([InlineKeyboardButton(text="🔗 Открыть Vinted", url=entry.url)])
    rows.append([InlineKeyboardButton(text="⬅️ Radar", callback_data="av:radar:all:0")])
    return "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=rows)


async def _vinted_watch_scan(bot_obj: Bot, chat_id: int, message_id: int, scan_id: int) -> None:
    """Best-effort live percentage updates. Manual Refresh remains the durable fallback."""
    try:
        last_text = ""
        while True:
            await asyncio.sleep(4)
            text_value, scan = await _vinted_scan_text(scan_id)
            if scan is None:
                return
            if text_value != last_text:
                try:
                    await bot_obj.edit_message_text(
                        chat_id=chat_id,
                        message_id=message_id,
                        text=text_value,
                        parse_mode=ParseMode.HTML,
                        reply_markup=_vinted_scan_keyboard(scan),
                    )
                    last_text = text_value
                except TelegramBadRequest:
                    pass
                except Exception:
                    pass
            if _vinted_terminal(scan.status):
                return
    finally:
        current = _VINTED_ADMIN_WATCHERS.get(int(scan_id))
        if current is asyncio.current_task():
            _VINTED_ADMIN_WATCHERS.pop(int(scan_id), None)


async def _vinted_start_watcher(callback: CallbackQuery, scan_id: int) -> None:
    message = callback.message
    if not message or not getattr(message, "chat", None) or not getattr(message, "message_id", None):
        return
    old = _VINTED_ADMIN_WATCHERS.pop(int(scan_id), None)
    if old and not old.done():
        old.cancel()
    _VINTED_ADMIN_WATCHERS[int(scan_id)] = asyncio.create_task(
        _vinted_watch_scan(callback.bot, int(message.chat.id), int(message.message_id), int(scan_id))
    )


@dp.callback_query(F.data.startswith("av:"))
async def vinted_admin_lab_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    data = str(callback.data or "")
    parts = data.split(":")
    action = parts[1] if len(parts) > 1 else "home"
    await callback.answer()

    if action == "home":
        await _edit_or_answer(callback.message, await _vinted_home_text(), reply_markup=_vinted_home_keyboard())
        return

    if action == "session":
        text_value, keyboard = await _vinted_session_screen()
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "sessionnew":
        service = await get_session_service()
        public_url = str(service.get("public_url") or "").strip().rstrip("/")
        if not service.get("online") or not public_url.startswith("https://"):
            await callback.answer("Vinted Session Worker offline или для него не создан Railway Domain.", show_alert=True)
            return
        token = await create_session_ticket(callback.from_user.id, ttl_minutes=15)
        setup_url = f"{public_url}/setup#token={token}"
        text_value = (
            "<b>🔐 Локальный вход в Vinted</b>\n\n"
            "Ссылка одноразовая и действует <b>15 минут</b>. Она откроет страницу DT Session, а сам Vinted затем "
            "запустится <b>в твоём обычном Chrome</b> с твоего IP.\n\n"
            "При первом использовании нужно один раз установить DT Vinted Local Helper. После этого достаточно просто войти в Vinted — "
            "сессия сохранится автоматически.\n\n"
            "<i>Пароль и 2FA DT не получает. Передаются только cookies подтверждённой Vinted-сессии.</i>"
        )
        await _edit_or_answer(callback.message, text_value, reply_markup=_vinted_session_open_keyboard(setup_url))
        return

    if action == "sessionclearask":
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🗑 Да, удалить", callback_data="av:sessionclear")],
            [InlineKeyboardButton(text="⬅️ Отмена", callback_data="av:session")],
        ])
        await _edit_or_answer(callback.message, "<b>Удалить сохранённую Vinted Session?</b>\n\nMetrics Worker перейдут в UNKNOWN до нового входа.", reply_markup=keyboard)
        return

    if action == "sessionclear":
        await clear_vinted_session()
        await callback.answer("Vinted Session удалена", show_alert=True)
        text_value, keyboard = await _vinted_session_screen()
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "radar":
        filter_name = parts[2] if len(parts) > 2 else "all"
        page = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 0
        text_value, keyboard = await _vinted_radar_screen(filter_name, page)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "ri":
        item_id = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        text_value, keyboard = await _vinted_radar_item_screen(item_id)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "radarstop":
        await disable_vinted_radar()
        text_value, keyboard = await _vinted_radar_screen("all", 0)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "new":
        cfg = _vinted_cfg(callback.from_user.id)
        cfg["selected"] = {}
        cfg["pages"] = 3
        cfg["mode"] = "radar" if len(parts) > 2 and parts[2] == "r" else "manual"
        text_value, keyboard = await _vinted_category_screen(callback.from_user.id, 0, 0)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "cat":
        node_id = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        page = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 0
        text_value, keyboard = await _vinted_category_screen(callback.from_user.id, node_id, page)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "pick":
        catalog_id = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        error = await _vinted_toggle_catalog(callback.from_user.id, catalog_id)
        if error:
            await callback.answer(error, show_alert=True)
        cfg = _vinted_cfg(callback.from_user.id)
        text_value, keyboard = await _vinted_category_screen(callback.from_user.id, int(cfg.get("node") or 0), 0)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "pg":
        value = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 3
        _vinted_cfg(callback.from_user.id)["pages"] = value if value in {1, 3, 5, 10} else 3
        cfg = _vinted_cfg(callback.from_user.id)
        text_value, keyboard = await _vinted_category_screen(callback.from_user.id, int(cfg.get("node") or 0), 0)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "clear":
        cfg = _vinted_cfg(callback.from_user.id)
        cfg["selected"] = {}
        text_value, keyboard = await _vinted_category_screen(callback.from_user.id, int(cfg.get("node") or 0), 0)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "start":
        cfg = _vinted_cfg(callback.from_user.id)
        selected: dict[int, str] = dict(cfg.get("selected") or {})
        if not selected:
            await callback.answer("Сначала выбери хотя бы одну категорию.", show_alert=True)
            return
        if not VINTED_QUEUE.enabled:
            await callback.answer("Vinted Lab: REDIS_URL не подключён к Parser.", show_alert=True)
            return
        scan = await create_vinted_scan(
            admin_user_id=callback.from_user.id,
            mode=str(cfg.get("mode") or "manual"),
            categories=[(int(cid), str(name)) for cid, name in selected.items()],
            pages=int(cfg.get("pages") or 3),
        )
        enqueue_ok = False
        try:
            queued = await enqueue_vinted_scan(scan.id)
            enqueue_ok = queued > 0
        except Exception as exc:
            await cancel_vinted_scan(scan.id)
            await callback.answer(f"Не удалось поставить в очередь: {type(exc).__name__}", show_alert=True)
        if enqueue_ok and str(cfg.get("mode") or "manual") == "radar":
            await enable_vinted_radar(
                admin_user_id=callback.from_user.id,
                categories=[(int(cid), str(name)) for cid, name in selected.items()],
                pages=int(cfg.get("pages") or 3),
                initial_scan_id=scan.id,
                initial_scan_at=scan.created_at,
            )
        cfg["selected"] = {}
        text_value, fresh = await _vinted_scan_text(scan.id)
        fresh = fresh or scan
        await _edit_or_answer(callback.message, text_value, reply_markup=_vinted_scan_keyboard(fresh))
        await _vinted_start_watcher(callback, scan.id)
        return

    if action == "scan":
        scan_id = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        text_value, scan = await _vinted_scan_text(scan_id)
        keyboard = _vinted_scan_keyboard(scan) if scan else _vinted_home_keyboard()
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        if scan and not _vinted_terminal(scan.status):
            await _vinted_start_watcher(callback, scan.id)
        return

    if action == "stop":
        scan_id = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        await cancel_vinted_scan(scan_id)
        text_value, scan = await _vinted_scan_text(scan_id)
        await _edit_or_answer(callback.message, text_value, reply_markup=_vinted_scan_keyboard(scan) if scan else _vinted_home_keyboard())
        return

    if action == "res":
        scan_id = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        offset = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 0
        text_value, keyboard = await _vinted_result_screen(scan_id, offset)
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "history":
        text_value, keyboard = await _vinted_history_screen()
        await _edit_or_answer(callback.message, text_value, reply_markup=keyboard)
        return

    if action == "workers":
        await _edit_or_answer(callback.message, await _vinted_workers_text(), reply_markup=_vinted_workers_keyboard())
        return

    await _edit_or_answer(callback.message, await _vinted_home_text(), reply_markup=_vinted_home_keyboard())
async def vinted_radar_autoscan_scheduler() -> None:
    """Persistent Vinted Radar 1.0 cadence.

    The global category snapshot repeats hourly by default. Individual products
    participate in Live scoring only for their first 24h; older observations remain
    available to the seven-day learning/reference pool.
    """
    while True:
        try:
            scan = await maybe_start_vinted_radar_round()
            if scan is not None:
                log.info(
                    "Vinted Radar 1.0 AutoScan queued | scan=%s categories=%s pages=%s",
                    scan.id, scan.total_categories, scan.pages_per_category,
                )
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Vinted Radar 1.0 AutoScan scheduler error")
        await asyncio.sleep(30)


# ---- end Vinted Lab ---------------------------------------------------------------


@dp.message(Command("admin"))
async def admin_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not _is_admin(message.from_user.id):
        await message.answer("Нет доступа.")
        return
    await message.answer(await _admin_dashboard_text(), parse_mode=ParseMode.HTML, reply_markup=await _admin_live_keyboard())


@dp.callback_query(F.data == "adminhome")
async def admin_home_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(callback.message, await _admin_dashboard_text(), reply_markup=await _admin_live_keyboard())


@dp.callback_query(F.data == "adminstats")
async def admin_stats_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(callback.message, await _admin_dashboard_text(), reply_markup=await _admin_live_keyboard())


async def _admin_trial_text() -> tuple[str, bool]:
    stats, radar_funnel = await asyncio.gather(
        free_trial_stats(), free_radar_funnel_stats()
    )
    enabled = bool(stats["enabled"])
    used_one = int(stats["used_one"] or 0)
    used_all = int(stats["used_all"] or 0)
    converted = int(stats["converted"] or 0)
    conversion = (converted / used_one * 100.0) if used_one else 0.0
    radar_opened = int(radar_funnel.get("opened", 0) or 0)
    radar_upgrade = int(radar_funnel.get("upgrade_click", 0) or 0)
    radar_converted = int(radar_funnel.get("converted", 0) or 0)
    radar_conversion = (radar_converted / radar_opened * 100.0) if radar_opened else 0.0
    text = (
        "<b>🎁 Бесплатные сканы · стартовая акция</b>\n\n"
        f"Статус: <b>{'✅ ВКЛ' if enabled else '⏸ ВЫКЛ'}</b>\n"
        f"Лимит: <b>{FREE_TRIAL_SCAN_LIMIT} скана</b> на нового пользователя\n"
        f"Пробный запуск: <b>{FREE_TRIAL_MAX_CATEGORIES} категория · до {FREE_TRIAL_MAX_PAGES} страниц</b>\n"
        "Включено: реальные просмотры · TOP · XLSX\n"
        "По подписке: 50 страниц · несколько категорий · повторы · обновление просмотров · автозамеры\n\n"
        f"👥 Использовали ≥1: <b>{used_one}</b>\n"
        f"🎁 Использовали {FREE_TRIAL_SCAN_LIMIT}/{FREE_TRIAL_SCAN_LIMIT}: <b>{used_all}</b>\n"
        f"💳 Купили после пробника: <b>{converted}</b>\n"
        f"📈 Конверсия в оплату: <b>{conversion:.1f}%</b>\n\n"
        "<b>📡 Бесплатный DT Radar</b>\n"
        f"👥 Зашли посмотреть: <b>{radar_opened}</b>\n"
        f"💎 Нажали полный доступ: <b>{radar_upgrade}</b>\n"
        f"💳 Купили после Radar: <b>{radar_converted}</b>\n"
        f"📈 Radar → оплата: <b>{radar_conversion:.1f}%</b>"
    )
    return text, enabled


async def _admin_radar_funnel_text(page: int = 0) -> tuple[str, list[dict], int, int]:
    page = max(0, int(page or 0))
    day_since = datetime.utcnow() - timedelta(hours=24)
    all_stats, day_stats, visitor_result = await asyncio.gather(
        free_radar_funnel_stats(),
        free_radar_funnel_stats(day_since),
        free_radar_recent_visitors(page, 6),
    )
    visitors, total_visitors = visitor_result
    pages = max(1, (int(total_visitors) + 5) // 6)
    if page >= pages:
        page = pages - 1
        visitors, total_visitors = await free_radar_recent_visitors(page, 6)

    def conversion(stats: dict) -> float:
        opened = int(stats.get("opened", 0) or 0)
        converted = int(stats.get("converted", 0) or 0)
        return (converted / opened * 100.0) if opened else 0.0

    lines = [
        "<b>📡 Бесплатный DT Radar · воронка</b>",
        "",
        "<b>За 24 часа / за всё время</b>",
        f"👥 Открыли Radar: <b>{int(day_stats['opened'])} / {int(all_stats['opened'])}</b>",
        f"📨 Пришли из Daily Radar: <b>{int(day_stats.get('daily_digest_open', 0))} / {int(all_stats.get('daily_digest_open', 0))}</b>",
        f"🔥 Открыли «Лучшие сейчас»: <b>{int(day_stats['best'])} / {int(all_stats['best'])}</b>",
        f"📂 Выбрали режим: <b>{int(day_stats['mode_opened'])} / {int(all_stats['mode_opened'])}</b>",
        f"👁 Открыли хотя бы 1 товар: <b>{int(day_stats['viewed_item'])} / {int(all_stats['viewed_item'])}</b>",
        f"✅ Посмотрели все {FREE_RADAR_PREVIEW_LIMIT}: <b>{int(day_stats['completed_five'])} / {int(all_stats['completed_five'])}</b>",
        f"💎 Нажали полный доступ: <b>{int(day_stats['upgrade_click'])} / {int(all_stats['upgrade_click'])}</b>",
        f"💳 Купили после Radar: <b>{int(day_stats['converted'])} / {int(all_stats['converted'])}</b>",
        f"📈 Radar → оплата: <b>{conversion(day_stats):.1f}% / {conversion(all_stats):.1f}%</b>",
        "",
        "<b>👥 Последние посетители</b>",
    ]
    if not visitors:
        lines.append("Пока никто не открывал бесплатный Radar.")
    else:
        feature_labels = {
            "search": "Поиск", "categories": "Категории",
            "favorites": "Мой Radar", "records": "Рекорды",
        }
        for visitor in visitors:
            uid = int(visitor.get("user_id") or 0)
            username = str(visitor.get("username") or "").strip()
            first_name = str(visitor.get("first_name") or "").strip()
            label = f"@{username}" if username else (first_name or f"ID {uid}")
            products = visitor.get("preview_products") or {}
            hot = min(FREE_RADAR_PREVIEW_LIMIT, len(products.get("hot", set())))
            rising = min(FREE_RADAR_PREVIEW_LIMIT, len(products.get("rising", set())))
            locked = visitor.get("locked_features") or set()
            locked_text = ", ".join(feature_labels.get(x, x) for x in sorted(locked))
            trial_used = int(visitor.get("trial_scans_used", 0) or 0)
            converted_after_radar = bool(visitor.get("converted_after_radar"))
            last_at = visitor.get("last_event_at")
            lines.extend([
                "",
                f"👤 <b>{html.escape(label)}</b> · <code>{uid}</code>",
                f"🕐 {_utc_to_msk_text(last_at)} МСК · Radar ×{int(visitor.get('radar_opens', 0) or 0)} · Лучшие ×{int(visitor.get('best_opens', 0) or 0)}",
                f"🔥 {hot}/{FREE_RADAR_PREVIEW_LIMIT} · 🚀 {rising}/{FREE_RADAR_PREVIEW_LIMIT}",
                f"💎 Полный доступ: <b>{'нажал' if int(visitor.get('upgrade_clicks', 0) or 0) > 0 else '—'}</b> · "
                f"🎁 сканы {trial_used}/{FREE_TRIAL_SCAN_LIMIT} · 💳 после Radar: <b>{'купил' if converted_after_radar else '—'}</b>",
            ])
            if locked_text:
                lines.append(f"🔒 Интересовался: {html.escape(locked_text)}")
        if pages > 1:
            lines += ["", f"Страница <b>{page + 1}/{pages}</b> · посетителей <b>{total_visitors}</b>"]
    return "\n".join(lines), visitors, page, pages


@dp.callback_query(F.data == "admintrial")
async def admin_trial_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    text, enabled = await _admin_trial_text()
    await callback.answer()
    await _edit_or_answer(callback.message, text, reply_markup=admin_trial_keyboard(enabled))


@dp.callback_query(F.data == "admintrial:toggle")
async def admin_trial_toggle_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    enabled = await free_trial_enabled()
    await set_free_trial_enabled(not enabled)
    text, current = await _admin_trial_text()
    await callback.answer("Акция включена" if current else "Акция выключена")
    await _edit_or_answer(callback.message, text, reply_markup=admin_trial_keyboard(current))


async def _admin_referral_text() -> tuple[str, bool]:
    enabled, stats = await asyncio.gather(
        referral_promo_enabled(), referral_admin_stats()
    )
    text = (
        "<b>👥 Реферальная акция</b>\n\n"
        f"Статус: <b>{'🟢 ВКЛ' if enabled else '⏸ ВЫКЛ'}</b>\n"
        "Условие: <b>2 новых пользователя → +1 день</b>\n"
        "Механика повторяется: 4 → +2 дня, 6 → +3 дня и т.д.\n\n"
        f"🔗 Уникальных входов по реферальным ссылкам: <b>{int(stats.get('total', 0) or 0)}</b>\n"
        f"✅ Входов во время активной акции: <b>{int(stats.get('eligible', 0) or 0)}</b>\n"
        f"👤 Пользователей, которые приглашали: <b>{int(stats.get('referrers', 0) or 0)}</b>\n"
        f"💎 Выдано бонусных дней: <b>{int(stats.get('days_earned', 0) or 0)}</b>\n\n"
        "<i>Считается только первый вход нового Telegram-пользователя в бот. "
        "Один человек может быть засчитан только одному пригласившему. "
        "Когда акция выключена, новые входы сохраняются в общей статистике, но не участвуют в бонусе.</i>"
    )
    return text, bool(enabled)


@dp.callback_query(F.data == "adminreferral")
async def admin_referral_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    text, enabled = await _admin_referral_text()
    await callback.answer()
    await _edit_or_answer(callback.message, text, reply_markup=admin_referral_keyboard(enabled))


@dp.callback_query(F.data == "adminreferral:toggle")
async def admin_referral_toggle_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    current = await referral_promo_enabled()
    await set_referral_promo_enabled(not current)
    text, enabled = await _admin_referral_text()
    await callback.answer("Реферальная акция включена" if enabled else "Реферальная акция выключена")
    await _edit_or_answer(callback.message, text, reply_markup=admin_referral_keyboard(enabled))


@dp.callback_query(F.data.startswith("adminradarfunnel:"))
async def admin_radar_funnel_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        page = max(0, int(str(callback.data or "").split(":", 1)[1]))
    except Exception:
        page = 0
    text, visitors, page, pages = await _admin_radar_funnel_text(page)
    await callback.answer()
    await _edit_or_answer(
        callback.message, text,
        reply_markup=admin_radar_funnel_keyboard(visitors, page, pages),
    )


@dp.callback_query(F.data == "adminradarauto")
async def admin_radar_autoscan_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    # Live Status is deliberately lightweight: one AppSetting read, no Radar analytics.
    # This screen must never wait for percentile/category aggregation.
    try:
        text, state = await asyncio.wait_for(_radar_autoscan_text(), timeout=2.0)
        await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))
    except Exception as exc:
        log.exception("DT Radar live status failed")
        await _edit_or_answer(
            callback.message,
            "<b>📡 DT Radar 3.2 · ADAPTIVE LIVE</b>\n\n⚠️ Live Status временно недоступен. AutoScan продолжает работать в фоне.\n"
            f"Диагностика: <code>{html.escape(type(exc).__name__)}</code>",
            reply_markup=admin_radar_autoscan_loading_keyboard(),
        )


@dp.callback_query(F.data == "adminradarauto:analytics")
async def admin_radar_analytics_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "<b>📊 DT Radar 3.2 · ADAPTIVE ANALYTICS</b>\n\nСчитаю глубокую статистику… ⏳\nLive AutoScan от этого не блокируется.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Live Status", callback_data="adminradarauto")]
        ]),
    )
    text = await _radar3_analytics_text()
    await _edit_or_answer(
        callback.message, text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Обновить аналитику", callback_data="adminradarauto:analytics")],
            [InlineKeyboardButton(text="⬅️ Live Status", callback_data="adminradarauto")],
        ]),
    )


@dp.callback_query(F.data == "adminradarauto:start")
async def admin_radar_autoscan_start_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    async with _radar_autoscan_guard:
        state = await load_radar_autoscan_state()
        if state.get("status") == "running":
            await callback.answer("Круг уже идёт", show_alert=True)
            return
        state = _radar_autoscan_new_round(state, "manual")
        state = await save_radar_autoscan_state(state)
    _radar_autoscan_stop_event.clear()
    _radar_autoscan_wakeup.set()
    _kick_radar_autoscan(callback.bot, "manual-start")
    _schedule_radar_autoscan_launch_watchdog(callback.bot, str(state.get("round_id") or ""))
    running, queued = await _radar_foreground_counts()
    if running or queued:
        await callback.answer(f"Круг запущен · ждёт пользовательские сканы: {running + queued}")
    else:
        await callback.answer("Круг запущен · первая категория стартует сейчас")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data.startswith("adminradarauto:errors:"))
async def admin_radar_autoscan_errors_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        page = int(str(callback.data or "").rsplit(":", 1)[-1])
    except Exception:
        page = 0
    text, state, page, pages = await _radar_autoscan_errors_text(page)
    await callback.answer()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_errors_keyboard(state, page, pages))


@dp.callback_query(F.data == "adminradarauto:retry")
async def admin_radar_autoscan_retry_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    async with _radar_autoscan_guard:
        state = await load_radar_autoscan_state()
        if state.get("status") != "idle":
            await callback.answer("Сначала останови текущий круг", show_alert=True)
            return
        retry_state = _radar_autoscan_retry_round(state)
        if retry_state is None:
            await callback.answer("Нет сохранённого списка ошибок для повтора", show_alert=True)
            return
        state = await save_radar_autoscan_state(retry_state)
    _radar_autoscan_stop_event.clear()
    _radar_autoscan_wakeup.set()
    _kick_radar_autoscan(callback.bot, "retry")
    _schedule_radar_autoscan_launch_watchdog(callback.bot, str(state.get("round_id") or ""))
    await callback.answer(f"Повторяю проблемные категории: {len(state.get('category_keys') or [])}")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data == "adminradarauto:stop")
async def admin_radar_autoscan_stop_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    async with _radar_autoscan_guard:
        state = await load_radar_autoscan_state()
        if state.get("status") != "running":
            await callback.answer("Сейчас круг не идёт", show_alert=True)
            return
        # Persist paused immediately so a Railway restart cannot resurrect the
        # round, then signal the process-local waiter to cancel the current child.
        state["status"] = "paused"
        state["stop_requested"] = False
        state["waiting_for_users"] = False
        state["current_stage"] = "stopping"
        state = await save_radar_autoscan_state(state)
    _radar_autoscan_stop_event.set()
    _radar_autoscan_wakeup.set()
    await callback.answer("Останавливаю текущую категорию сейчас")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data == "adminradarauto:resume")
async def admin_radar_autoscan_resume_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    async with _radar_autoscan_guard:
        state = await load_radar_autoscan_state()
        if state.get("status") != "paused":
            await callback.answer("Нет остановленного круга", show_alert=True)
            return
        state["status"] = "running"
        state["stop_requested"] = False
        state["waiting_for_users"] = False
        state["current_stage"] = "starting"
        state = await save_radar_autoscan_state(state)
    _radar_autoscan_stop_event.clear()
    _radar_autoscan_wakeup.set()
    _kick_radar_autoscan(callback.bot, "resume")
    _schedule_radar_autoscan_launch_watchdog(callback.bot, str(state.get("round_id") or ""))
    await callback.answer("Круг продолжен")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data == "adminradarauto:daily")
async def admin_radar_autoscan_daily_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    async with _radar_autoscan_guard:
        state = await load_radar_autoscan_state()
        state["daily_enabled"] = not bool(state.get("daily_enabled"))
        state = await save_radar_autoscan_state(state)
    _radar_autoscan_wakeup.set()
    await callback.answer("Ежедневный круг включён" if state.get("daily_enabled") else "Ежедневный круг выключен")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data == "adminradarauto:skipday")
async def admin_radar_autoscan_skipday_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    async with _radar_autoscan_guard:
        state = await load_radar_autoscan_state()
        state["skip_daily_if_completed_today"] = not bool(state.get("skip_daily_if_completed_today", True))
        state = await save_radar_autoscan_state(state)
    _radar_autoscan_wakeup.set()
    await callback.answer("Настройка сохранена")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data == "adminradarauto:time")
async def admin_radar_autoscan_time_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    state = await load_radar_autoscan_state()
    current = str(state.get("daily_time") or RADAR_AUTOSCAN_DEFAULT_TIME)
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "<b>🕐 Время ежедневного DT Radar AutoScan</b>\n\nВыбери время запуска по Москве.",
        reply_markup=admin_radar_autoscan_time_keyboard(current),
    )


@dp.callback_query(F.data.startswith("adminradarauto:settime:"))
async def admin_radar_autoscan_settime_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    value = str(callback.data or "").removeprefix("adminradarauto:settime:")
    if value not in RADAR_AUTOSCAN_TIME_CHOICES:
        await callback.answer("Некорректное время", show_alert=True)
        return
    async with _radar_autoscan_guard:
        state = await load_radar_autoscan_state()
        state["daily_time"] = value
        state = await save_radar_autoscan_state(state)
    _radar_autoscan_wakeup.set()
    await callback.answer(f"Время: {value} МСК")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data == "adminradarauto:history")
async def admin_radar_autoscan_history_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        await _radar_autoscan_history_text(),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Radar AutoScan", callback_data="adminradarauto")]
        ]),
    )


@dp.callback_query(F.data == "adminworkers")
async def admin_workers_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        await _admin_workers_text(),
        reply_markup=admin_workers_keyboard(),
    )


@dp.callback_query(F.data == "adminactive")
async def admin_active_scans_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        await _admin_active_scans_text(),
        reply_markup=admin_active_scans_keyboard(),
    )


@dp.callback_query(F.data == "adminai")
async def admin_ai_handler(callback: CallbackQuery) -> None:
    """Compatibility redirect for old admin messages after legacy AI retirement."""
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer("Открываю DT Radar 3.0")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data.startswith("adminai:"))
async def admin_ai_section_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer("Раздел удалён · DT Radar 3.0")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data.startswith("aic:"))
async def admin_ai_candidate_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer("Старая AI-карточка удалена · DT Radar 3.0")
    text, state = await _radar_autoscan_text()
    await _edit_or_answer(callback.message, text, reply_markup=admin_radar_autoscan_keyboard(state))


@dp.callback_query(F.data == "adminviews")
async def admin_views_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        await _admin_view_worker_text(),
        reply_markup=admin_view_worker_keyboard(),
    )


@dp.callback_query(F.data == "admindates")
async def admin_dates_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        await _admin_date_worker_text(),
        reply_markup=admin_date_worker_keyboard(),
    )


@dp.callback_query(F.data == "adminpages")
async def admin_pages_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        await _admin_page_worker_text(),
        reply_markup=admin_page_worker_keyboard(),
    )


def _normalize_daily_digest_time(value: object) -> str | None:
    raw = str(value or "").strip().replace(".", ":")
    if not RADAR_DAILY_DIGEST_TIME_RE.fullmatch(raw):
        return None
    hour_text, minute_text = raw.split(":", 1)
    try:
        hour = int(hour_text)
        minute = int(minute_text)
    except ValueError:
        return None
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return f"{hour:02d}:{minute:02d}"


def _daily_digest_default_state(*, initialize: bool = False) -> dict:
    now = datetime.now(MOSCOW)
    hh, mm = [int(x) for x in RADAR_DAILY_DIGEST_DEFAULT_TIME.split(":", 1)]
    passed_default = now.time() >= now.replace(hour=hh, minute=mm, second=0, microsecond=0).time()
    return {
        "enabled": True,
        "time": RADAR_DAILY_DIGEST_DEFAULT_TIME,
        "last_sent_date": now.date().isoformat() if (initialize and passed_default) else "",
        "last_sent_at": "",
        "last_delivered": 0,
        "last_blocked": 0,
        "last_failed": 0,
        "last_metrics": {},
        "last_metrics_at": "",
    }


async def load_radar_daily_digest_state() -> dict:
    async with SessionLocal() as session:
        row = await session.get(AppSetting, RADAR_DAILY_DIGEST_SETTING_KEY)
        if row is None:
            state = _daily_digest_default_state(initialize=True)
            session.add(AppSetting(
                key=RADAR_DAILY_DIGEST_SETTING_KEY,
                value=json.dumps(state, ensure_ascii=False, separators=(",", ":")),
                updated_at=datetime.utcnow(),
            ))
            await session.commit()
            return state
        try:
            raw = json.loads(row.value or "{}")
            state = raw if isinstance(raw, dict) else {}
        except Exception:
            state = {}
    default = _daily_digest_default_state()
    default.update(state)
    default["enabled"] = bool(default.get("enabled", True))
    normalized_time = _normalize_daily_digest_time(default.get("time"))
    default["time"] = normalized_time or RADAR_DAILY_DIGEST_DEFAULT_TIME
    return default


async def save_radar_daily_digest_state(state: dict) -> dict:
    clean = _daily_digest_default_state()
    clean.update(dict(state or {}))
    clean["enabled"] = bool(clean.get("enabled", True))
    normalized_time = _normalize_daily_digest_time(clean.get("time"))
    clean["time"] = normalized_time or RADAR_DAILY_DIGEST_DEFAULT_TIME
    payload = json.dumps(clean, ensure_ascii=False, separators=(",", ":"))
    async with SessionLocal() as session:
        row = await session.get(AppSetting, RADAR_DAILY_DIGEST_SETTING_KEY)
        if row is None:
            row = AppSetting(key=RADAR_DAILY_DIGEST_SETTING_KEY, value=payload, updated_at=datetime.utcnow())
            session.add(row)
        else:
            row.value = payload
            row.updated_at = datetime.utcnow()
        await session.commit()
    return clean


def _daily_digest_due(state: dict, now: datetime | None = None) -> bool:
    now = now or datetime.now(MOSCOW)
    if not bool(state.get("enabled", True)):
        return False
    if str(state.get("last_sent_date") or "") == now.date().isoformat():
        return False
    try:
        hh, mm = [int(x) for x in str(state.get("time") or RADAR_DAILY_DIGEST_DEFAULT_TIME).split(":", 1)]
    except Exception:
        hh, mm = 20, 0
    return now >= now.replace(hour=hh, minute=mm, second=0, microsecond=0)


def _daily_digest_next_run_text(state: dict) -> str:
    if not bool(state.get("enabled", True)):
        return "выключена"
    now = datetime.now(MOSCOW)
    try:
        hh, mm = [int(x) for x in str(state.get("time") or RADAR_DAILY_DIGEST_DEFAULT_TIME).split(":", 1)]
    except Exception:
        hh, mm = 20, 0
    candidate = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
    if str(state.get("last_sent_date") or "") == now.date().isoformat() or candidate <= now:
        candidate += timedelta(days=1)
    return candidate.strftime("%d.%m %H:%M МСК")


def _daily_digest_msk_date(value: str) -> str:
    try:
        dt = datetime.fromisoformat(str(value or ""))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=MOSCOW)
        return dt.astimezone(MOSCOW).date().isoformat()
    except Exception:
        return ""


def _radar_daily_digest_zero_metrics() -> dict[str, int]:
    return {
        "listings_seen": 0,
        "new_listings": 0,
        "pages_verified": 0,
        "categories_processed": 0,
        "signals_today": 0,
        "products_today": 0,
        "best_score_today": 0,
        "radar_total": 0,
        "hot": 0,
        "rising": 0,
        "ai_picks": 0,
        "categories": 0,
    }


def _radar_daily_digest_cached_metrics(state: dict | None) -> dict[str, int]:
    result = _radar_daily_digest_zero_metrics()
    raw = (state or {}).get("last_metrics") or {}
    if isinstance(raw, dict):
        for key in result:
            try:
                result[key] = max(0, int(raw.get(key) or 0))
            except Exception:
                pass
    return result


async def _load_radar_daily_digest_state_bounded() -> tuple[dict, bool]:
    try:
        state = await asyncio.wait_for(
            load_radar_daily_digest_state(),
            timeout=RADAR_DAILY_DIGEST_STATE_TIMEOUT_SECONDS,
        )
        return state, True
    except Exception:
        log.exception("Daily Radar state load failed/timeout")
        return _daily_digest_default_state(), False


async def _radar_daily_digest_metrics_bounded(
    state: dict | None,
    *,
    timeout_seconds: float = RADAR_DAILY_DIGEST_UI_METRICS_TIMEOUT_SECONDS,
    cache_success: bool = True,
) -> tuple[dict[str, int], bool]:
    """Return live metrics without allowing an admin callback to hang indefinitely."""
    try:
        metrics = await asyncio.wait_for(
            radar_daily_digest_metrics(),
            timeout=max(1.0, float(timeout_seconds)),
        )
    except Exception:
        log.exception("Daily Radar metrics failed/timeout timeout=%.1fs", float(timeout_seconds))
        return _radar_daily_digest_cached_metrics(state), False

    clean = _radar_daily_digest_zero_metrics()
    for key in clean:
        try:
            clean[key] = max(0, int(metrics.get(key) or 0))
        except Exception:
            pass

    if cache_success and state is not None:
        try:
            state["last_metrics"] = clean
            state["last_metrics_at"] = datetime.now(MOSCOW).replace(microsecond=0).isoformat()
            await asyncio.wait_for(
                save_radar_daily_digest_state(state),
                timeout=RADAR_DAILY_DIGEST_STATE_TIMEOUT_SECONDS,
            )
        except Exception:
            # Metric caching is best-effort and must never break the admin UI.
            log.warning("Daily Radar metric cache save failed", exc_info=True)
    return clean, True


async def radar_daily_digest_metrics() -> dict[str, int]:
    """Factual same-day metrics used by the marketing digest."""
    now_msk = datetime.now(MOSCOW)
    today = now_msk.date().isoformat()
    start_msk = now_msk.replace(hour=0, minute=0, second=0, microsecond=0)
    start_utc_naive = start_msk.astimezone(timezone.utc).replace(tzinfo=None)

    autoscan = await load_radar_autoscan_state()
    listings_seen = 0
    new_listings = 0
    pages_verified = 0
    categories_processed = 0
    counted_rounds: set[str] = set()

    for summary in list(autoscan.get("history") or []):
        if not isinstance(summary, dict) or _daily_digest_msk_date(str(summary.get("finished_at") or "")) != today:
            continue
        rid = str(summary.get("round_id") or "")
        if rid and rid in counted_rounds:
            continue
        if rid:
            counted_rounds.add(rid)
        listings_seen += max(0, int(summary.get("listings_seen") or 0))
        new_listings += max(0, int(summary.get("new_listings") or 0))
        pages_verified += max(0, int(summary.get("pages_verified") or 0))
        categories_processed += max(0, int(summary.get("processed") or 0))

    if str(autoscan.get("status") or "") in {"running", "paused"} and _daily_digest_msk_date(str(autoscan.get("started_at") or "")) == today:
        rid = str(autoscan.get("round_id") or "")
        if not rid or rid not in counted_rounds:
            listings_seen += max(0, int(autoscan.get("listings_seen") or 0))
            new_listings += max(0, int(autoscan.get("new_listings") or 0))
            pages_verified += max(0, int(autoscan.get("pages_verified") or 0))
            categories_processed += max(0, int(autoscan.get("processed") or 0))

    async with SessionLocal() as session:
        signals_today = int((await session.execute(
            select(func.count(RadarSnapshot.id)).where(RadarSnapshot.recorded_at >= start_utc_naive)
        )).scalar_one() or 0)
        products_today = int((await session.execute(
            select(func.count(RadarProduct.id)).where(RadarProduct.first_radar_at >= start_utc_naive)
        )).scalar_one() or 0)
        best_score_today = int((await session.execute(
            select(func.max(RadarSnapshot.score)).where(RadarSnapshot.recorded_at >= start_utc_naive)
        )).scalar_one() or 0)

    stats = await radar_stats()
    return {
        "listings_seen": listings_seen,
        "new_listings": new_listings,
        "pages_verified": pages_verified,
        "categories_processed": categories_processed,
        "signals_today": signals_today,
        "products_today": products_today,
        "best_score_today": best_score_today,
        "radar_total": int(stats.total),
        "hot": int(stats.hot),
        "rising": int(stats.rising),
        "ai_picks": int(stats.ai_picks),
        "categories": int(stats.categories),
    }


def _digest_n(value: int) -> str:
    return f"{max(0, int(value or 0)):,}".replace(",", " ")


def radar_daily_digest_text(metrics: dict[str, int], *, paid: bool) -> str:
    lines = [
        "📡 <b>DT RADAR · СЕГОДНЯ</b>",
        "",
        "Пока Kleinanzeigen обновляется, DT Radar продолжает проверять рынок автоматически.",
        "",
    ]
    if int(metrics.get("listings_seen") or 0) > 0:
        lines.append(f"🔎 Проверено объявлений: <b>{_digest_n(metrics['listings_seen'])}</b>")
    if int(metrics.get("new_listings") or 0) > 0:
        lines.append(f"🆕 Новых объявлений в сканах: <b>{_digest_n(metrics['new_listings'])}</b>")
    if int(metrics.get("categories_processed") or 0) > 0:
        lines.append(f"🗂 Категорий обработано: <b>{_digest_n(metrics['categories_processed'])}</b>")
    lines += [
        f"📡 Radar-сигналов за сегодня: <b>+{_digest_n(metrics.get('signals_today', 0))}</b>",
        f"✨ Новых товаров в Radar: <b>+{_digest_n(metrics.get('products_today', 0))}</b>",
        "",
        f"🔥 Горячих сейчас: <b>{_digest_n(metrics.get('hot', 0))}</b>",
        f"🚀 Набирают: <b>{_digest_n(metrics.get('rising', 0))}</b>",
                f"📦 В базе Radar: <b>{_digest_n(metrics.get('radar_total', 0))}</b> товаров",
    ]
    if int(metrics.get("best_score_today") or 0) > 0:
        lines.append(f"🏆 Лучший DT Score сегодня: <b>{int(metrics['best_score_today'])}/100</b>")
    lines += [
        "",
        "Пока другие листают объявления вручную — Radar уже отбирает то, на что стоит обратить внимание.",
        "",
        "💎 Полный DT Radar уже открыт для тебя. Новые находки ждут внутри." if paid else
        f"🎁 Первые <b>{FREE_RADAR_PREVIEW_LIMIT}</b> реальных находок можно посмотреть бесплатно.",
    ]
    return "\n".join(lines)


def radar_daily_digest_keyboard(*, paid: bool) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text="📡 Открыть DT Radar", callback_data="radardaily_open")]]
    if not paid:
        rows.append([InlineKeyboardButton(text="💎 Открыть полный доступ", callback_data="radar_upgrade:daily")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _send_radar_daily_digest(bot: Bot, metrics: dict[str, int], recipients: list[int] | None = None) -> dict[str, int]:
    recipients = list(recipients) if recipients is not None else await _broadcast_recipient_ids()
    sent = blocked = failed = 0
    for uid in recipients:
        paid = allowed(int(uid))
        try:
            try:
                await bot.send_message(
                    int(uid), radar_daily_digest_text(metrics, paid=paid),
                    parse_mode=ParseMode.HTML, reply_markup=radar_daily_digest_keyboard(paid=paid),
                )
            except TelegramRetryAfter as exc:
                await asyncio.sleep(float(exc.retry_after) + 0.2)
                await bot.send_message(
                    int(uid), radar_daily_digest_text(metrics, paid=paid),
                    parse_mode=ParseMode.HTML, reply_markup=radar_daily_digest_keyboard(paid=paid),
                )
            sent += 1
        except TelegramForbiddenError:
            blocked += 1
        except TelegramBadRequest:
            failed += 1
        except Exception:
            failed += 1
            log.warning("Daily Radar digest delivery failed user=%s", uid, exc_info=True)
        await asyncio.sleep(0.055)
    return {"recipients": len(recipients), "sent": sent, "blocked": blocked, "failed": failed}


async def radar_daily_digest_scheduler(bot: Bot) -> None:
    """Send at most one factual Radar digest per Moscow calendar day."""
    await asyncio.sleep(5)
    while True:
        try:
            # Re-check due state only after acquiring the shared send lock so a manual
            # send and the scheduler cannot race into two deliveries.
            async with _RADAR_DAILY_DIGEST_SEND_LOCK:
                state = await load_radar_daily_digest_state()
                now = datetime.now(MOSCOW)
                if _daily_digest_due(state, now) and current_access_mode() == "subscription":
                    # v4.12.2: do not mark the day as sent until fresh metrics were
                    # obtained and the delivery loop actually completed. v4.12.1
                    # marked last_sent_date before aggregates, so a metric failure
                    # could silently suppress the whole day's digest.
                    metrics, metrics_fresh = await _radar_daily_digest_metrics_bounded(
                        state,
                        timeout_seconds=RADAR_DAILY_DIGEST_SEND_METRICS_TIMEOUT_SECONDS,
                    )
                    if not metrics_fresh:
                        log.warning("Daily Radar scheduler skipped: fresh metrics unavailable")
                    else:
                        result = await _send_radar_daily_digest(bot, metrics)
                        state = await load_radar_daily_digest_state()
                        state["last_sent_date"] = now.date().isoformat()
                        state["last_sent_at"] = now.replace(microsecond=0).isoformat()
                        state["last_delivered"] = int(result["sent"])
                        state["last_blocked"] = int(result["blocked"])
                        state["last_failed"] = int(result["failed"])
                        state["last_metrics"] = metrics
                        state["last_metrics_at"] = datetime.now(MOSCOW).replace(microsecond=0).isoformat()
                        await save_radar_daily_digest_state(state)
                        log.info(
                            "Daily Radar digest sent date=%s recipients=%s delivered=%s blocked=%s failed=%s listings=%s signals=%s",
                            now.date().isoformat(), result["recipients"], result["sent"], result["blocked"], result["failed"],
                            metrics.get("listings_seen", 0), metrics.get("signals_today", 0),
                        )
                        for admin_id in sorted(ADMIN_IDS):
                            try:
                                await bot.send_message(
                                    int(admin_id),
                                    "✅ <b>Daily Radar отправлен</b>\n\n"
                                    f"Доставлено: <b>{result['sent']}</b> / {result['recipients']}\n"
                                    f"Недоступны: <b>{result['blocked']}</b> · ошибки: <b>{result['failed']}</b>",
                                    parse_mode=ParseMode.HTML,
                                )
                            except Exception:
                                pass
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Daily Radar digest scheduler failed")
        await asyncio.sleep(RADAR_DAILY_DIGEST_POLL_SECONDS)


async def _broadcast_recipient_ids() -> list[int]:
    """All registered, non-banned bot users; expired subscribers are included."""
    async with SessionLocal() as session:
        rows = (await session.execute(
            select(BotUser.user_id)
            .where(BotUser.is_banned.is_(False))
            .order_by(BotUser.user_id.asc())
        )).scalars().all()
    return [int(uid) for uid in rows]


@dp.callback_query(F.data == "admindailyradar")
async def admin_daily_radar_handler(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    # v4.12.2: acknowledge before any database aggregates. In v4.12.1 slow/busy
    # Radar queries ran first, so Telegram looked as if the button did nothing.
    await callback.answer("Открываю Daily Radar…")
    await _edit_or_answer(
        callback.message,
        "<b>📨 Daily Radar</b>\n\n⏳ Загружаю живые цифры…",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="⬅️ Админ-панель", callback_data="adminhome")
        ]]),
    )
    log.info("Daily Radar admin panel open requested admin=%s", callback.from_user.id)
    digest_state, state_fresh = await _load_radar_daily_digest_state_bounded()
    metrics, metrics_fresh = await _radar_daily_digest_metrics_bounded(digest_state)
    status = "✅ ВКЛ" if digest_state.get("enabled", True) else "⏸ ВЫКЛ"
    last = str(digest_state.get("last_sent_at") or "—")
    if last != "—":
        try:
            last_dt = datetime.fromisoformat(last)
            if last_dt.tzinfo is None:
                last_dt = last_dt.replace(tzinfo=MOSCOW)
            last = last_dt.astimezone(MOSCOW).strftime("%d.%m.%Y %H:%M МСК")
        except Exception:
            pass
    freshness = ""
    if not state_fresh or not metrics_fresh:
        cached_at = str(digest_state.get("last_metrics_at") or "")
        suffix = f" · кэш {html.escape(cached_at)}" if cached_at else ""
        freshness = f"\n\n⚠️ <b>Живые цифры временно недоступны</b>{suffix}. Управление рассылкой работает; нажми «Обновить цифры» позже."
    text = (
        "<b>📨 Daily Radar</b>\n\n"
        "Ежедневная продающая сводка с живыми цифрами DT Radar. Отправляется всем зарегистрированным пользователям, кроме заблокированных.\n\n"
        f"Статус: <b>{status}</b>\n"
        f"Время: <b>{html.escape(str(digest_state.get('time') or RADAR_DAILY_DIGEST_DEFAULT_TIME))} МСК</b>\n"
        f"Следующая: <b>{html.escape(_daily_digest_next_run_text(digest_state))}</b>\n"
        f"Последняя: <b>{html.escape(last)}</b>\n"
        f"Доставлено в прошлый раз: <b>{int(digest_state.get('last_delivered') or 0)}</b>\n\n"
        "<b>Живые цифры для сегодняшнего поста:</b>\n"
        f"🔎 Проверено: <b>{_digest_n(metrics.get('listings_seen', 0))}</b>\n"
        f"📡 Сигналов сегодня: <b>+{_digest_n(metrics.get('signals_today', 0))}</b>\n"
        f"🔥 Горячих: <b>{_digest_n(metrics.get('hot', 0))}</b> · 🚀 Набирают: <b>{_digest_n(metrics.get('rising', 0))}</b>\n"
        f"📦 В базе Radar: <b>{_digest_n(metrics.get('radar_total', 0))}</b>"
        + freshness
    )
    await _edit_or_answer(callback.message, text, reply_markup=admin_daily_radar_keyboard(digest_state))


@dp.callback_query(F.data == "admindailyradar:toggle")
async def admin_daily_radar_toggle(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer("Меняю режим…")
    state, _ = await _load_radar_daily_digest_state_bounded()
    state["enabled"] = not bool(state.get("enabled", True))
    try:
        state = await asyncio.wait_for(save_radar_daily_digest_state(state), timeout=RADAR_DAILY_DIGEST_STATE_TIMEOUT_SECONDS)
    except Exception:
        log.exception("Daily Radar toggle save failed")
        await _edit_or_answer(
            callback.message,
            "❌ <b>Не удалось сохранить настройку Daily Radar.</b>\n\nПопробуй ещё раз.",
            reply_markup=admin_daily_radar_keyboard(state),
        )
        return
    metrics, _ = await _radar_daily_digest_metrics_bounded(state)
    status = "✅ ВКЛ" if state["enabled"] else "⏸ ВЫКЛ"
    await _edit_or_answer(
        callback.message,
        "<b>📨 Daily Radar</b>\n\n"
        f"Статус: <b>{status}</b>\n"
        f"Время: <b>{state['time']} МСК</b>\n"
        f"Следующая: <b>{html.escape(_daily_digest_next_run_text(state))}</b>\n\n"
        f"Сегодня: 🔎 <b>{_digest_n(metrics.get('listings_seen', 0))}</b> · 📡 <b>+{_digest_n(metrics.get('signals_today', 0))}</b>",
        reply_markup=admin_daily_radar_keyboard(state),
    )


@dp.callback_query(F.data == "admindailyradar:time")
async def admin_daily_radar_time(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await callback.answer("Открываю выбор времени…")
    digest_state, _ = await _load_radar_daily_digest_state_bounded()
    await _edit_or_answer(
        callback.message,
        "<b>🕐 Время Daily Radar</b>\n\n"
        f"Сейчас: <b>{html.escape(str(digest_state.get('time') or RADAR_DAILY_DIGEST_DEFAULT_TIME))} МСК</b>\n\n"
        "Выбери готовое время или нажми <b>«Ввести своё время»</b>.\n"
        "Можно установить любое время от <b>00:00</b> до <b>23:59</b> по Москве.",
        reply_markup=admin_daily_radar_time_keyboard(str(digest_state.get("time") or "")),
    )


@dp.callback_query(F.data.startswith("admindailyradar:settime:"))
async def admin_daily_radar_set_time(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    token = str(callback.data or "").removeprefix("admindailyradar:settime:")
    value = f"{token[:2]}:{token[2:]}" if len(token) == 4 and token.isdigit() else token
    value = _normalize_daily_digest_time(value)
    if value is None:
        await callback.answer("Некорректное время", show_alert=True)
        return
    await state.clear()
    await callback.answer(f"Сохраняю {value} МСК…")
    digest_state, _ = await _load_radar_daily_digest_state_bounded()
    digest_state["time"] = value
    try:
        digest_state = await asyncio.wait_for(save_radar_daily_digest_state(digest_state), timeout=RADAR_DAILY_DIGEST_STATE_TIMEOUT_SECONDS)
    except Exception:
        log.exception("Daily Radar set-time save failed value=%s", value)
        await _edit_or_answer(
            callback.message,
            "❌ <b>Не удалось сохранить время.</b>\n\nПопробуй ещё раз.",
            reply_markup=admin_daily_radar_time_keyboard(str(digest_state.get("time") or "")),
        )
        return
    await _edit_or_answer(
        callback.message,
        "<b>📨 Daily Radar</b>\n\n"
        f"Время изменено: <b>{value} МСК</b>\n"
        f"Следующая рассылка: <b>{html.escape(_daily_digest_next_run_text(digest_state))}</b>",
        reply_markup=admin_daily_radar_keyboard(digest_state),
    )


@dp.callback_query(F.data == "admindailyradar:customtime")
async def admin_daily_radar_custom_time(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminInput.daily_radar_time)
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "<b>✍️ Своё время Daily Radar</b>\n\n"
        "Отправь время одним сообщением в формате <b>ЧЧ:ММ</b>.\n\n"
        "Например: <code>09:35</code>, <code>16:10</code> или <code>23:45</code>.\n"
        "Часовой пояс: <b>Москва</b>.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Отмена", callback_data="admindailyradar")
        ]]),
    )


@dp.message(AdminInput.daily_radar_time)
async def admin_daily_radar_custom_time_message(message: Message, state: FSMContext) -> None:
    if not _is_admin(message.from_user.id):
        await state.clear()
        return
    value = _normalize_daily_digest_time(message.text)
    if value is None:
        await message.answer(
            "❌ <b>Некорректное время.</b>\n\n"
            "Отправь время в формате <b>ЧЧ:ММ</b>, например <code>09:35</code> или <code>21:10</code>.",
            parse_mode=ParseMode.HTML,
        )
        return
    digest_state, _ = await _load_radar_daily_digest_state_bounded()
    digest_state["time"] = value
    try:
        digest_state = await asyncio.wait_for(save_radar_daily_digest_state(digest_state), timeout=RADAR_DAILY_DIGEST_STATE_TIMEOUT_SECONDS)
    except Exception:
        log.exception("Daily Radar custom-time save failed value=%s", value)
        await message.answer("❌ Не удалось сохранить время. Попробуй ещё раз.")
        return
    await state.clear()
    await message.answer(
        "✅ <b>Время Daily Radar сохранено</b>\n\n"
        f"Каждый день: <b>{value} МСК</b>\n"
        f"Следующая рассылка: <b>{html.escape(_daily_digest_next_run_text(digest_state))}</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_daily_radar_keyboard(digest_state),
    )


@dp.callback_query(F.data == "admindailyradar:sendnow")
async def admin_daily_radar_send_now_preview(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer("Готовлю предпросмотр…")
    await _edit_or_answer(
        callback.message,
        "<b>📣 Daily Radar</b>\n\n⏳ Считаю живые цифры и получателей…",
    )
    state, _ = await _load_radar_daily_digest_state_bounded()
    metrics, metrics_fresh = await _radar_daily_digest_metrics_bounded(state)
    try:
        recipients = await asyncio.wait_for(
            _broadcast_recipient_ids(),
            timeout=RADAR_DAILY_DIGEST_RECIPIENTS_TIMEOUT_SECONDS,
        )
    except Exception:
        log.exception("Daily Radar recipient count failed/timeout")
        await _edit_or_answer(
            callback.message,
            "❌ <b>Не удалось получить список получателей.</b>\n\nНичего не отправлено. Попробуй ещё раз.",
            reply_markup=admin_daily_radar_keyboard(state),
        )
        return
    if not metrics_fresh:
        await _edit_or_answer(
            callback.message,
            "⚠️ <b>Не удалось обновить живые цифры.</b>\n\nРассылка не запущена, чтобы не отправить устаревшую статистику. Нажми «Обновить цифры» и повтори.",
            reply_markup=admin_daily_radar_keyboard(state),
        )
        return
    await _edit_or_answer(
        callback.message,
        "<b>📣 Отправить Daily Radar сейчас?</b>\n\n"
        f"Получателей: <b>{len(recipients)}</b>\n"
        f"🔎 Проверено сегодня: <b>{_digest_n(metrics.get('listings_seen', 0))}</b>\n"
        f"📡 Сигналов сегодня: <b>+{_digest_n(metrics.get('signals_today', 0))}</b>\n"
        f"🔥 Горячих: <b>{_digest_n(metrics.get('hot', 0))}</b> · 🚀 Набирают: <b>{_digest_n(metrics.get('rising', 0))}</b>\n\n"
        "После подтверждения пост уйдёт всем незаблокированным пользователям. "
        "Ручную рассылку можно запускать в любое время.",
        reply_markup=admin_daily_radar_send_confirm_keyboard(),
    )


@dp.callback_query(F.data == "admindailyradar:sendconfirm")
async def admin_daily_radar_send_now_confirm(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    if _RADAR_DAILY_DIGEST_SEND_LOCK.locked():
        await callback.answer("Daily Radar уже отправляется", show_alert=True)
        return
    await callback.answer("Запускаю рассылку")
    await _edit_or_answer(
        callback.message,
        "<b>📣 Daily Radar отправляется…</b>\n\nСобираю живые цифры и отправляю пост пользователям.",
    )
    async with _RADAR_DAILY_DIGEST_SEND_LOCK:
        digest_state, _ = await _load_radar_daily_digest_state_bounded()
        metrics, metrics_fresh = await _radar_daily_digest_metrics_bounded(
            digest_state,
            timeout_seconds=RADAR_DAILY_DIGEST_SEND_METRICS_TIMEOUT_SECONDS,
        )
        if not metrics_fresh:
            await _edit_or_answer(
                callback.message,
                "❌ <b>Daily Radar не отправлен.</b>\n\nНе удалось получить свежие цифры. Попробуй ещё раз через минуту.",
                reply_markup=admin_daily_radar_keyboard(digest_state),
            )
            return
        result = await _send_radar_daily_digest(callback.bot, metrics)
        now = datetime.now(MOSCOW)
        # A manual send counts as today's digest so the scheduler does not duplicate it later.
        # The admin may still press "Send now" again explicitly if another blast is desired.
        digest_state["last_sent_date"] = now.date().isoformat()
        digest_state["last_sent_at"] = now.replace(microsecond=0).isoformat()
        digest_state["last_delivered"] = int(result["sent"])
        digest_state["last_blocked"] = int(result["blocked"])
        digest_state["last_failed"] = int(result["failed"])
        try:
            digest_state = await asyncio.wait_for(
                save_radar_daily_digest_state(digest_state),
                timeout=RADAR_DAILY_DIGEST_STATE_TIMEOUT_SECONDS,
            )
        except Exception:
            log.warning("Daily Radar manual-send state save failed", exc_info=True)
    log.info(
        "Daily Radar digest manual send admin=%s recipients=%s delivered=%s blocked=%s failed=%s listings=%s signals=%s",
        callback.from_user.id, result["recipients"], result["sent"], result["blocked"], result["failed"],
        metrics.get("listings_seen", 0), metrics.get("signals_today", 0),
    )
    await _edit_or_answer(
        callback.message,
        "✅ <b>Daily Radar отправлен вручную</b>\n\n"
        f"Получателей: <b>{result['recipients']}</b>\n"
        f"Доставлено: <b>{result['sent']}</b>\n"
        f"Недоступны: <b>{result['blocked']}</b> · ошибки: <b>{result['failed']}</b>\n\n"
        "Автоматическая рассылка сегодня повторно не сработает. "
        "При необходимости ты можешь снова нажать «Отправить сейчас» вручную.",
        reply_markup=admin_daily_radar_keyboard(digest_state),
    )


@dp.callback_query(F.data == "admindailyradar:test")
async def admin_daily_radar_test(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer("Готовлю тест…")
    state, _ = await _load_radar_daily_digest_state_bounded()
    metrics, fresh = await _radar_daily_digest_metrics_bounded(state)
    if not fresh:
        await callback.message.answer("⚠️ Не удалось получить свежие цифры Daily Radar. Тест не отправлен.")
        return
    paid = allowed(callback.from_user.id)
    await callback.bot.send_message(
        int(callback.from_user.id),
        radar_daily_digest_text(metrics, paid=paid),
        parse_mode=ParseMode.HTML,
        reply_markup=radar_daily_digest_keyboard(paid=paid),
    )


@dp.callback_query(F.data == "adminbroadcast")
async def admin_broadcast_begin(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminInput.broadcast_content)
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "<b>📣 Рассылка пользователям</b>\n\n"
        "Отправь сюда готовый пост в одном из форматов:\n"
        "• обычный текст;\n"
        "• фотография;\n"
        "• фотография + подпись.\n\n"
        "Я покажу предпросмотр и ничего не отправлю без отдельного подтверждения. "
        "Рассылка идёт всем зарегистрированным пользователям бота, включая пользователей с истёкшей подпиской; заблокированные администратором аккаунты пропускаются.",
        reply_markup=admin_broadcast_back_keyboard(),
    )


@dp.callback_query(F.data == "adminbroadcast:replace")
async def admin_broadcast_replace(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminInput.broadcast_content)
    await callback.answer()
    await callback.message.answer(
        "✏️ Отправь новый текст, фото или фото с подписью.",
        reply_markup=admin_broadcast_back_keyboard(),
    )


@dp.callback_query(F.data == "adminbroadcast:cancel")
async def admin_broadcast_cancel(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await callback.answer("Рассылка отменена")
    await _edit_or_answer(callback.message, await _admin_dashboard_text(), reply_markup=await _admin_live_keyboard())


@dp.message(AdminInput.broadcast_content)
async def admin_broadcast_content_message(message: Message, state: FSMContext) -> None:
    if not _is_admin(message.from_user.id):
        await state.clear()
        return
    # v4.8.7 deliberately supports the three launch formats only. Keeping the
    # source Telegram message lets copy_message preserve entities/caption/photo
    # quality without re-uploading media or adding a "Forwarded from" header.
    if not message.text and not message.photo:
        await message.answer("⚠️ Поддерживается текст, фото или фото с подписью. Отправь пост ещё раз.")
        return
    await state.update_data(
        broadcast_source_chat_id=int(message.chat.id),
        broadcast_source_message_id=int(message.message_id),
        broadcast_kind="photo" if message.photo else "text",
    )
    recipients = await _broadcast_recipient_ids()
    await message.answer("<b>👁 Предпросмотр</b>", parse_mode=ParseMode.HTML)
    try:
        await message.bot.copy_message(
            chat_id=message.chat.id,
            from_chat_id=message.chat.id,
            message_id=message.message_id,
        )
    except Exception:
        log.warning("Broadcast preview copy failed", exc_info=True)
    await message.answer(
        f"<b>📣 Готово к рассылке</b>\n\nПолучателей: <b>{len(recipients)}</b>.\n"
        "Проверь пост выше и подтверди отправку.",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_broadcast_preview_keyboard(),
    )


@dp.callback_query(F.data == "adminbroadcast:send")
async def admin_broadcast_send(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    data = await state.get_data()
    source_chat_id = int(data.get("broadcast_source_chat_id") or 0)
    source_message_id = int(data.get("broadcast_source_message_id") or 0)
    if not source_chat_id or not source_message_id:
        await callback.answer("Сначала подготовь пост", show_alert=True)
        await state.clear()
        return

    recipients = await _broadcast_recipient_ids()
    await state.clear()
    await callback.answer("Рассылка запущена")
    progress = await callback.message.answer(
        f"📣 <b>Рассылка запущена</b>\nПолучателей: <b>{len(recipients)}</b>\nОтправлено: <b>0</b>",
        parse_mode=ParseMode.HTML,
    )

    sent = 0
    blocked = 0
    failed = 0
    for index, uid in enumerate(recipients, start=1):
        try:
            try:
                await callback.bot.copy_message(
                    chat_id=uid,
                    from_chat_id=source_chat_id,
                    message_id=source_message_id,
                )
            except TelegramRetryAfter as exc:
                await asyncio.sleep(float(exc.retry_after) + 0.2)
                await callback.bot.copy_message(
                    chat_id=uid,
                    from_chat_id=source_chat_id,
                    message_id=source_message_id,
                )
            sent += 1
        except TelegramForbiddenError:
            blocked += 1
        except TelegramBadRequest:
            failed += 1
        except Exception:
            failed += 1
            log.warning("Broadcast delivery failed user=%s", uid, exc_info=True)

        # Keep comfortably below Telegram's common broadcast-rate ceiling and
        # avoid making the bot less responsive while a campaign is running.
        await asyncio.sleep(0.055)
        if index % 25 == 0 and index < len(recipients):
            try:
                await progress.edit_text(
                    f"📣 <b>Рассылка выполняется</b>\n"
                    f"Обработано: <b>{index}/{len(recipients)}</b>\n"
                    f"Отправлено: <b>{sent}</b> · недоступны: <b>{blocked}</b> · ошибки: <b>{failed}</b>",
                    parse_mode=ParseMode.HTML,
                )
            except Exception:
                pass

    await progress.edit_text(
        "✅ <b>Рассылка завершена</b>\n\n"
        f"Всего получателей: <b>{len(recipients)}</b>\n"
        f"Доставлено: <b>{sent}</b>\n"
        f"Бот заблокирован / чат недоступен: <b>{blocked}</b>\n"
        f"Другие ошибки: <b>{failed}</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_back_keyboard(),
    )


@dp.callback_query(F.data == "adminusers")
async def admin_users_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    users = await recent_users(20)
    await callback.answer()
    text = "<b>👥 Пользователи</b>\n\nПоследние по активности. Нажми на пользователя для управления доступом."
    if not users:
        text += "\n\nПока никого нет."
    await _edit_or_answer(callback.message, text, reply_markup=admin_users_keyboard(users))


@dp.callback_query(F.data == "adminusersearch")
async def admin_user_search_begin(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(AdminInput.user_search)
    await callback.answer()
    await callback.message.answer(
        "🔎 Отправь <b>Telegram ID</b>, <b>@username</b> или имя пользователя.",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_back_keyboard(),
    )


@dp.message(AdminInput.user_search)
async def admin_user_search_message(message: Message, state: FSMContext) -> None:
    if not _is_admin(message.from_user.id):
        await state.clear()
        return
    query = (message.text or "").strip()
    users = await find_users(query, 20)
    await state.clear()
    text = f"<b>🔎 Результаты поиска</b>\n\nЗапрос: <code>{html.escape(query)}</code>"
    if not users:
        text += "\n\nНичего не найдено. Пользователь должен хотя бы раз открыть бота."
    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=admin_users_keyboard(users))


@dp.callback_query(F.data.startswith("adminuser:"))
async def admin_user_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        uid = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Некорректный ID", show_alert=True)
        return
    rendered = await render_admin_user(uid)
    if rendered is None:
        await callback.answer("Пользователь не найден", show_alert=True)
        return
    await callback.answer()
    await _edit_or_answer(callback.message, rendered[0], reply_markup=rendered[1])


@dp.callback_query(F.data.startswith("admincustom:"))
async def admin_custom_days_begin(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        uid = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Некорректный ID", show_alert=True)
        return
    await state.set_state(AdminInput.custom_days)
    await state.update_data(admin_target_user=uid)
    await callback.answer()
    await callback.message.answer(
        f"➕ Сколько дней добавить пользователю <code>{uid}</code>?\n\nОтправь число от <b>1</b> до <b>3650</b>.",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_user_back_keyboard(uid),
    )


@dp.message(AdminInput.custom_days)
async def admin_custom_days_message(message: Message, state: FSMContext) -> None:
    if not _is_admin(message.from_user.id):
        await state.clear()
        return
    data = await state.get_data()
    uid = int(data.get("admin_target_user") or 0)
    try:
        days = int((message.text or "").strip())
        if days < 1 or days > 3650:
            raise ValueError
    except Exception:
        await message.answer("⚠️ Отправь целое число от 1 до 3650.")
        return
    until = await grant_access_days(uid, days)
    await state.clear()
    try:
        await message.bot.send_message(
            uid,
            f"✅ <b>Доступ продлён на {days} дн.</b>\nАктивен до <b>{_utc_to_msk_text(until)} МСК</b>.",
            parse_mode=ParseMode.HTML,
        )
    except Exception:
        pass
    rendered = await render_admin_user(uid)
    if rendered:
        await message.answer(rendered[0], parse_mode=ParseMode.HTML, reply_markup=rendered[1])


@dp.callback_query(F.data.startswith("adminuserpayments:"))
async def admin_user_payments_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    uid = int(callback.data.split(":", 1)[1])
    payments = await user_payments(uid, 20)
    lines = [f"<b>💳 Платежи пользователя</b>", f"ID: <code>{uid}</code>", ""]
    if not payments:
        lines.append("Платежей нет.")
    else:
        for p in payments[:15]:
            plan = await get_plan(p.plan_key)
            title = plan.title if plan else p.plan_key
            when = _utc_to_msk_text(p.paid_at or p.created_at)
            lines.append(
                f"{_payment_status_label(p.status)} · <b>{html.escape(title)}</b> · "
                f"{p.amount_usdt:g} USDT · {_provider_label(p.provider)} · {when}"
            )
    await callback.answer()
    await _edit_or_answer(callback.message, "\n".join(lines), reply_markup=admin_user_back_keyboard(uid))


@dp.callback_query(F.data.startswith("adminuserscans:"))
async def admin_user_scans_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    uid = int(callback.data.split(":", 1)[1])
    async with SessionLocal() as session:
        scans = list((await session.execute(
            select(UserScan).where(UserScan.user_id == uid).order_by(UserScan.created_at.desc()).limit(12)
        )).scalars().all())
    lines = ["<b>📊 Последние сканы пользователя</b>", f"ID: <code>{uid}</code>", ""]
    if not scans:
        lines.append("Сканов нет.")
    else:
        for scan in scans:
            icon = {"done": "✅", "partial": "⚠️", "failed": "❌", "cancelled": "⏹", "running": "🔄", "queued": "⏳"}.get(scan.status, "▫️")
            recovered = f" · ♻️{scan.resumed_count}" if int(scan.resumed_count or 0) else ""
            lines.append(
                f"{icon} <b>#{scan.id}</b> · {_date_label(scan.target_date)} · {html.escape(scan.title[:45])}\n"
                f"результат: {scan.result_count} · качество: {scan.quality_score}/100{recovered}"
            )
    await callback.answer()
    await _edit_or_answer(callback.message, "\n\n".join(lines), reply_markup=admin_user_back_keyboard(uid))


@dp.callback_query(F.data.startswith("adminusererrors:"))
async def admin_user_errors_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    uid = int(callback.data.split(":", 1)[1])
    async with SessionLocal() as session:
        runs = list((await session.execute(
            select(ParserRun)
            .where(ParserRun.user_id == uid, ParserRun.success.is_(False))
            .order_by(ParserRun.started_at.desc())
            .limit(8)
        )).scalars().all())
        scan_errors = list((await session.execute(
            select(UserScan)
            .where(UserScan.user_id == uid, UserScan.last_error.is_not(None))
            .order_by(UserScan.created_at.desc())
            .limit(5)
        )).scalars().all())
    lines = ["<b>⚠️ Последние ошибки</b>", f"ID: <code>{uid}</code>", ""]
    if not runs and not scan_errors:
        lines.append("Ошибок не зафиксировано.")
    for scan in scan_errors:
        lines.append(f"Скан #{scan.id}: <code>{html.escape((scan.last_error or '')[:220])}</code>")
    for run in runs:
        lines.append(
            f"{_utc_to_msk_text(run.started_at)} · {html.escape(run.category_name[:40])}\n"
            f"<code>{html.escape((run.error_text or run.stop_reason or 'ошибка')[:220])}</code>"
        )
    await callback.answer()
    await _edit_or_answer(callback.message, "\n\n".join(lines), reply_markup=admin_user_back_keyboard(uid))


@dp.callback_query(F.data.startswith("admingrant:"))
async def admin_grant_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        _, uid_raw, days_raw = callback.data.split(":", 2)
        uid, days = int(uid_raw), int(days_raw)
    except Exception:
        await callback.answer("Некорректные данные", show_alert=True)
        return
    until = await grant_access_days(uid, days)
    await callback.answer(f"Добавлено {days} дн.")
    try:
        await callback.bot.send_message(
            uid,
            f"✅ <b>Доступ продлён на {days} дн.</b>\nАктивен до <b>{_utc_to_msk_text(until)} МСК</b>.",
            parse_mode=ParseMode.HTML,
        )
    except Exception:
        pass
    rendered = await render_admin_user(uid)
    if rendered:
        await _edit_or_answer(callback.message, rendered[0], reply_markup=rendered[1])


@dp.callback_query(F.data.startswith("adminrevoke:"))
async def admin_revoke_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        uid = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Некорректный ID", show_alert=True)
        return
    await revoke_access(uid)
    await callback.answer("Доступ отозван")
    rendered = await render_admin_user(uid)
    if rendered:
        await _edit_or_answer(callback.message, rendered[0], reply_markup=rendered[1])


@dp.callback_query(F.data.startswith("adminban:"))
async def admin_ban_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        uid = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Некорректный ID", show_alert=True)
        return
    user = await get_commerce_user(uid)
    new_value = not bool(user and user.is_banned)
    await set_banned(uid, new_value)
    await callback.answer("Заблокирован" if new_value else "Разблокирован")
    rendered = await render_admin_user(uid)
    if rendered:
        await _edit_or_answer(callback.message, rendered[0], reply_markup=rendered[1])


@dp.callback_query(F.data == "adminpayments")
async def admin_payments_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    payments = await recent_payments(20)
    providers = providers_status()
    lines = [
        "<b>💳 Платежи</b>",
        "",
        f"CryptoBot: <b>{'✅ настроен' if providers['cryptobot'] else '▫️ нет токена'}</b>",
        f"xRocket: <b>{'✅ настроен' if providers['xrocket'] else '▫️ нет API key'}</b>",
    ]
    if payments:
        lines.extend(["", "<b>Последние счета</b>"])
        for p in payments[:15]:
            user = await get_commerce_user(p.user_id)
            who = f"@{user.username}" if user and user.username else str(p.user_id)
            lines.append(
                f"{_payment_status_label(p.status)} · <b>{p.amount_usdt:g} USDT</b> · "
                f"{html.escape(who)} · {_provider_label(p.provider)} · {_utc_to_msk_text(p.created_at)}"
            )
    else:
        lines.extend(["", "Платежей пока нет."])
    await callback.answer()
    await _edit_or_answer(callback.message, "\n".join(lines), reply_markup=admin_back_keyboard())


@dp.callback_query(F.data == "adminplans")
async def admin_plans_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    plans = await get_plans()
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "<b>🎟 Тарифы</b>\n\nЦена меняется здесь и сразу применяется к новым счетам.",
        reply_markup=admin_plans_keyboard(plans),
    )


@dp.callback_query(F.data.startswith("adminplan:"))
async def admin_plan_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    key = callback.data.split(":", 1)[1]
    plan = await get_plan(key)
    if not plan:
        await callback.answer("Тариф не найден", show_alert=True)
        return
    text = (
        f"<b>🎟 {html.escape(plan.title)}</b>\n\n"
        f"Срок: <b>{plan.days} дн.</b>\n"
        f"Цена: <b>{plan.price_usdt:g} USDT</b>\n"
        f"Статус: <b>{'✅ включён' if plan.is_active else '⏸ выключен'}</b>"
    )
    await callback.answer()
    await _edit_or_answer(callback.message, text, reply_markup=admin_plan_keyboard(plan))


@dp.callback_query(F.data.startswith("adminplanprice:"))
async def admin_plan_price_begin(callback: CallbackQuery, state: FSMContext) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    key = callback.data.split(":", 1)[1]
    plan = await get_plan(key)
    if not plan:
        await callback.answer("Тариф не найден", show_alert=True)
        return
    await state.set_state(AdminInput.plan_price)
    await state.update_data(admin_plan_key=key)
    await callback.answer()
    await callback.message.answer(
        f"💰 Новая цена для <b>{html.escape(plan.title)}</b> в USDT.\nНапример: <code>9.99</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_back_keyboard(),
    )


@dp.message(AdminInput.plan_price)
async def admin_plan_price_message(message: Message, state: FSMContext) -> None:
    if not _is_admin(message.from_user.id):
        await state.clear()
        return
    data = await state.get_data()
    key = data.get("admin_plan_key")
    raw = (message.text or "").strip().replace(",", ".")
    try:
        price = float(raw)
        plan = await update_plan_price(str(key), price)
    except Exception:
        await message.answer("⚠️ Нужна положительная цена, например <code>9.99</code>.", parse_mode=ParseMode.HTML)
        return
    await state.clear()
    if not plan:
        await message.answer("Тариф не найден.", reply_markup=admin_back_keyboard())
        return
    await message.answer(
        f"✅ Цена <b>{html.escape(plan.title)}</b> изменена на <b>{plan.price_usdt:g} USDT</b>.",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_plan_keyboard(plan),
    )


@dp.callback_query(F.data.startswith("adminplantoggle:"))
async def admin_plan_toggle_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    key = callback.data.split(":", 1)[1]
    plan = await toggle_plan(key)
    if not plan:
        await callback.answer("Тариф не найден", show_alert=True)
        return
    await callback.answer("Включён" if plan.is_active else "Выключен")
    text = (
        f"<b>🎟 {html.escape(plan.title)}</b>\n\n"
        f"Срок: <b>{plan.days} дн.</b>\n"
        f"Цена: <b>{plan.price_usdt:g} USDT</b>\n"
        f"Статус: <b>{'✅ включён' if plan.is_active else '⏸ выключен'}</b>"
    )
    await _edit_or_answer(callback.message, text, reply_markup=admin_plan_keyboard(plan))


@dp.callback_query(F.data == "adminmode")
async def admin_mode_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    text = (
        "<b>🔐 Режим доступа</b>\n\n"
        "🔒 <b>Только админы</b> — безопасный режим для тестов.\n"
        "💎 <b>По подписке</b> — без активной подписки доступна только оплата.\n"
        "🌍 <b>Открытый доступ</b> — бот доступен всем без оплаты.\n\n"
        f"Сейчас: <b>{_access_mode_label()}</b>"
    )
    await callback.answer()
    await _edit_or_answer(callback.message, text, reply_markup=admin_mode_keyboard())


@dp.callback_query(F.data.startswith("adminsetmode:"))
async def admin_set_mode_handler(callback: CallbackQuery) -> None:
    if not _is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    mode = callback.data.split(":", 1)[1]
    try:
        await set_access_mode(mode)
    except ValueError:
        await callback.answer("Некорректный режим", show_alert=True)
        return
    await callback.answer("Режим изменён")
    await _edit_or_answer(
        callback.message,
        f"<b>🔐 Режим доступа</b>\n\nСейчас: <b>{_access_mode_label(mode)}</b>",
        reply_markup=admin_mode_keyboard(),
    )


async def payment_scheduler(bot: Bot) -> None:
    """Poll pending invoices. No webhook service is required for the first commercial build."""
    while True:
        try:
            payments = await pending_payments(100)
            for payment in payments:
                try:
                    refreshed, just_paid = await refresh_payment(payment.id)
                    if not refreshed or not just_paid:
                        continue
                    user = await get_commerce_user(refreshed.user_id)
                    plan = await get_plan(refreshed.plan_key)
                    until = user.access_until if user else None
                    try:
                        await bot.send_message(
                            refreshed.user_id,
                            "✅ <b>Оплата подтверждена</b>\n\n"
                            f"{html.escape(plan.title) if plan else 'Подписка'} активна до "
                            f"<b>{_utc_to_msk_text(until)} МСК</b>.\nТеперь сервис доступен.",
                            parse_mode=ParseMode.HTML,
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                                [InlineKeyboardButton(text="🏠 Открыть сервис", callback_data="home")]
                            ]),
                        )
                    except Exception:
                        log.exception("Could not notify paid user %s", refreshed.user_id)
                    for admin_id in ADMIN_IDS:
                        try:
                            await bot.send_message(
                                admin_id,
                                "💳 <b>Новая оплата</b>\n"
                                f"User: <code>{refreshed.user_id}</code>\n"
                                f"Тариф: <b>{html.escape(plan.title) if plan else refreshed.plan_key}</b>\n"
                                f"Сумма: <b>{refreshed.amount_usdt:g} USDT</b>\n"
                                f"Способ: <b>{_provider_label(refreshed.provider)}</b>",
                                parse_mode=ParseMode.HTML,
                            )
                        except Exception:
                            pass
                except Exception:
                    log.exception("Payment polling failed for payment=%s", payment.id)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Payment scheduler loop failed")
        await asyncio.sleep(PAYMENT_POLL_SECONDS)


async def subscription_lifecycle_scheduler(bot: Bot) -> None:
    """Notify users once 24h before expiry and once right after expiry."""
    while True:
        try:
            warnings, expired = await subscription_notice_candidates(100)
            for user in warnings:
                until = user.access_until
                if until is None:
                    continue
                try:
                    await bot.send_message(
                        user.user_id,
                        "⏳ <b>Подписка закончится меньше чем через 24 часа.</b>\n\n"
                        f"Доступ до <b>{_utc_to_msk_text(until)} МСК</b>. Продлить можно заранее — дни прибавятся к текущему сроку.",
                        parse_mode=ParseMode.HTML,
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                            [InlineKeyboardButton(text="🔄 Продлить подписку", callback_data="subscription")]
                        ]),
                    )
                    await mark_subscription_notice(user.user_id, until, "warning")
                except Exception:
                    log.debug("Could not send subscription warning user=%s", user.user_id, exc_info=True)

            for user in expired:
                until = user.access_until
                if until is None:
                    continue
                try:
                    await bot.send_message(
                        user.user_id,
                        "⌛ <b>Подписка закончилась.</b>\n\n"
                        "Сохранённые сканы и история не удалены. Продли доступ, чтобы снова запускать парсер и обновлять данные.",
                        parse_mode=ParseMode.HTML,
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                            [InlineKeyboardButton(text="💎 Продлить доступ", callback_data="subscription")]
                        ]),
                    )
                    await mark_subscription_notice(user.user_id, until, "expired")
                except Exception:
                    log.debug("Could not send subscription expiry user=%s", user.user_id, exc_info=True)
        except asyncio.CancelledError:
            raise
        except Exception:
            log.exception("Subscription lifecycle scheduler failed")
        await asyncio.sleep(SUBSCRIPTION_NOTICE_POLL_SECONDS)





def user_commands_for_language(language: str = LANG_RU) -> list[BotCommand]:
    if language == LANG_EN:
        return [
            BotCommand(command="menu", description="🏠 Main menu"),
            BotCommand(command="new_scan", description="▶️ New scan"),
            BotCommand(command="stop", description="⏹ Stop parser"),
            BotCommand(command="my_scans", description="📊 My scans"),
            BotCommand(command="popular", description="🔥 Popular"),
            BotCommand(command="radar", description="📡 DT Radar"),
            BotCommand(command="categories", description="🗂 Categories"),
            BotCommand(command="settings", description="⚙️ Settings"),
            BotCommand(command="subscription", description="💎 Subscription"),
            BotCommand(command="language", description="🌐 Language"),
            BotCommand(command="help", description="ℹ️ Help"),
        ]
    return [
        BotCommand(command="menu", description="🏠 Главное меню"),
        BotCommand(command="new_scan", description="▶️ Новый скан"),
        BotCommand(command="stop", description="⏹ Остановить парсер"),
        BotCommand(command="my_scans", description="📊 Мои сканы"),
        BotCommand(command="popular", description="🔥 Популярное"),
        BotCommand(command="radar", description="📡 DT Radar"),
        BotCommand(command="categories", description="🗂 Категории"),
        BotCommand(command="settings", description="⚙️ Настройки"),
        BotCommand(command="subscription", description="💎 Подписка"),
        BotCommand(command="language", description="🌐 Язык"),
        BotCommand(command="help", description="ℹ️ Помощь"),
    ]


async def set_user_command_language(bot: Bot, user_id: int, language: str) -> None:
    if int(user_id) in ADMIN_IDS:
        return
    try:
        await bot.set_my_commands(
            user_commands_for_language(language),
            scope=BotCommandScopeChat(chat_id=int(user_id)),
        )
    except Exception:
        log.debug("Could not update command language user=%s", user_id, exc_info=True)


async def setup_bot_commands(bot: Bot) -> None:
    """Configure Telegram's bottom-left Menu button and command list."""
    user_commands = user_commands_for_language(LANG_RU)
    admin_commands = user_commands + [
        BotCommand(command="admin", description="🛠 Админ-панель"),
    ]

    # Default menu is Russian; Telegram-native English clients also get an
    # English command list.  The in-bot language selector additionally installs
    # a per-chat command list so the user's explicit choice wins.
    await bot.set_my_commands(user_commands, scope=BotCommandScopeDefault())
    try:
        await bot.set_my_commands(user_commands_for_language(LANG_EN), scope=BotCommandScopeDefault(), language_code="en")
    except Exception:
        log.debug("Could not configure Telegram English command list", exc_info=True)
    # Admin chats get the same menu plus /admin.
    for admin_id in sorted(ADMIN_IDS):
        try:
            await bot.set_my_commands(admin_commands, scope=BotCommandScopeChat(chat_id=admin_id))
        except Exception:
            log.exception("Could not set admin command menu for chat=%s", admin_id)

    # Force Telegram to render the standard Commands menu button instead of
    # requiring users to type slash commands manually.
    await bot.set_chat_menu_button(menu_button=MenuButtonCommands())


# Legacy body replaced by v4.6.5; marker kept out intentionally.
@dp.message(Command("language"))
async def language_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    current = await get_user_language(message.from_user.id)
    await _send_language_picker(message, current)


@dp.callback_query(F.data == "language_settings")
async def language_settings(callback: CallbackQuery) -> None:
    current = await get_user_language(callback.from_user.id)
    await callback.answer()
    token = _UI_LANGUAGE.set(LANG_RU)
    try:
        await _edit_or_answer(
            callback.message,
            language_prompt_text(current),
            reply_markup=language_keyboard(current),
        )
    finally:
        _UI_LANGUAGE.reset(token)


@dp.callback_query(F.data.startswith("language:"))
async def choose_language(callback: CallbackQuery, state: FSMContext) -> None:
    language = (callback.data or "").split(":", 1)[1].strip().lower()
    if language not in {LANG_RU, LANG_EN}:
        await callback.answer("Unknown language", show_alert=True)
        return
    await set_user_language(callback.from_user.id, language)
    await set_user_command_language(callback.bot, callback.from_user.id, language)
    token = _UI_LANGUAGE.set(language)
    try:
        await callback.answer("Language saved" if language == LANG_EN else "Язык сохранён")
        await state.clear()
        if not allowed(callback.from_user.id):
            if callback.message:
                await _send_home_message(callback.message, callback.from_user.id)
            return
        user = await get_commerce_user(callback.from_user.id)
        if user is not None and not bool(user.onboarding_completed):
            if callback.message:
                await _show_onboarding(callback.message, callback.from_user.id, 1)
            return
        if callback.message:
            await _send_home_message(callback.message, callback.from_user.id)
    finally:
        _UI_LANGUAGE.reset(token)


def onboarding_keyboard(step: int) -> InlineKeyboardMarkup:
    if step == 1:
        rows = [
            [InlineKeyboardButton(text="➡️ Как пользоваться", callback_data="onboard:2")],
            [InlineKeyboardButton(text="Пропустить", callback_data="onboard:skip")],
        ]
    elif step == 2:
        rows = [
            [InlineKeyboardButton(text="➡️ Что будет после скана", callback_data="onboard:3")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="onboard:1")],
        ]
    else:
        rows = [
            [InlineKeyboardButton(text="🗂 Выбрать категории", callback_data="onboard:categories")],
            [InlineKeyboardButton(text="🏠 Открыть главное меню", callback_data="onboard:finish")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="onboard:2")],
        ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


def onboarding_text(step: int) -> str:
    if step == 1:
        return (
            f"<b>👋 Kleinanzeigen Analytics</b>\n\n"
            "Бот помогает найти товары, которые реально привлекают внимание на Kleinanzeigen, "
            "и затем автоматически измеряет рост просмотров.\n\n"
            "Настройка первого запуска занимает меньше минуты."
        )
    if step == 2:
        return (
            "<b>1/2 · Как запустить первый скан</b>\n\n"
            "1. 🗂 Выбери одну или несколько категорий.\n"
            "2. ⚙️ В «Настройках» выбери режим результата и фильтры. Если сомневаешься — открой «Что выбрать?».\n"
            "3. ▶️ Нажми «Новый скан» и выбери дату, цену и глубину 15 / 25 / 50 страниц.\n\n"
            "Во время работы парсер можно полностью остановить кнопкой ⏹."
        )
    return (
        "<b>2/2 · Что будет после скана</b>\n\n"
        "📊 Скан сохранится в «Мои сканы».\n"
        "⏱ Автозамеры 3 / 6 / 12 ч выключены по умолчанию — включи их на главном экране, если нужны.\n"
        "🔥 В «Популярное» появятся лидеры по просмотрам и росту.\n"
        "📦 Через 24 часа карточка уйдёт в Архив, но данные не удалятся.\n\n"
        "Готово — можно выбирать категории."
    )


async def _show_onboarding(message: Message, user_id: int, step: int = 1) -> None:
    step = max(1, min(3, int(step)))
    await message.answer(
        onboarding_text(step),
        parse_mode=ParseMode.HTML,
        reply_markup=onboarding_keyboard(step),
    )


def home_text(
    selected_count: int, auto_observations: bool = False, access_active: bool = True, trial_remaining: int = 0
) -> str:
    if not access_active and trial_remaining > 0:
        return (
            "<b>🚀 DT PARSER · стартовая акция</b>\n\n"
            f"🎁 У тебя <b>{trial_remaining} бесплатных скан(а)</b> из {FREE_TRIAL_SCAN_LIMIT}.\n"
            f"Пробный запуск: <b>1 категория · до {FREE_TRIAL_MAX_PAGES} страниц</b>.\n\n"
            "👁 Реальные просмотры, TOP и XLSX доступны полностью — это не урезанная демо-таблица.\n\n"
            "Выбери категорию и попробуй DT PARSER на реальной выдаче Kleinanzeigen. "
            "Повторные замеры, несколько категорий и 50 страниц открываются по подписке."
        )
    if not access_active:
        return (
            "<b>🔎 Kleinanzeigen Analytics</b>\n\n"
            "⌛ <b>Подписка не активна</b>\n\n"
            "Твои прошлые сканы и сохранённые результаты остаются доступны. "
            "Можно открыть карточку скана, TOP, историю и скачать XLSX.\n\n"
            "🔒 Для нового скана, повторного запуска или обновления просмотров нужна активная подписка."
        )
    return (
        "<b>📡 DT PARSER — MARKET ANALYTICS</b>\n\n"
        "Сканируй рынок. Находи спрос. <b>Заходи раньше.</b>\n\n"
        "🔎 <b>Сканирование</b>\n"
        "Анализируй Kleinanzeigen и находи объявления, которые быстрее остальных набирают просмотры.\n\n"
        "📡 <b>DT Radar 3.0</b>\n"
        "Radar анализирует рынок и выцепляет <b>нестандартную товарку с растущим интересом</b>, "
        "пока её ещё не начали продавать все.\n\n"
        "<b>Выбери режим 👇</b>"
    )


async def _send_home_message(message: Message, user_id: int, *, intro: bool = False) -> None:
    """Send the branded DT PARSER home card with low-latency navigation."""
    global _MENU_IMAGE_FILE_ID
    selected, user_settings, trial, referral_enabled = await asyncio.gather(
        get_selected(user_id),
        get_settings(user_id),
        get_trial_status(user_id),
        referral_promo_enabled(),
    )
    auto_enabled = bool(getattr(user_settings, "auto_observations", False))
    access_active = allowed(user_id)
    trial_remaining = int(trial.remaining if (not access_active and trial.eligible) else 0)
    markup = main_keyboard(
        len(selected), admin=_is_admin(user_id), auto_observations=auto_enabled,
        access_active=access_active, trial_remaining=trial_remaining,
        referral_enabled=bool(referral_enabled),
    )
    caption = home_text(
        len(selected), auto_enabled, access_active=access_active, trial_remaining=trial_remaining
    )

    if _MENU_IMAGE_FILE_ID:
        try:
            await message.answer_photo(
                photo=_MENU_IMAGE_FILE_ID,
                caption=caption,
                reply_markup=markup,
                parse_mode=ParseMode.HTML,
            )
            return
        except Exception:
            # File IDs can theoretically become unusable after a bot/token
            # replacement. Drop the cache and upload the local asset once.
            _MENU_IMAGE_FILE_ID = None

    if MENU_IMAGE_PATH.exists():
        sent = await message.answer_photo(
            photo=FSInputFile(MENU_IMAGE_PATH),
            caption=caption,
            reply_markup=markup,
            parse_mode=ParseMode.HTML,
        )
        if getattr(sent, "photo", None):
            try:
                _MENU_IMAGE_FILE_ID = sent.photo[-1].file_id
            except Exception:
                pass
        return

    # Safe fallback if the asset was accidentally omitted from a deployment.
    log.error("Main menu image is missing: %s", MENU_IMAGE_PATH)
    await message.answer(caption, reply_markup=markup, parse_mode=ParseMode.HTML)


async def _send_popular_message(message: Message, user_id: int) -> None:
    items = await get_user_popular_categories(user_id)
    if not items:
        text = (
            "🔥 <b>Популярное</b>\n\n"
            "После первого успешного скана здесь появится актуальный TOP по категории."
        )
    else:
        text = (
            "🔥 <b>Популярное</b>\n\n"
            "Выбери категорию — показываем только её <b>последний успешный скан</b>.\n"
            "TOP роста доступен по замерам 3 / 6 / 12 часов; автозамеры включаются по желанию."
        )
    await message.answer(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=popular_categories_keyboard(items),
        disable_web_page_preview=True,
    )


async def _send_my_scans_message(message: Message, user_id: int) -> None:
    scans, archive_count = await get_user_scans_overview(user_id, 10)
    if not scans:
        text = (
            "<b>📊 Мои сканы</b>\n\nСвежих сканов пока нет. После завершения они будут храниться здесь 24 часа, затем уйдут в Архив."
        )
    else:
        text = (
            "<b>📊 Мои сканы</b>\n\nТекущие и свежие запуски за последние <b>24 часа</b>. Старые автоматически уходят в Архив; данные не удаляются."
        )
    await message.answer(
        text, parse_mode=ParseMode.HTML,
        reply_markup=my_scans_keyboard(scans, archive_count),
    )


async def _begin_scan_from_message(message: Message, state: FSMContext) -> None:
    user_id = message.from_user.id
    selected_keys, has_active_scan = await asyncio.gather(
        get_selected(user_id),
        user_has_active_scan(user_id),
    )
    selected_cats = [CATEGORIES[k] for k in CATEGORIES if k in selected_keys]
    if not selected_cats:
        await message.answer(
            "⚠️ <b>Сначала выбери хотя бы одну категорию.</b>",
            parse_mode=ParseMode.HTML,
            reply_markup=groups_keyboard(
                selected_keys,
                max_selected=(FREE_TRIAL_MAX_CATEGORIES if not allowed(user_id) else MAX_SELECTED_CATEGORIES),
            ),
        )
        return
    if not allowed(user_id) and len(selected_cats) > FREE_TRIAL_MAX_CATEGORIES:
        await message.answer(
            "🎁 <b>В бесплатном скане доступна 1 категория.</b>\n\n"
            "Оставь одну категорию и запусти скан снова.",
            parse_mode=ParseMode.HTML,
            reply_markup=groups_keyboard(selected_keys, max_selected=FREE_TRIAL_MAX_CATEGORIES),
        )
        return

    if has_active_scan:
        await message.answer("⏳ У тебя уже идёт парсинг.")
        return

    await state.set_state(ScanInput.target_date)
    await message.answer(
        "<b>▶️ Новый скан</b>\n\n"
        "<b>1/3 · Дата объявлений</b>\n"
        "Выбери одну из 5 доступных дат: сегодня или один из 4 предыдущих дней. "
        "Время считаем по Москве.",
        parse_mode=ParseMode.HTML,
        reply_markup=scan_date_keyboard(),
    )


@dp.callback_query(F.data.startswith("onboard:"))
async def onboarding_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not await trial_or_paid_scan_access(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    action = callback.data.split(":", 1)[1]
    if action in {"1", "2", "3"}:
        await callback.answer()
        await _edit_or_answer(callback.message, onboarding_text(int(action)), reply_markup=onboarding_keyboard(int(action)))
        return
    if action in {"finish", "skip"}:
        await set_onboarding_completed(callback.from_user.id, True)
        await callback.answer("Готово")
        await _send_home_message(callback.message, callback.from_user.id)
        return
    if action == "categories":
        await set_onboarding_completed(callback.from_user.id, True)
        selected = await get_selected(callback.from_user.id)
        await callback.answer()
        await _edit_or_answer(
            callback.message,
            "<b>🗂 Выбери категории</b>\n\nОтметь товары, которые хочешь анализировать. Потом вернись и запускай первый скан.",
            reply_markup=groups_keyboard(selected),
        )
        return
    await callback.answer("Неизвестное действие", show_alert=True)


@dp.message(CommandStart())
async def start(message: Message, state: FSMContext) -> None:
    await state.clear()
    try:
        await touch_user(message.from_user, force=True)
    except Exception:
        # Never leave a new user with a silent /start if PostgreSQL/profile storage
        # is temporarily unavailable. The outer middleware already logs its own
        # attempt, but this handler used to raise before sending any Telegram reply.
        log.exception("Could not persist /start user=%s", message.from_user.id)
        await message.answer(
            "⚠️ <b>Не удалось зарегистрировать профиль</b>\n\n"
            "База данных временно не приняла запрос. Попробуй нажать /start ещё раз через несколько секунд.",
            parse_mode=ParseMode.HTML,
        )
        return
    language = await get_user_language(message.from_user.id)
    if language is None and message.from_user.id not in ADMIN_IDS:
        await _send_language_picker(message)
        return
    if language is not None:
        await set_user_command_language(message.bot, message.from_user.id, language)
    if not allowed(message.from_user.id):
        await _send_home_message(message, message.from_user.id, intro=True)
        return
    user = await get_commerce_user(message.from_user.id)
    if user is not None and not bool(user.onboarding_completed):
        await _show_onboarding(message, message.from_user.id, 1)
        return
    await _send_home_message(message, message.from_user.id, intro=True)


@dp.message(Command("menu"))
async def menu_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    await _send_home_message(message, message.from_user.id)


@dp.message(Command("new_scan"))
async def new_scan_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    await _begin_scan_from_message(message, state)


@dp.message(Command("my_scans"))
async def my_scans_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    await _send_my_scans_message(message, message.from_user.id)


@dp.message(Command("popular"))
async def popular_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    await _send_popular_message(message, message.from_user.id)


@dp.message(Command("categories"))
async def categories_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    selected = await get_selected(message.from_user.id)
    trial_mode = bool(not allowed(message.from_user.id) and (await get_trial_status(message.from_user.id)).eligible)
    limit = FREE_TRIAL_MAX_CATEGORIES if trial_mode else MAX_SELECTED_CATEGORIES
    await message.answer(
        f"<b>🗂 Категории</b>\n\nВыбери до <b>{limit}</b> категорий на один скан.",
        reply_markup=groups_keyboard(selected, max_selected=limit),
        parse_mode=ParseMode.HTML,
    )


@dp.message(Command("settings"))
async def settings_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    s = await get_settings(message.from_user.id)
    await message.answer(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.message(Command("subscription"))
async def subscription_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        await subscription_text(message.from_user.id),
        parse_mode=ParseMode.HTML,
        reply_markup=await subscription_keyboard(message.from_user.id),
    )


@dp.message(Command("result"))
async def result_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    selected = await get_selected(message.from_user.id)
    await send_smart_export(message, message.from_user.id, len(selected))


@dp.message(Command("help"))
async def help_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    selected = await get_selected(message.from_user.id)
    await message.answer(
        "<b>ℹ️ Помощь</b>\n\n"
        "Основные разделы всегда доступны через кнопку <b>Menu</b>.\n\n"
        "▶️ Новый скан — выбрать дату, цену, глубину и запустить сбор\n"
        "🔥 Популярное — актуальный TOP последнего успешного скана\n"
        "📊 Мои сканы — свежие запуски и архив\n"
        "🗂 Категории — что анализировать\n"
        "⚙️ Настройки — фильтры результата\n"
        "⏱ Автозамеры — включить/выключить контрольные 3/6/12 ч\n"
        "💎 Подписка — доступ и платежи",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🎓 Показать обучение", callback_data="onboard:1")],
            [InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
        ]),
    )


@dp.callback_query(F.data == "home")
async def home(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not read_only_history_allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    await callback.answer()
    if not allowed(callback.from_user.id):
        await _send_home_message(callback.message, callback.from_user.id)
        return
    user = await get_commerce_user(callback.from_user.id)
    if user is not None and not bool(user.onboarding_completed):
        await _edit_or_answer(callback.message, onboarding_text(1), reply_markup=onboarding_keyboard(1))
        return
    await _send_home_message(callback.message, callback.from_user.id)


@dp.callback_query(F.data == "post_settings")
async def post_settings(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not await trial_or_paid_scan_access(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    s = await get_settings(callback.from_user.id)
    await callback.answer()
    await callback.message.answer(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "post_home")
async def post_home(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not read_only_history_allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    await callback.answer()
    await _send_home_message(callback.message, callback.from_user.id)


def auto_observations_keyboard(enabled: bool) -> InlineKeyboardMarkup:
    toggle_text = "⛔ Выключить автозамеры" if enabled else "✅ Включить автозамеры"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=toggle_text, callback_data="toggle_auto_obs")],
        [InlineKeyboardButton(text="🏠 Меню", callback_data="home")],
    ])


def auto_observations_text(enabled: bool) -> str:
    status = "✅ <b>ВКЛЮЧЕНЫ</b>" if enabled else "⛔ <b>ВЫКЛЮЧЕНЫ</b>"
    return (
        "<b>⏱ Автозамеры просмотров</b>\n\n"
        f"Сейчас: {status}\n\n"
        "Если включить, после <b>каждого нового завершённого скана</b> бот автоматически "
        "сделает контрольные замеры через <b>3 / 6 / 12 часов</b>.\n\n"
        "Если выключить, новые автоматические замеры не создаются, а уже ожидающие "
        "3/6/12ч отменяются. Готовая история не удаляется.\n\n"
        "👁 <b>Ручное «Обновить» работает всегда</b> — независимо от этой настройки."
    )


@dp.callback_query(F.data == "auto_obs_menu")
async def auto_obs_menu(callback: CallbackQuery) -> None:
    if not allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    s = await get_settings(callback.from_user.id)
    enabled = bool(getattr(s, "auto_observations", False))
    await callback.answer()
    await _edit_or_answer(
        callback.message, auto_observations_text(enabled),
        reply_markup=auto_observations_keyboard(enabled),
    )


@dp.callback_query(F.data == "toggle_auto_obs")
async def toggle_auto_obs(callback: CallbackQuery) -> None:
    if not allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    current = await get_settings(callback.from_user.id)
    enabled = not bool(getattr(current, "auto_observations", False))
    await set_auto_observations(callback.from_user.id, enabled)
    await callback.answer("Автозамеры включены" if enabled else "Автозамеры выключены")
    await _edit_or_answer(
        callback.message, auto_observations_text(enabled),
        reply_markup=auto_observations_keyboard(enabled),
    )


@dp.callback_query(F.data == "settings")
async def settings(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not await trial_or_paid_scan_access(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    await callback.answer()
    s = await get_settings(callback.from_user.id)
    await _edit_or_answer(callback.message, settings_text(s), reply_markup=settings_keyboard(s))


@dp.callback_query(F.data == "mode_help")
async def mode_help(callback: CallbackQuery) -> None:
    await callback.answer()
    text = (
        "<b>ℹ️ Что выбрать в настройках</b>\n\n"
        "Настройки <b>не меняют дату, цену и глубину скана</b>. Они определяют, какие найденные объявления останутся в XLSX и TOP. Чем сильнее фильтры — тем меньше итоговый список.\n\n"
        "<b>📦 РЕЖИМ РЕЗУЛЬТАТА</b>\n"
        "🆕 <b>Самые новые</b> — базовый режим для первого запуска: обычный список объявлений выбранного дня. Рекомендуем новичкам.\n\n"
        "📚 <b>Все</b> — полный список после остальных фильтров, без аналитического отбора. На скане одного дня близок к «Самые новые»; порядок задаёт сортировка.\n\n"
        "💎 <b>Уникальные</b> — оставляет модели/варианты, встретившиеся только один раз. Полезно для поиска редких товаров.\n\n"
        "🔥 <b>Часто публикуемые</b> — показывает группы минимум из 3 разных объявлений одной модели. Полезно, чтобы понять, что часто появляется на площадке.\n\n"
        "💰 <b>Ниже рынка</b> — ищет объявления минимум на 20% дешевле медианы похожих. Нужны хотя бы 5 цен в одной уверенной группе.\n\n"
        "⚡ <b>Быстро исчезающие</b> — товары, которые исчезли примерно за 12 часов или быстрее. Нужна история повторных проверок; для первого скана режим почти ничего не покажет.\n\n"
        "📉 <b>Снижение цены</b> — тот же ID подешевел минимум на 5 € и 5%. Нужен предыдущий замер/скан этого объявления.\n\n"
        "<b>🎛 ДОПОЛНИТЕЛЬНЫЕ ФИЛЬТРЫ</b>\n"
        "👁 <b>Минимум просмотров</b> — например 50+ или 100+. В результат попадут только объявления с подтверждённым счётчиком не ниже порога. <b>Не ускоряет скан</b>: просмотры всё равно сначала собираются.\n\n"
        "🧠 <b>Умные дубли</b> — скрывает почти одинаковые объявления с одинаковой ценой. Обычно лучше держать <b>Вкл</b>. В режиме «Часто публикуемые» повторы специально не скрываются.\n\n"
        "🧹 <b>Очистка шума</b> — убирает заголовки вроде Suche, Ankauf, Reparatur, Service, Vermietung, Tausch. Для обычных товаров рекомендуем <b>Вкл</b>.\n\n"
        "↕️ <b>Сортировка</b> — меняет только порядок: новые сверху или цена ↑/↓.\n"
        "🔎 <b>Ключевые слова</b> — оставляет объявления с нужными словами в заголовке.\n"
        "🚫 <b>Исключения</b> — убирает объявления со стоп-словами из заголовка.\n\n"
        "<b>✅ Рекомендуемый старт</b>\n"
        "Самые новые · просмотры без порога · дубли Вкл · шум Вкл · сортировка «Сначала новые» · слова пустые.\n\n"
        "После первого скана уже можно экспериментировать: 50+/100+ для уже заметных объявлений, «Ниже рынка» для дешёвых, «Уникальные» для редких, «Часто публикуемые» для повторяющихся моделей."
    )
    await callback.message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="⬅️ К настройкам", callback_data="settings")]]
    ))


@dp.callback_query(F.data.startswith("quickmode:"))
async def quick_mode(callback: CallbackQuery) -> None:
    value = callback.data.split(":", 1)[1]
    if value not in MODE_LABELS:
        await callback.answer("Режим не найден", show_alert=True)
        return
    s = await update_setting(callback.from_user.id, "output_mode", value)
    await callback.answer(f"Выбрано: {MODE_LABELS[value]}")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "set_mode")
async def set_mode(callback: CallbackQuery) -> None:
    await callback.answer()
    opts = [(k, v) for k, v in MODE_LABELS.items()]
    await callback.message.edit_text("<b>Режим результата</b>\n\nВыбери, что попадёт в файл:", reply_markup=choice_keyboard("mode", opts), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data.startswith("mode:"))
async def choose_mode(callback: CallbackQuery) -> None:
    value = callback.data.split(":", 1)[1]
    if value not in MODE_LABELS: return
    s = await update_setting(callback.from_user.id, "output_mode", value)
    await callback.answer("Режим сохранён")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "toggle_dedupe")
async def toggle_dedupe(callback: CallbackQuery) -> None:
    old = await get_settings(callback.from_user.id)
    s = await update_setting(callback.from_user.id, "smart_dedupe", not old.smart_dedupe)
    await callback.answer("Обновлено")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "toggle_noise")
async def toggle_noise(callback: CallbackQuery) -> None:
    old = await get_settings(callback.from_user.id)
    s = await update_setting(callback.from_user.id, "clean_noise", not old.clean_noise)
    await callback.answer("Обновлено")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "set_period")
async def set_period(callback: CallbackQuery) -> None:
    # Compatibility for old Telegram messages created before v3.2.7.
    s = await get_settings(callback.from_user.id)
    await callback.answer("Период перенесён в запуск парсера")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data.startswith("period:"))
async def choose_period(callback: CallbackQuery) -> None:
    # Do not preserve a hidden legacy period: date is now selected only at scan start.
    s = await get_settings(callback.from_user.id)
    await callback.answer("Эта настройка больше не используется")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "set_price")
async def set_price(callback: CallbackQuery) -> None:
    # Compatibility with old Telegram messages created before v4.3.7.
    await callback.answer("Цена теперь выбирается при запуске скана", show_alert=True)


@dp.callback_query(F.data.startswith("price:"))
async def choose_price(callback: CallbackQuery) -> None:
    await callback.answer("Цена теперь выбирается при запуске скана", show_alert=True)


@dp.callback_query(F.data == "set_min_views")
async def set_min_views(callback: CallbackQuery) -> None:
    s = await get_settings(callback.from_user.id)
    await callback.answer()
    await callback.message.edit_text(
        "<b>👁 Минимум просмотров</b>\n\n"
        "В результат попадут только объявления, у которых текущее число просмотров не ниже выбранного порога.\n\n"
        "<i>Важно: бот всё равно должен сначала получить счётчик просмотров у объявления, поэтому эта настройка фильтрует результат, а не уменьшает число запросов к Kleinanzeigen.</i>",
        reply_markup=min_views_keyboard(int(getattr(s, "min_views", 0) or 0)),
        parse_mode=ParseMode.HTML,
    )


@dp.callback_query(F.data.startswith("minviews:"))
async def choose_min_views(callback: CallbackQuery, state: FSMContext) -> None:
    value = callback.data.split(":", 1)[1]
    if value == "custom":
        await state.set_state(SettingsInput.min_views)
        await callback.answer()
        await callback.message.answer(
            "👁 Пришли минимальное количество просмотров числом.\n\n"
            "Например: <code>75</code>\n"
            "Чтобы отключить порог — отправь <code>0</code>.",
            parse_mode=ParseMode.HTML,
        )
        return
    try:
        threshold = max(0, min(10_000_000, int(value)))
    except ValueError:
        await callback.answer("Некорректное значение", show_alert=True)
        return
    s = await update_setting(callback.from_user.id, "min_views", threshold)
    await callback.answer("Порог просмотров сохранён")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.message(SettingsInput.min_views)
async def save_custom_min_views(message: Message, state: FSMContext) -> None:
    raw = (message.text or "").strip().replace(" ", "")
    try:
        threshold = int(raw)
    except ValueError:
        await message.answer("Пришли целое число, например <code>75</code>.", parse_mode=ParseMode.HTML)
        return
    if threshold < 0 or threshold > 10_000_000:
        await message.answer("Укажи значение от 0 до 10 000 000.")
        return
    s = await update_setting(message.from_user.id, "min_views", threshold)
    await state.clear()
    await message.answer(
        "✅ Порог просмотров сохранён.\n\n" + settings_text(s),
        reply_markup=settings_keyboard(s),
        parse_mode=ParseMode.HTML,
    )


@dp.callback_query(F.data == "set_sort")
async def set_sort(callback: CallbackQuery) -> None:
    await callback.answer()
    await callback.message.edit_text("<b>↕️ Сортировка</b>", reply_markup=choice_keyboard("sort", [(k, v) for k, v in SORT_LABELS.items()]), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data.startswith("sort:"))
async def choose_sort(callback: CallbackQuery) -> None:
    value = callback.data.split(":", 1)[1]
    if value not in SORT_LABELS: return
    s = await update_setting(callback.from_user.id, "sort_mode", value)
    await callback.answer("Сортировка сохранена")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "set_include")
async def set_include(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(SettingsInput.include_words)
    await callback.message.answer("✅ Пришли слова через запятую. В результат попадут объявления, содержащие хотя бы одно из них.\n\nНапример: <code>apple tv, playstation, macbook</code>\n\nЧтобы очистить — отправь <code>-</code>.", parse_mode=ParseMode.HTML)


@dp.message(SettingsInput.include_words)
async def save_include(message: Message, state: FSMContext) -> None:
    value = (message.text or "").strip()
    if value == "-": value = ""
    s = await update_setting(message.from_user.id, "include_words", value[:1000])
    await state.clear()
    await message.answer("✅ Ключевые слова сохранены.\n\n" + settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "set_exclude")
async def set_exclude(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(SettingsInput.exclude_words)
    await callback.message.answer("🚫 Пришли исключаемые слова через запятую.\n\nНапример: <code>defekt, hülle, case</code>\n\nЧтобы очистить — отправь <code>-</code>.", parse_mode=ParseMode.HTML)


@dp.message(SettingsInput.exclude_words)
async def save_exclude(message: Message, state: FSMContext) -> None:
    value = (message.text or "").strip()
    if value == "-": value = ""
    s = await update_setting(message.from_user.id, "exclude_words", value[:1000])
    await state.clear()
    await message.answer("🚫 Исключения сохранены.\n\n" + settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "reset_settings")
async def reset_settings(callback: CallbackQuery) -> None:
    s = await reset_user_settings(callback.from_user.id)
    await callback.answer("Настройки сброшены")
    await callback.message.edit_text(settings_text(s), reply_markup=settings_keyboard(s), parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "view_test")
async def view_test(callback: CallbackQuery, state: FSMContext) -> None:
    if not allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await state.set_state(SettingsInput.view_test_url)
    await callback.answer()
    await callback.message.answer(
        "<b>⚡ Тест быстрого получения просмотров</b>\n\n"
        "Пришли одну публичную ссылку Kleinanzeigen. Бот сравнит два способа:\n"
        "1) прямой запрос счётчика без открытия карточки;\n"
        "2) обычное открытие страницы через Chromium как контроль.\n\n"
        "Так мы сразу увидим, сколько времени экономит быстрый режим.",
        parse_mode=ParseMode.HTML,
    )


@dp.message(SettingsInput.view_test_url)
async def run_view_test(message: Message, state: FSMContext) -> None:
    if not allowed(message.from_user.id):
        await state.clear()
        await message.answer("Нет доступа.")
        return
    url = (message.text or "").strip()
    if not (url.startswith("https://") and "kleinanzeigen.de/s-anzeige/" in url):
        await message.answer("⚠️ Пришли именно публичную ссылку на объявление Kleinanzeigen или нажми /start для выхода.")
        return

    await state.clear()
    status = await message.answer(
        "⏳ <b>Сравниваю быстрый запрос и Chromium…</b>\n\n"
        "Тест может добавить единичные просмотры к объявлению.",
        parse_mode=ParseMode.HTML,
    )
    parser = KleinanzeigenParser()
    try:
        t0 = time.perf_counter()
        mode, direct = await parser.probe_direct_view_mode(url, force=True)
        direct_time = time.perf_counter() - t0

        t1 = time.perf_counter()
        browser = await parser.fetch_public_view_count(url, http_fast_path=False)
        browser_time = time.perf_counter() - t1
    finally:
        await parser.close()

    selected = await get_selected(message.from_user.id)
    lines = ["⚡ <b>Тест быстрого счётчика</b>", ""]

    if direct.views is not None:
        lines += [
            f"🚀 Прямой способ: <b>{direct.views}</b> просмотров",
            f"Источник: <code>{html.escape(direct.source)}</code>",
            f"Время: <b>{direct_time:.2f} сек</b>",
        ]
    else:
        lines += [
            "🚀 Прямой способ: <b>не сработал</b>",
            f"Режим: <code>{html.escape(mode)}</code>",
            f"Время попытки: <b>{direct_time:.2f} сек</b>",
        ]
        if direct.error:
            lines.append(f"Ошибка: <code>{html.escape(direct.error[:180])}</code>")

    lines.append("")
    if browser.views is not None:
        lines += [
            f"🌐 Chromium: <b>{browser.views}</b> просмотров",
            f"Источник: <code>{html.escape(browser.source)}</code>",
            f"Время: <b>{browser_time:.2f} сек</b>",
        ]
    else:
        lines += [
            "🌐 Chromium: <b>не удалось получить</b>",
            f"Время: <b>{browser_time:.2f} сек</b>",
        ]

    if direct.views is not None and browser_time > 0:
        speedup = browser_time / max(direct_time, 0.01)
        lines += [
            "",
            f"📈 Ускорение на тесте: <b>примерно ×{speedup:.1f}</b>",
            "✅ Массовый сбор использует direct-счётчик первым; Chromium включается только для неподтверждённых объявлений.",
        ]
    else:
        lines += [
            "",
            "ℹ️ В массовом сборе неподтверждённые direct-счётчики точечно перепроверяются через Chromium.",
        ]

    await status.edit_text(
        "\n".join(lines),
        parse_mode=ParseMode.HTML,
        reply_markup=main_keyboard(len(selected)),
        disable_web_page_preview=True,
    )



def _radar_freshness(value: datetime | None) -> str:
    """Human freshness label for Radar cards, based on the latest signal."""
    if value is None:
        return "давно"
    try:
        if value.tzinfo is None:
            moment = value.replace(tzinfo=timezone.utc)
        else:
            moment = value.astimezone(timezone.utc)
        seconds = max(0, int((datetime.now(timezone.utc) - moment).total_seconds()))
    except Exception:
        return "давно"
    if seconds < 60:
        return "только что"
    minutes = seconds // 60
    if minutes < 60:
        return f"{minutes} мин назад"
    hours = minutes // 60
    if hours < 24:
        return f"{hours} ч назад"
    days = hours // 24
    if days == 1:
        return "вчера"
    if days < 7:
        return f"{days} дн назад"
    return _moscow_text(value)


def _fast_sold_lifetime_text(seconds: int | None) -> str:
    if seconds is None:
        return "—"
    total = max(0, int(seconds))
    minutes = max(1, int(round(total / 60.0)))
    if minutes < 60:
        return f"~{minutes} мин"
    hours, rest = divmod(minutes, 60)
    if rest == 0:
        return f"~{hours} ч"
    return f"~{hours} ч {rest} мин"


def _radar_added_label(value: datetime | None) -> str:
    """Compact category-feed label based on when the product first entered Radar.

    Unlike ``last_signal_at``, this cannot make an old product look new simply
    because another observation updated its score.
    """
    if value is None:
        return "давно"
    try:
        moment_utc = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
        now_utc = datetime.now(timezone.utc)
        age_seconds = max(0, int((now_utc - moment_utc).total_seconds()))
        if age_seconds < 3 * 3600:
            return f"🆕 Новое · {_radar_freshness(value)}"
        moment_msk = moment_utc.astimezone(MOSCOW)
        now_msk = now_utc.astimezone(MOSCOW)
        day_delta = (now_msk.date() - moment_msk.date()).days
        if day_delta <= 0:
            return "🟢 Сегодня"
        if day_delta == 1:
            return "вчера"
        if day_delta < 7:
            return f"{day_delta} дн назад"
        return moment_msk.strftime("%d.%m.%Y")
    except Exception:
        return "давно"


async def _radar_home_text(user_id: int | None = None) -> str:
    stats = await radar_stats()
    full = bool(user_id is not None and allowed(int(user_id)))
    if not full:
        return (
            "📡 <b>DT Radar</b>\n\n"
            "Посмотри, как Radar отбирает сильные товары из тысяч объявлений.\n\n"
            f"🎁 <b>Бесплатно:</b> первые {FREE_RADAR_PREVIEW_LIMIT} находок в каждом режиме «Лучшие сейчас».\n"
            "🔒 Поиск, Категории, Мой Radar и полные ленты открываются с подпиской.\n\n"
            f"Сейчас в Radar: <b>{stats.total}</b> товаров · 🔥 <b>{stats.hot}</b> горячих · 🚀 <b>{stats.rising}</b> набирают · ⚡ <b>{stats.fast_sold}</b> Fast Sold.\n\n"
            "👁 <b>Observed Score</b> строится только на росте просмотров, который DT увидел после своего baseline."
        )
    return (
        "📡 <b>DT Radar</b>\n\n"
        "Найди сильные товары без лишней сложности. Выбери, что хочешь сделать:\n\n"
        "🔥 <b>Лучшие сейчас</b> — горячие, набирающие и быстро исчезнувшие\n"
        "🔎 <b>Поиск</b> — если уже знаешь название товара\n"
        "🗂 <b>Категории</b> — если хочешь посмотреть по разделам\n"
        "⭐ <b>Мой Radar</b> — сохранённые товары\n\n"
        f"Сейчас в Radar: <b>{stats.total}</b> товаров · 🔥 <b>{stats.hot}</b> горячих · 🚀 <b>{stats.rising}</b> набирают · ⚡ <b>{stats.fast_sold}</b> Fast Sold.\n\n"
        "👁 <b>Observed Score</b>: первый счётчик не оценивается; Radar верит только собственным повторным замерам DT."
    )


async def _radar_list_payload(
    user_id: int, mode: str, page: int, category_key: str | None = None, *, preview: bool = False,
    price_filter: str = "any",
):
    if preview:
        page = 0
        rows, total = await list_radar_products(
            mode=mode, category_key=None, page=0, page_size=FREE_RADAR_PREVIEW_LIMIT, user_id=None
        )
    else:
        rows, total = await list_radar_products(
            mode=mode, category_key=category_key, page=page, user_id=user_id,
            price_filter=price_filter if category_key else "any",
        )
    titles = {
        "hot": "🔥 Горячие сейчас",
        "rising": "🚀 Набирают обороты",
        "ai": "🚀 Набирают обороты",
        "fastsold": "⚡ Быстро исчезли",
        "alltime": "🏆 Рекорды Radar",
        "favorites": "⭐ Мой Radar",
    }
    if mode == "alltime" and not category_key and not preview:
        lines_hint = True
    else:
        lines_hint = False
    if category_key:
        cat = CATEGORIES.get(category_key)
        if cat is not None and cat.group in GROUPS:
            group = GROUPS[cat.group]
            heading = f"{group.icon} {group.name} · {cat.name}"
        else:
            heading = f"🗂 {cat.name if cat is not None else category_key}"
    else:
        heading = titles.get(mode, "📡 DT Radar")
    lines = [f"<b>{html.escape(heading)}</b>", ""]
    if lines_hint:
        lines += [
            "Здесь хранится история сильных сигналов. Истёкший live-сигнал не обнуляет доказанный Score.",
            "В кнопке показывается 🏆 Peak Score; live-категории и поиск историю не смешивают.",
            "",
        ]
    fast_infos = {}
    if mode == "fastsold" and rows and not preview:
        fast_infos = await get_fast_sold_infos([int(product.id) for product in rows])
        lines += [
            "Radar автоматически проверяет сильные свежие объявления через 15 / 30 / 60 / 120 / 180 минут.",
            "В список попадает только исчезновение, подтверждённое двумя прямыми проверками.",
            "ℹ️ Исчезновение обычно означает продажу или снятие объявления; Kleinanzeigen не всегда раскрывает причину.",
            "",
        ]
    if preview:
        shown = min(len(rows), FREE_RADAR_PREVIEW_LIMIT)
        lines += [
            f"🎁 <b>Бесплатный просмотр: {shown}/{FREE_RADAR_PREVIEW_LIMIT}</b>",
            "Это реальные актуальные находки из полной базы DT Radar.",
            "",
        ]
    elif category_key:
        if mode == "category_best":
            lines += [
                "🔥 Сортировка: <b>сначала самые сильные подтверждённые сигналы</b>.",
                f"💶 Цена: <b>{html.escape(price_filter_label(price_filter))}</b>.",
                "В списке остаются только <b>актуальные live-сигналы</b>; устаревшие товары уходят в 🏆 Рекорды Radar.",
                "",
            ]
        else:
            lines += [
                "🆕 Сортировка: <b>сначала новые</b>.",
                f"💶 Цена: <b>{html.escape(price_filter_label(price_filter))}</b>.",
                "В списке — только <b>актуальные live-сигналы</b>; история не засоряет категорию.",
                "",
            ]
    if not rows:
        lines.append("Пока здесь нет товаров. База пополняется автоматически после завершённых сканов и AI-анализа.")
    else:
        start_index = 0 if preview else page * RADAR_PAGE_SIZE
        for index, product in enumerate(rows, start_index + 1):
            icon = RADAR_STATUS_ICON.get(str(product.status or ""), "📡")
            cat = CATEGORIES.get(str(product.category_key or ""))
            cat_name = cat.name if cat is not None else str(product.category_key or "—")
            freshness = (
                _radar_added_label(product.first_radar_at)
                if category_key and not preview
                else _radar_freshness(product.last_signal_at)
            )
            if mode == "fastsold" and not preview:
                info = fast_infos.get(int(product.id))
                if info is not None:
                    price = f"{int(info.last_price_eur)} €" if info.last_price_eur is not None else "—"
                    views = str(int(info.last_views)) if info.last_views is not None else "—"
                    lines.append(
                        f"⚡ <b>{index}. {html.escape(str(product.title or info.title or 'Товар')[:70])}</b>\n"
                        f"⏱ Исчезло за <b>{html.escape(_fast_sold_lifetime_text(info.lifetime_seconds))}</b> · "
                        f"🕐 {html.escape(_moscow_text(info.disappeared_at))}\n"
                        f"💶 <b>{html.escape(price)}</b> · 👀 последний замер <b>{views}</b> · "
                        f"🏆 Peak <b>{max(int(product.peak_score or 0), int(info.peak_score or 0))}</b>\n"
                        f"📂 {html.escape(cat_name)}"
                    )
                    continue
            score_line = (
                f"🕒 Последний Score <b>{int(product.current_score or 0)}</b>/100 · Peak <b>{int(product.peak_score or 0)}</b> · "
                if str(product.status or "") == "historical"
                else f"⭐ <b>{int(product.current_score or 0)}</b>/100 · Peak <b>{int(product.peak_score or 0)}</b> · "
            )
            lines.append(
                f"{icon} <b>{index}. {html.escape(str(product.title or 'Товар')[:70])}</b>\n"
                + score_line
                + f"📂 {html.escape(cat_name)}\n"
                + f"💶 <b>{html.escape(_radar_product_price_text(product))}</b> · 🕐 <b>{html.escape(freshness)}</b> · "
                + f"🔁 сигналов: <b>{int(product.signal_count or 0)}</b> · объявлений: <b>{int(product.listing_count or 0)}</b>"
            )
    if preview:
        hidden = max(0, int(total) - len(rows))
        if hidden:
            lines += [
                "",
                f"🔒 Ещё <b>{hidden}</b> товаров в этом режиме доступны в полном DT Radar.",
                "Полный доступ открывает все результаты, 🔎 Поиск, 🗂 Категории и ⭐ Мой Radar.",
            ]
        trial = await get_trial_status(user_id)
        return "\n\n".join(lines), radar_preview_list_keyboard(
            rows, mode=mode, total=total, trial_remaining=(trial.remaining if trial.eligible else 0)
        )
    if total:
        pages = max(1, (total + RADAR_PAGE_SIZE - 1) // RADAR_PAGE_SIZE)
        lines += ["", f"Страница <b>{page + 1}/{pages}</b> · всего <b>{total}</b>"]
    return "\n\n".join(lines), radar_list_keyboard(
        rows, mode=mode, page=page, total=total, category_key=category_key,
        price_filter=price_filter,
    )


async def _radar_product_payload(
    user_id: int, product_id: int, *, preview_mode: str | None = None,
    return_callback: str | None = None, return_text: str | None = None,
):
    product, listing, snapshots = await get_radar_product(product_id)
    if product is None:
        return None, None
    fast_info = await get_fast_sold_info(product_id)
    favorite = False if preview_mode is not None else await is_radar_favorite(user_id, product_id)
    status = RADAR_STATUS_LABEL.get(str(product.status or ""), str(product.status or "—"))
    status_icon = RADAR_STATUS_ICON.get(str(product.status or ""), "📡")
    cat = CATEGORIES.get(str(product.category_key or ""))
    cat_name = cat.name if cat is not None else str(product.category_key or "—")
    type_label = RADAR_TYPE_LABEL.get(str(product.opportunity_type or ""), str(product.opportunity_type or "—"))
    price = "—"
    if product.min_price_eur is not None and product.max_price_eur is not None:
        if int(product.min_price_eur) == int(product.max_price_eur):
            price = f"{int(product.min_price_eur)} €"
        else:
            price = f"{int(product.min_price_eur)}–{int(product.max_price_eur)} €"
    lines = [
        f"📡 <b>{html.escape(str(product.title or 'Товар'))}</b>",
        "",
        f"{status_icon} Статус: <b>{html.escape(status)}</b>",
        (f"🕒 Последний подтверждённый Score: <b>{int(product.current_score or 0)}/100</b> · Peak <b>{int(product.peak_score or 0)}</b>" if str(product.status or "") == "historical" else f"⭐ Observed Score: <b>{int(product.current_score or 0)}/100</b> · Peak <b>{int(product.peak_score or 0)}</b>"),
        f"👁 DT-наблюдаемый прирост: <b>+{int(getattr(product, 'demand_views', 0) or 0)}</b> просмотров",
        f"🧠 Доказательство: <b>{html.escape(type_label)}</b> · уверенность <b>{int(product.confidence or 0)}%</b>",
        f"🗂 Категория: <b>{html.escape(cat_name)}</b>",
        f"💶 Диапазон замеченных цен: <b>{html.escape(price)}</b>",
        f"👀 Текущий счётчик: <b>{int(product.best_views or 0)}</b> просмотров (baseline не оценивается)",
        f"⚡ Лучший DT-наблюдаемый рост: <b>{float(product.best_views_per_hour or 0.0):.1f}/ч</b>",
        f"🔁 Подтверждённых замеров: <b>{int(product.signal_count or 0)}</b>",
        f"📦 Независимых объявлений: <b>{int(product.listing_count or 0)}</b>",
                f"🕒 В Radar с: <b>{html.escape(_moscow_text(product.first_radar_at))}</b>",
        f"📡 Последний сигнал: <b>{html.escape(_moscow_text(product.last_signal_at))}</b>",
        f"🕐 Свежесть: <b>{html.escape(_radar_freshness(product.last_signal_at))}</b>",
    ]
    if fast_info is not None:
        lines += [
            "",
            "⚡ <b>Fast Sold / быстро исчезло</b>",
            f"⏱ Было доступно примерно: <b>{html.escape(_fast_sold_lifetime_text(fast_info.lifetime_seconds))}</b>",
            f"❌ Исчезновение замечено: <b>{html.escape(_moscow_text(fast_info.disappeared_at))}</b>",
            f"✅ Повторно подтверждено: <b>{html.escape(_moscow_text(fast_info.confirmed_at))}</b>",
        ]
        if fast_info.last_views is not None:
            lines.append(f"👀 Последний известный замер: <b>{int(fast_info.last_views)}</b>")
        lines.append("ℹ️ Kleinanzeigen не всегда сообщает, было ли объявление продано или снято продавцом.")
    if product.latest_reason:
        lines += ["", f"💡 <b>Почему в Radar:</b> {html.escape(str(product.latest_reason)[:500])}"]
    if snapshots:
        lines += ["", "<b>Последние изменения рейтинга:</b>"]
        for snap in snapshots[:5]:
            icon = "🧠" if str(snap.source or "").startswith("ai") else "👀"
            lines.append(
                f"{icon} {_moscow_text(snap.recorded_at)} · ⭐ <b>{int(snap.score or 0)}</b> · "
                f"{html.escape(RADAR_TYPE_LABEL.get(str(snap.opportunity_type or ''), str(snap.stage or 'signal')))}"
            )
    listing_url = str(listing.url) if listing is not None and listing.url else None
    if preview_mode is not None:
        lines += ["", "🎁 <b>Бесплатная находка DT Radar</b> · полный доступ открывает всю базу, поиск и категории."]
    return "\n".join(lines), radar_product_keyboard(
        product_id, favorite=favorite, listing_url=listing_url, preview_mode=preview_mode,
        return_callback=return_callback, return_text=return_text,
    )


async def _radar_preview_product_allowed(product_id: int, mode: str) -> bool:
    if mode not in {"hot", "rising", "ai"}:
        return False
    rows, _total = await list_radar_products(mode=mode, page=0, page_size=FREE_RADAR_PREVIEW_LIMIT)
    return any(int(product.id) == int(product_id) for product in rows)


async def _radar_search_payload(query: str, page: int = 0, *, price_filter: str = "any"):
    clean = " ".join(str(query or "").split())[:80]
    rows, total = await search_radar_products(clean, page=page, price_filter=price_filter)
    lines = [
        f"🔎 <b>Поиск DT Radar</b>",
        f"Запрос: <b>{html.escape(clean)}</b>",
        f"💶 Цена: <b>{html.escape(price_filter_label(price_filter))}</b>",
        "",
    ]
    if not rows:
        lines.append("Ничего не нашёл. Попробуй более короткое название или модель.")
    else:
        start = page * RADAR_PAGE_SIZE
        for index, product in enumerate(rows, start + 1):
            icon = RADAR_STATUS_ICON.get(str(product.status or ""), "📡")
            cat = CATEGORIES.get(str(product.category_key or ""))
            cat_name = cat.name if cat is not None else str(product.category_key or "—")
            lines.append(
                f"{icon} <b>{index}. {html.escape(str(product.title or 'Товар')[:70])}</b>\n"
                f"⭐ <b>{int(product.current_score or 0)}</b>/100 · 💶 <b>{html.escape(_radar_product_price_text(product))}</b> · "
                f"📂 {html.escape(cat_name)}\n"
                f"🕐 <b>{html.escape(_radar_freshness(product.last_signal_at))}</b>"
            )
    if total:
        pages = max(1, (total + RADAR_PAGE_SIZE - 1) // RADAR_PAGE_SIZE)
        lines += ["", f"Страница <b>{page + 1}/{pages}</b> · найдено <b>{total}</b>"]
    return "\n\n".join(lines), radar_search_keyboard(
        rows, page=page, total=total, price_filter=price_filter
    )


async def _render_radar_context(message: Message, user_id: int, state: FSMContext) -> None:
    data = await state.get_data()
    kind = str(data.get("radar_context_kind") or "")
    page = max(0, int(data.get("radar_context_page") or 0))
    price_filter = _radar_price_filter(data)
    if kind == "category":
        category_key = str(data.get("radar_context_category") or "")
        mode = str(data.get("radar_context_mode") or "category_new")
        if category_key in CATEGORIES:
            text, markup = await _radar_list_payload(
                user_id, mode, page, category_key=category_key, price_filter=price_filter
            )
            await _edit_or_answer(message, text, reply_markup=markup)
            return
    if kind == "search":
        query = str(data.get("radar_search_query") or "").strip()
        if query:
            text, markup = await _radar_search_payload(query, page, price_filter=price_filter)
            await _edit_or_answer(message, text, reply_markup=markup)
            return
    if kind == "list":
        mode = str(data.get("radar_context_mode") or "hot")
        if mode in {"hot", "rising", "ai", "fastsold", "alltime", "favorites"}:
            text, markup = await _radar_list_payload(user_id, mode, page)
            await _edit_or_answer(message, text, reply_markup=markup)
            return
    await _edit_or_answer(
        message, await _radar_home_text(user_id), reply_markup=radar_home_keyboard(user_id)
    )


@dp.callback_query(F.data.startswith("radar_locked"))
async def radar_locked(callback: CallbackQuery) -> None:
    if allowed(callback.from_user.id):
        await callback.answer("Полный DT Radar уже открыт ✅")
        await _edit_or_answer(
            callback.message, await _radar_home_text(callback.from_user.id),
            reply_markup=radar_home_keyboard(callback.from_user.id),
        )
        return
    stats = await radar_stats()
    await callback.answer()
    feature = ""
    if ":" in str(callback.data or ""):
        feature = str(callback.data or "").split(":", 1)[1]
    labels = {
        "search": "🔎 Поиск",
        "categories": "🗂 Категории",
        "favorites": "⭐ Мой Radar",
        "records": "🏆 Рекорды Radar",
        "fastsold": "⚡ Быстро исчезли",
    }
    if free_radar_preview_allowed(callback.from_user.id):
        await record_free_radar_event(callback.from_user.id, "locked_feature", feature=feature)
    title = labels.get(feature, "Полный DT Radar")
    text = (
        f"🔒 <b>{html.escape(title)}</b>\n\n"
        f"Эта функция доступна в полном DT Radar. Бесплатно можно посмотреть первые <b>{FREE_RADAR_PREVIEW_LIMIT}</b> "
        "реальных находок в каждом режиме «Лучшие сейчас».\n\n"
        f"📦 В Radar уже: <b>{stats.total}</b> товаров\n"
        f"🔥 Горячих: <b>{stats.hot}</b> · 🚀 Набирают: <b>{stats.rising}</b>\n"
        f"⚡ Быстро исчезли: <b>{stats.fast_sold}</b>\n\n"
        "💎 Полный доступ открывает все результаты, поиск, категории, сохранения и историю Radar."
    )
    await _edit_or_answer(callback.message, text, reply_markup=radar_locked_keyboard())


@dp.callback_query(F.data == "radardaily_open")
async def radar_daily_open(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(None)
    if free_radar_preview_allowed(callback.from_user.id):
        await record_free_radar_event(callback.from_user.id, "daily_digest_open", feature="daily_digest")
        await record_free_radar_event(callback.from_user.id, "radar_open", feature="daily_digest")
    await callback.answer()
    await _edit_or_answer(
        callback.message, await _radar_home_text(callback.from_user.id),
        reply_markup=radar_home_keyboard(callback.from_user.id),
    )


@dp.callback_query(F.data == "radar_home")
async def radar_home(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(None)
    await state.update_data(radar_context_kind="", radar_context_page=0)
    if free_radar_preview_allowed(callback.from_user.id):
        await record_free_radar_event(callback.from_user.id, "radar_open", feature="home")
    await callback.answer()
    # Score cooling is maintained hourly in the background. Opening Radar stays a
    # small indexed read even after the accumulated base grows very large.
    await _edit_or_answer(
        callback.message, await _radar_home_text(callback.from_user.id),
        reply_markup=radar_home_keyboard(callback.from_user.id),
    )


@dp.message(Command("radar"))
async def radar_command(message: Message) -> None:
    if free_radar_preview_allowed(message.from_user.id):
        await record_free_radar_event(message.from_user.id, "radar_open", feature="command")
    await message.answer(
        await _radar_home_text(message.from_user.id), parse_mode=ParseMode.HTML,
        reply_markup=radar_home_keyboard(message.from_user.id),
    )


@dp.callback_query(F.data == "radarbest")
async def radar_best_handler(callback: CallbackQuery) -> None:
    stats = await radar_stats()
    full = allowed(callback.from_user.id)
    if not full and free_radar_preview_allowed(callback.from_user.id):
        await record_free_radar_event(callback.from_user.id, "best_open")
    await callback.answer()
    if full:
        text = (
            "🔥 <b>Лучшие сейчас</b>\n\n"
            "Выбери, какие сильные товары хочешь посмотреть:\n\n"
            f"🔥 Горячие: <b>{stats.hot}</b>\n"
            f"🚀 Набирают: <b>{stats.rising}</b>\n"
            
            f"⚡ Быстро исчезли: <b>{stats.fast_sold}</b>\n\n"
            "⚡ «Быстро исчезли» — объявления, которые Radar видел активными и затем подтвердил недоступными в первые 3 часа.\n\n"
            "🏆 Рекорды Radar оставлены ниже как дополнительная история."
        )
    else:
        text = (
            "🔥 <b>Лучшие сейчас</b>\n\n"
            "Мы уже отобрали сильные товары из DT Radar.\n"
            f"🎁 <b>Бесплатно покажем первые {FREE_RADAR_PREVIEW_LIMIT} находок</b> в выбранном режиме.\n\n"
            f"🔥 Горячие: <b>{stats.hot}</b>\n"
            f"🚀 Набирают: <b>{stats.rising}</b>\n"
            
            f"⚡ Быстро исчезли: <b>{stats.fast_sold}</b> · 🔒\n\n"
            "Выбери режим и посмотри реальные результаты Radar."
        )
    await _edit_or_answer(callback.message, text, reply_markup=radar_best_keyboard(callback.from_user.id))


@dp.callback_query(F.data == "radarsearch")
async def radar_search_begin(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(RadarInput.search)
    await state.update_data(radar_search_query="", radar_context_kind="search", radar_context_page=0)
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "🔎 <b>Поиск DT Radar</b>\n\nНапиши название товара или модели.\nНапример: <code>Apple TV</code> или <code>PlayStation Portal</code>.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ DT Radar", callback_data="radar_home")]]),
    )


@dp.message(RadarInput.search)
async def radar_search_message(message: Message, state: FSMContext) -> None:
    query = " ".join(str(message.text or "").split())
    if len(query) < 2:
        await message.answer("Напиши хотя бы 2 символа для поиска.")
        return
    query = query[:80]
    await state.update_data(
        radar_search_query=query, radar_context_kind="search", radar_context_page=0
    )
    await state.set_state(None)
    data = await state.get_data()
    text, markup = await _radar_search_payload(query, 0, price_filter=_radar_price_filter(data))
    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=markup)


@dp.callback_query(F.data.startswith("radarsearchpage:"))
async def radar_search_page(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    query = str(data.get("radar_search_query") or "").strip()
    if not query:
        await callback.answer("Сначала выполни новый поиск", show_alert=True)
        return
    try:
        page = max(0, int(callback.data.split(":", 1)[1]))
    except Exception:
        page = 0
    await state.update_data(radar_context_kind="search", radar_context_page=page)
    data = await state.get_data()
    await callback.answer()
    text, markup = await _radar_search_payload(query, page, price_filter=_radar_price_filter(data))
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data == "radarprice:open")
async def radar_price_open(callback: CallbackQuery, state: FSMContext) -> None:
    if not allowed(callback.from_user.id):
        await callback.answer("Доступно в полном DT Radar", show_alert=True)
        return
    data = await state.get_data()
    if str(data.get("radar_context_kind") or "") not in {"category", "search"}:
        await callback.answer("Открой категорию или результаты поиска", show_alert=True)
        return
    current = _radar_price_filter(data)
    await state.set_state(None)
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "💶 <b>Фильтр цены DT Radar</b>\n\n"
        f"Сейчас: <b>{html.escape(price_filter_label(current))}</b>.\n"
        "Выбери диапазон. Radar оставит товары, у которых есть реально замеченное объявление в этой цене.",
        reply_markup=radar_price_keyboard(current),
    )


@dp.callback_query(F.data.startswith("radarprice:set:"))
async def radar_price_set(callback: CallbackQuery, state: FSMContext) -> None:
    if not allowed(callback.from_user.id):
        await callback.answer("Доступно в полном DT Radar", show_alert=True)
        return
    value = str(callback.data or "").removeprefix("radarprice:set:").strip()
    allowed_values = {"any", "0_50", "50_100", "100_200", "200_500", "500_plus"}
    if value not in allowed_values:
        await callback.answer("Неизвестный диапазон", show_alert=True)
        return
    await state.set_state(None)
    await state.update_data(radar_price_filter=value, radar_context_page=0)
    await callback.answer(f"Цена: {price_filter_label(value)}")
    await _render_radar_context(callback.message, callback.from_user.id, state)


@dp.callback_query(F.data == "radarprice:custom")
async def radar_price_custom(callback: CallbackQuery, state: FSMContext) -> None:
    if not allowed(callback.from_user.id):
        await callback.answer("Доступно в полном DT Radar", show_alert=True)
        return
    await state.set_state(RadarInput.price)
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        "✍️ <b>Свой диапазон цены</b>\n\n"
        "Напиши, например:\n"
        "<code>120-250</code>\n"
        "<code>до 100</code>\n"
        "<code>500+</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="⬅️ Назад", callback_data="radarprice:open")
        ]]),
    )


@dp.message(RadarInput.price)
async def radar_price_custom_message(message: Message, state: FSMContext) -> None:
    value = parse_scan_price_input(message.text)
    if value is None:
        await message.answer(
            "Не понял диапазон. Напиши, например: <code>120-250</code>, <code>до 100</code> или <code>500+</code>.",
            parse_mode=ParseMode.HTML,
        )
        return
    await state.update_data(radar_price_filter=value, radar_context_page=0)
    await state.set_state(None)
    data = await state.get_data()
    kind = str(data.get("radar_context_kind") or "")
    price_filter = _radar_price_filter(data)
    if kind == "category":
        category_key = str(data.get("radar_context_category") or "")
        mode = str(data.get("radar_context_mode") or "category_new")
        text, markup = await _radar_list_payload(
            message.from_user.id, mode, 0, category_key=category_key, price_filter=price_filter
        )
    elif kind == "search" and str(data.get("radar_search_query") or "").strip():
        text, markup = await _radar_search_payload(
            str(data.get("radar_search_query") or ""), 0, price_filter=price_filter
        )
    else:
        text, markup = await _radar_home_text(message.from_user.id), radar_home_keyboard(message.from_user.id)
    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=markup)


@dp.callback_query(F.data == "radarprice:back")
async def radar_price_back(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(None)
    await callback.answer()
    await _render_radar_context(callback.message, callback.from_user.id, state)


@dp.callback_query(F.data.startswith("radarlist:"))
async def radar_list_handler(callback: CallbackQuery, state: FSMContext) -> None:
    parts = callback.data.split(":")
    mode = parts[1] if len(parts) > 1 else "hot"
    if mode == "ai":
        # Compatibility for old inline messages after legacy AI Picks retirement.
        mode = "rising"
    try:
        page = max(0, int(parts[2])) if len(parts) > 2 else 0
    except Exception:
        page = 0
    if mode not in {"hot", "rising", "ai", "fastsold", "alltime", "favorites"}:
        mode = "hot"
    full = allowed(callback.from_user.id)
    if not full:
        if not free_radar_preview_allowed(callback.from_user.id) or mode not in {"hot", "rising", "ai"} or page != 0:
            await callback.answer("Доступно в полном DT Radar", show_alert=True)
            if callback.message:
                stats = await radar_stats()
                text = (
                    "🔒 <b>Полный DT Radar</b>\n\n"
                    f"Бесплатно доступны первые <b>{FREE_RADAR_PREVIEW_LIMIT}</b> находок в Горячих и Набирают.\n"
                    f"В полной базе уже <b>{stats.total}</b> товаров.\n\n"
                    "💎 Подписка открывает все результаты, поиск, категории и Мой Radar."
                )
                await _edit_or_answer(callback.message, text, reply_markup=radar_locked_keyboard())
            return
        await callback.answer()
        text, markup = await _radar_list_payload(callback.from_user.id, mode, 0, preview=True)
        await record_free_radar_event(callback.from_user.id, "mode_open", mode=mode, item_count=FREE_RADAR_PREVIEW_LIMIT)
        await _edit_or_answer(callback.message, text, reply_markup=markup)
        return
    await state.update_data(
        radar_context_kind="list", radar_context_mode=mode, radar_context_page=page,
        radar_context_category="",
    )
    await callback.answer()
    text, markup = await _radar_list_payload(callback.from_user.id, mode, page)
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data.startswith("radarpreviewitem:"))
async def radar_preview_item_handler(callback: CallbackQuery) -> None:
    parts = str(callback.data or "").split(":")
    if len(parts) != 3:
        await callback.answer("Товар не найден", show_alert=True)
        return
    mode = parts[1]
    if mode == "ai":
        mode = "rising"
    try:
        product_id = int(parts[2])
    except Exception:
        await callback.answer("Товар не найден", show_alert=True)
        return
    if allowed(callback.from_user.id):
        text, markup = await _radar_product_payload(callback.from_user.id, product_id)
    else:
        if not free_radar_preview_allowed(callback.from_user.id) or not await _radar_preview_product_allowed(product_id, mode):
            await callback.answer("Эта находка доступна в полном DT Radar", show_alert=True)
            return
        text, markup = await _radar_product_payload(callback.from_user.id, product_id, preview_mode=mode)
    if text is None:
        await callback.answer("Товар не найден", show_alert=True)
        return
    if not allowed(callback.from_user.id) and free_radar_preview_allowed(callback.from_user.id):
        await record_free_radar_event(
            callback.from_user.id, "preview_item", mode=mode, product_id=product_id
        )
    await callback.answer()
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data.startswith("radarcats:"))
async def radar_categories_handler(callback: CallbackQuery) -> None:
    items = await radar_categories()
    await callback.answer()
    text = (
        "🗂 <b>DT Radar · Категории</b>\n\n"
        "Сначала выбери <b>большой раздел</b>. Затем Radar покажет только его подкатегории.\n\n"
        "Число справа — сколько <b>всего отобранных Radar-товаров</b> хранится в разделе."
    )
    await _edit_or_answer(callback.message, text, reply_markup=radar_groups_keyboard(items))


@dp.callback_query(F.data.startswith("radargroup:"))
async def radar_group_handler(callback: CallbackQuery) -> None:
    group_key = callback.data.split(":", 1)[1] if ":" in callback.data else ""
    group = GROUPS.get(group_key)
    if group is None:
        await callback.answer("Раздел не найден", show_alert=True)
        return
    items = await radar_categories()
    stats = _radar_category_stats(items)
    total = sum(
        stats.get(cat.key, (0, 0, 0))[0]
        for cat in categories_for_group(group_key)
        if not cat.is_group
    )
    new_today = sum(
        stats.get(cat.key, (0, 0, 0))[1]
        for cat in categories_for_group(group_key)
        if not cat.is_group
    )
    await callback.answer()
    text = (
        f"{group.icon} <b>{html.escape(group.name)}</b>\n\n"
        "Выбери нужную <b>подкатегорию</b>.\n"
        f"Всего отобрано: <b>{int(total)}</b> · 🆕 новых сегодня: <b>{int(new_today)}</b>.\n\n"
        "Внутри подкатегории новые товары идут первыми, но вся отобранная база остаётся доступна."
    )
    await _edit_or_answer(callback.message, text, reply_markup=radar_group_keyboard(group_key, items))


@dp.callback_query(F.data.startswith("radarcat:"))
async def radar_category_handler(callback: CallbackQuery, state: FSMContext) -> None:
    parts = callback.data.split(":")
    if len(parts) < 2:
        await callback.answer("Категория не найдена", show_alert=True); return
    category_key = parts[1]
    try:
        page = max(0, int(parts[2])) if len(parts) > 2 else 0
    except Exception:
        page = 0
    await state.update_data(
        radar_context_kind="category", radar_context_mode="category_new",
        radar_context_category=category_key, radar_context_page=page,
    )
    data = await state.get_data()
    await callback.answer()
    text, markup = await _radar_list_payload(
        callback.from_user.id, "category_new", page, category_key=category_key,
        price_filter=_radar_price_filter(data),
    )
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data.startswith("radarcatbest:"))
async def radar_category_best_handler(callback: CallbackQuery, state: FSMContext) -> None:
    parts = callback.data.split(":")
    if len(parts) < 2:
        await callback.answer("Категория не найдена", show_alert=True); return
    category_key = parts[1]
    try:
        page = max(0, int(parts[2])) if len(parts) > 2 else 0
    except Exception:
        page = 0
    await state.update_data(
        radar_context_kind="category", radar_context_mode="category_best",
        radar_context_category=category_key, radar_context_page=page,
    )
    data = await state.get_data()
    await callback.answer()
    text, markup = await _radar_list_payload(
        callback.from_user.id, "category_best", page, category_key=category_key,
        price_filter=_radar_price_filter(data),
    )
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data.startswith("radaritem:"))
async def radar_item_handler(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        product_id = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Товар не найден", show_alert=True); return
    data = await state.get_data()
    return_callback, return_text = _radar_context_back(data)
    text, markup = await _radar_product_payload(
        callback.from_user.id, product_id, return_callback=return_callback, return_text=return_text
    )
    if text is None:
        await callback.answer("Товар не найден", show_alert=True); return
    await callback.answer()
    await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data.startswith("radarfav:"))
async def radar_favorite_handler(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        product_id = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Товар не найден", show_alert=True); return
    favorite = await toggle_radar_favorite(callback.from_user.id, product_id)
    await callback.answer("Добавлено в Мой Radar ⭐" if favorite else "Убрано из Моего Radar")
    data = await state.get_data()
    return_callback, return_text = _radar_context_back(data)
    text, markup = await _radar_product_payload(
        callback.from_user.id, product_id, return_callback=return_callback, return_text=return_text
    )
    if text is not None:
        await _edit_or_answer(callback.message, text, reply_markup=markup)


@dp.callback_query(F.data == "popular_now")
async def popular_now(callback: CallbackQuery) -> None:
    if not allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    await callback.answer()
    items = await get_user_popular_categories(callback.from_user.id)
    if not items:
        text = (
            "🔥 <b>Популярное</b>\n\n"
            "После первого успешного скана здесь появится актуальный TOP по категории."
        )
    else:
        text = (
            "🔥 <b>Популярное</b>\n\n"
            "Выбери категорию — показываем только её <b>последний успешный скан</b>.\n"
            "TOP роста доступен по замерам 3 / 6 / 12 часов; автозамеры включаются по желанию."
        )
    await _edit_or_answer(
        callback.message, text, reply_markup=popular_categories_keyboard(items)
    )


@dp.callback_query(F.data.startswith("popularcat:"))
async def popular_category(callback: CallbackQuery) -> None:
    category_key = callback.data.split(":", 1)[1]
    cat = CATEGORIES.get(category_key)
    rows, scans = await get_category_scan_rows(callback.from_user.id, category_key)
    if cat is None or not scans:
        await callback.answer("Категория или сканы не найдены", show_alert=True); return
    scan_settings = await get_settings_for_scan(scans[0])
    rows = apply_listing_settings(rows, scan_settings, exact_date_scan=True, apply_output_mode=True)
    scan = scans[0]
    viewed = sum(1 for row in rows if row.view_count is not None)
    await callback.answer()
    finished_label = _moscow_text(scan.finished_at or scan.created_at)
    text = (
        f"🔥 <b>{html.escape(cat.name)}</b>\n\n"
        f"📅 <b>{html.escape(_date_label(scan.target_date))}</b>\n"
        f"🕒 Последний скан: <b>{html.escape(finished_label)}</b>\n"
        f"📦 Объявлений: <b>{len(rows)}</b> · 👁 С просмотрами: <b>{viewed}</b>\n\n"
        "Актуальный TOP по последнему успешному скану."
    )
    await callback.message.answer(
        text, parse_mode=ParseMode.HTML,
        reply_markup=popular_category_keyboard(scan.id, category_key),
    )


@dp.callback_query(F.data.startswith("pcv:"))
async def popular_category_views(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Некорректный запрос", show_alert=True); return
    try:
        representative_scan_id, category_key = int(parts[1]), parts[2]
    except Exception:
        await callback.answer("Некорректный запрос", show_alert=True); return
    cat = CATEGORIES.get(category_key)
    rows, scans = await get_category_scan_rows(callback.from_user.id, category_key)
    if cat is None or not scans:
        await callback.answer("Сканы категории не найдены", show_alert=True); return
    # Prefer the newest owned scan for navigation even when an old Telegram
    # message contains a stale representative scan id.
    scan_id = scans[0].id
    scan_settings = await get_settings_for_scan(scans[0])
    rows = apply_listing_settings(rows, scan_settings, exact_date_scan=True, apply_output_mode=True)
    rows = [row for row in rows if row.view_count is not None]
    rows.sort(key=lambda row: (row.view_count or 0, row.first_seen_at), reverse=True)
    await callback.answer()
    if not rows:
        text = f"👁 <b>{html.escape(cat.name)}</b>\n\nВ последнем успешном скане пока нет объявлений с подходящими данными просмотров."
    else:
        lines = [
            f"👁 <b>Самые просматриваемые · {html.escape(cat.name)}</b>",
            f"Последний успешный скан · дата объявлений: <b>{html.escape(_date_label(scans[0].target_date))}</b>",
            "",
        ]
        for i, row in enumerate(rows[:GROWTH_TELEGRAM_LIMIT], 1):
            lines.append(
                f"<b>{i}. {html.escape(row.title[:60])}</b>\n"
                f"📅 {_date_label(row.posted_date_msk)} · 👁 <b>{row.view_count}</b> · "
                f"💶 {html.escape(_price_display(row.price_text, row.price_eur))}\n"
                f'<a href="{html.escape(row.url)}">Открыть</a>'
            )
        text = "\n\n".join(lines)
    await callback.message.answer(
        text, parse_mode=ParseMode.HTML, disable_web_page_preview=True,
        reply_markup=popular_category_keyboard(scan_id, category_key),
    )


@dp.callback_query(F.data.startswith("pcg:"))
async def popular_category_growth(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    if len(parts) != 4:
        await callback.answer("Некорректный запрос", show_alert=True); return
    try:
        representative_scan_id, category_key, period_hours = int(parts[1]), parts[2], int(parts[3])
    except Exception:
        await callback.answer("Некорректный запрос", show_alert=True); return
    if period_hours not in OBSERVATION_HOURS:
        period_hours = 3
    scans = await get_user_category_scans(callback.from_user.id, category_key)
    cat = CATEGORIES.get(category_key)
    if not scans or cat is None:
        await callback.answer("Сканы категории не найдены", show_alert=True); return
    scan_id = scans[0].id
    growth, scan_count, rounds = await get_category_growth_rows(
        callback.from_user.id, category_key, period_hours
    )
    scan_settings = await get_settings_for_scan(scans[0])
    allowed_growth_rows = apply_listing_settings(
        [item.listing for item in growth], scan_settings, exact_date_scan=True, apply_output_mode=True
    )
    allowed_growth_ids = {row.external_id for row in allowed_growth_rows}
    growth = [item for item in growth if item.listing.external_id in allowed_growth_ids]
    await callback.answer()
    period_label = f"{period_hours} ч"
    if not growth:
        text = (
            f"🚀 <b>TOP роста · {html.escape(cat.name)} · {period_label}</b>\n\n"
            "Проверен последний успешный скан категории. "
            "Контрольные замеры для этого периода ещё не готовы или прироста пока нет. "
            "Автозамеры 3 / 6 / 12 часов выполняются только если они включены в главном меню."
        )
    else:
        lines = [
            f"🚀 <b>TOP роста · {html.escape(cat.name)} · {period_label}</b>",
            "Последний успешный скан категории.",
            "Сортировка: <b>кто набрал больше всего новых просмотров</b>.",
            "",
        ]
        for i, item in enumerate(growth[:GROWTH_TELEGRAM_LIMIT], 1):
            row = item.listing
            lines.append(
                f"<b>{i}. {html.escape(row.title[:60])}</b>\n"
                f"📅 {_date_label(row.posted_date_msk)} · 👁 {item.base_views} → <b>{item.current_views}</b> · "
                f"🚀 <b>+{item.delta}</b> · ⚡ {item.per_hour:.1f}/ч\n"
                f"💶 {html.escape(_price_display(row.price_text, row.price_eur))} · "
                f'<a href="{html.escape(row.url)}">Открыть</a>'
            )
        lines += ["", f"📊 Полный рейтинг: до <b>{GROWTH_TOP_LIMIT}</b> товаров в таблице."]
        text = "\n\n".join(lines)
    await callback.message.answer(
        text, parse_mode=ParseMode.HTML, disable_web_page_preview=True,
        reply_markup=growth_period_keyboard(scan_id, period_hours, category_key=category_key),
    )


@dp.callback_query(F.data == "my_scans")
async def my_scans(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not read_only_history_allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    await callback.answer()
    scans, archive_count = await get_user_scans_overview(callback.from_user.id, 10)
    if not scans:
        text = (
            "<b>📊 Мои сканы</b>\n\nСвежих сканов пока нет. После завершения они будут храниться здесь 24 часа, затем уйдут в Архив."
        )
    else:
        text = (
            "<b>📊 Мои сканы</b>\n\nТекущие и свежие запуски за последние <b>24 часа</b>. Старые автоматически уходят в Архив; данные не удаляются."
        )
    await _edit_or_answer(
        callback.message,
        text,
        reply_markup=my_scans_keyboard(scans, archive_count),
    )


@dp.callback_query(F.data == "archive_my_scans")
async def archive_my_scans(callback: CallbackQuery) -> None:
    if not read_only_history_allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    moved = await archive_active_finished_scans(callback.from_user.id)
    scans = await get_user_scans(callback.from_user.id, 10)
    archive_count = await get_archive_count(callback.from_user.id)
    await callback.answer(f"В архив перемещено: {moved}")
    await callback.message.edit_text(
        "<b>📊 Мои сканы</b>\n\n"
        f"📦 Перемещено в архив: <b>{moved}</b>.\n"
        "Активный/ожидающий парсинг остаётся здесь. Данные сканов не удаляются.",
        parse_mode=ParseMode.HTML,
        reply_markup=my_scans_keyboard(scans, archive_count),
    )


@dp.callback_query(F.data.startswith("scan_archive:"))
async def scan_archive(callback: CallbackQuery) -> None:
    if not read_only_history_allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True); return
    try:
        page = max(0, int(callback.data.split(":", 1)[1]))
    except Exception:
        page = 0
    scans, total = await get_user_archive(callback.from_user.id, page)
    if total and not scans and page > 0:
        page = max(0, (total - 1) // SCAN_ARCHIVE_PAGE_SIZE)
        scans, total = await get_user_archive(callback.from_user.id, page)
    await callback.answer()
    text = (
        "<b>📦 Архив сканов</b>\n\n"
        f"Всего: <b>{total}</b>. Здесь хранятся сканы старше 24 часов и те, "
        "которые ты убрал вручную.\n\n"
        "Архив <b>не удаляет данные</b>: история просмотров и аналитика «Популярное сейчас» сохраняются."
    )
    await callback.message.edit_text(
        text, parse_mode=ParseMode.HTML,
        reply_markup=scan_archive_keyboard(scans, page, total),
    )


@dp.callback_query(F.data == "archive_noop")
async def archive_noop(callback: CallbackQuery) -> None:
    await callback.answer()


async def render_scan_detail(scan: UserScan) -> str:
    """Compact product-style scan card. Show diagnostics only when actionable."""
    pairs = await get_scan_rows(scan.id)
    scan_settings = await get_settings_for_scan(scan)
    allowed_rows = apply_listing_settings(
        [listing for listing, _ in pairs], scan_settings, exact_date_scan=True, apply_output_mode=True
    )
    allowed_ids = {row.external_id for row in allowed_rows}
    pairs = [pair for pair in pairs if pair[0].external_id in allowed_ids]
    rows = [listing for listing, _ in pairs]
    viewed = sum(1 for row in rows if row.view_count is not None)
    disappeared = sum(1 for row in rows if not row.is_active)

    growers = 0
    total_growth = 0
    for listing, snap in pairs:
        if listing.view_count is not None and snap.initial_view_count is not None:
            delta = listing.view_count - snap.initial_view_count
            if delta > 0:
                growers += 1
                total_growth += delta

    history_rounds = await get_scan_history_rounds(scan.id, limit=50)
    observation_statuses = await get_scan_observation_statuses(scan.id)
    auto_enabled = bool(getattr(scan_settings, "auto_observations", False))
    status_icons = {
        "done": "✅", "pending": "⏳", "running": "🔄", "missed": "▫️",
        "error": "⚠️", "cancelled": "⛔",
    }
    observation_line = (
        " · ".join(
            f"{hours}ч{status_icons.get(observation_statuses.get(hours, 'pending'), '⏳')}"
            for hours in OBSERVATION_HOURS
        )
        if auto_enabled else "⛔ выключены"
    )

    quality_value = int(getattr(scan, "quality_score", 0) or 0)
    status_label = {
        "done": "✅ Завершён",
        "partial": "⚠️ Частичный результат",
        "running": "🔄 Выполняется",
        "queued": "⏳ Ожидает",
        "cancelling": "⏹ Останавливается",
        "cancelled": "⏹ Остановлен",
        "failed": "❌ Ошибка",
    }.get(scan.status, scan.status)
    depth = scan.page_limit if scan.page_limit in PAGE_LIMIT_CHOICES else 50

    lines = [
        f"<b>📊 {html.escape(scan.title)}</b>",
        status_label,
        "",
        f"📅 <b>{_date_label(scan.target_date)}</b> · 📄 <b>{depth} стр.</b>",
        f"💶 Цена: <b>{html.escape(price_filter_label(getattr(scan, 'price_filter', 'any')))}</b>",
        f"📦 Объявлений: <b>{len(rows)}</b> · 👁 С просмотрами: <b>{viewed}</b>",
    ]
    if scan.total_categories > 1:
        lines.append(f"🗂 Категории: <b>{scan.completed_categories}/{scan.total_categories}</b>")

    if growers > 0:
        lines.append(f"🚀 Набирают просмотры: <b>{growers}</b> · всего <b>+{total_growth}</b>")
    elif len(history_rounds) < 2:
        lines.append("🚀 Рост появится после первого контрольного замера")
    if disappeared > 0:
        lines.append(f"▫️ Исчезли: <b>{disappeared}</b>")

    lines += [
        "",
        f"🔔 Автозамеры: <b>{observation_line}</b>",
    ]
    if scan.last_view_refresh_at:
        lines.append(f"🕒 Обновлено: <b>{_moscow_text(scan.last_view_refresh_at)} МСК</b>")

    if scan.status == "partial":
        incomplete_keys = [
            key for key in (getattr(scan, "incomplete_category_keys", "") or "").split(",") if key in CATEGORIES
        ]
        if incomplete_keys:
            names = ", ".join(CATEGORIES[key].name for key in incomplete_keys[:5])
            lines += ["", f"⚠️ Допроверка: <b>{len(incomplete_keys)}</b> · {html.escape(names)}"]
        else:
            lines += ["", "⚠️ Часть скана требует допроверки."]
    elif quality_value and quality_value < 90 and getattr(scan, "quality_note", ""):
        lines += ["", f"⚠️ Проверка качества: {html.escape(scan.quality_note)}"]

    return "\n".join(lines)


@dp.callback_query(F.data.startswith("scan:"))
async def scan_detail(callback: CallbackQuery) -> None:
    try:
        scan_id = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Скан не найден", show_alert=True); return
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return
    await callback.answer()
    text = await render_scan_detail(scan)
    # A scan can be opened both from a normal text message and from the result-file caption.
    # Telegram cannot edit a document into text, so open a fresh card for document callbacks.
    if callback.message.text:
        await callback.message.edit_text(
            text, parse_mode=ParseMode.HTML, reply_markup=scan_detail_keyboard(scan.id, archived=scan.archived_at is not None, recheck=bool(getattr(scan, "incomplete_category_keys", ""))), disable_web_page_preview=True
        )
    else:
        await callback.message.answer(
            text, parse_mode=ParseMode.HTML, reply_markup=scan_detail_keyboard(scan.id, archived=scan.archived_at is not None, recheck=bool(getattr(scan, "incomplete_category_keys", ""))), disable_web_page_preview=True
        )


@dp.callback_query(F.data.startswith("scanproducts:"))
async def scan_products(callback: CallbackQuery) -> None:
    """Legacy callback from old messages. The model section was removed in v3.1.6."""
    try:
        scan_id = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Скан не найден", show_alert=True); return
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return
    await callback.answer("Раздел «Модели» убран — открыл карточку скана")
    await callback.message.answer(
        await render_scan_detail(scan),
        parse_mode=ParseMode.HTML,
        reply_markup=scan_detail_keyboard(scan_id, archived=scan.archived_at is not None, recheck=bool(getattr(scan, "incomplete_category_keys", ""))),
        disable_web_page_preview=True,
    )


async def _scan_top_pairs(user_id: int, scan_id: int) -> tuple[UserScan | None, list[tuple[Listing, ScanListing]]]:
    scan = await get_user_scan(user_id, scan_id)
    if scan is None:
        return None, []
    pairs = await get_scan_rows(scan_id)
    scan_settings = await get_settings_for_scan(scan)
    allowed_rows = apply_listing_settings(
        [p[0] for p in pairs], scan_settings, exact_date_scan=True, apply_output_mode=True
    )
    allowed_ids = {row.external_id for row in allowed_rows}
    pairs = [
        p for p in pairs
        if p[0].external_id in allowed_ids and p[0].view_count is not None
    ]
    pairs.sort(key=lambda p: p[0].view_count or 0, reverse=True)
    return scan, pairs


def _top_entry(index: int, row: Listing, snap: ScanListing) -> str:
    delta = (row.view_count - snap.initial_view_count) if snap.initial_view_count is not None else None
    growth = f" · 🚀 +{delta}" if delta is not None and delta > 0 else ""
    return (
        f"<b>{index}. {html.escape(row.title[:55])}</b>\n"
        f"👁 {row.view_count}{growth} · 💶 {html.escape(_price_display(row.price_text, row.price_eur))}\n"
        f'<a href="{html.escape(row.url)}">Открыть</a>'
    )


def scan_top12_keyboard(scan_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Показать TOP-50", callback_data=f"scantop50:{scan_id}:0")],
        [InlineKeyboardButton(text="📊 Открыть скан", callback_data=f"scan:{scan_id}")],
    ])


def scan_top50_keyboard(scan_id: int, page: int, total_pages: int) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"scantop50:{scan_id}:{page - 1}"))
    nav.append(InlineKeyboardButton(text=f"{page + 1}/{max(1, total_pages)}", callback_data="archive_noop"))
    if page + 1 < total_pages:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"scantop50:{scan_id}:{page + 1}"))
    rows.append(nav)
    rows.append([
        InlineKeyboardButton(text="🔥 TOP-12", callback_data=f"scantop:{scan_id}"),
        InlineKeyboardButton(text="📊 Скан", callback_data=f"scan:{scan_id}"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@dp.callback_query(F.data.startswith("scantop:"))
async def scan_top(callback: CallbackQuery) -> None:
    scan_id = int(callback.data.split(":", 1)[1])
    scan, pairs = await _scan_top_pairs(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return
    await callback.answer()
    if not pairs:
        text = "🔥 <b>Самые просматриваемые</b>\n\nПока нет данных просмотров."
    else:
        lines = [f"🔥 <b>TOP-12: {html.escape(scan.title)}</b>", ""]
        for i, (row, snap) in enumerate(pairs[:12], 1):
            lines.append(_top_entry(i, row, snap))
        if len(pairs) > 12:
            lines += ["", f"📋 Доступен полный <b>TOP-{min(50, len(pairs))}</b> по кнопке ниже."]
        text = "\n\n".join(lines)
    await callback.message.answer(
        text, parse_mode=ParseMode.HTML, disable_web_page_preview=True,
        reply_markup=scan_top12_keyboard(scan_id),
    )


@dp.callback_query(F.data.startswith("scantop50:"))
async def scan_top50(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    try:
        scan_id = int(parts[1])
        requested_page = int(parts[2]) if len(parts) > 2 else 0
    except Exception:
        await callback.answer("Некорректный запрос", show_alert=True); return

    scan, pairs = await _scan_top_pairs(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return
    top = pairs[:50]
    if not top:
        await callback.answer("Пока нет данных просмотров", show_alert=True)
        return

    page_size = 10
    total_pages = max(1, (len(top) + page_size - 1) // page_size)
    page = max(0, min(requested_page, total_pages - 1))
    start_idx = page * page_size
    chunk = top[start_idx:start_idx + page_size]
    lines = [
        f"📋 <b>TOP-{len(top)} по просмотрам</b>",
        f"<b>{html.escape(scan.title)}</b> · позиции {start_idx + 1}–{start_idx + len(chunk)}",
        "",
    ]
    for i, (row, snap) in enumerate(chunk, start_idx + 1):
        lines.append(_top_entry(i, row, snap))
    text = "\n\n".join(lines)
    markup = scan_top50_keyboard(scan_id, page, total_pages)
    await callback.answer(f"TOP-{len(top)} · страница {page + 1}/{total_pages}")
    if callback.message.text:
        await callback.message.edit_text(
            text, parse_mode=ParseMode.HTML, disable_web_page_preview=True, reply_markup=markup
        )
    else:
        await callback.message.answer(
            text, parse_mode=ParseMode.HTML, disable_web_page_preview=True, reply_markup=markup
        )


@dp.callback_query(F.data.startswith("scangrowth:"))
async def scan_growth(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    try:
        scan_id = int(parts[1])
        period_hours = int(parts[2]) if len(parts) > 2 else 3
    except Exception:
        await callback.answer("Скан не найден", show_alert=True); return
    if period_hours not in OBSERVATION_HOURS:
        period_hours = 3
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return

    growth, rounds = await get_scan_growth_rows(scan_id, period_hours)
    scan_settings = await get_settings_for_scan(scan)
    allowed_growth_rows = apply_listing_settings(
        [item.listing for item in growth], scan_settings, exact_date_scan=True, apply_output_mode=True
    )
    allowed_growth_ids = {row.external_id for row in allowed_growth_rows}
    growth = [item for item in growth if item.listing.external_id in allowed_growth_ids]
    await callback.answer()
    period_label = f"{period_hours} ч"
    if not growth:
        text = (
            f"🚀 <b>TOP роста за {period_label}</b>\n\n"
            "Контрольный замер для этого периода ещё не готов или прироста пока нет. "
            "Автозамеры 3 / 6 / 12 часов выполняются только если они включены в главном меню."
        )
    else:
        lines = [
            f"🚀 <b>TOP роста · {period_label}</b>",
            f"<b>{html.escape(scan.title)}</b>",
            "Сортировка: <b>по реальному приросту просмотров</b>.",
            "",
        ]
        for i, item in enumerate(growth[:GROWTH_TELEGRAM_LIMIT], 1):
            row = item.listing
            lines.append(
                f"<b>{i}. {html.escape(row.title[:60])}</b>\n"
                                f"👁 {item.base_views} → <b>{item.current_views}</b> · "
                f"🚀 <b>+{item.delta}</b> · ⚡ {item.per_hour:.1f}/ч\n"
                f"💶 {html.escape(_price_display(row.price_text, row.price_eur))} · "
                f'<a href="{html.escape(row.url)}">Открыть</a>'
            )
        lines += ["", f"📊 Полный рейтинг: до <b>{GROWTH_TOP_LIMIT}</b> товаров в таблице."]
        text = "\n\n".join(lines)
    await callback.message.answer(
        text, parse_mode=ParseMode.HTML, disable_web_page_preview=True,
        reply_markup=growth_period_keyboard(scan_id, period_hours),
    )


def build_growth_top_xlsx(
    scan: UserScan, period_hours: int, growth: list[GrowthMetric], category_key: str | None = None
) -> Path:
    """Build the downloadable TOP-50 table."""
    cat = CATEGORIES.get(category_key) if category_key else None
    wb = Workbook()
    ws = wb.active
    ws.title = "TOP growth"
    title = f"TOP-{min(GROWTH_TOP_LIMIT, len(growth))} роста за {period_hours}ч"
    if cat is not None:
        title += f" · {cat.name}"
    ws.append([title])
    ws.merge_cells("A1:L1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([
        "#", "Товар", "Категория", "Цена €",
        "Было просмотров", "Сейчас просмотров", "Прирост", "Просмотров/час",
        "Фактический интервал, ч", "Дата объявления", "ID", "Ссылка",
    ])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(growth[:GROWTH_TOP_LIMIT], 1):
        row = item.listing
        ws.append([
            idx, row.title, row.category, row.price_eur,
            item.base_views, item.current_views, item.delta, round(item.per_hour, 2),
            round(item.elapsed_hours, 2), row.posted_date_msk or row.posted_text or "",
            row.external_id, row.url,
        ])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:L{max(2, ws.max_row)}"
    widths = {
        "A": 6, "B": 44, "C": 26, "D": 11, "E": 16, "F": 17,
        "G": 12, "H": 16, "I": 19, "J": 16, "K": 16, "L": 52,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_cells in ws.iter_rows(min_row=3):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    growth_fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in ws["G"][2:]:
        cell.fill = growth_fill
        cell.font = Font(bold=True)
    for cell in ws["L"][2:]:
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"

    out_dir = Path(tempfile.mkdtemp(prefix="growth_top_"))
    safe_cat = re.sub(r"[^A-Za-z0-9_-]+", "_", category_key or "scan")
    out = out_dir / f"TOP50_{safe_cat}_{period_hours}h_scan_{scan.id}.xlsx"
    wb.save(out)
    return out


async def send_growth_xlsx(
    message: Message, scan: UserScan, period_hours: int, category_key: str | None = None
) -> None:
    growth, _ = await get_scan_growth_rows(scan.id, period_hours, category_key=category_key)
    scan_settings = await get_settings_for_scan(scan)
    allowed_rows = apply_listing_settings(
        [item.listing for item in growth], scan_settings, exact_date_scan=True, apply_output_mode=True
    )
    allowed_ids = {row.external_id for row in allowed_rows}
    growth = [item for item in growth if item.listing.external_id in allowed_ids]
    if not growth:
        await message.answer(
            f"📊 TOP-{GROWTH_TOP_LIMIT} за {period_hours}ч пока нельзя сформировать: "
            "контрольный замер ещё не готов или прироста нет."
        )
        return
    path = build_growth_top_xlsx(scan, period_hours, growth, category_key=category_key)
    try:
        cat = CATEGORIES.get(category_key) if category_key else None
        suffix = f" · {cat.name}" if cat else ""
        await message.answer_document(
            FSInputFile(path),
            caption=f"📊 TOP-{min(GROWTH_TOP_LIMIT, len(growth))} роста за {period_hours}ч{suffix}",
        )
    finally:
        shutil.rmtree(path.parent, ignore_errors=True)


async def send_category_growth_xlsx(
    message: Message, user_id: int, category_key: str, period_hours: int
) -> None:
    scans = await get_user_category_scans(user_id, category_key)
    growth, scan_count, _ = await get_category_growth_rows(user_id, category_key, period_hours)
    if not scans:
        await message.answer(
            f"📊 TOP-{GROWTH_TOP_LIMIT} за {period_hours}ч пока нельзя сформировать: "
            "контрольные замеры ещё не готовы или прироста нет."
        )
        return
    representative = scans[0]
    scan_settings = await get_settings_for_scan(representative)
    allowed_rows = apply_listing_settings(
        [item.listing for item in growth], scan_settings, exact_date_scan=True, apply_output_mode=True
    )
    allowed_ids = {row.external_id for row in allowed_rows}
    growth = [item for item in growth if item.listing.external_id in allowed_ids]
    if not growth:
        await message.answer(
            f"📊 TOP-{GROWTH_TOP_LIMIT} за {period_hours}ч пока нельзя сформировать: "
            "контрольные замеры ещё не готовы или прироста нет."
        )
        return
    path = build_growth_top_xlsx(
        representative, period_hours, growth, category_key=category_key
    )
    try:
        cat = CATEGORIES.get(category_key)
        suffix = f" · {cat.name}" if cat else ""
        await message.answer_document(
            FSInputFile(path),
            caption=(
                f"📊 TOP-{min(GROWTH_TOP_LIMIT, len(growth))} роста за {period_hours}ч{suffix} "
                "· последний успешный скан"
            ),
        )
    finally:
        shutil.rmtree(path.parent, ignore_errors=True)


@dp.callback_query(F.data.startswith("scangrowthexport:"))
async def scan_growth_export(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    try:
        scan_id, period_hours = int(parts[1]), int(parts[2])
    except Exception:
        await callback.answer("Некорректный запрос", show_alert=True); return
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None or period_hours not in OBSERVATION_HOURS:
        await callback.answer("Скан не найден", show_alert=True); return
    await callback.answer("Формирую TOP-50")
    await send_growth_xlsx(callback.message, scan, period_hours)


@dp.callback_query(F.data.startswith("pce:"))
async def popular_growth_export(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    if len(parts) != 4:
        await callback.answer("Некорректный запрос", show_alert=True); return
    try:
        representative_scan_id, category_key, period_hours = int(parts[1]), parts[2], int(parts[3])
    except Exception:
        await callback.answer("Некорректный запрос", show_alert=True); return
    if period_hours not in OBSERVATION_HOURS or category_key not in CATEGORIES:
        await callback.answer("Некорректный запрос", show_alert=True); return
    scans = await get_user_category_scans(callback.from_user.id, category_key)
    if not scans:
        await callback.answer("Сканы категории не найдены", show_alert=True); return
    await callback.answer("Формирую TOP-50 последнего скана")
    await send_category_growth_xlsx(
        callback.message, callback.from_user.id, category_key, period_hours
    )


@dp.callback_query(F.data.startswith("scanhistory:"))
async def scan_history(callback: CallbackQuery) -> None:
    try:
        scan_id = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Скан не найден", show_alert=True); return
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return
    rounds = await get_scan_history_rounds(scan_id, limit=10)
    await callback.answer()
    if not rounds:
        text = "🕘 <b>История просмотров</b>\n\nПока нет сохранённых точек наблюдения."
    else:
        lines = [f"🕘 <b>История: {html.escape(scan.title)}</b>", "", "Последние точки по московскому времени:"]
        for idx, (recorded_at, count, total_views) in enumerate(rounds, 1):
            marker = "🟢" if idx == 1 else "▫️"
            lines.append(
                f"{marker} <b>{_moscow_text(recorded_at)} МСК</b> · "
                f"{count} объявл. · суммарно 👁 {total_views}"
            )
        lines += ["", "Каждое ручное «Обновить просмотры» добавляет новую точку для сравнения."]
        text = "\n".join(lines)
    await callback.message.answer(text, parse_mode=ParseMode.HTML, reply_markup=scan_detail_keyboard(scan_id, archived=scan.archived_at is not None, recheck=bool(getattr(scan, "incomplete_category_keys", ""))))


async def _manual_view_refresh_job(
    bot: Bot, user_id: int, scan_id: int, progress_message: Message | None = None
) -> None:
    try:
        scan = await get_user_scan(user_id, scan_id)
        if scan is None:
            return
        pairs = await get_scan_rows(scan_id)
        rows = [row for row, _ in pairs]
        if not rows:
            if progress_message is not None:
                await progress_message.edit_text("ℹ️ В этом скане пока нет объявлений для обновления.")
            else:
                await bot.send_message(user_id, "ℹ️ В этом скане пока нет объявлений для обновления.")
            return

        title = html.escape(scan.title)
        measurement_started = datetime.utcnow() - timedelta(seconds=VIEW_MEASUREMENT_REUSE_SECONDS)

        # v4.3.6: a button press is foreground/manual work, not a scheduled
        # background checkpoint. In Multi-User Stable mode background view traffic
        # is intentionally paused while ANY scan job is active. Treating this
        # explicit user action as background therefore left the card at 0/N until
        # every other user's scan finished. Manual refreshes now use the normal
        # adaptive view lane directly; TRAFFIC still reserves scan slots, so this
        # cannot starve category/date parsing. Same-scan duplicate clicks are
        # already coalesced by manual_view_tasks.
        requested, updated, failed = await refresh_view_counts(
            rows, None, force=False, max_age_seconds=VIEW_MEASUREMENT_REUSE_SECONDS,
            traffic_priority="manual",
            progress_message=progress_message,
            progress_title=f"👁 Повторный замер · {scan.title}",
        )
        recorded = await update_scan_view_refresh(
            scan_id, fresh_after=measurement_started
        )

        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🚀 Открыть динамику", callback_data=f"scangrowth:{scan_id}:3"),
             InlineKeyboardButton(text="🔥 Топ", callback_data=f"scantop:{scan_id}")],
            [InlineKeyboardButton(text="📊 Открыть этот скан", callback_data=f"scan:{scan_id}")],
        ])

        if recorded <= 0:
            text = (
                f"⚠️ <b>Замер не выполнен</b>\n\n"
                f"Скан: <b>{title}</b>\n"
                f"📦 Объявлений: <b>{len(rows)}</b>\n"
                f"👁 Свежих значений получить не удалось.\n\n"
                "Новая точка наблюдения не создана. Остальными разделами бота можно пользоваться как обычно."
            )
            if progress_message is not None:
                await progress_message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=keyboard)
            else:
                await bot.send_message(user_id, text, parse_mode=ParseMode.HTML, reply_markup=keyboard)
            return

        reused = max(0, recorded - updated)
        grew_count, max_delta, total_delta = await get_latest_manual_growth_summary(scan_id)
        text = (
            f"✅ <b>Просмотры обновлены</b>\n\n"
            f"Скан: <b>{title}</b>\n"
            f"📦 Объявлений: <b>{len(rows)}</b>\n"
            f"👁 Свежих значений: <b>{recorded}</b>\n"
            f"⚡ Получено direct-запросами: <b>{updated}</b>\n"
            f"♻️ Переиспользовано одновременно свежих: <b>{reused}</b>\n"
            f"▫️ Без данных: <b>{max(0, len(rows) - recorded)}</b>\n\n"
            f"🚀 Выросли с прошлого замера: <b>{grew_count}</b>\n"
            f"🔥 Максимальный прирост: <b>+{max_delta}</b>\n"
            f"📈 Суммарный прирост: <b>+{total_delta}</b>\n"
            f"🕐 Замер: <b>{_moscow_text(datetime.utcnow())} МСК</b>\n\n"
            "TOP и динамика уже пересчитаны по этой реальной точке наблюдения."
        )
        if progress_message is not None:
            await progress_message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=keyboard)
        else:
            await bot.send_message(user_id, text, parse_mode=ParseMode.HTML, reply_markup=keyboard)
    except Exception:
        log.exception("Manual background view refresh failed scan=%s", scan_id)
        try:
            text = (
                "⚠️ Не удалось полностью обновить просмотры. Новая точка наблюдения не будет считаться готовой; "
                "остальные разделы бота продолжают работать."
            )
            if progress_message is not None:
                await progress_message.edit_text(text)
            else:
                await bot.send_message(user_id, text)
        except Exception:
            pass
    finally:
        async with manual_view_tasks_guard:
            current = manual_view_tasks.get(scan_id)
            if current is asyncio.current_task() or (current is not None and current.done()):
                manual_view_tasks.pop(scan_id, None)


@dp.callback_query(F.data.startswith("scanviews:"))
async def scan_refresh_views(callback: CallbackQuery, bot: Bot) -> None:
    scan_id = int(callback.data.split(":", 1)[1])
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True)
        return
    pairs = await get_scan_rows(scan_id)
    if not pairs:
        await callback.answer("В этом скане пока нет объявлений", show_alert=True)
        return

    async with manual_view_tasks_guard:
        existing = manual_view_tasks.get(scan_id)
        if existing is not None and not existing.done():
            await callback.answer("👁 Просмотры этого скана уже обновляются в фоне")
            return

        # This is a separate message, so navigating menus cannot overwrite it.
        progress_message = await callback.message.answer(
            f"<b>👁 Повторный замер · {html.escape(scan.title)}</b>\n\n"
            f"{_progress_bar(0)} <b>0%</b>\n"
            f"📦 Проверено: <b>0/{len(pairs)}</b>\n\n"
            "Можно сразу переходить в другие разделы — замер работает в фоне.",
            parse_mode=ParseMode.HTML,
        )
        task = asyncio.create_task(
            _manual_view_refresh_job(bot, callback.from_user.id, scan_id, progress_message),
            name=f"manual-view-refresh-{scan_id}",
        )
        manual_view_tasks[scan_id] = task

    # Handler returns immediately; the progress message is edited at most every ~1.5 s.
    await callback.answer("👁 Замер запущен в фоне")


@dp.callback_query(F.data.startswith("scanexport:"))
async def scan_export(callback: CallbackQuery) -> None:
    scan_id = int(callback.data.split(":", 1)[1])
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return
    pairs = await get_scan_rows(scan_id)
    rows = [row for row, _ in pairs]
    await callback.answer()
    await send_smart_export(
        callback.message, callback.from_user.id, scan.total_categories,
        category_keys_override=set(_scan_category_keys(scan)), rows_override=rows,
        price_filter_override=(getattr(scan, "price_filter", "any") or "any"),
    )


@dp.callback_query(F.data.startswith("scanrepeat:"))
async def scan_repeat(callback: CallbackQuery) -> None:
    scan_id = int(callback.data.split(":", 1)[1])
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True); return
    if await user_has_active_scan(callback.from_user.id):
        await callback.answer("У тебя уже идёт скан", show_alert=True); return
    if await queue_is_full():
        await callback.answer("Сервис сейчас сильно загружен. Попробуй чуть позже.", show_alert=True); return
    keys = [k for k in _scan_category_keys(scan) if k in CATEGORIES]
    if not keys:
        await callback.answer("Категории этого скана больше недоступны", show_alert=True); return
    if len(keys) > MAX_SELECTED_CATEGORIES:
        await callback.answer(
            f"Этот старый скан содержит {len(keys)} категорий. Сейчас лимит — {MAX_SELECTED_CATEGORIES}; выбери нужные категории для нового запуска.",
            show_alert=True,
        )
        return
    repeat_depth = scan.page_limit if scan.page_limit in PAGE_LIMIT_CHOICES else 50
    await callback.answer("Повторяю скан")
    await enqueue_user_scan(
        callback.message, callback.from_user.id, keys, repeat_depth, scan.target_date or _moscow_today_iso(),
        price_filter=(getattr(scan, "price_filter", "any") or "any"),
    )


@dp.callback_query(F.data.startswith("scanrecheck:"))
async def scan_recheck_partial(callback: CallbackQuery) -> None:
    """Re-run only categories that were not fully verified in a partial scan."""
    try:
        scan_id = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Скан не найден", show_alert=True)
        return
    scan = await get_user_scan(callback.from_user.id, scan_id)
    if scan is None:
        await callback.answer("Скан не найден", show_alert=True)
        return

    keys = [
        key for key in (getattr(scan, "incomplete_category_keys", "") or "").split(",")
        if key in CATEGORIES
    ]
    keys = list(dict.fromkeys(keys))
    if not keys:
        await callback.answer("У этого скана нет категорий, требующих допроверки", show_alert=True)
        return
    if len(keys) > MAX_SELECTED_CATEGORIES:
        keys = keys[:MAX_SELECTED_CATEGORIES]

    if await user_has_active_scan(callback.from_user.id):
        await callback.answer("У тебя уже идёт парсинг", show_alert=True)
        return
    if await queue_is_full():
        await callback.answer("Сервис сейчас сильно загружен. Попробуй чуть позже.", show_alert=True)
        return

    depth = scan.page_limit if scan.page_limit in PAGE_LIMIT_CHOICES else 50
    target_date = scan.target_date or _moscow_today_iso()
    # Do not immediately reuse the partial 5-minute category result cache.
    for key in keys:
        progress_key = _progress_key(key, target_date, depth)
        category_result_cache.pop(progress_key, None)
        if DISTRIBUTED_WORKERS:
            try:
                await COORDINATOR.delete_category_result(progress_key)
            except Exception:
                log.debug("Could not clear distributed partial cache key=%s", progress_key, exc_info=True)

    await callback.answer(f"Допроверяю категорий: {len(keys)}")
    await enqueue_user_scan(
        callback.message, callback.from_user.id, keys, depth, target_date,
        price_filter=(getattr(scan, "price_filter", "any") or "any"),
    )


@dp.callback_query(F.data == "groups")
async def groups(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not await trial_or_paid_scan_access(callback.from_user.id): await callback.answer("Нет доступа", show_alert=True); return
    selected = await get_selected(callback.from_user.id)
    trial_mode = bool(not allowed(callback.from_user.id) and (await get_trial_status(callback.from_user.id)).eligible)
    limit = FREE_TRIAL_MAX_CATEGORIES if trial_mode else MAX_SELECTED_CATEGORIES
    await callback.answer()
    await _edit_or_answer(
        callback.message,
        f"<b>🗂 Категории</b>\n\nВыбери до <b>{limit}</b> категорий на один скан.",
        reply_markup=groups_keyboard(selected, max_selected=limit),
    )


@dp.callback_query(F.data.startswith("grp:"))
async def open_group(callback: CallbackQuery) -> None:
    group_key = callback.data.split(":", 1)[1]
    if group_key not in GROUPS: await callback.answer("Раздел не найден", show_alert=True); return
    selected = await get_selected(callback.from_user.id)
    group = GROUPS[group_key]
    trial_mode = bool(not allowed(callback.from_user.id) and (await get_trial_status(callback.from_user.id)).eligible)
    limit = FREE_TRIAL_MAX_CATEGORIES if trial_mode else MAX_SELECTED_CATEGORIES
    await callback.answer()
    await callback.message.edit_text(
        f"<b>{group.icon} {html.escape(group.name)}</b>\n\n"
        f"Отметь нужные подкатегории. Максимум за один запуск: <b>{limit}</b>.",
        reply_markup=category_keyboard(group_key, selected, max_selected=limit),
        parse_mode=ParseMode.HTML,
    )


@dp.callback_query(F.data.startswith("cat:"))
async def toggle_cat(callback: CallbackQuery) -> None:
    key = callback.data.split(":", 1)[1]
    if key not in CATEGORIES: await callback.answer("Категория не найдена", show_alert=True); return
    if CATEGORIES[key].is_group:
        await callback.answer("Выбор всего раздела убран. Выбери нужные подкатегории.", show_alert=True)
        return
    trial_mode = bool(not allowed(callback.from_user.id) and (await get_trial_status(callback.from_user.id)).eligible)
    if trial_mode:
        selected = await get_selected(callback.from_user.id)
        if key in selected:
            selected = await _replace_selected_categories(callback.from_user.id, set())
        else:
            # One-tap replacement keeps the launch offer frictionless: choosing a
            # different category simply replaces the previous free-trial choice.
            selected = await _replace_selected_categories(callback.from_user.id, {key})
        await callback.answer("Категория выбрана" if selected else "Категория снята")
        await callback.message.edit_reply_markup(
            reply_markup=category_keyboard(
                CATEGORIES[key].group, selected, max_selected=FREE_TRIAL_MAX_CATEGORIES
            )
        )
        return

    selected, limit_reached = await toggle_category(callback.from_user.id, key)
    if limit_reached:
        await callback.answer(
            f"Можно выбрать максимум {MAX_SELECTED_CATEGORIES} категорий за один запуск. Сними одну галочку и выбери другую.",
            show_alert=True,
        )
    else:
        await callback.answer(f"Выбрано: {len(selected)}/{MAX_SELECTED_CATEGORIES}")
    await callback.message.edit_reply_markup(reply_markup=category_keyboard(CATEGORIES[key].group, selected))


@dp.callback_query(F.data.startswith("grpall:"))
async def toggle_all_children(callback: CallbackQuery) -> None:
    # Legacy buttons from old Telegram messages are disabled too, not only hidden
    # from the new UI. This prevents accidental whole/bulk section selection.
    await callback.answer(
        "Массовый выбор раздела убран. Выбери нужные подкатегории вручную.",
        show_alert=True,
    )


@dp.callback_query(F.data == "clear_all")
async def clear_all(callback: CallbackQuery) -> None:
    await clear_selected(callback.from_user.id)
    trial_mode = bool(not allowed(callback.from_user.id) and (await get_trial_status(callback.from_user.id)).eligible)
    limit = FREE_TRIAL_MAX_CATEGORIES if trial_mode else MAX_SELECTED_CATEGORIES
    await callback.answer("Выбор очищен")
    await callback.message.edit_reply_markup(reply_markup=groups_keyboard(set(), max_selected=limit))


@dp.callback_query(F.data == "selected")
async def selected(callback: CallbackQuery) -> None:
    keys = await get_selected(callback.from_user.id)
    cats = [CATEGORIES[k] for k in CATEGORIES if k in keys]
    trial_mode = bool(not allowed(callback.from_user.id) and (await get_trial_status(callback.from_user.id)).eligible)
    limit = FREE_TRIAL_MAX_CATEGORIES if trial_mode else MAX_SELECTED_CATEGORIES
    if not cats:
        text = "<b>Категории пока не выбраны.</b>"
    else:
        counter = f"{len(cats)}/{limit}"
        lines = [f"<b>Выбрано категорий: {counter}</b>", ""]
        if len(cats) > limit:
            lines += [f"⚠️ Для нового запуска оставь максимум {limit} категорий.", ""]
        for group in GROUPS.values():
            chosen = [cat for cat in cats if cat.group == group.key and not cat.is_group]
            if not chosen:
                continue
            lines.append(f"{group.icon} <b>{html.escape(group.name)}</b>")
            lines.extend(f"  • {html.escape(cat.name)}" for cat in chosen)
            lines.append("")
        text = "\n".join(lines).rstrip()
    await callback.answer()
    await callback.message.answer(text, parse_mode=ParseMode.HTML, reply_markup=main_keyboard(len(keys)))


@dp.callback_query(F.data == "stats")
async def stats(callback: CallbackQuery) -> None:
    await callback.answer()
    keys = await get_selected(callback.from_user.id)
    await callback.message.answer(await stats_text(), parse_mode=ParseMode.HTML, reply_markup=main_keyboard(len(keys)))


@dp.callback_query(F.data == "export_smart")
async def export_smart(callback: CallbackQuery) -> None:
    if not allowed(callback.from_user.id): await callback.answer("Нет доступа", show_alert=True); return
    await callback.answer()
    selected = await get_selected(callback.from_user.id)
    await send_smart_export(callback.message, callback.from_user.id, len(selected))


@dp.callback_query(F.data == "queue_status")
async def queue_status(callback: CallbackQuery) -> None:
    if not read_only_history_allowed(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()

    if DISTRIBUTED_WORKERS:
        scan = await get_active_user_scan(callback.from_user.id)
        if scan is not None:
            worker_count = 0
            try:
                worker_count = await COORDINATOR.worker_count("parser")
            except Exception:
                pass
            if scan.status == "queued":
                text = (
                    "⏳ <b>Скан стоит в распределённой очереди</b>\n\n"
                    f"📅 <b>{_date_label(scan.target_date)}</b> · 📄 <b>{scan.page_limit} стр.</b>\n"
                    f"🗂 Категорий: <b>{scan.total_categories}</b>\n"
                    f"⚙️ Активных parser-worker: <b>{worker_count}</b>\n\n"
                    "Как только освободится воркер, прогресс продолжится в карточке скана."
                )
            elif scan.status == "cancelling":
                text = "⏹ <b>Останавливаю активный скан…</b>"
            else:
                text = (
                    "⚙️ <b>Скан выполняется отдельным parser-worker</b>\n\n"
                    f"📅 <b>{_date_label(scan.target_date)}</b> · 📄 <b>{scan.page_limit} стр.</b>\n"
                    f"🗂 Категорий: <b>{scan.total_categories}</b>\n"
                    f"⚙️ Активных parser-worker: <b>{worker_count}</b>\n\n"
                    "Живой процент обновляется в основной карточке запуска."
                )
            markup = job_keyboard(str(scan.job_uid), queued=(scan.status == "queued"))
        else:
            text = "✅ Сейчас у тебя нет активного парсинга."
            selected = await get_selected(callback.from_user.id)
            markup = main_keyboard(len(selected))
        await callback.message.answer(text, parse_mode=ParseMode.HTML, reply_markup=markup)
        return

    async with job_guard:
        job = active_jobs.get(callback.from_user.id)
    if job and job.state in {"queued", "running"}:
        text = render_user_job_status(job)
        markup = job_keyboard(job.job_id, queued=(job.state == "queued"))
    else:
        text = "✅ Сейчас у тебя нет активного парсинга."
        selected = await get_selected(callback.from_user.id)
        markup = main_keyboard(len(selected))
    await callback.message.answer(text, parse_mode=ParseMode.HTML, reply_markup=markup)


async def request_user_scan_stop(user_id: int, job_id: str | None = None) -> tuple[ScanJob | None, str]:
    """Set and persist the hard-stop signal used by both the button and /stop."""
    if DISTRIBUTED_WORKERS:
        scan = await get_active_user_scan(user_id)
        if scan is None or (job_id is not None and str(scan.job_uid) != str(job_id)):
            return None, "missing"
        if scan.status not in ACTIVE_SCAN_STATUSES:
            return None, "finished"
        previous = "queued" if scan.status == "queued" else "running"
        job = scan_job_from_record(scan)
        job.cancel_requested = True
        job.stop_event.set()
        async with SessionLocal() as session:
            db_scan = await session.get(UserScan, int(scan.id))
            if db_scan is not None and db_scan.status in ACTIVE_SCAN_STATUSES:
                if db_scan.status == "queued":
                    # No worker owns it yet: cancel immediately so the user can
                    # start another scan without waiting for the stream item to be consumed.
                    db_scan.status = "cancelled"
                    db_scan.finished_at = datetime.utcnow()
                else:
                    db_scan.status = "cancelling"
                db_scan.last_error = "Остановка запрошена пользователем"
                await session.commit()
        try:
            await COORDINATOR.request_cancel(str(scan.job_uid))
        except Exception:
            log.exception("Could not publish distributed cancel job=%s", scan.job_uid)
        return job, previous

    scan_id: int | None = None
    async with job_guard:
        job = active_jobs.get(user_id)
        if job is None or (job_id is not None and job.job_id != job_id):
            return None, "missing"
        if job.state not in {"queued", "running"}:
            return None, "finished"
        previous = job.state
        job.cancel_requested = True
        job.stop_event.set()
        scan_id = job.scan_id
        if job.job_id in queued_job_ids:
            queued_job_ids.remove(job.job_id)

    if scan_id is not None:
        async with SessionLocal() as session:
            scan = await session.get(UserScan, int(scan_id))
            if scan is not None and scan.status in {"queued", "running"}:
                scan.status = "cancelling"
                scan.last_error = "Остановка запрошена пользователем"
                await session.commit()
    return job, previous


@dp.message(Command("stop"))
async def stop_scan_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not read_only_history_allowed(message.from_user.id):
        await message.answer("Нет доступа.")
        return
    job, previous_state = await request_user_scan_stop(message.from_user.id)
    if job is None:
        await message.answer(
            "✅ Сейчас активного парсинга нет.",
            reply_markup=main_keyboard(len(await get_selected(message.from_user.id))),
        )
        return
    if previous_state == "queued":
        refunded = bool(job.scan_id and await _refund_trial_credit_for_scan(job.scan_id))
        await message.answer(
            "❌ <b>Скан удалён из очереди.</b> Задание отменено до начала сетевого сканирования."
            + ("\n🎁 Бесплатный запуск возвращён на баланс." if refunded else ""),
            parse_mode=ParseMode.HTML,
            reply_markup=stopped_job_keyboard(),
        )
    else:
        await message.answer(
            "⏹ <b>Останавливаю парсер прямо сейчас…</b>\n\n"
            "Новые страницы и объявления больше не будут запускаться. Можно сразу выбрать другую категорию.",
            parse_mode=ParseMode.HTML,
            reply_markup=stopped_job_keyboard(),
        )


@dp.callback_query(F.data.startswith("cancel_scan:"))
async def cancel_scan(callback: CallbackQuery) -> None:
    job_id = callback.data.split(":", 1)[1]
    job, previous_state = await request_user_scan_stop(callback.from_user.id, job_id)
    if job is None:
        await callback.answer("Активная задача уже не найдена", show_alert=True)
        return

    if previous_state == "queued":
        refunded = bool(job.scan_id and await _refund_trial_credit_for_scan(job.scan_id))
        await callback.answer("Убрано из очереди")
        await callback.message.edit_text(
            "❌ <b>Скан удалён из очереди</b>\n\nЗадание отменено до начала сканирования."
            + ("\n🎁 Бесплатный запуск возвращён на баланс." if refunded else ""),
            parse_mode=ParseMode.HTML,
            reply_markup=stopped_job_keyboard(),
        )
    else:
        await callback.answer("Останавливаю парсер")
        await callback.message.edit_text(
            "⏹ <b>Останавливаю парсер прямо сейчас…</b>\n\n"
            "Текущий сетевой скан отменяется. Можно сразу выбрать другую категорию.",
            parse_mode=ParseMode.HTML,
            reply_markup=stopped_job_keyboard(),
        )


@dp.callback_query(F.data == "start_scan")
async def start_scan(callback: CallbackQuery, state: FSMContext) -> None:
    if not await trial_or_paid_scan_access(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    selected_keys, has_active_scan = await asyncio.gather(
        get_selected(callback.from_user.id),
        user_has_active_scan(callback.from_user.id),
    )
    selected_cats = [CATEGORIES[k] for k in CATEGORIES if k in selected_keys]
    if not selected_cats:
        await callback.answer("Сначала выбери хотя бы одну категорию", show_alert=True)
        return
    if not allowed(callback.from_user.id) and len(selected_cats) > FREE_TRIAL_MAX_CATEGORIES:
        await callback.answer("🎁 В бесплатном скане можно выбрать 1 категорию", show_alert=True)
        return
    if len(selected_cats) > MAX_SELECTED_CATEGORIES:
        await callback.answer(
            f"Сейчас выбрано {len(selected_cats)} категорий. В новой версии максимум {MAX_SELECTED_CATEGORIES} за один запуск — убери лишние или очисти выбор.",
            show_alert=True,
        )
        return

    if has_active_scan:
        await callback.answer("У тебя уже идёт парсинг", show_alert=True)
        return

    await state.set_state(ScanInput.target_date)
    await callback.answer()
    await callback.message.answer(
        "<b>▶️ Новый скан</b>\n\n"
        "<b>1/3 · Дата объявлений</b>\n"
        "Выбери одну из 5 доступных дат: сегодня или один из 4 предыдущих дней. "
        "Время считаем по Москве.",
        parse_mode=ParseMode.HTML,
        reply_markup=scan_date_keyboard(),
    )


async def _show_scan_price_choice(message: Message, state: FSMContext, user_id: int, target_date: str) -> None:
    await state.update_data(target_date=target_date)
    await state.set_state(None)
    selected = await get_selected(user_id)
    selected_cats = [CATEGORIES[k] for k in CATEGORIES if k in selected]
    if not selected_cats:
        await state.clear()
        await message.answer("Сначала выбери хотя бы одну категорию.")
        return
    await message.answer(
        "<b>▶️ Новый скан</b>\n\n"
        "<b>2/3 · Цена</b>\n"
        f"📅 Дата: <b>{_date_label(target_date)}</b>\n\n"
        "Выбери минимальную цену или укажи свой диапазон.",
        parse_mode=ParseMode.HTML,
        reply_markup=scan_price_keyboard(),
    )


async def _show_scan_depth_choice(
    message: Message, state: FSMContext, user_id: int, target_date: str, price_filter: str
) -> None:
    await state.update_data(target_date=target_date, price_filter=price_filter)
    await state.set_state(None)
    selected = await get_selected(user_id)
    selected_cats = [CATEGORIES[k] for k in CATEGORIES if k in selected]
    if not selected_cats:
        await state.clear()
        await message.answer("Сначала выбери хотя бы одну категорию.")
        return

    trial = await get_trial_status(user_id) if not allowed(user_id) else TrialStatus(False, False, 0, 0)
    trial_mode = bool(not allowed(user_id) and trial.eligible)
    category_limit = FREE_TRIAL_MAX_CATEGORIES if trial_mode else MAX_SELECTED_CATEGORIES
    if len(selected_cats) > category_limit:
        await state.clear()
        if trial_mode:
            text = (
                "🎁 <b>В бесплатном скане доступна 1 категория.</b>\n\n"
                "Оставь одну категорию и вернись к запуску."
            )
        else:
            text = (
                f"⚠️ Сейчас выбрано <b>{len(selected_cats)}</b> категорий, а максимум для одного запуска — "
                f"<b>{MAX_SELECTED_CATEGORIES}</b>. Убери лишние категории и запусти снова."
            )
        await message.answer(
            text, parse_mode=ParseMode.HTML, reply_markup=groups_keyboard(selected)
        )
        return

    scan_settings = await get_settings(user_id)
    include = html.escape(scan_settings.include_words) if scan_settings.include_words else ""
    exclude = html.escape(scan_settings.exclude_words) if scan_settings.exclude_words else ""
    extra_lines = []
    if include:
        extra_lines.append(f"🔎 Ключевые: <b>{include}</b>")
    if exclude:
        extra_lines.append(f"🚫 Исключения: <b>{exclude}</b>")
    extras = ("\n" + "\n".join(extra_lines)) if extra_lines else ""
    trial_line = (
        f"\n🎁 Пробный режим: <b>до {FREE_TRIAL_MAX_PAGES} страниц</b> · "
        f"после запуска останется <b>{max(0, trial.remaining - 1)}</b> бесплатн. скан(а)"
        if trial_mode else ""
    )
    await message.answer(
        "<b>▶️ Новый скан</b>\n\n"
        "<b>3/3 · Глубина сканирования</b>\n"
        f"📅 Дата: <b>{_date_label(target_date)}</b>\n"
        f"🗂 Категорий: <b>{len(selected_cats)}/{category_limit}</b>\n"
        f"Режим: <b>{MODE_LABELS.get(scan_settings.output_mode, scan_settings.output_mode)}</b>\n"
        f"💶 Цена: <b>{price_filter_label(price_filter)}</b> · "
        f"👁 <b>{min_views_label(getattr(scan_settings, 'min_views', 0))}</b>\n"
        f"🧠 Дубли: <b>{'Вкл' if scan_settings.smart_dedupe else 'Выкл'}</b> · "
        f"🧹 Шум: <b>{'Вкл' if scan_settings.clean_noise else 'Выкл'}</b>"
        f"{extras}{trial_line}\n\n"
        + (
            "Выбери глубину бесплатного скана."
            if trial_mode else
            "Выбери максимальное количество страниц общей выдачи для выбранной даты."
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=page_limit_keyboard(trial=trial_mode),
    )


@dp.callback_query(F.data.startswith("scan_date:"))
async def choose_scan_date(callback: CallbackQuery, state: FSMContext) -> None:
    if not await trial_or_paid_scan_access(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    choice = callback.data.split(":", 1)[1]
    today = datetime.now(MOSCOW).date()
    if choice == "today":
        target_date = today.isoformat()
        await callback.answer()
        await _show_scan_price_choice(callback.message, state, callback.from_user.id, target_date)
        return
    if choice == "yesterday":
        target_date = (today - timedelta(days=1)).isoformat()
        await callback.answer()
        await _show_scan_price_choice(callback.message, state, callback.from_user.id, target_date)
        return
    if choice == "custom":
        # v4.3.38: manual date entry is no longer part of the product flow.
        # Keep this guard for stale Telegram callback buttons from older messages.
        await state.set_state(None)
        await callback.answer("Выбери одну из 5 доступных дат", show_alert=True)
        await callback.message.answer(
            "<b>📅 Выбор даты</b>\n\n"
            "Доступны только 5 дат: сегодня и 4 предыдущих дня.",
            parse_mode=ParseMode.HTML,
            reply_markup=scan_date_keyboard(),
        )
        return
    try:
        parsed = datetime.strptime(choice, "%Y-%m-%d").date()
    except ValueError:
        await callback.answer("Неизвестная дата", show_alert=True)
        return
    if parsed > today or parsed < today - timedelta(days=DATE_MAX_AGE_DAYS):
        await callback.answer(f"Можно выбрать только последние {DATE_MAX_AGE_DAYS + 1} дней", show_alert=True)
        return
    target_date = parsed.isoformat()

    await callback.answer()
    await _show_scan_price_choice(callback.message, state, callback.from_user.id, target_date)


@dp.message(ScanInput.target_date)
async def receive_scan_date(message: Message, state: FSMContext) -> None:
    # v4.3.38: typed dates are deliberately disabled. This handler only catches
    # stale FSM state left by a pre-upgrade message and returns the bounded picker.
    if not await trial_or_paid_scan_access(message.from_user.id):
        await state.clear()
        await message.answer("Нет доступа.")
        return
    await state.set_state(None)
    await message.answer(
        "<b>📅 Выбери дату кнопкой</b>\n\n"
        "Доступны только 5 дат: сегодня и 4 предыдущих дня.",
        parse_mode=ParseMode.HTML,
        reply_markup=scan_date_keyboard(),
    )


@dp.callback_query(F.data == "scanprice_menu")
async def reopen_scan_price(callback: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    target_date = data.get("target_date") or _moscow_today_iso()
    await callback.answer()
    await _show_scan_price_choice(callback.message, state, callback.from_user.id, target_date)


@dp.callback_query(F.data.startswith("scanprice:"))
async def choose_scan_price(callback: CallbackQuery, state: FSMContext) -> None:
    if not await trial_or_paid_scan_access(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    choice = callback.data.split(":", 1)[1]
    data = await state.get_data()
    target_date = data.get("target_date") or _moscow_today_iso()
    if choice == "custom":
        await state.set_state(ScanInput.custom_price)
        await callback.answer()
        await callback.message.answer(
            "<b>💶 Свой диапазон цены</b>\n\n"
            "Отправь, например:\n"
            "<code>200-300</code> — от 200 до 300 €\n"
            "<code>500+</code> — от 500 €\n"
            "<code>-200</code> — до 200 €",
            parse_mode=ParseMode.HTML,
        )
        return
    if choice not in {"any", "50_plus", "100_plus", "200_plus", "500_plus"}:
        await callback.answer("Неизвестный диапазон", show_alert=True)
        return
    await callback.answer(f"Цена: {price_filter_label(choice)}")
    await _show_scan_depth_choice(callback.message, state, callback.from_user.id, target_date, choice)


@dp.message(ScanInput.custom_price)
async def receive_scan_price(message: Message, state: FSMContext) -> None:
    if not await trial_or_paid_scan_access(message.from_user.id):
        await state.clear()
        await message.answer("Нет доступа.")
        return
    price_filter = parse_scan_price_input(message.text)
    if price_filter is None:
        await message.answer(
            "⚠️ Не понял диапазон. Используй <code>200-300</code>, <code>500+</code> или <code>-200</code>.",
            parse_mode=ParseMode.HTML,
        )
        return
    data = await state.get_data()
    target_date = data.get("target_date") or _moscow_today_iso()
    await _show_scan_depth_choice(message, state, message.from_user.id, target_date, price_filter)


@dp.callback_query(F.data.startswith("scanpages:"))
async def start_scan_with_pages(callback: CallbackQuery, state: FSMContext) -> None:
    if not await trial_or_paid_scan_access(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        page_limit = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("Некорректный лимит", show_alert=True)
        return
    if page_limit not in PAGE_LIMIT_CHOICES:
        await callback.answer("Выбери 15, 25 или 50 страниц", show_alert=True)
        return

    is_trial = not allowed(callback.from_user.id)
    trial = await get_trial_status(callback.from_user.id) if is_trial else TrialStatus(False, False, 0, 0)
    if is_trial and not trial.eligible:
        await callback.answer("Бесплатные сканы закончились", show_alert=True)
        if callback.message:
            await callback.message.answer(
                await subscription_text(callback.from_user.id),
                parse_mode=ParseMode.HTML,
                reply_markup=await subscription_keyboard(callback.from_user.id),
            )
        return
    if is_trial and page_limit > FREE_TRIAL_MAX_PAGES:
        await callback.answer(f"В пробном режиме максимум {FREE_TRIAL_MAX_PAGES} страниц", show_alert=True)
        return

    data = await state.get_data()
    target_date = data.get("target_date") or _moscow_today_iso()
    price_filter = (data.get("price_filter") or "any").strip()
    selected_keys = await get_selected(callback.from_user.id)
    selected_cats = [CATEGORIES[k] for k in CATEGORIES if k in selected_keys]
    if not selected_cats:
        await callback.answer("Сначала выбери хотя бы одну категорию", show_alert=True)
        return
    category_limit = FREE_TRIAL_MAX_CATEGORIES if is_trial else MAX_SELECTED_CATEGORIES
    if len(selected_cats) > category_limit:
        await callback.answer(
            "🎁 В бесплатном скане можно выбрать только 1 категорию"
            if is_trial else
            f"Сейчас выбрано {len(selected_cats)} категорий. Максимум {MAX_SELECTED_CATEGORIES} за один запуск — убери лишние или очисти выбор.",
            show_alert=True,
        )
        return

    if await user_has_active_scan(callback.from_user.id):
        await callback.answer("У тебя уже идёт парсинг", show_alert=True)
        return
    if await queue_is_full():
        await callback.answer("Сервис сейчас сильно загружен. Попробуй чуть позже.", show_alert=True)
        return

    await update_setting(callback.from_user.id, "page_limit", page_limit)
    await state.clear()
    await callback.answer("🎁 Бесплатный скан запущен" if is_trial else "Скан запущен")
    await enqueue_user_scan(
        callback.message, callback.from_user.id, [cat.key for cat in selected_cats], page_limit, target_date,
        price_filter=price_filter, is_trial=is_trial,
    )


async def _start_embedded_fleet_fallback(bot: Bot) -> tuple[list[asyncio.Task], object | None]:
    """Start one browser-backed Redis consumer inside the Telegram service.

    v4.1.7 deliberately does *not* suppress this reserve lane when another
    parser heartbeat is visible. Railway redeploys leave the old process heartbeat
    in Redis until its TTL expires; v4.1.5/v4.1.6 could mistake that stale key for
    a healthy external fleet, skip bootstrap, and end up with zero workers seconds
    later. One embedded reserve is cheap, safe with dedicated replicas, and makes
    the main parser service self-starting.
    """
    if not (DISTRIBUTED_WORKERS and EMBEDDED_FLEET_FALLBACK):
        return [], None

    # parser.py was imported before Railway role detection. Switch its process-local
    # runtime explicitly for this single reserve lane; KleinanzeigenParser reads
    # these module globals when each job instance is created.
    import parser as parser_module
    parser_module.SCAN_TRANSPORT = "browser"
    parser_module.SHARED_BROWSER_RUNTIME = True
    os.environ["SCAN_TRANSPORT"] = "browser"
    os.environ["SHARED_BROWSER_RUNTIME"] = "1"

    worker_id = f"parser-embedded-{os.getenv('RAILWAY_SERVICE_ID', 'service')}-{os.getpid()}"

    # Register synchronously before Telegram polling starts. This closes the small
    # race where the user could press 15/25/50 before the heartbeat task got its
    # first event-loop turn.
    await COORDINATOR.heartbeat(worker_id, "parser")

    tasks = [
        asyncio.create_task(
            distributed_worker_heartbeat(worker_id, "parser"),
            name="embedded-fleet-heartbeat",
        ),
        asyncio.create_task(
            distributed_scan_worker(bot, f"{worker_id}-1"),
            name="embedded-fleet-worker",
        ),
        asyncio.create_task(progress_ticker(bot), name="embedded-fleet-progress-ticker"),
    ]
    log.warning(
        "Embedded Browser Fleet reserve online | id=%s | lanes=1 | transport=browser | "
        "dedicated fleet-worker replicas may run alongside it",
        worker_id,
    )
    return tasks, parser_module


async def main() -> None:
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN is not set")
    # v4.9.1 fail-fast guard. If a future change breaks the four-lane contract,
    # Railway should restart visibly instead of silently serving only 1–2 users.
    if (
        not STABLE_SINGLE_SERVICE_MODE
        or DISTRIBUTED_WORKERS
        or not MULTIUSER_STABLE_MODE
        or MAX_CONCURRENT_JOBS != GUARANTEED_LOCAL_PARSER_LANES
    ):
        raise RuntimeError(
            "v4.9.1 four-lane guarantee violated: "
            f"stable={STABLE_SINGLE_SERVICE_MODE} distributed={DISTRIBUTED_WORKERS} "
            f"multiuser={MULTIUSER_STABLE_MODE} lanes={MAX_CONCURRENT_JOBS}"
        )
    if DISTRIBUTED_CONFIG_ERROR:
        raise RuntimeError(
            "Railway production requires REDIS_URL. Add a Railway Redis service and "
            "reference its REDIS_URL in the bot service. Local scan-worker fallback is "
            "disabled on Railway so production cannot silently run in mode=local."
        )
    await init_db()
    # v4.15.6 correctness-first: hide the pre-v4.15.6 Radar immediately, before
    # Telegram polling starts. The sweep only marks historical links checked; a
    # fresh v4.15.6 demand-safe signal is required to certify/rebuild a family.
    await prepare_bump_resurrection_sweep_once()
    organic_cleanup = await purge_nonorganic_analytics(
        infer_historical_price_drops=True
    )
    velocity_repair = await prepare_verified_organic_velocity_once()
    if any(int(v or 0) for v in velocity_repair.values()):
        log.warning("v4.15.7 Verified Organic Velocity startup repair: %s", velocity_repair)
    unified_repair = await prepare_unified_48h_ranking_once()
    if any(int(v or 0) for v in unified_repair.values()):
        log.warning("v4.20.0 Unified 48H Radar startup repair: %s", unified_repair)
    # Radar startup preservation guard: never destructively clear evidence on deploy.
    await prepare_radar_v3_once()
    history_score_repair = await repair_radar_v3_historical_scores_once()
    if history_score_repair:
        log.warning("v4.21.14 Radar history score repair: restored %s stale scores", history_score_repair)
    if organic_cleanup.get("dirty_listings", 0):
        log.warning(
            "v4.15.3 Strict Organic Radar Gate cleanup: %s",
            organic_cleanup,
        )
    invalidated_views = await invalidate_untrusted_view_analytics_once()
    if invalidated_views:
        log.warning(
            "v4.2.2 Accurate Views reset: invalidated %s legacy listing counters and old growth history",
            invalidated_views,
        )
    stale_distributed = await cleanup_stale_distributed_queue_rows()
    if stale_distributed:
        log.info("v4.1.7 stale distributed queue cleanup: %s", stale_distributed)
    await initialize_commerce()
    backfilled = await backfill_product_identities()
    if backfilled:
        log.info("v3.0 product identity backfill: %s listings", backfilled)
    if DISTRIBUTED_WORKERS:
        # The dedicated views-worker owns observation recovery/scheduling. Running
        # those routines in the Telegram service too can reset a live checkpoint.
        obsolete_observations = recovered_observations = planned = disabled_observations = 0
    else:
        obsolete_observations = await cleanup_obsolete_observation_plans()
        disabled_observations = await cleanup_disabled_observation_plans()
        recovered_observations = await recover_running_observations()
        # After the one-time Accurate Views reset, old scans no longer have a
        # trustworthy baseline. Do not schedule new growth points for them; only
        # scans completed under v4.2.2 create fresh 0/+3/+6/+12 plans.
        planned = 0 if invalidated_views else await backfill_recent_observation_plans()
    archived = await archive_expired_scans()
    if recovered_observations or planned or obsolete_observations or disabled_observations:
        log.info(
            "v4.3.19 observations: removed_old=%s disabled_removed=%s recovered=%s recent_scans_planned=%s",
            obsolete_observations, disabled_observations, recovered_observations, planned,
        )
    if archived:
        log.info("v3.3.0 initial scan archive: %s moved", archived)

    bot = LocalizedBot(BOT_TOKEN)
    try:
        await setup_bot_commands(bot)
    except Exception:
        # A Telegram menu configuration error must never keep the parser offline.
        log.exception("Could not configure Telegram command menu")

    if DISTRIBUTED_WORKERS:
        recovered_scans = 0
        try:
            await COORDINATOR.connect()
            await COORDINATOR.ensure_group()
        except Exception as exc:
            # Production must fail closed. An apparently healthy Telegram bot that has
            # no queue is worse than a Railway restart because users would launch jobs
            # that can never reach the browser fleet.
            await bot.session.close()
            raise RuntimeError("Distributed mode requires a healthy Redis connection") from exc
    else:
        recovered_scans = await recover_interrupted_user_scans(bot)
        if recovered_scans:
            log.warning("v4.2.2 closed %s interrupted user scan(s); no automatic resurrection", recovered_scans)

    embedded_fleet_tasks: list[asyncio.Task] = []
    embedded_parser_module = None
    if DISTRIBUTED_WORKERS:
        embedded_fleet_tasks, embedded_parser_module = await _start_embedded_fleet_fallback(bot)

    me = await bot.get_me()
    traffic = await TRAFFIC.snapshot()
    log.info(
        "Starting @%s | version=%s | mode=%s source=%s railway=%s redis=%s | local_workers=%s embedded_fleet=%s cache_ttl=%ss | traffic scan=%s view=%s browser=%s global=%s",
        me.username, APP_VERSION, "distributed" if DISTRIBUTED_WORKERS else "local",
        DISTRIBUTED_MODE_SOURCE, IS_RAILWAY, bool(REDIS_URL),
        0 if DISTRIBUTED_WORKERS else MAX_CONCURRENT_JOBS, bool(embedded_fleet_tasks), CATEGORY_CACHE_TTL_SECONDS,
        traffic.scan_limit, traffic.view_limit, traffic.browser_limit, traffic.global_limit,
    )
    log.info("Database backend: %s", DATABASE_BACKEND)
    if STABLE_SINGLE_SERVICE_MODE:
        log.warning(
            "v4.3.2 Multi-User Stable active | parser_lanes=%s | shared_chromium=%s | "
            "isolated_context_per_job=%s | view_pool=%s | view_interval=%.2fs | transport=browser | "
            "redis_in_scan_path=False | accurate_views=%s | category_watchdog=%ss | page_retries=%s",
            MAX_CONCURRENT_JOBS, bool(MULTIUSER_STABLE_MODE), bool(MULTIUSER_STABLE_MODE),
            MULTIUSER_VIEW_POOL_SIZE, TRAFFIC.view_min_interval, ACCURATE_VIEWS_MODE,
            int(SCAN_CATEGORY_HARD_TIMEOUT_SECONDS), STABLE_PAGE_RETRIES,
        )

    worker_tasks = [] if DISTRIBUTED_WORKERS else [
        asyncio.create_task(scan_worker(bot, i), name=f"scan-worker-{i}")
        for i in range(1, MAX_CONCURRENT_JOBS + 1)
    ]
    if len(worker_tasks) != GUARANTEED_LOCAL_PARSER_LANES:
        raise RuntimeError(
            f"v4.9.1 expected {GUARANTEED_LOCAL_PARSER_LANES} scan workers, got {len(worker_tasks)}"
        )
    log.warning(
        "v4.14.0 Fast Sold Lifecycle + Referral Promo + Daily Radar FSM Hotfix online | parser_lanes=%s | fifth_plus=FIFO | "
        "trial_and_paid_same_queue=True | railway_lane_overrides_ignored=True",
        GUARANTEED_LOCAL_PARSER_LANES,
    )
    log.warning(
        "DT Radar 3.0 Live Today online | category_watchdog=%ss | view_recovery_watchdog=%ss | background_pause=round | radar_checkpoint=throttled-during-autoscan | detail_lanes=foreground+background",
        int(RADAR_AUTOSCAN_CATEGORY_TIMEOUT_SECONDS), int(RADAR_AUTOSCAN_VIEW_RECOVERY_TIMEOUT_SECONDS),
    )
    ticker_task = None if DISTRIBUTED_WORKERS else asyncio.create_task(
        progress_ticker(bot), name="user-progress-ticker"
    )
    distributed_queue_ticker_task = asyncio.create_task(
        distributed_queue_ui_ticker(bot), name="distributed-queue-ui-ticker"
    ) if DISTRIBUTED_WORKERS else None
    payment_task = asyncio.create_task(payment_scheduler(bot), name="payment-scheduler")
    subscription_task = asyncio.create_task(
        subscription_lifecycle_scheduler(bot), name="subscription-lifecycle-scheduler"
    )
    archive_task = asyncio.create_task(scan_archive_scheduler(), name="scan-archive-scheduler")
    radar_task = asyncio.create_task(radar_maintenance_scheduler(), name="dt-radar-maintenance")
    radar_v3_observation_task = asyncio.create_task(radar_v3_observation_scheduler(), name="dt-radar-v3-observations")
    organic_velocity_task = asyncio.create_task(organic_velocity_scheduler(), name="verified-organic-velocity")
    radar_autoscan_task = asyncio.create_task(radar_autoscan_scheduler(bot), name="dt-radar-autoscan")
    radar_daily_digest_task = asyncio.create_task(radar_daily_digest_scheduler(bot), name="dt-radar-daily-digest")
    vinted_radar_task = asyncio.create_task(vinted_radar_autoscan_scheduler(), name="vinted-radar-1-autoscan")
    observation_tasks = [] if DISTRIBUTED_WORKERS else [
        asyncio.create_task(observation_scheduler(bot, i), name=f"view-observation-worker-{i}")
        for i in range(1, OBSERVATION_CONCURRENCY + 1)
    ]
    try:
        await dp.start_polling(bot)
    finally:
        if ticker_task is not None:
            ticker_task.cancel()
        if distributed_queue_ticker_task is not None:
            distributed_queue_ticker_task.cancel()
        payment_task.cancel()
        subscription_task.cancel()
        archive_task.cancel()
        radar_task.cancel()
        radar_v3_observation_task.cancel()
        organic_velocity_task.cancel()
        radar_autoscan_task.cancel()
        radar_daily_digest_task.cancel()
        vinted_radar_task.cancel()
        for task in observation_tasks:
            task.cancel()
        for task in worker_tasks:
            task.cancel()
        shutdown_tasks = [payment_task, subscription_task, archive_task, radar_task, radar_v3_observation_task, organic_velocity_task, radar_autoscan_task, radar_daily_digest_task, vinted_radar_task, *observation_tasks, *worker_tasks]
        if distributed_queue_ticker_task is not None:
            shutdown_tasks.insert(0, distributed_queue_ticker_task)
        if ticker_task is not None:
            shutdown_tasks.insert(0, ticker_task)
        await asyncio.gather(*shutdown_tasks, return_exceptions=True)
        async with category_inflight_guard:
            inflight = list(category_inflight.values())
        for task in inflight:
            task.cancel()
        if inflight:
            await asyncio.gather(*inflight, return_exceptions=True)
        for task in embedded_fleet_tasks:
            task.cancel()
        if embedded_fleet_tasks:
            await asyncio.gather(*embedded_fleet_tasks, return_exceptions=True)
        if embedded_parser_module is not None:
            try:
                await embedded_parser_module.shutdown_shared_browser_runtime()
            except Exception:
                log.debug("Embedded browser runtime shutdown failed", exc_info=True)
        if DISTRIBUTED_WORKERS:
            await COORDINATOR.close()
        if STABLE_SINGLE_SERVICE_MODE and MULTIUSER_STABLE_MODE:
            try:
                await _stable_parser_module.shutdown_shared_browser_runtime()
            except Exception:
                log.debug("Shared local Chromium shutdown failed", exc_info=True)
        try:
            await REMOTE_PAGE_MANAGER.close()
        except Exception:
            log.debug("Page Manager Redis shutdown failed", exc_info=True)
        await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())
